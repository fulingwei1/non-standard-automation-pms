# -*- coding: utf-8 -*-
"""
预警与异常管理 API endpoints
包含：预警规则、预警记录、异常事件、项目健康度
"""

from typing import Any, List, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, status, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import or_, and_, func, case

from app.api import deps
from app.core import security
from app.core.config import settings
from app.models.user import User
from app.models.project import Project, Machine
from app.models.issue import Issue
from app.models.alert import (
    AlertRule, AlertRuleTemplate, AlertRecord, AlertNotification,
    ExceptionEvent, ExceptionAction, ExceptionEscalation,
    AlertStatistics, ProjectHealthSnapshot, AlertSubscription
)
from app.schemas.alert import (
    AlertRuleCreate, AlertRuleUpdate, AlertRuleResponse,
    AlertRecordHandle, AlertRecordResponse, AlertRecordListResponse,
    ExceptionEventCreate, ExceptionEventUpdate, ExceptionEventResolve,
    ExceptionEventVerify, ExceptionEventResponse, ExceptionEventListResponse,
    ProjectHealthResponse, AlertStatisticsResponse,
    AlertSubscriptionCreate, AlertSubscriptionUpdate, AlertSubscriptionResponse
)
from app.schemas.common import ResponseModel, PaginatedResponse

router = APIRouter()


# ==================== 预警规则管理 ====================

@router.get("/alert-rules", response_model=PaginatedResponse, status_code=status.HTTP_200_OK)
def read_alert_rules(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索（规则编码/名称）"),
    rule_type: Optional[str] = Query(None, description="规则类型筛选"),
    target_type: Optional[str] = Query(None, description="监控对象类型筛选"),
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警规则列表（支持分页和筛选）
    """
    query = db.query(AlertRule)
    
    # 关键词搜索
    if keyword:
        query = query.filter(
            or_(
                AlertRule.rule_code.like(f"%{keyword}%"),
                AlertRule.rule_name.like(f"%{keyword}%"),
            )
        )
    
    # 规则类型筛选
    if rule_type:
        query = query.filter(AlertRule.rule_type == rule_type)
    
    # 监控对象类型筛选
    if target_type:
        query = query.filter(AlertRule.target_type == target_type)
    
    # 启用状态筛选
    if is_enabled is not None:
        query = query.filter(AlertRule.is_enabled == is_enabled)
    
    # 计算总数
    total = query.count()
    
    # 分页
    offset = (page - 1) * page_size
    rules = query.order_by(AlertRule.created_at.desc()).offset(offset).limit(page_size).all()
    
    return PaginatedResponse(
        items=rules,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/alert-rules/{rule_id}", response_model=AlertRuleResponse, status_code=status.HTTP_200_OK)
def read_alert_rule(
    rule_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警规则详情
    """
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="预警规则不存在")
    
    return rule


@router.post("/alert-rules", response_model=AlertRuleResponse, status_code=status.HTTP_201_CREATED)
def create_alert_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_in: AlertRuleCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建预警规则
    """
    # 检查规则编码是否已存在
    existing = db.query(AlertRule).filter(AlertRule.rule_code == rule_in.rule_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="规则编码已存在")
    
    # 验证规则编码格式（字母、数字、下划线）
    import re
    if not re.match(r'^[A-Za-z0-9_]+$', rule_in.rule_code):
        raise HTTPException(status_code=400, detail="规则编码只能包含字母、数字和下划线")
    
    # 验证阈值格式（如果是数值类型）
    if rule_in.threshold_value:
        try:
            float(rule_in.threshold_value)
        except ValueError:
            # 如果不是纯数字，可能是表达式，允许通过
            pass
    
    # 验证阈值范围（如果有 min 和 max）
    if rule_in.threshold_min and rule_in.threshold_max:
        try:
            min_val = float(rule_in.threshold_min)
            max_val = float(rule_in.threshold_max)
            if min_val >= max_val:
                raise HTTPException(status_code=400, detail="阈值下限必须小于阈值上限")
        except ValueError:
            pass
    
    # 验证通知渠道
    valid_channels = ['SYSTEM', 'EMAIL', 'WECHAT', 'SMS']
    if rule_in.notify_channels:
        for channel in rule_in.notify_channels:
            if channel.upper() not in valid_channels:
                raise HTTPException(
                    status_code=400,
                    detail=f"无效的通知渠道: {channel}。支持的渠道: {', '.join(valid_channels)}"
                )
    
    # 验证检查频率
    valid_frequencies = ['REALTIME', 'HOURLY', 'DAILY', 'WEEKLY']
    if rule_in.check_frequency and rule_in.check_frequency.upper() not in valid_frequencies:
        raise HTTPException(
            status_code=400,
            detail=f"无效的检查频率: {rule_in.check_frequency}。支持的频率: {', '.join(valid_frequencies)}"
        )
    
    # 验证预警级别
    valid_levels = ['INFO', 'WARNING', 'CRITICAL', 'URGENT']
    if rule_in.alert_level and rule_in.alert_level.upper() not in valid_levels:
        raise HTTPException(
            status_code=400,
            detail=f"无效的预警级别: {rule_in.alert_level}。支持的级别: {', '.join(valid_levels)}"
        )
    
    rule = AlertRule(**rule_in.model_dump(), created_by=current_user.id)
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    return rule


@router.put("/alert-rules/{rule_id}", response_model=AlertRuleResponse, status_code=status.HTTP_200_OK)
def update_alert_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    rule_in: AlertRuleUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新预警规则
    """
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="预警规则不存在")
    
    # 系统预置规则不允许修改某些字段
    if rule.is_system:
        raise HTTPException(status_code=400, detail="系统预置规则不允许修改")
    
    update_data = rule_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule, field, value)
    
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    return rule


@router.put("/alert-rules/{rule_id}/toggle", response_model=AlertRuleResponse, status_code=status.HTTP_200_OK)
def toggle_alert_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    启用/禁用预警规则
    """
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="预警规则不存在")
    
    # 系统预置规则可以启用/禁用，但不能删除
    rule.is_enabled = not rule.is_enabled
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    return rule


@router.delete("/alert-rules/{rule_id}", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def delete_alert_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除预警规则
    """
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="预警规则不存在")
    
    # 系统预置规则不允许删除
    if rule.is_system:
        raise HTTPException(status_code=400, detail="系统预置规则不允许删除")
    
    # 检查是否有预警记录使用此规则
    alert_count = db.query(AlertRecord).filter(AlertRecord.rule_id == rule_id).count()
    if alert_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"该规则已被 {alert_count} 条预警记录使用，无法删除。请先处理相关预警记录。"
        )
    
    db.delete(rule)
    db.commit()
    
    return ResponseModel(code=200, message="预警规则已删除")


@router.get("/alert-rule-templates", response_model=List[dict], status_code=status.HTTP_200_OK)
def read_alert_rule_templates(
    category: Optional[str] = Query(None, description="模板分类筛选"),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警规则模板列表
    """
    query = db.query(AlertRuleTemplate).filter(AlertRuleTemplate.is_active == True)
    
    if category:
        query = query.filter(AlertRuleTemplate.template_category == category)
    
    templates = query.order_by(AlertRuleTemplate.template_code).all()
    
    return [
        {
            "id": t.id,
            "template_code": t.template_code,
            "template_name": t.template_name,
            "template_category": t.template_category,
            "description": t.description,
            "usage_guide": t.usage_guide,
            "rule_config": t.rule_config
        }
        for t in templates
    ]


# ==================== 预警记录管理 ====================

@router.get("/alerts", response_model=PaginatedResponse, status_code=status.HTTP_200_OK)
def read_alert_records(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    machine_id: Optional[int] = Query(None, description="机台ID筛选"),
    alert_level: Optional[str] = Query(None, description="预警级别筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    target_type: Optional[str] = Query(None, description="对象类型筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    date_from: Optional[date] = Query(None, description="开始日期（别名）"),
    date_to: Optional[date] = Query(None, description="结束日期（别名）"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警记录列表
    """
    query = db.query(AlertRecord)
    
    # 项目筛选
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    
    # 机台筛选
    if machine_id:
        query = query.filter(AlertRecord.machine_id == machine_id)
    
    # 预警级别筛选
    if alert_level:
        query = query.filter(AlertRecord.alert_level == alert_level)
    
    # 状态筛选
    if status:
        query = query.filter(AlertRecord.status == status)
    
    # 对象类型筛选
    if target_type:
        query = query.filter(AlertRecord.target_type == target_type)
    
    # 日期范围筛选（支持 start_date/end_date 和 date_from/date_to）
    date_from_value = date_from or start_date
    date_to_value = date_to or end_date
    if date_from_value:
        query = query.filter(AlertRecord.triggered_at >= datetime.combine(date_from_value, datetime.min.time()))
    if date_to_value:
        query = query.filter(AlertRecord.triggered_at <= datetime.combine(date_to_value, datetime.max.time()))
    
    # 计算总数
    total = query.count()
    
    # 分页 - 使用 eager loading 避免 N+1 查询
    offset = (page - 1) * page_size
    alerts = query.options(
        joinedload(AlertRecord.rule),
        joinedload(AlertRecord.project),
        joinedload(AlertRecord.machine)
    ).order_by(AlertRecord.triggered_at.desc()).offset(offset).limit(page_size).all()
    
    # 批量获取处理人信息（避免循环查询）
    handler_ids = [alert.handler_id for alert in alerts if alert.handler_id]
    handlers_map = {}
    if handler_ids:
        handlers = db.query(User).filter(User.id.in_(handler_ids)).all()
        handlers_map = {h.id: h for h in handlers}
    
    # 补充关联信息
    items = []
    for alert in alerts:
        rule_name = alert.rule.rule_name if alert.rule else None
        project_name = alert.project.project_name if alert.project else None
        
        handler_name = None
        if alert.handler_id and alert.handler_id in handlers_map:
            handler = handlers_map[alert.handler_id]
            handler_name = handler.real_name or handler.username
        
        items.append({
            "id": alert.id,
            "alert_no": alert.alert_no,
            "alert_level": alert.alert_level,
            "alert_title": alert.alert_title,
            "target_type": alert.target_type,
            "target_name": alert.target_name,
            "project_name": project_name,
            "triggered_at": alert.triggered_at,
            "status": alert.status,
            "handler_name": handler_name
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/alerts/{alert_id}", response_model=AlertRecordResponse, status_code=status.HTTP_200_OK)
def read_alert_record(
    alert_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警详情
    """
    alert = db.query(AlertRecord).options(
        joinedload(AlertRecord.rule),
        joinedload(AlertRecord.project),
        joinedload(AlertRecord.machine)
    ).filter(AlertRecord.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="预警记录不存在")
    
    # 补充关联信息（已通过 eager loading 加载）
    rule_name = alert.rule.rule_name if alert.rule else None
    project_name = alert.project.project_name if alert.project else None
    
    handler_name = None
    if alert.handler_id:
        handler = db.query(User).filter(User.id == alert.handler_id).first()
        handler_name = handler.real_name or handler.username if handler else None
    
    return {
        "id": alert.id,
        "alert_no": alert.alert_no,
        "rule_id": alert.rule_id,
        "rule_name": rule_name,
        "target_type": alert.target_type,
        "target_id": alert.target_id,
        "target_no": alert.target_no,
        "target_name": alert.target_name,
        "project_id": alert.project_id,
        "project_name": project_name,
        "alert_level": alert.alert_level,
        "alert_title": alert.alert_title,
        "alert_content": alert.alert_content,
        "triggered_at": alert.triggered_at,
        "trigger_value": alert.trigger_value,
        "threshold_value": alert.threshold_value,
        "status": alert.status,
        "acknowledged_by": alert.acknowledged_by,
        "acknowledged_at": alert.acknowledged_at,
        "handler_id": alert.handler_id,
        "handler_name": handler_name,
        "handle_start_at": alert.handle_start_at,
        "handle_end_at": alert.handle_end_at,
        "handle_result": alert.handle_result,
        "created_at": alert.created_at,
        "updated_at": alert.updated_at
    }


@router.put("/alerts/{alert_id}/acknowledge", response_model=AlertRecordResponse, status_code=status.HTTP_200_OK)
def acknowledge_alert(
    *,
    db: Session = Depends(deps.get_db),
    alert_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    确认预警
    """
    alert = db.query(AlertRecord).filter(AlertRecord.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="预警记录不存在")
    
    if alert.status != "PENDING":
        raise HTTPException(status_code=400, detail="只能确认待处理状态的预警")
    
    alert.status = "ACKNOWLEDGED"
    alert.acknowledged_by = current_user.id
    alert.acknowledged_at = datetime.now()
    
    db.add(alert)
    db.commit()
    db.refresh(alert)
    
    # 返回详情（需要重新查询以获取关联信息）
    return read_alert_record(alert_id, db, current_user)


@router.put("/alerts/{alert_id}/resolve", response_model=AlertRecordResponse, status_code=status.HTTP_200_OK)
def resolve_alert(
    *,
    db: Session = Depends(deps.get_db),
    alert_id: int,
    handle_in: AlertRecordHandle,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    处理预警
    """
    alert = db.query(AlertRecord).filter(AlertRecord.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="预警记录不存在")
    
    if alert.status == "RESOLVED":
        raise HTTPException(status_code=400, detail="预警已处理")
    
    # 如果还未开始处理，设置开始时间
    if not alert.handle_start_at:
        alert.handle_start_at = datetime.now()
        alert.handler_id = current_user.id
    
    alert.status = "RESOLVED"
    alert.handle_end_at = datetime.now()
    alert.handle_result = handle_in.handle_result
    alert.handle_note = handle_in.handle_note
    
    db.add(alert)
    db.commit()
    db.refresh(alert)
    
    return read_alert_record(alert_id, db, current_user)


@router.put("/alerts/{alert_id}/close", response_model=AlertRecordResponse, status_code=status.HTTP_200_OK)
def close_alert(
    *,
    db: Session = Depends(deps.get_db),
    alert_id: int,
    handle_in: AlertRecordHandle,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    关闭预警
    """
    alert = db.query(AlertRecord).filter(AlertRecord.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="预警记录不存在")
    
    if alert.status == "CLOSED":
        raise HTTPException(status_code=400, detail="预警已关闭")
    
    alert.status = "CLOSED"
    alert.handler_id = current_user.id
    alert.handle_end_at = datetime.now()
    if handle_in.handle_result:
        alert.handle_result = handle_in.handle_result
    
    db.add(alert)
    db.commit()
    db.refresh(alert)
    
    return read_alert_record(alert_id, db, current_user)


@router.put("/alerts/{alert_id}/ignore", response_model=AlertRecordResponse, status_code=status.HTTP_200_OK)
def ignore_alert(
    *,
    db: Session = Depends(deps.get_db),
    alert_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    忽略预警
    """
    alert = db.query(AlertRecord).filter(AlertRecord.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="预警记录不存在")
    
    if alert.status == "RESOLVED":
        raise HTTPException(status_code=400, detail="已处理的预警不能忽略")
    
    alert.status = "IGNORED"
    alert.handler_id = current_user.id
    alert.handle_end_at = datetime.now()
    alert.handle_result = "已忽略"
    
    db.add(alert)
    db.commit()
    db.refresh(alert)
    
    return read_alert_record(alert_id, db, current_user)


@router.get("/alert-notifications", response_model=PaginatedResponse, status_code=status.HTTP_200_OK)
def read_alert_notifications(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    is_read: Optional[bool] = Query(None, description="是否已读"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警通知列表（当前用户的）
    """
    from app.services.notification_service import AlertNotificationService
    
    service = AlertNotificationService(db)
    offset = (page - 1) * page_size
    result = service.get_user_notifications(
        user_id=current_user.id,
        is_read=is_read,
        limit=page_size,
        offset=offset
    )
    
    if not result.get('success'):
        raise HTTPException(status_code=500, detail=result.get('message', '获取通知列表失败'))
    
    return PaginatedResponse(
        items=result.get('items', []),
        total=result.get('total', 0),
        page=page,
        page_size=page_size,
        pages=(result.get('total', 0) + page_size - 1) // page_size
    )


@router.put("/alert-notifications/{notification_id}/read", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def mark_notification_read(
    notification_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    标记通知已读
    """
    from app.services.notification_service import AlertNotificationService
    
    service = AlertNotificationService(db)
    success = service.mark_notification_read(notification_id, current_user.id)
    
    if not success:
        raise HTTPException(status_code=404, detail="通知不存在或无权操作")
    
    return ResponseModel(code=200, message="已标记为已读")


@router.get("/alert-notifications/unread-count", response_model=dict, status_code=status.HTTP_200_OK)
def get_unread_notification_count(
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取未读通知数量
    """
    from app.services.notification_service import AlertNotificationService
    
    service = AlertNotificationService(db)
    count = service.get_unread_count(current_user.id)
    
    return {
        "unread_count": count,
        "user_id": current_user.id
    }


@router.post("/alert-notifications/batch-read", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def batch_mark_notifications_read(
    notification_ids: List[int] = Body(..., description="通知ID列表"),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    批量标记通知为已读
    """
    from app.services.notification_service import AlertNotificationService
    
    service = AlertNotificationService(db)
    result = service.batch_mark_read(notification_ids, current_user.id)
    
    if not result.get('success'):
        raise HTTPException(status_code=500, detail=result.get('message', '批量标记失败'))
    
    return ResponseModel(
        code=200,
        message=f"成功标记 {result.get('success_count', 0)} 条通知为已读",
        data=result
    )


# ==================== 异常事件管理 ====================

def generate_exception_no(db: Session) -> str:
    """生成异常事件编号：EXC-yymmdd-xxx"""
    today = datetime.now().strftime("%y%m%d")
    max_event = (
        db.query(ExceptionEvent)
        .filter(ExceptionEvent.event_no.like(f"EXC-{today}-%"))
        .order_by(ExceptionEvent.event_no.desc())
        .first()
    )
    
    if max_event:
        seq = int(max_event.event_no.split("-")[-1]) + 1
    else:
        seq = 1
    
    return f"EXC-{today}-{seq:03d}"


@router.get("/exceptions", response_model=PaginatedResponse, status_code=status.HTTP_200_OK)
def read_exception_events(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索（异常编号/标题）"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    event_type: Optional[str] = Query(None, description="异常类型筛选"),
    severity: Optional[str] = Query(None, description="严重程度筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    responsible_user_id: Optional[int] = Query(None, description="责任人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取异常事件列表（支持分页和筛选）
    """
    query = db.query(ExceptionEvent)
    
    # 关键词搜索
    if keyword:
        query = query.filter(
            or_(
                ExceptionEvent.event_no.like(f"%{keyword}%"),
                ExceptionEvent.event_title.like(f"%{keyword}%"),
            )
        )
    
    # 项目筛选
    if project_id:
        query = query.filter(ExceptionEvent.project_id == project_id)
    
    # 异常类型筛选
    if event_type:
        query = query.filter(ExceptionEvent.event_type == event_type)
    
    # 严重程度筛选
    if severity:
        query = query.filter(ExceptionEvent.severity == severity)
    
    # 状态筛选
    if status:
        query = query.filter(ExceptionEvent.status == status)
    
    # 责任人筛选
    if responsible_user_id:
        query = query.filter(ExceptionEvent.responsible_user_id == responsible_user_id)
    
    # 计算总数
    total = query.count()
    
    # 分页
    offset = (page - 1) * page_size
    events = query.order_by(ExceptionEvent.created_at.desc()).offset(offset).limit(page_size).all()
    
    # 构建响应数据
    items = []
    for event in events:
        discovered_by_name = None
        if event.discovered_by:
            user = db.query(User).filter(User.id == event.discovered_by).first()
            discovered_by_name = user.real_name if user else None
        
        items.append({
            "id": event.id,
            "event_no": event.event_no,
            "source_type": event.source_type,
            "project_id": event.project_id,
            "project_name": event.project.project_name if event.project else None,
            "machine_id": event.machine_id,
            "machine_name": event.machine.machine_name if event.machine else None,
            "event_type": event.event_type,
            "severity": event.severity,
            "event_title": event.event_title,
            "status": event.status,
            "discovered_at": event.discovered_at.isoformat() if event.discovered_at else None,
            "discovered_by_name": discovered_by_name,
            "schedule_impact": event.schedule_impact or 0,
            "cost_impact": float(event.cost_impact) if event.cost_impact else 0,
            "responsible_user_id": event.responsible_user_id,
            "due_date": event.due_date.isoformat() if event.due_date else None,
            "is_overdue": event.is_overdue or False,
            "created_at": event.created_at.isoformat() if event.created_at else None,
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/exceptions", response_model=ExceptionEventResponse, status_code=status.HTTP_201_CREATED)
def create_exception_event(
    *,
    db: Session = Depends(deps.get_db),
    event_in: ExceptionEventCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建异常事件
    """
    # 验证项目
    if event_in.project_id:
        project = db.query(Project).filter(Project.id == event_in.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
    
    # 验证设备
    if event_in.machine_id:
        machine = db.query(Machine).filter(Machine.id == event_in.machine_id).first()
        if not machine:
            raise HTTPException(status_code=404, detail="设备不存在")
    
    # 生成异常编号
    event_no = generate_exception_no(db)
    
    # 创建异常事件
    event = ExceptionEvent(
        event_no=event_no,
        source_type=event_in.source_type,
        source_id=event_in.source_id,
        alert_id=event_in.alert_id,
        project_id=event_in.project_id,
        machine_id=event_in.machine_id,
        event_type=event_in.event_type,
        severity=event_in.severity,
        event_title=event_in.event_title,
        event_description=event_in.event_description,
        discovered_at=datetime.now(),
        discovered_by=current_user.id,
        discovery_location=event_in.discovery_location,
        impact_scope=event_in.impact_scope,
        impact_description=event_in.impact_description,
        schedule_impact=event_in.schedule_impact,
        cost_impact=event_in.cost_impact or 0,
        responsible_dept=event_in.responsible_dept,
        responsible_user_id=event_in.responsible_user_id,
        due_date=event_in.due_date,
        status="OPEN",
        attachments=event_in.attachments,
        created_by=current_user.id,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    
    return read_exception_event(event.id, db, current_user)


@router.get("/exceptions/{event_id}", response_model=ExceptionEventResponse, status_code=status.HTTP_200_OK)
def read_exception_event(
    event_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取异常事件详情
    """
    event = db.query(ExceptionEvent).filter(ExceptionEvent.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="异常事件不存在")
    
    # 获取发现人姓名
    discovered_by_name = None
    if event.discovered_by:
        user = db.query(User).filter(User.id == event.discovered_by).first()
        discovered_by_name = user.real_name if user else None
    
    # 获取处理记录
    actions = event.actions.order_by(ExceptionAction.created_at.desc()).all()
    action_list = []
    for action in actions:
        action_user = db.query(User).filter(User.id == action.created_by).first()
        action_list.append({
            "id": action.id,
            "action_type": action.action_type,
            "action_content": action.action_content,
            "old_status": action.old_status,
            "new_status": action.new_status,
            "action_user_id": action.created_by,
            "action_user_name": action_user.real_name if action_user else None,
            "created_at": action.created_at.isoformat() if action.created_at else None,
        })
    
    return {
        "id": event.id,
        "event_no": event.event_no,
        "source_type": event.source_type,
        "alert_id": event.alert_id,
        "project_id": event.project_id,
        "project_name": event.project.project_name if event.project else None,
        "machine_id": event.machine_id,
        "machine_name": event.machine.machine_name if event.machine else None,
        "event_type": event.event_type,
        "severity": event.severity,
        "event_title": event.event_title,
        "event_description": event.event_description,
        "discovered_at": event.discovered_at,
        "discovered_by": event.discovered_by,
        "discovered_by_name": discovered_by_name,
        "discovery_location": event.discovery_location,
        "impact_scope": event.impact_scope,
        "impact_description": event.impact_description,
        "schedule_impact": event.schedule_impact or 0,
        "cost_impact": event.cost_impact or 0,
        "status": event.status,
        "responsible_dept": event.responsible_dept,
        "responsible_user_id": event.responsible_user_id,
        "due_date": event.due_date,
        "is_overdue": event.is_overdue or False,
        "root_cause": event.root_cause,
        "cause_category": event.cause_category,
        "solution": event.solution,
        "preventive_measures": event.preventive_measures,
        "resolved_at": event.resolved_at,
        "resolved_by": event.resolved_by,
        "resolution_note": event.resolution_note,
        "verified_at": event.verified_at,
        "verified_by": event.verified_by,
        "verification_result": event.verification_result,
        "attachments": event.attachments,
        "actions": action_list,
        "created_at": event.created_at,
        "updated_at": event.updated_at,
    }


@router.put("/exceptions/{event_id}/status", response_model=ExceptionEventResponse, status_code=status.HTTP_200_OK)
def update_exception_status(
    *,
    db: Session = Depends(deps.get_db),
    event_id: int,
    status: str = Query(..., description="新状态"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新异常状态
    """
    event = db.query(ExceptionEvent).filter(ExceptionEvent.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="异常事件不存在")
    
    old_status = event.status
    event.status = status
    
    # 如果状态为RESOLVED，记录解决时间
    if status == "RESOLVED" and not event.resolved_at:
        event.resolved_at = datetime.now()
        event.resolved_by = current_user.id
    
    db.add(event)
    db.commit()
    db.refresh(event)
    
    return read_exception_event(event_id, db, current_user)


@router.post("/exceptions/{event_id}/actions", response_model=ResponseModel, status_code=status.HTTP_201_CREATED)
def add_exception_action(
    *,
    db: Session = Depends(deps.get_db),
    event_id: int,
    action_type: str = Query(..., description="操作类型"),
    action_content: str = Query(..., description="操作内容"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    添加处理记录
    """
    event = db.query(ExceptionEvent).filter(ExceptionEvent.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="异常事件不存在")
    
    action = ExceptionAction(
        event_id=event_id,
        action_type=action_type,
        action_content=action_content,
        old_status=event.status,
        new_status=event.status,
        created_by=current_user.id,
    )
    db.add(action)
    db.commit()
    
    return ResponseModel(
        code=200,
        message="处理记录已添加",
        data={
            "action_id": action.id,
            "event_id": event_id,
        }
    )


@router.post("/exceptions/{event_id}/escalate", response_model=ExceptionEventResponse, status_code=status.HTTP_200_OK)
def escalate_exception(
    *,
    db: Session = Depends(deps.get_db),
    event_id: int,
    escalate_to_user_id: Optional[int] = Query(None, description="升级到用户ID"),
    escalate_to_dept: Optional[str] = Query(None, description="升级到部门"),
    escalation_reason: Optional[str] = Query(None, description="升级原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    异常升级
    """
    event = db.query(ExceptionEvent).filter(ExceptionEvent.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="异常事件不存在")
    
    # 创建升级记录
    escalation = ExceptionEscalation(
        event_id=event_id,
        escalated_from=current_user.id,
        escalated_to=escalate_to_user_id or current_user.id,  # 如果没有指定，使用当前用户
        escalated_at=datetime.now(),
        escalation_reason=escalation_reason,
        escalation_level=1,  # 升级层级，可以根据实际情况计算
    )
    db.add(escalation)
    
    # 更新责任人
    if escalate_to_user_id:
        event.responsible_user_id = escalate_to_user_id
    if escalate_to_dept:
        event.responsible_dept = escalate_to_dept
    
    # 如果严重程度不是最高，可以提升严重程度
    if event.severity == "MINOR":
        event.severity = "MAJOR"
    elif event.severity == "MAJOR":
        event.severity = "CRITICAL"
    
    db.add(event)
    db.commit()
    db.refresh(event)
    
    return read_exception_event(event_id, db, current_user)


@router.post("/exceptions/from-issue", response_model=ExceptionEventResponse, status_code=status.HTTP_201_CREATED)
def create_exception_from_issue(
    *,
    db: Session = Depends(deps.get_db),
    issue_id: int = Query(..., description="问题ID"),
    event_type: str = Query(..., description="异常类型"),
    severity: str = Query(..., description="严重程度"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    从问题创建异常事件
    """
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="问题不存在")
    
    # 生成异常编号
    event_no = generate_exception_no(db)
    
    # 创建异常事件
    event = ExceptionEvent(
        event_no=event_no,
        source_type="ISSUE",
        source_id=issue_id,
        project_id=issue.project_id,
        machine_id=issue.machine_id,
        event_type=event_type,
        severity=severity,
        event_title=f"问题转异常：{issue.issue_title}",
        event_description=issue.issue_description or "",
        discovered_at=issue.created_at or datetime.now(),
        discovered_by=issue.reporter_id or current_user.id,
        impact_scope="LOCAL",
        schedule_impact=0,
        cost_impact=0,
        status="OPEN",
        created_by=current_user.id,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    
    return read_exception_event(event.id, db, current_user)


# ==================== 项目健康度快照 ====================

@router.get("/projects/{project_id}/health-history", response_model=List[dict], status_code=status.HTTP_200_OK)
def get_project_health_history(
    project_id: int,
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    健康度趋势查询
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    query = db.query(ProjectHealthSnapshot).filter(ProjectHealthSnapshot.project_id == project_id)
    
    if start_date:
        query = query.filter(ProjectHealthSnapshot.snapshot_date >= start_date)
    if end_date:
        query = query.filter(ProjectHealthSnapshot.snapshot_date <= end_date)
    
    snapshots = query.order_by(ProjectHealthSnapshot.snapshot_date.desc()).all()
    
    result = []
    for snapshot in snapshots:
        result.append({
            "id": snapshot.id,
            "snapshot_date": snapshot.snapshot_date.isoformat() if snapshot.snapshot_date else None,
            "overall_health": snapshot.overall_health,
            "schedule_health": snapshot.schedule_health,
            "cost_health": snapshot.cost_health,
            "quality_health": snapshot.quality_health,
            "resource_health": snapshot.resource_health,
            "health_score": snapshot.health_score or 0,
            "open_alerts": snapshot.open_alerts or 0,
            "open_exceptions": snapshot.open_exceptions or 0,
            "blocking_issues": snapshot.blocking_issues or 0,
            "schedule_variance": float(snapshot.schedule_variance) if snapshot.schedule_variance else 0,
            "cost_variance": float(snapshot.cost_variance) if snapshot.cost_variance else 0,
            "budget_used_pct": float(snapshot.budget_used_pct) if snapshot.budget_used_pct else 0,
            "milestone_on_track": snapshot.milestone_on_track or 0,
            "milestone_delayed": snapshot.milestone_delayed or 0,
            "top_risks": snapshot.top_risks or [],
            "created_at": snapshot.created_at.isoformat() if snapshot.created_at else None,
        })
    
    return result


@router.get("/alerts/statistics", response_model=dict, status_code=status.HTTP_200_OK)
def get_alert_statistics(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    预警统计分析
    """
    query = db.query(AlertRecord)
    
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    
    if start_date:
        query = query.filter(AlertRecord.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(AlertRecord.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    alerts = query.all()
    
    # 按级别统计
    by_level = {}
    for alert in alerts:
        level = alert.alert_level or "UNKNOWN"
        by_level[level] = by_level.get(level, 0) + 1
    
    # 按类型统计
    by_type = {}
    for alert in alerts:
        rule_type = alert.rule_type or "UNKNOWN"
        by_type[rule_type] = by_type.get(rule_type, 0) + 1
    
    # 按状态统计
    by_status = {}
    for alert in alerts:
        status = alert.status or "UNKNOWN"
        by_status[status] = by_status.get(status, 0) + 1
    
    # 按项目统计
    by_project = {}
    for alert in alerts:
        if alert.project_id:
            project = db.query(Project).filter(Project.id == alert.project_id).first()
            project_name = project.project_name if project else f"项目{alert.project_id}"
            by_project[project_name] = by_project.get(project_name, 0) + 1
    
    # 趋势统计（按日期）
    by_date = {}
    for alert in alerts:
        if alert.created_at:
            date_key = alert.created_at.date().isoformat()
            by_date[date_key] = by_date.get(date_key, 0) + 1
    
    return {
        "total_alerts": len(alerts),
        "by_level": by_level,
        "by_type": by_type,
        "by_status": by_status,
        "by_project": by_project,
        "by_date": dict(sorted(by_date.items())),
        "summary": {
            "open_count": by_status.get("OPEN", 0),
            "acknowledged_count": by_status.get("ACKNOWLEDGED", 0),
            "resolved_count": by_status.get("RESOLVED", 0),
            "critical_count": by_level.get("CRITICAL", 0),
            "high_count": by_level.get("HIGH", 0),
        }
    }


@router.get("/alerts/statistics/trends", response_model=dict, status_code=status.HTTP_200_OK)
def get_alert_trends(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    period: str = Query("DAILY", description="统计周期: DAILY/WEEKLY/MONTHLY"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    预警趋势分析数据
    
    返回多维度趋势数据：
    - 按日期统计（日/周/月）
    - 按级别统计趋势
    - 按类型统计趋势
    - 按状态统计趋势
    """
    from datetime import timedelta
    from app.services.alert_trend_service import (
        build_trend_statistics,
        build_summary_statistics,
        generate_date_range
    )
    
    # 默认时间范围：最近30天
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    query = db.query(AlertRecord).filter(
        AlertRecord.triggered_at.isnot(None)
    )
    
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    
    query = query.filter(
        AlertRecord.triggered_at >= datetime.combine(start_date, datetime.min.time()),
        AlertRecord.triggered_at <= datetime.combine(end_date, datetime.max.time())
    )
    
    alerts = query.all()
    
    # 构建趋势统计
    trend_stats = build_trend_statistics(alerts, period)
    date_trends = trend_stats['date_trends']
    level_trends = trend_stats['level_trends']
    type_trends = trend_stats['type_trends']
    status_trends = trend_stats['status_trends']
    
    # 生成完整的时间序列
    date_range = generate_date_range(start_date, end_date, period)
    
    # 构建趋势数据数组
    trends_data = []
    for date_key in date_range:
        trends_data.append({
            "date": date_key,
            "total": date_trends.get(date_key, 0),
            "by_level": level_trends.get(date_key, {}),
            "by_type": type_trends.get(date_key, {}),
            "by_status": status_trends.get(date_key, {}),
        })
    
    # 汇总统计
    summary_stats = build_summary_statistics(alerts)
    
    return {
        "period": period,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "trends": trends_data,
        "summary": {
            "total": len(alerts),
            "by_level": summary_stats['by_level'],
            "by_type": summary_stats['by_type'],
            "by_status": summary_stats['by_status'],
        }
    }


@router.get("/alerts/statistics/dashboard", response_model=dict, status_code=status.HTTP_200_OK)
def get_alert_dashboard(
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    预警中心仪表板统计数据
    返回：活跃预警统计、今日新增/关闭数量等
    """
    today = datetime.now().date()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    
    # 活跃预警统计（按级别）
    active_query = db.query(AlertRecord).filter(
        AlertRecord.status.in_(["PENDING", "ACKNOWLEDGED"])
    )
    
    total_active = active_query.count()
    
    urgent_count = active_query.filter(AlertRecord.alert_level == "URGENT").count()
    critical_count = active_query.filter(AlertRecord.alert_level == "CRITICAL").count()
    warning_count = active_query.filter(AlertRecord.alert_level == "WARNING").count()
    info_count = active_query.filter(AlertRecord.alert_level == "INFO").count()
    
    # 今日新增预警
    today_new = db.query(AlertRecord).filter(
        AlertRecord.triggered_at >= today_start,
        AlertRecord.triggered_at <= today_end
    ).count()
    
    # 今日关闭的预警
    today_closed = db.query(AlertRecord).filter(
        AlertRecord.status == "CLOSED",
        AlertRecord.handle_end_at >= today_start,
        AlertRecord.handle_end_at <= today_end
    ).count()
    
    # 今日处理的预警（包括已解决和已关闭）
    today_processed = db.query(AlertRecord).filter(
        AlertRecord.status.in_(["RESOLVED", "CLOSED"]),
        AlertRecord.handle_end_at >= today_start,
        AlertRecord.handle_end_at <= today_end
    ).count()
    
    return {
        "active_alerts": {
            "total": total_active,
            "urgent": urgent_count,
            "critical": critical_count,
            "warning": warning_count,
            "info": info_count,
        },
        "today_new": today_new,
        "today_closed": today_closed,
        "today_processed": today_processed,
    }


@router.get("/alerts/statistics/response-metrics", response_model=dict, status_code=status.HTTP_200_OK)
def get_response_metrics(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    预警响应时效分析
    
    返回：
    - 平均响应时间（确认时间 - 触发时间）
    - 平均解决时间（处理完成时间 - 确认时间）
    - 响应时效分布（<1小时、1-4小时、4-8小时、>8小时）
    - 按级别统计响应时效
    - 按类型统计响应时效
    - 响应时效排行榜（最快/最慢的项目/责任人）
    """
    from app.services.alert_response_service import (
        calculate_response_times,
        calculate_resolve_times,
        calculate_response_distribution,
        calculate_level_metrics,
        calculate_type_metrics,
        calculate_project_metrics,
        calculate_handler_metrics,
        generate_response_rankings
    )
    
    query = db.query(AlertRecord).filter(
        AlertRecord.triggered_at.isnot(None)
    )
    
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    
    if start_date:
        query = query.filter(AlertRecord.triggered_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(AlertRecord.triggered_at <= datetime.combine(end_date, datetime.max.time()))
    
    # 查询已确认的预警（用于计算响应时间）
    acknowledged_alerts = query.filter(
        AlertRecord.acknowledged_at.isnot(None),
        AlertRecord.triggered_at.isnot(None)
    ).all()
    
    # 查询已处理的预警（用于计算解决时间）
    resolved_alerts = query.filter(
        AlertRecord.acknowledged_at.isnot(None),
        AlertRecord.handle_end_at.isnot(None)
    ).all()
    
    # 计算响应时间和解决时间
    response_times = calculate_response_times(acknowledged_alerts)
    resolve_times = calculate_resolve_times(resolved_alerts)
    
    # 平均响应时间和解决时间
    avg_response_time = sum(rt['hours'] for rt in response_times) / len(response_times) if response_times else 0
    avg_resolve_time = sum(rt['hours'] for rt in resolve_times) / len(resolve_times) if resolve_times else 0
    
    # 计算各项指标
    response_distribution = calculate_response_distribution(response_times)
    level_metrics = calculate_level_metrics(response_times)
    type_metrics = calculate_type_metrics(response_times)
    project_metrics = calculate_project_metrics(response_times, db)
    handler_metrics = calculate_handler_metrics(response_times, db)
    rankings = generate_response_rankings(project_metrics, handler_metrics)
    
    return {
        'summary': {
            'total_acknowledged': len(acknowledged_alerts),
            'total_resolved': len(resolved_alerts),
            'avg_response_time_hours': round(avg_response_time, 2),
            'avg_resolve_time_hours': round(avg_resolve_time, 2),
        },
        'response_distribution': response_distribution,
        'by_level': {
            level: {
                'avg_hours': round(metrics['avg_hours'], 2),
                'min_hours': round(metrics['min_hours'], 2),
                'max_hours': round(metrics['max_hours'], 2),
                'count': metrics['count'],
            }
            for level, metrics in level_metrics.items()
        },
        'by_type': {
            rule_type: {
                'avg_hours': round(metrics['avg_hours'], 2),
                'min_hours': round(metrics['min_hours'], 2),
                'max_hours': round(metrics['max_hours'], 2),
                'count': metrics['count'],
            }
            for rule_type, metrics in type_metrics.items()
        },
        'by_project': {
            project_name: {
                'project_id': data['project_id'],
                'avg_hours': round(data['avg_hours'], 2),
                'min_hours': round(data['min_hours'], 2),
                'max_hours': round(data['max_hours'], 2),
                'count': data['count'],
            }
            for project_name, data in project_metrics.items()
        },
        'by_handler': {
            handler_name: {
                'user_id': data['user_id'],
                'avg_hours': round(data['avg_hours'], 2),
                'min_hours': round(data['min_hours'], 2),
                'max_hours': round(data['max_hours'], 2),
                'count': data['count'],
            }
            for handler_name, data in handler_metrics.items()
        },
        'rankings': rankings,
    }


@router.get("/alerts/statistics/efficiency-metrics", response_model=dict, status_code=status.HTTP_200_OK)
def get_efficiency_metrics(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    预警处理效率分析
    
    返回：
    - 处理率（已处理数 / 总数）
    - 及时处理率（在响应时限内处理的比例）
    - 升级率（升级预警数 / 总数）
    - 重复预警率（重复预警数 / 总数）
    - 按项目、责任人、类型统计处理效率
    - 处理效率排行榜
    """
    from app.services.alert_rule_engine import AlertRuleEngine
    from app.services.alert_efficiency_service import (
        calculate_basic_metrics,
        calculate_project_metrics,
        calculate_handler_metrics,
        calculate_type_metrics,
        generate_rankings
    )
    
    query = db.query(AlertRecord).filter(
        AlertRecord.triggered_at.isnot(None)
    )
    
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    
    if start_date:
        query = query.filter(AlertRecord.triggered_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(AlertRecord.triggered_at <= datetime.combine(end_date, datetime.max.time()))
    
    all_alerts = query.all()
    total_count = len(all_alerts)
    
    if total_count == 0:
        return {
            'summary': {
                'total': 0,
                'processed_count': 0,
                'processing_rate': 0,
                'timely_processing_rate': 0,
                'escalation_rate': 0,
                'duplicate_rate': 0,
            },
            'by_project': {},
            'by_handler': {},
            'by_type': {},
            'rankings': {
                'best_projects': [],
                'worst_projects': [],
                'best_handlers': [],
                'worst_handlers': [],
            },
        }
    
    # 初始化引擎
    engine = AlertRuleEngine(db)
    
    # 计算基础指标
    basic_metrics = calculate_basic_metrics(all_alerts, engine)
    processed_count = len([a for a in all_alerts if a.status in ['RESOLVED', 'CLOSED']])
    
    # 按维度统计
    project_metrics = calculate_project_metrics(all_alerts, db, engine)
    handler_metrics = calculate_handler_metrics(all_alerts, db, engine)
    type_metrics = calculate_type_metrics(all_alerts, engine)
    
    # 生成排行榜
    rankings = generate_rankings(project_metrics, handler_metrics)
    
    # 格式化返回数据
    return {
        'summary': {
            'total': total_count,
            'processed_count': processed_count,
            'processing_rate': round(basic_metrics['processing_rate'], 4),
            'timely_processing_rate': round(basic_metrics['timely_processing_rate'], 4),
            'escalation_rate': round(basic_metrics['escalation_rate'], 4),
            'duplicate_rate': round(basic_metrics['duplicate_rate'], 4),
        },
        'by_project': {
            project_name: {
                'project_id': data['project_id'],
                'total': data['total'],
                'processing_rate': round(data['processing_rate'], 4),
                'timely_processing_rate': round(data['timely_processing_rate'], 4),
                'escalation_rate': round(data['escalation_rate'], 4),
                'efficiency_score': round(data['efficiency_score'], 2),
            }
            for project_name, data in project_metrics.items()
        },
        'by_handler': {
            handler_name: {
                'user_id': data['user_id'],
                'total': data['total'],
                'processing_rate': round(data['processing_rate'], 4),
                'timely_processing_rate': round(data['timely_processing_rate'], 4),
                'escalation_rate': round(data['escalation_rate'], 4),
                'efficiency_score': round(data['efficiency_score'], 2),
            }
            for handler_name, data in handler_metrics.items()
        },
        'by_type': {
            rule_type: {
                'total': data['total'],
                'processing_rate': round(data['processing_rate'], 4),
                'timely_processing_rate': round(data['timely_processing_rate'], 4),
                'escalation_rate': round(data['escalation_rate'], 4),
            }
            for rule_type, data in type_metrics.items()
        },
        'rankings': rankings,
    }


# ==================== 预警订阅配置 ====================

@router.get("/alerts/subscriptions", response_model=PaginatedResponse, status_code=status.HTTP_200_OK)
def read_alert_subscriptions(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    alert_type: Optional[str] = Query(None, description="预警类型筛选"),
    is_active: Optional[bool] = Query(None, description="是否启用"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取我的预警订阅配置列表
    """
    query = db.query(AlertSubscription).filter(AlertSubscription.user_id == current_user.id)
    
    # 预警类型筛选
    if alert_type:
        query = query.filter(
            or_(
                AlertSubscription.alert_type == alert_type,
                AlertSubscription.alert_type.is_(None)  # 全部类型
            )
        )
    
    # 启用状态筛选
    if is_active is not None:
        query = query.filter(AlertSubscription.is_active == is_active)
    
    # 计算总数
    total = query.count()
    
    # 分页
    offset = (page - 1) * page_size
    subscriptions = query.order_by(AlertSubscription.created_at.desc()).offset(offset).limit(page_size).all()
    
    # 转换为响应格式
    items = []
    for sub in subscriptions:
        item = {
            "id": sub.id,
            "user_id": sub.user_id,
            "alert_type": sub.alert_type,
            "project_id": sub.project_id,
            "project_name": sub.project.project_name if sub.project else None,
            "min_level": sub.min_level,
            "notify_channels": sub.notify_channels or ["SYSTEM"],
            "quiet_start": sub.quiet_start,
            "quiet_end": sub.quiet_end,
            "is_active": sub.is_active,
            "created_at": sub.created_at.isoformat() if sub.created_at else None,
            "updated_at": sub.updated_at.isoformat() if sub.updated_at else None,
        }
        items.append(item)
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/alerts/subscriptions/{subscription_id}", response_model=AlertSubscriptionResponse, status_code=status.HTTP_200_OK)
def read_alert_subscription(
    subscription_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取预警订阅配置详情
    """
    subscription = db.query(AlertSubscription).filter(
        AlertSubscription.id == subscription_id,
        AlertSubscription.user_id == current_user.id
    ).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="订阅配置不存在")
    
    return {
        "id": subscription.id,
        "user_id": subscription.user_id,
        "alert_type": subscription.alert_type,
        "project_id": subscription.project_id,
        "project_name": subscription.project.project_name if subscription.project else None,
        "min_level": subscription.min_level,
        "notify_channels": subscription.notify_channels or ["SYSTEM"],
        "quiet_start": subscription.quiet_start,
        "quiet_end": subscription.quiet_end,
        "is_active": subscription.is_active,
        "created_at": subscription.created_at.isoformat() if subscription.created_at else None,
        "updated_at": subscription.updated_at.isoformat() if subscription.updated_at else None,
    }


@router.post("/alerts/subscriptions", response_model=AlertSubscriptionResponse, status_code=status.HTTP_201_CREATED)
def create_alert_subscription(
    *,
    db: Session = Depends(deps.get_db),
    subscription_in: AlertSubscriptionCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建预警订阅配置
    """
    # 验证预警类型（如果提供）
    if subscription_in.alert_type:
        from app.models.enums import AlertRuleTypeEnum
        valid_types = [e.value for e in AlertRuleTypeEnum]
        if subscription_in.alert_type not in valid_types:
            raise HTTPException(
                status_code=400,
                detail=f"无效的预警类型: {subscription_in.alert_type}。支持的类型: {', '.join(valid_types)}"
            )
    
    # 验证预警级别
    from app.models.enums import AlertLevelEnum
    valid_levels = [e.value for e in AlertLevelEnum]
    if subscription_in.min_level not in valid_levels:
        raise HTTPException(
            status_code=400,
            detail=f"无效的预警级别: {subscription_in.min_level}。支持的级别: {', '.join(valid_levels)}"
        )
    
    # 验证通知渠道
    valid_channels = ['SYSTEM', 'EMAIL', 'WECHAT', 'SMS']
    if subscription_in.notify_channels:
        for channel in subscription_in.notify_channels:
            if channel.upper() not in valid_channels:
                raise HTTPException(
                    status_code=400,
                    detail=f"无效的通知渠道: {channel}。支持的渠道: {', '.join(valid_channels)}"
                )
    
    # 验证免打扰时间格式
    if subscription_in.quiet_start or subscription_in.quiet_end:
        import re
        time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'
        if subscription_in.quiet_start and not re.match(time_pattern, subscription_in.quiet_start):
            raise HTTPException(status_code=400, detail="免打扰开始时间格式错误，应为 HH:mm 格式")
        if subscription_in.quiet_end and not re.match(time_pattern, subscription_in.quiet_end):
            raise HTTPException(status_code=400, detail="免打扰结束时间格式错误，应为 HH:mm 格式")
    
    # 检查是否已存在相同的订阅（相同用户、类型、项目）
    existing = db.query(AlertSubscription).filter(
        AlertSubscription.user_id == current_user.id,
        AlertSubscription.alert_type == (subscription_in.alert_type or None),
        AlertSubscription.project_id == (subscription_in.project_id or None)
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="已存在相同的订阅配置")
    
    # 验证项目ID（如果提供）
    if subscription_in.project_id:
        project = db.query(Project).filter(Project.id == subscription_in.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
    
    subscription = AlertSubscription(
        user_id=current_user.id,
        alert_type=subscription_in.alert_type,
        project_id=subscription_in.project_id,
        min_level=subscription_in.min_level,
        notify_channels=subscription_in.notify_channels or ["SYSTEM"],
        quiet_start=subscription_in.quiet_start,
        quiet_end=subscription_in.quiet_end,
        is_active=subscription_in.is_active
    )
    
    db.add(subscription)
    db.commit()
    db.refresh(subscription)
    
    return {
        "id": subscription.id,
        "user_id": subscription.user_id,
        "alert_type": subscription.alert_type,
        "project_id": subscription.project_id,
        "project_name": subscription.project.project_name if subscription.project else None,
        "min_level": subscription.min_level,
        "notify_channels": subscription.notify_channels or ["SYSTEM"],
        "quiet_start": subscription.quiet_start,
        "quiet_end": subscription.quiet_end,
        "is_active": subscription.is_active,
        "created_at": subscription.created_at.isoformat() if subscription.created_at else None,
        "updated_at": subscription.updated_at.isoformat() if subscription.updated_at else None,
    }


@router.put("/alerts/subscriptions/{subscription_id}", response_model=AlertSubscriptionResponse, status_code=status.HTTP_200_OK)
def update_alert_subscription(
    *,
    db: Session = Depends(deps.get_db),
    subscription_id: int,
    subscription_in: AlertSubscriptionUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新预警订阅配置
    """
    subscription = db.query(AlertSubscription).filter(
        AlertSubscription.id == subscription_id,
        AlertSubscription.user_id == current_user.id
    ).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="订阅配置不存在")
    
    # 验证预警类型（如果提供）
    if subscription_in.alert_type is not None:
        from app.models.enums import AlertRuleTypeEnum
        valid_types = [e.value for e in AlertRuleTypeEnum]
        if subscription_in.alert_type and subscription_in.alert_type not in valid_types:
            raise HTTPException(
                status_code=400,
                detail=f"无效的预警类型: {subscription_in.alert_type}"
            )
    
    # 验证预警级别（如果提供）
    if subscription_in.min_level is not None:
        from app.models.enums import AlertLevelEnum
        valid_levels = [e.value for e in AlertLevelEnum]
        if subscription_in.min_level not in valid_levels:
            raise HTTPException(
                status_code=400,
                detail=f"无效的预警级别: {subscription_in.min_level}"
            )
    
    # 验证通知渠道（如果提供）
    if subscription_in.notify_channels is not None:
        valid_channels = ['SYSTEM', 'EMAIL', 'WECHAT', 'SMS']
        for channel in subscription_in.notify_channels:
            if channel.upper() not in valid_channels:
                raise HTTPException(
                    status_code=400,
                    detail=f"无效的通知渠道: {channel}"
                )
    
    # 验证免打扰时间格式（如果提供）
    if subscription_in.quiet_start or subscription_in.quiet_end:
        import re
        time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'
        if subscription_in.quiet_start and not re.match(time_pattern, subscription_in.quiet_start):
            raise HTTPException(status_code=400, detail="免打扰开始时间格式错误")
        if subscription_in.quiet_end and not re.match(time_pattern, subscription_in.quiet_end):
            raise HTTPException(status_code=400, detail="免打扰结束时间格式错误")
    
    # 验证项目ID（如果提供）
    if subscription_in.project_id is not None and subscription_in.project_id:
        project = db.query(Project).filter(Project.id == subscription_in.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查是否与其他订阅冲突（如果修改了类型或项目）
    if subscription_in.alert_type is not None or subscription_in.project_id is not None:
        new_type = subscription_in.alert_type if subscription_in.alert_type is not None else subscription.alert_type
        new_project_id = subscription_in.project_id if subscription_in.project_id is not None else subscription.project_id
        
        existing = db.query(AlertSubscription).filter(
            AlertSubscription.id != subscription_id,
            AlertSubscription.user_id == current_user.id,
            AlertSubscription.alert_type == (new_type or None),
            AlertSubscription.project_id == (new_project_id or None)
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="已存在相同的订阅配置")
    
    # 更新字段
    update_data = subscription_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(subscription, field, value)
    
    db.add(subscription)
    db.commit()
    db.refresh(subscription)
    
    return {
        "id": subscription.id,
        "user_id": subscription.user_id,
        "alert_type": subscription.alert_type,
        "project_id": subscription.project_id,
        "project_name": subscription.project.project_name if subscription.project else None,
        "min_level": subscription.min_level,
        "notify_channels": subscription.notify_channels or ["SYSTEM"],
        "quiet_start": subscription.quiet_start,
        "quiet_end": subscription.quiet_end,
        "is_active": subscription.is_active,
        "created_at": subscription.created_at.isoformat() if subscription.created_at else None,
        "updated_at": subscription.updated_at.isoformat() if subscription.updated_at else None,
    }


@router.delete("/alerts/subscriptions/{subscription_id}", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def delete_alert_subscription(
    *,
    db: Session = Depends(deps.get_db),
    subscription_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除预警订阅配置
    """
    subscription = db.query(AlertSubscription).filter(
        AlertSubscription.id == subscription_id,
        AlertSubscription.user_id == current_user.id
    ).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="订阅配置不存在")
    
    db.delete(subscription)
    db.commit()
    
    return ResponseModel(code=200, message="订阅配置已删除")


@router.put("/alerts/subscriptions/{subscription_id}/toggle", response_model=AlertSubscriptionResponse, status_code=status.HTTP_200_OK)
def toggle_alert_subscription(
    *,
    db: Session = Depends(deps.get_db),
    subscription_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    启用/禁用预警订阅配置
    """
    subscription = db.query(AlertSubscription).filter(
        AlertSubscription.id == subscription_id,
        AlertSubscription.user_id == current_user.id
    ).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="订阅配置不存在")
    
    subscription.is_active = not subscription.is_active
    db.add(subscription)
    db.commit()
    db.refresh(subscription)
    
    return {
        "id": subscription.id,
        "user_id": subscription.user_id,
        "alert_type": subscription.alert_type,
        "project_id": subscription.project_id,
        "project_name": subscription.project.project_name if subscription.project else None,
        "min_level": subscription.min_level,
        "notify_channels": subscription.notify_channels or ["SYSTEM"],
        "quiet_start": subscription.quiet_start,
        "quiet_end": subscription.quiet_end,
        "is_active": subscription.is_active,
        "created_at": subscription.created_at.isoformat() if subscription.created_at else None,
        "updated_at": subscription.updated_at.isoformat() if subscription.updated_at else None,
    }



# ==================== 预警报表导出 ====================

@router.get("/alerts/export/excel", response_class=StreamingResponse, status_code=status.HTTP_200_OK)
def export_alerts_excel(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    alert_level: Optional[str] = Query(None, description="预警级别筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    rule_type: Optional[str] = Query(None, description="预警类型筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    group_by: Optional[str] = Query("none", description="分组方式: none/level/type"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出预警数据到 Excel
    
    支持筛选参数（与列表接口一致）
    支持多 Sheet（按级别或类型分组）
    """
    try:
        import pandas as pd
        import openpyxl
        import io
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl: pip install pandas openpyxl"
        )
    
    # 构建查询（与列表接口一致）
    query = db.query(AlertRecord).filter(AlertRecord.triggered_at.isnot(None))
    
    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    if alert_level:
        query = query.filter(AlertRecord.alert_level == alert_level)
    if status:
        query = query.filter(AlertRecord.status == status)
    if rule_type:
        query = query.join(AlertRule).filter(AlertRule.rule_type == rule_type)
    if start_date:
        query = query.filter(AlertRecord.triggered_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(AlertRecord.triggered_at <= datetime.combine(end_date, datetime.max.time()))
    
    alerts = query.order_by(AlertRecord.triggered_at.desc()).all()
    
    if not alerts:
        raise HTTPException(status_code=404, detail="没有符合条件的数据")
    
    # 准备数据
    data = []
    for alert in alerts:
        rule = alert.rule
        project = alert.project
        handler = None
        if alert.handler_id:
            handler = db.query(User).filter(User.id == alert.handler_id).first()
        elif alert.acknowledged_by:
            handler = db.query(User).filter(User.id == alert.acknowledged_by).first()
        
        data.append({
            '预警编号': alert.alert_no,
            '预警级别': alert.alert_level,
            '预警标题': alert.alert_title,
            '预警类型': rule.rule_type if rule else 'UNKNOWN',
            '项目名称': project.project_name if project else '',
            '项目编码': project.project_code if project else '',
            '触发时间': alert.triggered_at.strftime('%Y-%m-%d %H:%M:%S') if alert.triggered_at else '',
            '状态': alert.status,
            '处理人': handler.username if handler else '',
            '确认时间': alert.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S') if alert.acknowledged_at else '',
            '处理完成时间': alert.handle_end_at.strftime('%Y-%m-%d %H:%M:%S') if alert.handle_end_at else '',
            '是否升级': '是' if alert.is_escalated else '否',
            '处理结果': alert.handle_result or '',
        })
    
    # 创建 Excel 文件
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if group_by == 'level':
            # 按级别分组
            level_groups = {}
            for item in data:
                level = item['预警级别']
                if level not in level_groups:
                    level_groups[level] = []
                level_groups[level].append(item)
            
            for level, items in level_groups.items():
                df = pd.DataFrame(items)
                sheet_name = f"{level}级预警"[:31]  # Excel sheet名称限制31字符
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 设置样式
                worksheet = writer.sheets[sheet_name]
                _format_alert_excel_sheet(worksheet, level)
        
        elif group_by == 'type':
            # 按类型分组
            type_groups = {}
            for item in data:
                rule_type = item['预警类型']
                if rule_type not in type_groups:
                    type_groups[rule_type] = []
                type_groups[rule_type].append(item)
            
            for rule_type, items in type_groups.items():
                df = pd.DataFrame(items)
                sheet_name = rule_type[:31]  # Excel sheet名称限制31字符
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 设置样式
                worksheet = writer.sheets[sheet_name]
                _format_alert_excel_sheet(worksheet, None)
        
        else:
            # 不分组，单个Sheet
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='预警列表', index=False)
            
            # 设置样式
            worksheet = writer.sheets['预警列表']
            _format_alert_excel_sheet(worksheet, None)
    
    output.seek(0)
    
    # 生成文件名
    filename = f"预警报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


def _format_alert_excel_sheet(worksheet, level: Optional[str] = None):
    """格式化 Excel Sheet"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # 级别颜色映射
    level_colors = {
        'URGENT': 'FF0000',      # 红色
        'CRITICAL': 'FF8C00',    # 橙色
        'WARNING': 'FFD700',     # 黄色
        'INFO': '4169E1',        # 蓝色
    }
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置列宽
    column_widths = {
        'A': 20,  # 预警编号
        'B': 12,  # 预警级别
        'C': 40,  # 预警标题
        'D': 20,  # 预警类型
        'E': 25,  # 项目名称
        'F': 15,  # 项目编码
        'G': 18,  # 触发时间
        'H': 12,  # 状态
        'I': 12,  # 处理人
        'J': 18,  # 确认时间
        'K': 18,  # 处理完成时间
        'L': 10,  # 是否升级
        'M': 40,  # 处理结果
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # 根据级别设置行颜色
    if level and level in level_colors:
        fill_color = level_colors[level]
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # 只对级别列（B列）设置颜色
            if row[1].value == level:
                row[1].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                row[1].font = Font(color="FFFFFF", bold=True)
    
    # 设置数据行对齐
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['G', 'J', 'K']:  # 时间列
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


@router.get("/alerts/export/pdf", response_class=StreamingResponse, status_code=status.HTTP_200_OK)
def export_alerts_pdf(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    alert_level: Optional[str] = Query(None, description="预警级别筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    rule_type: Optional[str] = Query(None, description="预警类型筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出预警数据到 PDF
    
    包含统计摘要和预警列表
    支持分页
    """
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate
    
    from app.services.alert_pdf_service import (
        build_alert_query,
        get_pdf_styles,
        build_pdf_content
    )
    
    try:
        # 构建查询
        query = build_alert_query(
            db, project_id, alert_level, status, rule_type, start_date, end_date
        )
        alerts = query.all()
        
        if not alerts:
            raise HTTPException(status_code=404, detail="没有符合条件的数据")
        
        # 创建PDF缓冲区
        buffer = io.BytesIO()
        
        # 创建PDF文档
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # 获取样式
        title_style, heading_style, normal_style, _ = get_pdf_styles()
        
        # 构建PDF内容
        story = build_pdf_content(db, alerts, title_style, heading_style, normal_style)
        
        # 生成PDF
        doc.build(story)
        buffer.seek(0)
        
        # 生成文件名
        filename = f"预警报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
            }
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF处理库未安装，请安装reportlab: pip install reportlab"
        )

# -*- coding: utf-8 -*-
"""
EXPORTS - 自动生成
从 alerts.py 拆分
"""

import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, List, Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, case, func, or_
from sqlalchemy.orm import Session, joinedload, selectinload

from app.api import deps
from app.core import security
from app.core.config import settings
from app.services.import_export_engine import ExcelExportEngine
from app.models.alert import (
    AlertNotification,
    AlertRecord,
    AlertRule,
    AlertRuleTemplate,
    AlertStatistics,
    AlertSubscription,
    ExceptionAction,
    ExceptionEscalation,
    ExceptionEvent,
    ProjectHealthSnapshot,
)
from app.models.issue import Issue
from app.models.project import Machine, Project
from app.models.user import User
from app.schemas.alert import (
    AlertRecordHandle,
    AlertRecordListResponse,
    AlertRecordResponse,
    AlertRuleCreate,
    AlertRuleResponse,
    AlertRuleUpdate,
    AlertStatisticsResponse,
    AlertSubscriptionCreate,
    AlertSubscriptionResponse,
    AlertSubscriptionUpdate,
    ExceptionEventCreate,
    ExceptionEventListResponse,
    ExceptionEventResolve,
    ExceptionEventResponse,
    ExceptionEventUpdate,
    ExceptionEventVerify,
    ProjectHealthResponse,
)
from app.schemas.common import PaginatedResponse, ResponseModel

router = APIRouter(tags=["exports"])

# ==================== 路由定义 ====================
# 共 2 个路由

@router.get("/alerts/export/excel", response_class=StreamingResponse, status_code=status.HTTP_200_OK)
def export_alerts_excel(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    alert_level: Optional[str] = Query(None, description="预警级别筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    rule_type: Optional[str] = Query(None, description="预警类型筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    group_by: Optional[str] = Query("none", description="分组方式: none/level/type"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出预警数据到 Excel

    支持筛选参数（与列表接口一致）
    支持多 Sheet（按级别或类型分组）
    """
    # 构建查询（与列表接口一致）
    query = db.query(AlertRecord).filter(AlertRecord.triggered_at.isnot(None))

    if project_id:
        query = query.filter(AlertRecord.project_id == project_id)
    if alert_level:
        query = query.filter(AlertRecord.alert_level == alert_level)
    if status:
        query = query.filter(AlertRecord.status == status)
    if rule_type:
        query = query.join(AlertRule).filter(AlertRule.rule_type == rule_type)
    if start_date:
        query = query.filter(AlertRecord.triggered_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(AlertRecord.triggered_at <= datetime.combine(end_date, datetime.max.time()))

    alerts = query.order_by(AlertRecord.triggered_at.desc()).all()

    if not alerts:
        raise HTTPException(status_code=404, detail="没有符合条件的数据")

    # 准备数据
    labels = [
        "预警编号",
        "预警级别",
        "预警标题",
        "预警类型",
        "项目名称",
        "项目编码",
        "触发时间",
        "状态",
        "处理人",
        "确认时间",
        "处理完成时间",
        "是否升级",
        "处理结果",
    ]
    widths = [20, 12, 40, 20, 25, 15, 18, 12, 12, 18, 18, 10, 40]
    columns = ExcelExportEngine.build_columns(labels, widths=widths)
    data = []
    for alert in alerts:
        rule = alert.rule
        project = alert.project
        handler = None
        if alert.handler_id:
            handler = db.query(User).filter(User.id == alert.handler_id).first()
        elif alert.acknowledged_by:
            handler = db.query(User).filter(User.id == alert.acknowledged_by).first()

        data.append({
            '预警编号': alert.alert_no,
            '预警级别': alert.alert_level,
            '预警标题': alert.alert_title,
            '预警类型': rule.rule_type if rule else 'UNKNOWN',
            '项目名称': project.project_name if project else '',
            '项目编码': project.project_code if project else '',
            '触发时间': alert.triggered_at.strftime('%Y-%m-%d %H:%M:%S') if alert.triggered_at else '',
            '状态': alert.status,
            '处理人': handler.username if handler else '',
            '确认时间': alert.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S') if alert.acknowledged_at else '',
            '处理完成时间': alert.handle_end_at.strftime('%Y-%m-%d %H:%M:%S') if alert.handle_end_at else '',
            '是否升级': '是' if alert.is_escalated else '否',
            '处理结果': alert.handle_result or '',
        })

    def post_process(worksheet, sheet_config):
        _format_alert_excel_sheet(worksheet, sheet_config.get("level"))

    if group_by == 'level':
        # 按级别分组
        level_groups = {}
        for item in data:
            level = item['预警级别']
            if level not in level_groups:
                level_groups[level] = []
            level_groups[level].append(item)

        sheets = []
        for level, items in level_groups.items():
            sheet_name = f"{level}级预警"[:31]
            sheets.append({
                "name": sheet_name,
                "data": items,
                "columns": columns,
                "level": level,
            })

        output = ExcelExportEngine.export_multi_sheet(sheets, sheet_post_process=post_process)

    elif group_by == 'type':
        # 按类型分组
        type_groups = {}
        for item in data:
            rule_type = item['预警类型']
            if rule_type not in type_groups:
                type_groups[rule_type] = []
            type_groups[rule_type].append(item)

        sheets = []
        for rule_type, items in type_groups.items():
            sheet_name = rule_type[:31]
            sheets.append({
                "name": sheet_name,
                "data": items,
                "columns": columns,
            })

        output = ExcelExportEngine.export_multi_sheet(sheets, sheet_post_process=post_process)

    else:
        # 不分组，单个Sheet
        output = ExcelExportEngine.export_multi_sheet(
            [{"name": "预警列表", "data": data, "columns": columns}],
            sheet_post_process=post_process,
        )

    # 生成文件名
    filename = f"预警报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


def _format_alert_excel_sheet(worksheet, level: Optional[str] = None):
    """格式化 Excel Sheet"""
    from openpyxl.styles import Alignment, Font, PatternFill

    # 级别颜色映射
    level_colors = {
        'URGENT': 'FF0000',      # 红色
        'CRITICAL': 'FF8C00',    # 橙色
        'WARNING': 'FFD700',     # 黄色
        'INFO': '4169E1',        # 蓝色
    }

    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 设置列宽
    column_widths = {
        'A': 20,  # 预警编号
        'B': 12,  # 预警级别
        'C': 40,  # 预警标题
        'D': 20,  # 预警类型
        'E': 25,  # 项目名称
        'F': 15,  # 项目编码
        'G': 18,  # 触发时间
        'H': 12,  # 状态
        'I': 12,  # 处理人
        'J': 18,  # 确认时间
        'K': 18,  # 处理完成时间
        'L': 10,  # 是否升级
        'M': 40,  # 处理结果
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # 根据级别设置行颜色
    if level and level in level_colors:
        fill_color = level_colors[level]
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # 只对级别列（B列）设置颜色
            if row[1].value == level:
                row[1].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                row[1].font = Font(color="FFFFFF", bold=True)

    # 设置数据行对齐
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['G', 'J', 'K']:  # 时间列
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


@router.get("/alerts/export/pdf", response_class=StreamingResponse, status_code=status.HTTP_200_OK)
def export_alerts_pdf(
    db: Session = Depends(deps.get_db),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    alert_level: Optional[str] = Query(None, description="预警级别筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    rule_type: Optional[str] = Query(None, description="预警类型筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出预警数据到 PDF

    包含统计摘要和预警列表
    支持分页
    """
    import io

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate

    from app.services.alert_pdf_service import (
        build_alert_query,
        build_pdf_content,
        get_pdf_styles,
    )

    try:
        # 构建查询
        query = build_alert_query(
            db, project_id, alert_level, status, rule_type, start_date, end_date
        )
        alerts = query.all()

        if not alerts:
            raise HTTPException(status_code=404, detail="没有符合条件的数据")

        # 创建PDF缓冲区
        buffer = io.BytesIO()

        # 创建PDF文档
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # 获取样式
        title_style, heading_style, normal_style, _ = get_pdf_styles()

        # 构建PDF内容
        story = build_pdf_content(db, alerts, title_style, heading_style, normal_style)

        # 生成PDF
        doc.build(story)
        buffer.seek(0)

        # 生成文件名
        filename = f"预警报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
            }
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF处理库未安装，请安装reportlab: pip install reportlab"
        )

# -*- coding: utf-8 -*-
"""
奖金激励模块 API 端点
"""

from typing import Any, List, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
import os
import io
import uuid
from pathlib import Path

from app.api import deps
from app.core import security
from app.core.config import settings
from app.models.user import User
from app.models.bonus import (
    BonusRule, BonusCalculation, BonusDistribution, TeamBonusAllocation, BonusAllocationSheet
)
from app.models.performance import PerformanceResult, ProjectContribution, PerformancePeriod
from app.models.project import Project, ProjectMilestone
from app.models.sales import Contract, Invoice
from app.models.presale import PresaleSupportTicket
from app.schemas.bonus import (
    BonusRuleCreate, BonusRuleUpdate, BonusRuleResponse, BonusRuleListResponse,
    BonusCalculationCreate, BonusCalculationResponse, BonusCalculationListResponse,
    BonusCalculationApprove, BonusCalculationQuery,
    BonusDistributionCreate, BonusDistributionResponse, BonusDistributionListResponse,
    BonusDistributionPay, BonusDistributionQuery,
    TeamBonusAllocationCreate, TeamBonusAllocationResponse, TeamBonusAllocationListResponse,
    TeamBonusAllocationApprove,
    CalculatePerformanceBonusRequest, CalculateProjectBonusRequest,
    CalculateMilestoneBonusRequest, CalculateTeamBonusRequest,
    CalculateSalesBonusRequest, CalculateSalesDirectorBonusRequest, CalculatePresaleBonusRequest,
    MyBonusResponse, BonusStatisticsResponse,
    BonusAllocationSheetResponse, BonusAllocationSheetConfirm, BonusAllocationRow
)
from app.schemas.common import ResponseModel, PageParams
from app.services.bonus_calculator import BonusCalculator

router = APIRouter()


# ==================== 奖金规则管理 ====================
def paginate_items(items: List[Any], page: int, page_size: int) -> Tuple[List[Any], int, int]:
    """
    简单的列表分页工具，返回分页后的数据、总数以及总页数
    """
    total = len(items)
    start = (page - 1) * page_size
    end = start + page_size
    return items[start:end], total, (total + page_size - 1) // page_size

@router.post("/rules", response_model=ResponseModel[BonusRuleResponse], status_code=status.HTTP_201_CREATED)
def create_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_in: BonusRuleCreate,
    current_user: User = Depends(security.require_hr_access),
) -> Any:
    """
    创建奖金规则（仅人力资源经理可配置）
    """
    # 检查规则编码是否已存在
    existing = db.query(BonusRule).filter(BonusRule.rule_code == rule_in.rule_code).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"规则编码 {rule_in.rule_code} 已存在"
        )
    
    rule = BonusRule(**rule_in.model_dump())
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    return ResponseModel(code=200, message="创建成功", data=rule)


@router.get("/rules", response_model=BonusRuleListResponse, status_code=status.HTTP_200_OK)
def get_bonus_rules(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    bonus_type: Optional[str] = Query(None, description="奖金类型"),
    is_active: Optional[bool] = Query(None, description="是否启用"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金规则列表
    """
    query = db.query(BonusRule)
    
    if bonus_type:
        query = query.filter(BonusRule.bonus_type == bonus_type)
    if is_active is not None:
        query = query.filter(BonusRule.is_active == is_active)
    
    total = query.count()
    rules = query.order_by(desc(BonusRule.priority), desc(BonusRule.created_at)).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    
    return BonusRuleListResponse(
        items=rules,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/rules/{rule_id}", response_model=ResponseModel[BonusRuleResponse], status_code=status.HTTP_200_OK)
def get_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金规则详情
    """
    rule = db.query(BonusRule).filter(BonusRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    
    return ResponseModel(code=200, data=rule)


@router.put("/rules/{rule_id}", response_model=ResponseModel[BonusRuleResponse], status_code=status.HTTP_200_OK)
def update_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    rule_in: BonusRuleUpdate,
    current_user: User = Depends(security.require_hr_access),
) -> Any:
    """
    更新奖金规则（仅人力资源经理可配置）
    """
    rule = db.query(BonusRule).filter(BonusRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    
    update_data = rule_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule, field, value)
    
    db.commit()
    db.refresh(rule)
    
    return ResponseModel(code=200, message="更新成功", data=rule)


@router.delete("/rules/{rule_id}", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def delete_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.require_hr_access),
) -> Any:
    """
    删除奖金规则（仅人力资源经理可配置）
    """
    rule = db.query(BonusRule).filter(BonusRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    
    # 检查是否有计算记录
    calc_count = db.query(BonusCalculation).filter(BonusCalculation.rule_id == rule_id).count()
    if calc_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"该规则已有 {calc_count} 条计算记录，无法删除"
        )
    
    db.delete(rule)
    db.commit()
    
    return ResponseModel(code=200, message="删除成功")


@router.post("/rules/{rule_id}/activate", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def activate_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.require_hr_access),
) -> Any:
    """
    启用奖金规则（仅人力资源经理可配置）
    """
    rule = db.query(BonusRule).filter(BonusRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    
    rule.is_active = True
    db.commit()
    
    return ResponseModel(code=200, message="启用成功")


@router.post("/rules/{rule_id}/deactivate", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def deactivate_bonus_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.require_hr_access),
) -> Any:
    """
    停用奖金规则（仅人力资源经理可配置）
    """
    rule = db.query(BonusRule).filter(BonusRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")
    
    rule.is_active = False
    db.commit()
    
    return ResponseModel(code=200, message="停用成功")


# ==================== 奖金计算 ====================

@router.post("/calculate/performance", response_model=ResponseModel[List[BonusCalculationResponse]], status_code=status.HTTP_200_OK)
def calculate_performance_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculatePerformanceBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算绩效奖金
    """
    calculator = BonusCalculator(db)
    
    # 获取绩效结果
    query = db.query(PerformanceResult).filter(
        PerformanceResult.period_id == request.period_id
    )
    if request.user_id:
        query = query.filter(PerformanceResult.user_id == request.user_id)
    performance_results = query.all()
    
    if not performance_results:
        return ResponseModel(code=200, message="未找到绩效结果", data=[])
    
    # 获取规则
    if request.rule_id:
        rules = [db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()]
        if not rules[0]:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='PERFORMANCE_BASED')
    
    calculations = []
    for perf_result in performance_results:
        for rule in rules:
            calc = calculator.calculate_performance_bonus(perf_result, rule)
            if calc:
                db.add(calc)
                calculations.append(calc)
    
    db.commit()
    
    for calc in calculations:
        db.refresh(calc)
    
    return ResponseModel(code=200, message="计算完成", data=calculations)


@router.post("/calculate/project", response_model=ResponseModel[List[BonusCalculationResponse]], status_code=status.HTTP_200_OK)
def calculate_project_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculateProjectBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算项目奖金
    """
    project = db.query(Project).filter(Project.id == request.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    calculator = BonusCalculator(db)
    
    # 获取项目贡献记录
    contributions = db.query(ProjectContribution).filter(
        ProjectContribution.project_id == request.project_id
    ).all()
    
    if not contributions:
        return ResponseModel(code=200, message="未找到项目贡献记录", data=[])
    
    # 获取规则
    if request.rule_id:
        rules = [db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()]
        if not rules[0]:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='PROJECT_BASED')
    
    calculations = []
    for contrib in contributions:
        for rule in rules:
            calc = calculator.calculate_project_bonus(contrib, project, rule)
            if calc:
                db.add(calc)
                calculations.append(calc)
    
    db.commit()
    
    for calc in calculations:
        db.refresh(calc)
    
    return ResponseModel(code=200, message="计算完成", data=calculations)


@router.post("/calculate/milestone", response_model=ResponseModel[List[BonusCalculationResponse]], status_code=status.HTTP_200_OK)
def calculate_milestone_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculateMilestoneBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算里程碑奖金
    """
    milestone = db.query(ProjectMilestone).filter(ProjectMilestone.id == request.milestone_id).first()
    if not milestone:
        raise HTTPException(status_code=404, detail="里程碑不存在")
    
    project = db.query(Project).filter(Project.id == milestone.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    calculator = BonusCalculator(db)
    
    # 获取规则
    if request.rule_id:
        rules = [db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()]
        if not rules[0]:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='MILESTONE_BASED')
    
    calculations = []
    for rule in rules:
        calcs = calculator.calculate_milestone_bonus(milestone, project, rule)
        for calc in calcs:
            db.add(calc)
            calculations.append(calc)
    
    db.commit()
    
    for calc in calculations:
        db.refresh(calc)
    
    return ResponseModel(code=200, message="计算完成", data=calculations)


@router.post("/calculate/team", response_model=ResponseModel[TeamBonusAllocationResponse], status_code=status.HTTP_200_OK)
def calculate_team_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculateTeamBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算团队奖金
    """
    project = db.query(Project).filter(Project.id == request.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    calculator = BonusCalculator(db)
    
    # 获取规则
    if request.rule_id:
        rule = db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='TEAM_BASED')
        if not rules:
            raise HTTPException(status_code=404, detail="未找到团队奖金规则")
        rule = rules[0]
    
    allocation = calculator.calculate_team_bonus(project, rule, request.period_id)
    db.add(allocation)
    db.commit()
    db.refresh(allocation)
    
    return ResponseModel(code=200, message="计算完成", data=allocation)


# ==================== 销售奖金计算 ====================

@router.post("/calculate/sales", response_model=ResponseModel[BonusCalculationResponse], status_code=status.HTTP_200_OK)
def calculate_sales_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculateSalesBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算销售奖金
    
    支持两种计算方式：
    1. CONTRACT: 基于合同签订金额
    2. PAYMENT: 基于回款金额
    """
    contract = db.query(Contract).filter(Contract.id == request.contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    
    calculator = BonusCalculator(db)
    
    # 获取规则
    if request.rule_id:
        rule = db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='SALES_BASED')
        if not rules:
            raise HTTPException(status_code=404, detail="未找到销售奖金规则")
        rule = rules[0]
    
    calculation = calculator.calculate_sales_bonus(contract, rule, request.based_on)
    if not calculation:
        raise HTTPException(status_code=400, detail="不满足计算条件")
    
    db.add(calculation)
    db.commit()
    db.refresh(calculation)
    
    return ResponseModel(code=200, message="计算完成", data=calculation)


@router.post("/calculate/sales-director", response_model=ResponseModel[BonusCalculationResponse], status_code=status.HTTP_200_OK)
def calculate_sales_director_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculateSalesDirectorBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算销售总监奖金（基于团队业绩）
    """
    calculator = BonusCalculator(db)
    
    # 获取规则
    if request.rule_id:
        rule = db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='SALES_DIRECTOR_BASED')
        if not rules:
            raise HTTPException(status_code=404, detail="未找到销售总监奖金规则")
        rule = rules[0]
    
    calculation = calculator.calculate_sales_director_bonus(
        request.director_id,
        request.period_start,
        request.period_end,
        rule
    )
    
    if not calculation:
        raise HTTPException(status_code=400, detail="该周期内无团队业绩")
    
    db.add(calculation)
    db.commit()
    db.refresh(calculation)
    
    return ResponseModel(code=200, message="计算完成", data=calculation)


@router.post("/calculate/presale", response_model=ResponseModel[BonusCalculationResponse], status_code=status.HTTP_200_OK)
def calculate_presale_bonus(
    *,
    db: Session = Depends(deps.get_db),
    request: CalculatePresaleBonusRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    计算售前技术支持奖金
    
    支持两种计算方式：
    1. COMPLETION: 基于工单完成
    2. WON: 基于中标
    """
    ticket = db.query(PresaleSupportTicket).filter(
        PresaleSupportTicket.id == request.ticket_id
    ).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="售前支持工单不存在")
    
    calculator = BonusCalculator(db)
    
    # 获取规则
    if request.rule_id:
        rule = db.query(BonusRule).filter(BonusRule.id == request.rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="规则不存在")
    else:
        rules = calculator.get_active_rules(bonus_type='PRESALE_BASED')
        if not rules:
            raise HTTPException(status_code=404, detail="未找到售前技术支持奖金规则")
        rule = rules[0]
    
    calculation = calculator.calculate_presale_bonus(ticket, rule, request.based_on)
    if not calculation:
        raise HTTPException(status_code=400, detail="不满足计算条件")
    
    db.add(calculation)
    db.commit()
    db.refresh(calculation)
    
    return ResponseModel(code=200, message="计算完成", data=calculation)


@router.get("/calculations", response_model=BonusCalculationListResponse, status_code=status.HTTP_200_OK)
def get_bonus_calculations(
    *,
    db: Session = Depends(deps.get_db),
    query_params: BonusCalculationQuery = Depends(),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金计算记录列表
    """
    query = db.query(BonusCalculation)
    
    if query_params.rule_id:
        query = query.filter(BonusCalculation.rule_id == query_params.rule_id)
    if query_params.user_id:
        query = query.filter(BonusCalculation.user_id == query_params.user_id)
    if query_params.project_id:
        query = query.filter(BonusCalculation.project_id == query_params.project_id)
    if query_params.period_id:
        query = query.filter(BonusCalculation.period_id == query_params.period_id)
    if query_params.status:
        query = query.filter(BonusCalculation.status == query_params.status)
    if query_params.bonus_type:
        query = query.join(BonusRule).filter(BonusRule.bonus_type == query_params.bonus_type)
    if query_params.start_date:
        query = query.filter(BonusCalculation.calculated_at >= datetime.combine(query_params.start_date, datetime.min.time()))
    if query_params.end_date:
        query = query.filter(BonusCalculation.calculated_at <= datetime.combine(query_params.end_date, datetime.max.time()))
    
    total = query.count()
    calculations = query.order_by(desc(BonusCalculation.calculated_at)).offset(
        (query_params.page - 1) * query_params.page_size
    ).limit(query_params.page_size).all()
    
    return BonusCalculationListResponse(
        items=calculations,
        total=total,
        page=query_params.page,
        page_size=query_params.page_size,
        pages=(total + query_params.page_size - 1) // query_params.page_size
    )


@router.get("/calculations/{calc_id}", response_model=ResponseModel[BonusCalculationResponse], status_code=status.HTTP_200_OK)
def get_bonus_calculation(
    *,
    db: Session = Depends(deps.get_db),
    calc_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金计算详情
    """
    calculation = db.query(BonusCalculation).filter(BonusCalculation.id == calc_id).first()
    if not calculation:
        raise HTTPException(status_code=404, detail="计算记录不存在")
    
    return ResponseModel(code=200, data=calculation)


@router.post("/calculations/{calc_id}/approve", response_model=ResponseModel[BonusCalculationResponse], status_code=status.HTTP_200_OK)
def approve_bonus_calculation(
    *,
    db: Session = Depends(deps.get_db),
    calc_id: int,
    approve_in: BonusCalculationApprove,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批奖金计算
    """
    calculation = db.query(BonusCalculation).filter(BonusCalculation.id == calc_id).first()
    if not calculation:
        raise HTTPException(status_code=404, detail="计算记录不存在")
    
    if approve_in.approved:
        calculation.status = 'APPROVED'
        calculation.approved_by = current_user.id
        calculation.approved_at = datetime.now()
        calculation.approval_comment = approve_in.comment
    else:
        calculation.status = 'CANCELLED'
        calculation.approved_by = current_user.id
        calculation.approved_at = datetime.now()
        calculation.approval_comment = approve_in.comment
    
    db.commit()
    db.refresh(calculation)
    
    return ResponseModel(code=200, message="审批完成", data=calculation)


# ==================== 奖金发放 ====================

def generate_distribution_code() -> str:
    """生成发放单号"""
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return f"BD{timestamp}"


@router.post("/distribute", response_model=ResponseModel[BonusDistributionResponse], status_code=status.HTTP_201_CREATED)
def create_bonus_distribution(
    *,
    db: Session = Depends(deps.get_db),
    dist_in: BonusDistributionCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建奖金发放记录
    """
    # 检查计算记录是否存在且已审批
    calculation = db.query(BonusCalculation).filter(
        BonusCalculation.id == dist_in.calculation_id
    ).first()
    if not calculation:
        raise HTTPException(status_code=404, detail="计算记录不存在")
    
    if calculation.status != 'APPROVED':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="计算记录未审批，无法发放"
        )
    
    # 检查是否已发放
    existing = db.query(BonusDistribution).filter(
        BonusDistribution.calculation_id == dist_in.calculation_id,
        BonusDistribution.status == 'PAID'
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="该计算记录已发放"
        )
    
    distribution = BonusDistribution(
        distribution_code=generate_distribution_code(),
        **dist_in.model_dump()
    )
    db.add(distribution)
    
    # 更新计算记录状态
    calculation.status = 'DISTRIBUTED'
    
    db.commit()
    db.refresh(distribution)
    
    return ResponseModel(code=200, message="创建成功", data=distribution)


@router.get("/distributions", response_model=BonusDistributionListResponse, status_code=status.HTTP_200_OK)
def get_bonus_distributions(
    *,
    db: Session = Depends(deps.get_db),
    query_params: BonusDistributionQuery = Depends(),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金发放记录列表
    """
    query = db.query(BonusDistribution)
    
    if query_params.user_id:
        query = query.filter(BonusDistribution.user_id == query_params.user_id)
    if query_params.status:
        query = query.filter(BonusDistribution.status == query_params.status)
    if query_params.start_date:
        query = query.filter(BonusDistribution.distribution_date >= query_params.start_date)
    if query_params.end_date:
        query = query.filter(BonusDistribution.distribution_date <= query_params.end_date)
    
    total = query.count()
    distributions = query.order_by(desc(BonusDistribution.distribution_date)).offset(
        (query_params.page - 1) * query_params.page_size
    ).limit(query_params.page_size).all()
    
    return BonusDistributionListResponse(
        items=distributions,
        total=total,
        page=query_params.page,
        page_size=query_params.page_size,
        pages=(total + query_params.page_size - 1) // query_params.page_size
    )


@router.get("/distributions/{dist_id}", response_model=ResponseModel[BonusDistributionResponse], status_code=status.HTTP_200_OK)
def get_bonus_distribution(
    *,
    db: Session = Depends(deps.get_db),
    dist_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金发放记录详情
    """
    distribution = db.query(BonusDistribution).filter(BonusDistribution.id == dist_id).first()
    if not distribution:
        raise HTTPException(status_code=404, detail="发放记录不存在")
    
    return ResponseModel(code=200, data=distribution)


@router.post("/distributions/{dist_id}/pay", response_model=ResponseModel[BonusDistributionResponse], status_code=status.HTTP_200_OK)
def pay_bonus_distribution(
    *,
    db: Session = Depends(deps.get_db),
    dist_id: int,
    pay_in: BonusDistributionPay,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    确认发放奖金
    """
    distribution = db.query(BonusDistribution).filter(BonusDistribution.id == dist_id).first()
    if not distribution:
        raise HTTPException(status_code=404, detail="发放记录不存在")
    
    if distribution.status == 'PAID':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="该记录已发放"
        )
    
    distribution.status = 'PAID'
    distribution.paid_by = current_user.id
    distribution.paid_at = datetime.now()
    
    if pay_in.voucher_no:
        distribution.voucher_no = pay_in.voucher_no
    if pay_in.payment_account:
        distribution.payment_account = pay_in.payment_account
    if pay_in.payment_remark:
        distribution.payment_remark = pay_in.payment_remark
    
    db.commit()
    db.refresh(distribution)
    
    return ResponseModel(code=200, message="发放成功", data=distribution)


# ==================== 团队奖金分配 ====================

@router.get("/team-allocations", response_model=TeamBonusAllocationListResponse, status_code=status.HTTP_200_OK)
def get_team_bonus_allocations(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    project_id: Optional[int] = Query(None, description="项目ID"),
    period_id: Optional[int] = Query(None, description="周期ID"),
    status: Optional[str] = Query(None, description="状态"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取团队奖金分配列表
    """
    query = db.query(TeamBonusAllocation)
    
    if project_id:
        query = query.filter(TeamBonusAllocation.project_id == project_id)
    if period_id:
        query = query.filter(TeamBonusAllocation.period_id == period_id)
    if status:
        query = query.filter(TeamBonusAllocation.status == status)
    
    total = query.count()
    allocations = query.order_by(desc(TeamBonusAllocation.created_at)).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    
    return TeamBonusAllocationListResponse(
        items=allocations,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/team-allocations/{allocation_id}", response_model=ResponseModel[TeamBonusAllocationResponse], status_code=status.HTTP_200_OK)
def get_team_bonus_allocation(
    *,
    db: Session = Depends(deps.get_db),
    allocation_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取团队奖金分配详情
    """
    allocation = db.query(TeamBonusAllocation).filter(TeamBonusAllocation.id == allocation_id).first()
    if not allocation:
        raise HTTPException(status_code=404, detail="团队奖金分配记录不存在")
    
    return ResponseModel(code=200, data=allocation)


@router.post("/team-allocations/{allocation_id}/approve", response_model=ResponseModel[TeamBonusAllocationResponse], status_code=status.HTTP_200_OK)
def approve_team_bonus_allocation(
    *,
    db: Session = Depends(deps.get_db),
    allocation_id: int,
    approve_in: TeamBonusAllocationApprove,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批团队奖金分配
    """
    allocation = db.query(TeamBonusAllocation).filter(TeamBonusAllocation.id == allocation_id).first()
    if not allocation:
        raise HTTPException(status_code=404, detail="团队奖金分配记录不存在")
    
    if approve_in.approved:
        allocation.status = 'APPROVED'
        allocation.approved_by = current_user.id
        allocation.approved_at = datetime.now()
    else:
        allocation.status = 'REJECTED'
        allocation.approved_by = current_user.id
        allocation.approved_at = datetime.now()
    
    db.commit()
    db.refresh(allocation)
    
    return ResponseModel(code=200, message="审批完成", data=allocation)


# ==================== 我的奖金 ====================

@router.get("/my", response_model=ResponseModel[MyBonusResponse], status_code=status.HTTP_200_OK)
def get_my_bonus(
    *,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取我的奖金
    """
    # 获取计算记录
    calculations = db.query(BonusCalculation).filter(
        BonusCalculation.user_id == current_user.id
    ).order_by(desc(BonusCalculation.calculated_at)).all()
    
    # 获取发放记录
    distributions = db.query(BonusDistribution).filter(
        BonusDistribution.user_id == current_user.id
    ).order_by(desc(BonusDistribution.distribution_date)).all()
    
    # 计算总金额
    total_amount = sum(float(c.calculated_amount) for c in calculations)
    paid_amount = sum(float(d.distributed_amount) for d in distributions if d.status == 'PAID')
    pending_amount = total_amount - paid_amount
    
    return ResponseModel(
        code=200,
        data=MyBonusResponse(
            total_amount=Decimal(str(total_amount)),
            pending_amount=Decimal(str(pending_amount)),
            paid_amount=Decimal(str(paid_amount)),
            calculations=calculations,
            distributions=distributions
        )
    )


# ==================== 奖金统计 ====================

@router.get("/statistics", response_model=ResponseModel[BonusStatisticsResponse], status_code=status.HTTP_200_OK)
def get_bonus_statistics(
    *,
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取奖金统计
    """
    # 计算记录统计
    calc_query = db.query(BonusCalculation)
    if start_date:
        calc_query = calc_query.filter(BonusCalculation.calculated_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        calc_query = calc_query.filter(BonusCalculation.calculated_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_calculated = calc_query.with_entities(func.sum(BonusCalculation.calculated_amount)).scalar() or Decimal('0')
    calculation_count = calc_query.count()
    
    # 发放记录统计
    dist_query = db.query(BonusDistribution)
    if start_date:
        dist_query = dist_query.filter(BonusDistribution.distribution_date >= start_date)
    if end_date:
        dist_query = dist_query.filter(BonusDistribution.distribution_date <= end_date)
    
    total_distributed = dist_query.filter(
        BonusDistribution.status == 'PAID'
    ).with_entities(func.sum(BonusDistribution.distributed_amount)).scalar() or Decimal('0')
    
    total_pending = dist_query.filter(
        BonusDistribution.status == 'PENDING'
    ).with_entities(func.sum(BonusDistribution.distributed_amount)).scalar() or Decimal('0')
    
    distribution_count = dist_query.count()
    
    # 按类型统计
    by_type = {}
    type_query = calc_query.join(BonusRule).with_entities(
        BonusRule.bonus_type,
        func.sum(BonusCalculation.calculated_amount)
    ).group_by(BonusRule.bonus_type).all()
    
    for bonus_type, amount in type_query:
        by_type[bonus_type] = amount or Decimal('0')
    
    # 按部门统计（关联用户表）
    from app.models.user import User
    by_department = {}
    dept_query = calc_query.join(User, BonusCalculation.user_id == User.id).with_entities(
        User.department,
        func.sum(BonusCalculation.calculated_amount)
    ).group_by(User.department).all()
    
    for dept_name, amount in dept_query:
        if dept_name:  # 只统计有部门信息的记录
            by_department[dept_name] = amount or Decimal('0')
    
    return ResponseModel(
        code=200,
        data=BonusStatisticsResponse(
            total_calculated=total_calculated,
            total_distributed=total_distributed,
            total_pending=total_pending,
            calculation_count=calculation_count,
            distribution_count=distribution_count,
            by_type=by_type,
            by_department=by_department
        )
    )


# ==================== 奖金分配明细表 ====================

def generate_sheet_code() -> str:
    """生成分配明细表编号"""
    today = datetime.now().strftime('%y%m%d')
    return f"BAS-{today}-{uuid.uuid4().hex[:6].upper()}"


@router.get("/allocation-sheets/template", response_class=FileResponse)
def download_allocation_template(
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    下载奖金分配明细表模板（Excel）
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "奖金分配明细表"
    
    # 设置表头（支持两种模式）
    headers = [
        "计算记录ID",
        "团队奖金分配ID",
        "受益人ID*",
        "受益人姓名",
        "计算金额",
        "发放金额*",
        "发放日期*",
        "发放方式",
        "凭证号",
        "付款账户",
        "付款备注"
    ]
    
    # 写入表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # 添加说明行
    ws.insert_rows(1)
    ws.merge_cells(f'A1:K1')
    note_cell = ws.cell(row=1, column=1, value="说明：1. 带*的列为必填项；2. 必须提供'计算记录ID'或'团队奖金分配ID'之一；3. 如果使用团队奖金分配ID，系统会自动创建个人计算记录；4. 受益人ID必须为数字；5. 金额必须为数字；6. 发放日期格式：YYYY-MM-DD")
    note_cell.font = Font(size=10, italic=True)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 创建临时文件
    template_dir = os.path.join(settings.UPLOAD_DIR, "templates")
    os.makedirs(template_dir, exist_ok=True)
    template_path = os.path.join(template_dir, "奖金分配明细表模板.xlsx")
    with open(template_path, "wb") as f:
        f.write(output.getvalue())
    
    return FileResponse(
        path=template_path,
        filename="奖金分配明细表模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/allocation-sheets/upload", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_201_CREATED)
async def upload_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    file: UploadFile = File(..., description="分配明细表Excel文件"),
    sheet_name: str = Form(..., description="明细表名称"),
    project_id: Optional[int] = Form(None, description="项目ID（可选）"),
    period_id: Optional[int] = Form(None, description="考核周期ID（可选）"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    上传奖金分配明细表
    
    上传后会自动解析Excel文件，验证数据格式
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    from app.services.bonus_allocation_parser import (
        validate_file_type,
        save_uploaded_file,
        read_and_save_file,
        parse_excel_file,
        validate_required_columns,
        parse_allocation_sheet
    )
    
    file_path = None
    try:
        # 验证文件类型
        validate_file_type(file.filename)
        
        # 保存文件
        file_path, relative_path, _ = save_uploaded_file(file)
        file_content, file_size = await read_and_save_file(file, file_path)
        
        # 解析Excel
        df = parse_excel_file(file_content)
        validate_required_columns(df)
        
        # 解析数据
        valid_rows, parse_errors = parse_allocation_sheet(df, db)
        invalid_rows = list(parse_errors.keys())
        
        # 创建上传记录
        sheet_code = generate_sheet_code()
        allocation_sheet = BonusAllocationSheet(
            sheet_code=sheet_code,
            sheet_name=sheet_name,
            file_path=relative_path,
            file_name=file.filename,
            file_size=file_size,
            project_id=project_id,
            period_id=period_id,
            total_rows=len(df),
            valid_rows=len(valid_rows),
            invalid_rows=len(invalid_rows),
            status='PARSED' if len(invalid_rows) == 0 else 'UPLOADED',
            parse_result={'valid_rows': valid_rows},
            parse_errors=parse_errors if parse_errors else None,
            uploaded_by=current_user.id
        )
        
        db.add(allocation_sheet)
        db.commit()
        db.refresh(allocation_sheet)
        
        return ResponseModel(
            code=201,
            message=f"上传成功，有效行数：{len(valid_rows)}，无效行数：{len(invalid_rows)}",
            data=BonusAllocationSheetResponse.model_validate(allocation_sheet)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        # 删除已上传的文件
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"解析Excel文件失败: {str(e)}")


@router.post("/allocation-sheets/{sheet_id}/confirm", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_200_OK)
def confirm_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    confirm_in: BonusAllocationSheetConfirm,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    确认分配明细表（线下确认完成）
    
    记录财务部、人力资源部、总经理的线下确认状态
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")
    
    if sheet.status == 'DISTRIBUTED':
        raise HTTPException(status_code=400, detail="该明细表已发放，无法修改确认状态")
    
    sheet.finance_confirmed = confirm_in.finance_confirmed
    sheet.hr_confirmed = confirm_in.hr_confirmed
    sheet.manager_confirmed = confirm_in.manager_confirmed
    
    # 如果全部确认，更新确认时间
    if confirm_in.finance_confirmed and confirm_in.hr_confirmed and confirm_in.manager_confirmed:
        sheet.confirmed_at = datetime.now()
    
    db.commit()
    db.refresh(sheet)
    
    return ResponseModel(
        code=200,
        message="确认状态更新成功",
        data=BonusAllocationSheetResponse.model_validate(sheet)
    )


@router.post("/allocation-sheets/{sheet_id}/distribute", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def distribute_bonus_from_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    同意发放 - 根据分配明细表批量创建发放记录
    
    只有线下确认完成（财务、人力、总经理都确认）的明细表才能发放
    """
    from app.services.bonus_distribution_service import (
        validate_sheet_for_distribution,
        create_calculation_from_team_allocation,
        create_distribution_record,
        check_distribution_exists
    )
    from app.services.bonus_calculator import BonusCalculator
    
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")
    
    # 验证明细表
    is_valid, error_msg = validate_sheet_for_distribution(sheet)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    valid_rows = sheet.parse_result['valid_rows']
    
    # 批量创建发放记录
    distributions = []
    errors = []
    calculator = BonusCalculator(db)
    
    for row_data in valid_rows:
        try:
            calculation = None
            calculation_id = row_data.get('calculation_id')
            team_allocation_id = row_data.get('team_allocation_id')
            
            # 如果使用团队奖金分配ID，先创建个人计算记录
            if team_allocation_id:
                try:
                    calculation = create_calculation_from_team_allocation(
                        db, team_allocation_id, row_data['user_id'],
                        Decimal(str(row_data['calculated_amount'])), calculator
                    )
                    calculation_id = calculation.id
                except ValueError as e:
                    errors.append(str(e))
                    continue
            else:
                # 使用已有的计算记录
                calculation = db.query(BonusCalculation).filter(
                    BonusCalculation.id == calculation_id
                ).first()
                if not calculation:
                    errors.append(f"计算记录ID {calculation_id} 不存在")
                    continue
            
            # 检查是否已发放
            if check_distribution_exists(db, calculation_id, row_data['user_id']):
                errors.append(f"计算记录ID {calculation_id} 对用户ID {row_data['user_id']} 已发放")
                continue
            
            # 创建发放记录
            distribution = create_distribution_record(
                db, calculation_id, row_data['user_id'], row_data,
                current_user.id, generate_distribution_code
            )
            distributions.append(distribution)
            
            # 更新计算记录状态
            if calculation:
                calculation.status = 'DISTRIBUTED'
            
        except Exception as e:
            errors.append(f"处理行数据失败: {str(e)}")
            continue
    
    if errors and not distributions:
        raise HTTPException(
            status_code=400,
            detail=f"发放失败：{'; '.join(errors[:5])}"  # 只显示前5个错误
        )
    
    # 更新明细表状态
    sheet.status = 'DISTRIBUTED'
    sheet.distributed_at = datetime.now()
    sheet.distributed_by = current_user.id
    sheet.distribution_count = len(distributions)
    
    db.commit()
    
    return ResponseModel(
        code=200,
        message=f"发放成功，共创建 {len(distributions)} 条发放记录" + (f"，{len(errors)} 条失败" if errors else ""),
        data={
            "sheet_id": sheet_id,
            "sheet_code": sheet.sheet_code,
            "distributed_count": len(distributions),
            "error_count": len(errors),
            "errors": errors[:10] if errors else []  # 最多返回10个错误
        }
    )


@router.get("/allocation-sheets", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def get_allocation_sheets(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取分配明细表列表
    """
    query = db.query(BonusAllocationSheet)
    
    if status:
        query = query.filter(BonusAllocationSheet.status == status)
    
    total = query.count()
    offset = (page - 1) * page_size
    sheets = query.order_by(desc(BonusAllocationSheet.created_at)).offset(offset).limit(page_size).all()
    
    items = [BonusAllocationSheetResponse.model_validate(sheet) for sheet in sheets]
    
    return ResponseModel(
        code=200,
        data={
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size
        }
    )


@router.get("/allocation-sheets/{sheet_id}", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_200_OK)
def get_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取分配明细表详情
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")
    
    return ResponseModel(
        code=200,
        data=BonusAllocationSheetResponse.model_validate(sheet)
    )


@router.get("/allocation-sheets/{sheet_id}/download", response_class=FileResponse)
def download_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    下载已上传的分配明细表Excel文件，便于复核和留档
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")
    
    if not sheet.file_path:
        raise HTTPException(status_code=404, detail="明细表文件不存在")
    
    upload_dir = os.path.abspath(settings.UPLOAD_DIR)
    file_path = os.path.abspath(os.path.join(settings.UPLOAD_DIR, sheet.file_path))
    if not file_path.startswith(upload_dir):
        raise HTTPException(status_code=400, detail="文件路径非法，拒绝下载")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件已被删除或不存在")
    
    filename = sheet.file_name or os.path.basename(file_path)
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get(
    "/allocation-sheets/{sheet_id}/rows",
    response_model=ResponseModel,
    status_code=status.HTTP_200_OK
)
def get_allocation_sheet_rows(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    row_type: str = Query("valid", description="数据类型：valid 或 error"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(50, ge=1, le=200, description="每页条数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    查看已解析的分配明细表数据，支持分页查看有效行或错误行
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")
    
    normalized_type = row_type.lower()
    if normalized_type not in {"valid", "error"}:
        raise HTTPException(status_code=400, detail="row_type 仅支持 valid 或 error")
    
    if normalized_type == "valid":
        valid_rows = []
        if sheet.parse_result and sheet.parse_result.get('valid_rows'):
            for row_data in sheet.parse_result['valid_rows']:
                distribution_date = row_data.get('distribution_date')
                if isinstance(distribution_date, str):
                    try:
                        distribution_date = date.fromisoformat(distribution_date)
                    except ValueError:
                        distribution_date = datetime.fromisoformat(distribution_date).date()
                elif isinstance(distribution_date, datetime):
                    distribution_date = distribution_date.date()
                elif not isinstance(distribution_date, date):
                    try:
                        distribution_date = datetime.fromisoformat(str(distribution_date)).date()
                    except ValueError:
                        distribution_date = datetime.strptime(str(distribution_date), '%Y-%m-%d').date()
                
                valid_rows.append(
                    BonusAllocationRow(
                        calculation_id=int(row_data['calculation_id']),
                        user_id=int(row_data['user_id']),
                        user_name=row_data.get('user_name'),
                        calculated_amount=Decimal(str(row_data['calculated_amount'])),
                        distributed_amount=Decimal(str(row_data['distributed_amount'])),
                        distribution_date=distribution_date,
                        payment_method=row_data.get('payment_method'),
                        voucher_no=row_data.get('voucher_no'),
                        payment_account=row_data.get('payment_account'),
                        payment_remark=row_data.get('payment_remark')
                    ).model_dump()
                )
        
        page_items, total, pages = paginate_items(valid_rows, page, page_size)
        data = {
            "sheet_id": sheet_id,
            "row_type": normalized_type,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "items": page_items
        }
    else:
        error_rows = []
        errors = sheet.parse_errors or {}
        for row_no, messages in errors.items():
            try:
                row_number = int(row_no)
            except (TypeError, ValueError):
                row_number = row_no
            if not isinstance(messages, list):
                messages = [str(messages)]
            error_rows.append({
                "row_number": row_number,
                "errors": [str(m) for m in messages]
            })
        
        error_rows.sort(key=lambda x: x["row_number"])
        page_items, total, pages = paginate_items(error_rows, page, page_size)
        data = {
            "sheet_id": sheet_id,
            "row_type": normalized_type,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "items": page_items
        }
    
    return ResponseModel(code=200, data=data)

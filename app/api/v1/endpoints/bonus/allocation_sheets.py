# -*- coding: utf-8 -*-
"""
奖金分配明细表 - 分配表管理

包含分配表的CRUD、上传、确认、分发、下载等功能
"""

import io
import os
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.api import deps
from app.core import security
from app.core.config import settings
from app.models.bonus import BonusAllocationSheet, BonusCalculation
from app.models.user import User
from app.schemas.bonus import (
    BonusAllocationRow,
    BonusAllocationSheetConfirm,
    BonusAllocationSheetResponse,
)
from app.schemas.common import ResponseModel

from .allocation_helpers import generate_sheet_code
from .payment import generate_distribution_code
from .rules import paginate_items

router = APIRouter()


@router.get("/allocation-sheets/template", response_class=FileResponse)
def download_allocation_template(
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    下载奖金分配明细表模板（Excel）
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "奖金分配明细表"

    # 设置表头（支持两种模式）
    headers = [
        "计算记录ID",
        "团队奖金分配ID",
        "受益人ID*",
        "受益人姓名",
        "计算金额",
        "发放金额*",
        "发放日期*",
        "发放方式",
        "凭证号",
        "付款账户",
        "付款备注"
    ]

    # 写入表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # 添加说明行
    ws.insert_rows(1)
    ws.merge_cells(f'A1:K1')
    note_cell = ws.cell(row=1, column=1, value="说明：1. 带*的列为必填项；2. 必须提供'计算记录ID'或'团队奖金分配ID'之一；3. 如果使用团队奖金分配ID，系统会自动创建个人计算记录；4. 受益人ID必须为数字；5. 金额必须为数字；6. 发放日期格式：YYYY-MM-DD")
    note_cell.font = Font(size=10, italic=True)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 创建临时文件
    template_dir = os.path.join(settings.UPLOAD_DIR, "templates")
    os.makedirs(template_dir, exist_ok=True)
    template_path = os.path.join(template_dir, "奖金分配明细表模板.xlsx")
    with open(template_path, "wb") as f:
        f.write(output.getvalue())

    return FileResponse(
        path=template_path,
        filename="奖金分配明细表模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/allocation-sheets/upload", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_201_CREATED)
async def upload_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    file: UploadFile = File(..., description="分配明细表Excel文件"),
    sheet_name: str = Form(..., description="明细表名称"),
    project_id: Optional[int] = Form(None, description="项目ID（可选）"),
    period_id: Optional[int] = Form(None, description="考核周期ID（可选）"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    上传奖金分配明细表

    上传后会自动解析Excel文件，验证数据格式
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )

    from app.services.bonus_allocation_parser import (
        parse_allocation_sheet,
        parse_excel_file,
        read_and_save_file,
        save_uploaded_file,
        validate_file_type,
        validate_required_columns,
    )

    file_path = None
    try:
        # 验证文件类型
        validate_file_type(file.filename)

        # 保存文件
        file_path, relative_path, _ = save_uploaded_file(file)
        file_content, file_size = await read_and_save_file(file, file_path)

        # 解析Excel
        df = parse_excel_file(file_content)
        validate_required_columns(df)

        # 解析数据
        valid_rows, parse_errors = parse_allocation_sheet(df, db)
        invalid_rows = list(parse_errors.keys())

        # 创建上传记录
        sheet_code = generate_sheet_code()
        allocation_sheet = BonusAllocationSheet(
            sheet_code=sheet_code,
            sheet_name=sheet_name,
            file_path=relative_path,
            file_name=file.filename,
            file_size=file_size,
            project_id=project_id,
            period_id=period_id,
            total_rows=len(df),
            valid_rows=len(valid_rows),
            invalid_rows=len(invalid_rows),
            status='PARSED' if len(invalid_rows) == 0 else 'UPLOADED',
            parse_result={'valid_rows': valid_rows},
            parse_errors=parse_errors if parse_errors else None,
            uploaded_by=current_user.id
        )

        db.add(allocation_sheet)
        db.commit()
        db.refresh(allocation_sheet)

        return ResponseModel(
            code=201,
            message=f"上传成功，有效行数：{len(valid_rows)}，无效行数：{len(invalid_rows)}",
            data=BonusAllocationSheetResponse.model_validate(allocation_sheet)
        )

    except HTTPException:
        raise
    except Exception as e:
        # 删除已上传的文件
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"解析Excel文件失败: {str(e)}")


@router.get("/allocation-sheets", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def get_allocation_sheets(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取分配明细表列表
    """
    query = db.query(BonusAllocationSheet)

    if status:
        query = query.filter(BonusAllocationSheet.status == status)

    total = query.count()
    sheets = query.order_by(desc(BonusAllocationSheet.created_at)).offset((page - 1) * page_size).limit(page_size).all()

    items = [BonusAllocationSheetResponse.model_validate(sheet) for sheet in sheets]

    return ResponseModel(
        code=200,
        data={
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size
        }
    )


@router.get("/allocation-sheets/{sheet_id}", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_200_OK)
def get_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取分配明细表详情
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")

    return ResponseModel(
        code=200,
        data=BonusAllocationSheetResponse.model_validate(sheet)
    )


@router.post("/allocation-sheets/{sheet_id}/confirm", response_model=ResponseModel[BonusAllocationSheetResponse], status_code=status.HTTP_200_OK)
def confirm_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    confirm_in: BonusAllocationSheetConfirm,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    确认分配明细表（线下确认完成）

    记录财务部、人力资源部、总经理的线下确认状态
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")

    if sheet.status == 'DISTRIBUTED':
        raise HTTPException(status_code=400, detail="该明细表已发放，无法修改确认状态")

    sheet.finance_confirmed = confirm_in.finance_confirmed
    sheet.hr_confirmed = confirm_in.hr_confirmed
    sheet.manager_confirmed = confirm_in.manager_confirmed

    # 如果全部确认，更新确认时间
    if confirm_in.finance_confirmed and confirm_in.hr_confirmed and confirm_in.manager_confirmed:
        sheet.confirmed_at = datetime.now()

    db.add(sheet)
    db.commit()
    db.refresh(sheet)

    return ResponseModel(
        code=200,
        message="确认状态更新成功",
        data=BonusAllocationSheetResponse.model_validate(sheet)
    )


@router.post("/allocation-sheets/{sheet_id}/distribute", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def distribute_bonus_from_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    同意发放 - 根据分配明细表批量创建发放记录

    只有线下确认完成（财务、人力、总经理都确认）的明细表才能发放
    """
    from app.services.bonus import BonusCalculator
    from app.services.bonus_distribution_service import (
        check_distribution_exists,
        create_calculation_from_team_allocation,
        create_distribution_record,
        validate_sheet_for_distribution,
    )

    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")

    # 验证明细表
    is_valid, error_msg = validate_sheet_for_distribution(sheet)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    valid_rows = sheet.parse_result['valid_rows']

    # 批量创建发放记录
    distributions = []
    errors = []
    calculator = BonusCalculator(db)

    for row_data in valid_rows:
        try:
            calculation = None
            calculation_id = row_data.get('calculation_id')
            team_allocation_id = row_data.get('team_allocation_id')

            # 如果使用团队奖金分配ID，先创建个人计算记录
            if team_allocation_id:
                try:
                    calculation = create_calculation_from_team_allocation(
                        db, team_allocation_id, row_data['user_id'],
                        Decimal(str(row_data['calculated_amount'])), calculator
                    )
                    calculation_id = calculation.id
                except ValueError as e:
                    errors.append(str(e))
                    continue
            else:
                # 使用已有的计算记录
                calculation = db.query(BonusCalculation).filter(
                    BonusCalculation.id == calculation_id
                ).first()
                if not calculation:
                    errors.append(f"计算记录ID {calculation_id} 不存在")
                    continue

            # 检查是否已发放
            if check_distribution_exists(db, calculation_id, row_data['user_id']):
                errors.append(f"计算记录ID {calculation_id} 对用户ID {row_data['user_id']} 已发放")
                continue

            # 创建发放记录
            distribution = create_distribution_record(
                db, calculation_id, row_data['user_id'], row_data,
                current_user.id, generate_distribution_code
            )
            distributions.append(distribution)

            # 更新计算记录状态
            if calculation:
                calculation.status = 'DISTRIBUTED'

        except Exception as e:
            errors.append(f"处理行数据失败: {str(e)}")
            continue

    if errors and not distributions:
        raise HTTPException(
            status_code=400,
            detail=f"发放失败：{'; '.join(errors[:5])}"  # 只显示前5个错误
        )

    # 更新明细表状态
    sheet.status = 'DISTRIBUTED'
    sheet.distributed_at = datetime.now()
    sheet.distributed_by = current_user.id
    sheet.distribution_count = len(distributions)

    db.commit()

    return ResponseModel(
        code=200,
        message=f"发放成功，共创建 {len(distributions)} 条发放记录" + (f"，{len(errors)} 条失败" if errors else ""),
        data={
            "sheet_id": sheet_id,
            "sheet_code": sheet.sheet_code,
            "distributed_count": len(distributions),
            "error_count": len(errors),
            "errors": errors[:10] if errors else []  # 最多返回10个错误
        }
    )


@router.get("/allocation-sheets/{sheet_id}/download", response_class=FileResponse)
def download_allocation_sheet(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    下载已上传的分配明细表Excel文件，便于复核和留档
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")

    if not sheet.file_path:
        raise HTTPException(status_code=404, detail="明细表文件不存在")

    upload_dir = os.path.abspath(settings.UPLOAD_DIR)
    file_path = os.path.abspath(os.path.join(settings.UPLOAD_DIR, sheet.file_path))
    if not file_path.startswith(upload_dir):
        raise HTTPException(status_code=400, detail="文件路径非法，拒绝下载")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件已被删除或不存在")

    filename = sheet.file_name or os.path.basename(file_path)
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get(
    "/allocation-sheets/{sheet_id}/rows",
    response_model=ResponseModel,
    status_code=status.HTTP_200_OK
)
def get_allocation_sheet_rows(
    *,
    db: Session = Depends(deps.get_db),
    sheet_id: int,
    row_type: str = Query("valid", description="数据类型：valid 或 error"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(50, ge=1, le=200, description="每页条数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    查看已解析的分配明细表数据，支持分页查看有效行或错误行
    """
    sheet = db.query(BonusAllocationSheet).filter(BonusAllocationSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="分配明细表不存在")

    normalized_type = row_type.lower()
    if normalized_type not in {"valid", "error"}:
        raise HTTPException(status_code=400, detail="row_type 仅支持 valid 或 error")

    if normalized_type == "valid":
        valid_rows = []
        if sheet.parse_result and sheet.parse_result.get('valid_rows'):
            for row_data in sheet.parse_result['valid_rows']:
                distribution_date = row_data.get('distribution_date')
                if isinstance(distribution_date, str):
                    try:
                        distribution_date = date.fromisoformat(distribution_date)
                    except ValueError:
                        distribution_date = datetime.fromisoformat(distribution_date).date()
                elif isinstance(distribution_date, datetime):
                    distribution_date = distribution_date.date()
                elif not isinstance(distribution_date, date):
                    try:
                        distribution_date = datetime.fromisoformat(str(distribution_date)).date()
                    except ValueError:
                        distribution_date = datetime.strptime(str(distribution_date), '%Y-%m-%d').date()

                valid_rows.append(
                    BonusAllocationRow(
                        calculation_id=int(row_data['calculation_id']),
                        user_id=int(row_data['user_id']),
                        user_name=row_data.get('user_name'),
                        calculated_amount=Decimal(str(row_data['calculated_amount'])),
                        distributed_amount=Decimal(str(row_data['distributed_amount'])),
                        distribution_date=distribution_date,
                        payment_method=row_data.get('payment_method'),
                        voucher_no=row_data.get('voucher_no'),
                        payment_account=row_data.get('payment_account'),
                        payment_remark=row_data.get('payment_remark')
                    ).model_dump()
                )

        page_items, total, pages = paginate_items(valid_rows, page, page_size)
        data = {
            "sheet_id": sheet_id,
            "row_type": normalized_type,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "items": page_items
        }
    else:
        error_rows = []
        errors = sheet.parse_errors or {}
        for row_no, messages in errors.items():
            try:
                row_number = int(row_no)
            except (TypeError, ValueError):
                row_number = row_no
            if not isinstance(messages, list):
                messages = [str(messages)]
            error_rows.append({
                "row_number": row_number,
                "errors": [str(m) for m in messages]
            })

        error_rows.sort(key=lambda x: x["row_number"])
        page_items, total, pages = paginate_items(error_rows, page, page_size)
        data = {
            "sheet_id": sheet_id,
            "row_type": normalized_type,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "items": page_items
        }

    return ResponseModel(code=200, data=data)

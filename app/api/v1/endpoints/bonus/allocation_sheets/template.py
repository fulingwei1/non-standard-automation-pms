# -*- coding: utf-8 -*-
"""
奖金分配明细表 - 模板下载
"""
import io
import os
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse

from app.core import security
from app.core.config import settings
from app.models.user import User

router = APIRouter()


@router.get("/allocation-sheets/template", response_class=FileResponse)
def download_allocation_template(
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    下载奖金分配明细表模板（Excel）
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "奖金分配明细表"

    # 设置表头（支持两种模式）
    headers = [
        "计算记录ID",
        "团队奖金分配ID",
        "受益人ID*",
        "受益人姓名",
        "计算金额",
        "发放金额*",
        "发放日期*",
        "发放方式",
        "凭证号",
        "付款账户",
        "付款备注"
    ]

    # 写入表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # 添加说明行
    ws.insert_rows(1)
    ws.merge_cells(f'A1:K1')
    note_cell = ws.cell(row=1, column=1, value="说明：1. 带*的列为必填项；2. 必须提供'计算记录ID'或'团队奖金分配ID'之一；3. 如果使用团队奖金分配ID，系统会自动创建个人计算记录；4. 受益人ID必须为数字；5. 金额必须为数字；6. 发放日期格式：YYYY-MM-DD")
    note_cell.font = Font(size=10, italic=True)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 创建临时文件
    template_dir = os.path.join(settings.UPLOAD_DIR, "templates")
    os.makedirs(template_dir, exist_ok=True)
    template_path = os.path.join(template_dir, "奖金分配明细表模板.xlsx")
    with open(template_path, "wb") as f:
        f.write(output.getvalue())

    return FileResponse(
        path=template_path,
        filename="奖金分配明细表模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

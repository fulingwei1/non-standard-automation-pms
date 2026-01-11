# -*- coding: utf-8 -*-
"""
数据导入导出 API endpoints
核心功能：Excel模板导入、数据验证、多类型导出
"""

from typing import Any, List, Optional, Dict
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Body, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
import io

from app.api import deps
from app.core.config import settings
from app.core import security
from app.models.user import User
from app.models.project import Project
from app.models.report_center import (
    DataImportTask, DataExportTask, ImportTemplate
)
from app.schemas.common import ResponseModel, PaginatedResponse
from app.schemas.data_import_export import (
    ImportTemplateTypeResponse, ImportPreviewRequest, ImportPreviewResponse,
    ImportValidateRequest, ImportValidateResponse,
    ImportUploadRequest, ImportUploadResponse,
    ExportProjectListRequest, ExportProjectDetailRequest,
    ExportTaskListRequest, ExportTimesheetRequest, ExportWorkloadRequest
)

router = APIRouter()


# ==================== 数据导入 ====================

# ==================== 导入验证器 ====================

def _validate_project_row(row, row_num, errors):
    """验证项目导入行"""
    is_valid = True
    project_code = str(row.get('项目编码*', row.get('项目编码', ''))).strip()
    project_name = str(row.get('项目名称*', row.get('项目名称', ''))).strip()

    if not project_code:
        errors.append({"row": row_num, "field": "项目编码", "message": "项目编码不能为空"})
        is_valid = False
    if not project_name:
        errors.append({"row": row_num, "field": "项目名称", "message": "项目名称不能为空"})
        is_valid = False
    return is_valid


def _validate_user_row(row, row_num, errors):
    """验证用户导入行"""
    name = str(row.get('姓名', row.get('名字', ''))).strip()
    if not name:
        errors.append({"row": row_num, "field": "姓名", "message": "姓名不能为空"})
        return False
    return True


def _validate_timesheet_row(row, row_num, errors, pd):
    """验证工时导入行"""
    is_valid = True
    work_date = row.get('工作日期*') or row.get('工作日期')
    user_name = str(row.get('人员姓名*', row.get('人员姓名', ''))).strip()
    hours = row.get('工时(小时)*') or row.get('工时(小时)') or row.get('工时')

    if pd.isna(work_date) or not work_date:
        errors.append({"row": row_num, "field": "工作日期", "message": "工作日期不能为空"})
        is_valid = False
    if not user_name:
        errors.append({"row": row_num, "field": "人员姓名", "message": "人员姓名不能为空"})
        is_valid = False
    if pd.isna(hours):
        errors.append({"row": row_num, "field": "工时", "message": "工时不能为空"})
        is_valid = False
    return is_valid


def _validate_task_row(row, row_num, errors):
    """验证任务导入行"""
    is_valid = True
    task_name = str(row.get('任务名称*', row.get('任务名称', ''))).strip()
    project_code = str(row.get('项目编码*', row.get('项目编码', ''))).strip()

    if not task_name:
        errors.append({"row": row_num, "field": "任务名称", "message": "任务名称不能为空"})
        is_valid = False
    if not project_code:
        errors.append({"row": row_num, "field": "项目编码", "message": "项目编码不能为空"})
        is_valid = False
    return is_valid


def _validate_material_row(row, row_num, errors):
    """验证物料导入行"""
    is_valid = True
    material_code = str(row.get('物料编码*', row.get('物料编码', ''))).strip()
    material_name = str(row.get('物料名称*', row.get('物料名称', ''))).strip()

    if not material_code:
        errors.append({"row": row_num, "field": "物料编码", "message": "物料编码不能为空"})
        is_valid = False
    if not material_name:
        errors.append({"row": row_num, "field": "物料名称", "message": "物料名称不能为空"})
        is_valid = False
    return is_valid


def _validate_bom_row(row, row_num, errors, pd):
    """验证BOM导入行"""
    is_valid = True
    bom_code = str(row.get('BOM编码*', row.get('BOM编码', ''))).strip()
    project_code = str(row.get('项目编码*', row.get('项目编码', ''))).strip()
    material_code = str(row.get('物料编码*', row.get('物料编码', ''))).strip()
    quantity = row.get('用量*') or row.get('用量')

    if not bom_code:
        errors.append({"row": row_num, "field": "BOM编码", "message": "BOM编码不能为空"})
        is_valid = False
    if not project_code:
        errors.append({"row": row_num, "field": "项目编码", "message": "项目编码不能为空"})
        is_valid = False
    if not material_code:
        errors.append({"row": row_num, "field": "物料编码", "message": "物料编码不能为空"})
        is_valid = False
    if pd.isna(quantity):
        errors.append({"row": row_num, "field": "用量", "message": "用量不能为空"})
        is_valid = False
    return is_valid


def _validate_import_row(row, row_num, template_type, errors, pd):
    """根据模板类型验证导入行"""
    validators = {
        "PROJECT": lambda: _validate_project_row(row, row_num, errors),
        "USER": lambda: _validate_user_row(row, row_num, errors),
        "TIMESHEET": lambda: _validate_timesheet_row(row, row_num, errors, pd),
        "TASK": lambda: _validate_task_row(row, row_num, errors),
        "MATERIAL": lambda: _validate_material_row(row, row_num, errors),
        "BOM": lambda: _validate_bom_row(row, row_num, errors, pd),
    }
    validator = validators.get(template_type)
    return validator() if validator else True


@router.get("/templates", response_model=ImportTemplateTypeResponse, status_code=status.HTTP_200_OK)
def get_import_template_types(
    *,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    获取所有模板类型（项目/任务/人员/工时等）
    """
    types = [
        {"type": "PROJECT", "name": "项目", "description": "项目基本信息导入"},
        {"type": "TASK", "name": "任务", "description": "任务数据导入"},
        {"type": "USER", "name": "人员", "description": "人员信息导入"},
        {"type": "TIMESHEET", "name": "工时", "description": "工时数据导入"},
        {"type": "MATERIAL", "name": "物料", "description": "物料信息导入"},
        {"type": "BOM", "name": "BOM", "description": "BOM数据导入"}
    ]
    
    return ImportTemplateTypeResponse(types=types)


@router.get("/templates/{template_type}", response_class=StreamingResponse)
def download_import_template(
    *,
    db: Session = Depends(deps.get_db),
    template_type: str,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    下载导入模板（按类型下载）
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 使用模板生成服务
    from app.services.excel_template_service import (
        get_template_config,
        create_template_excel
    )
    
    config = get_template_config(template_type)
    if not config:
        raise HTTPException(status_code=400, detail=f"不支持的模板类型: {template_type}")
    
    return create_template_excel(
        template_data=config["template_data"],
        sheet_name=config["sheet_name"],
        column_widths=config["column_widths"],
        instructions=config["instructions"],
        filename_prefix=config["filename_prefix"]
    )


@router.post("/preview", response_model=ImportPreviewResponse, status_code=status.HTTP_200_OK)
def preview_import_data(
    *,
    db: Session = Depends(deps.get_db),
    file: UploadFile = File(...),
    template_type: str = Query(..., description="模板类型"),
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    预览导入数据（上传预览）

    支持多种数据类型预览：
    - PROJECT: 项目导入
    - USER: 用户导入
    - TIMESHEET: 工时导入
    - TASK: 任务导入
    - MATERIAL: 物料导入
    - BOM: BOM导入
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas"
        )

    # 各模板类型的必需列定义
    REQUIRED_COLUMNS = {
        "PROJECT": ['项目编码*', '项目名称*'],
        "USER": ['姓名'],
        "TIMESHEET": ['工作日期*', '人员姓名*', '工时(小时)*'],
        "TASK": ['任务名称*', '项目编码*'],
        "MATERIAL": ['物料编码*', '物料名称*'],
        "BOM": ['BOM编码*', '项目编码*', '物料编码*', '用量*']
    }

    try:
        file_content = file.file.read()
        df = pd.read_excel(io.BytesIO(file_content))

        # 去除空行
        df = df.dropna(how='all')
        total_rows = len(df)

        if total_rows == 0:
            return ImportPreviewResponse(
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                preview_data=[],
                errors=[{"row": 0, "field": "", "message": "文件中没有数据"}]
            )

        template_type_upper = template_type.upper()

        # 验证必需列
        required_columns = REQUIRED_COLUMNS.get(template_type_upper, [])
        missing_columns = []

        for req_col in required_columns:
            if req_col not in df.columns and req_col.replace('*', '') not in df.columns:
                missing_columns.append(req_col)

        if missing_columns:
            return ImportPreviewResponse(
                total_rows=total_rows,
                valid_rows=0,
                invalid_rows=total_rows,
                preview_data=[],
                errors=[{"row": 0, "field": "", "message": f"缺少必需的列：{', '.join(missing_columns)}"}]
            )

        # 预览前10行数据
        preview_rows = min(10, total_rows)
        preview_data = df.head(preview_rows).to_dict('records')

        # 根据类型进行验证
        errors = []
        valid_rows = 0

        for idx, row in df.iterrows():
            row_num = idx + 2
            is_valid = _validate_import_row(row, row_num, template_type_upper, errors, pd)
            if is_valid:
                valid_rows += 1

        return ImportPreviewResponse(
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=total_rows - valid_rows,
            preview_data=preview_data[:preview_rows],
            errors=errors[:20]  # 最多返回20个错误
        )

    except Exception as e:
        return ImportPreviewResponse(
            total_rows=0,
            valid_rows=0,
            invalid_rows=0,
            preview_data=[],
            errors=[{"row": 0, "field": "", "message": f"文件解析失败: {str(e)}"}]
        )


@router.post("/validate", response_model=ImportValidateResponse, status_code=status.HTTP_200_OK)
def validate_import_data(
    *,
    db: Session = Depends(deps.get_db),
    validate_in: ImportValidateRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    验证导入数据（格式校验）
    验证数据格式和业务规则

    支持多种数据类型验证：
    - PROJECT: 项目导入
    - USER: 用户导入
    - TIMESHEET: 工时导入
    - TASK: 任务导入
    - MATERIAL: 物料导入
    - BOM: BOM导入
    """
    errors = []
    valid_count = 0
    template_type = validate_in.template_type.upper()

    for idx, row_data in enumerate(validate_in.data, start=1):
        row_errors = []

        if template_type == "PROJECT":
            # 验证必填字段
            project_code = row_data.get('project_code', '').strip() if row_data.get('project_code') else ''
            project_name = row_data.get('project_name', '').strip() if row_data.get('project_name') else ''

            if not project_code:
                row_errors.append({"field": "project_code", "message": "项目编码不能为空"})
            if not project_name:
                row_errors.append({"field": "project_name", "message": "项目名称不能为空"})

            # 检查项目编码是否已存在
            if project_code:
                existing = db.query(Project).filter(Project.project_code == project_code).first()
                if existing:
                    row_errors.append({"field": "project_code", "message": f"项目编码 {project_code} 已存在"})

            # 验证日期格式
            for date_field in ['planned_start_date', 'planned_end_date']:
                if row_data.get(date_field):
                    try:
                        datetime.strptime(str(row_data[date_field]), '%Y-%m-%d')
                    except:
                        row_errors.append({"field": date_field, "message": "日期格式错误，应为YYYY-MM-DD"})

            # 验证日期逻辑
            if row_data.get('planned_start_date') and row_data.get('planned_end_date'):
                try:
                    start_date = datetime.strptime(str(row_data['planned_start_date']), '%Y-%m-%d').date()
                    end_date = datetime.strptime(str(row_data['planned_end_date']), '%Y-%m-%d').date()
                    if start_date > end_date:
                        row_errors.append({"field": "planned_end_date", "message": "计划结束日期不能早于计划开始日期"})
                except:
                    pass

            # 验证金额格式
            for amount_field in ['contract_amount', 'budget_amount']:
                if row_data.get(amount_field):
                    try:
                        float(row_data[amount_field])
                    except:
                        row_errors.append({"field": amount_field, "message": f"{amount_field} 必须是数字"})

        elif template_type == "USER":
            name = row_data.get('name', '').strip() if row_data.get('name') else ''
            if not name:
                row_errors.append({"field": "name", "message": "姓名不能为空"})

        elif template_type == "TIMESHEET":
            work_date = row_data.get('work_date')
            user_name = row_data.get('user_name', '').strip() if row_data.get('user_name') else ''
            hours = row_data.get('hours')

            if not work_date:
                row_errors.append({"field": "work_date", "message": "工作日期不能为空"})
            else:
                try:
                    datetime.strptime(str(work_date), '%Y-%m-%d')
                except:
                    row_errors.append({"field": "work_date", "message": "日期格式错误，应为YYYY-MM-DD"})

            if not user_name:
                row_errors.append({"field": "user_name", "message": "人员姓名不能为空"})

            if hours is None:
                row_errors.append({"field": "hours", "message": "工时不能为空"})
            else:
                try:
                    h = float(hours)
                    if h <= 0 or h > 24:
                        row_errors.append({"field": "hours", "message": "工时必须在0-24之间"})
                except:
                    row_errors.append({"field": "hours", "message": "工时格式错误"})

        elif template_type == "TASK":
            task_name = row_data.get('task_name', '').strip() if row_data.get('task_name') else ''
            project_code = row_data.get('project_code', '').strip() if row_data.get('project_code') else ''

            if not task_name:
                row_errors.append({"field": "task_name", "message": "任务名称不能为空"})
            if not project_code:
                row_errors.append({"field": "project_code", "message": "项目编码不能为空"})

            # 检查项目是否存在
            if project_code:
                project = db.query(Project).filter(Project.project_code == project_code).first()
                if not project:
                    row_errors.append({"field": "project_code", "message": f"项目 {project_code} 不存在"})

        elif template_type == "MATERIAL":
            material_code = row_data.get('material_code', '').strip() if row_data.get('material_code') else ''
            material_name = row_data.get('material_name', '').strip() if row_data.get('material_name') else ''

            if not material_code:
                row_errors.append({"field": "material_code", "message": "物料编码不能为空"})
            if not material_name:
                row_errors.append({"field": "material_name", "message": "物料名称不能为空"})

            # 检查物料编码是否已存在
            if material_code:
                from app.models.material import Material
                existing = db.query(Material).filter(Material.material_code == material_code).first()
                if existing:
                    row_errors.append({"field": "material_code", "message": f"物料编码 {material_code} 已存在"})

        elif template_type == "BOM":
            bom_code = row_data.get('bom_code', '').strip() if row_data.get('bom_code') else ''
            project_code = row_data.get('project_code', '').strip() if row_data.get('project_code') else ''
            material_code = row_data.get('material_code', '').strip() if row_data.get('material_code') else ''
            quantity = row_data.get('quantity')

            if not bom_code:
                row_errors.append({"field": "bom_code", "message": "BOM编码不能为空"})
            if not project_code:
                row_errors.append({"field": "project_code", "message": "项目编码不能为空"})
            if not material_code:
                row_errors.append({"field": "material_code", "message": "物料编码不能为空"})
            if quantity is None:
                row_errors.append({"field": "quantity", "message": "用量不能为空"})
            else:
                try:
                    q = float(quantity)
                    if q <= 0:
                        row_errors.append({"field": "quantity", "message": "用量必须大于0"})
                except:
                    row_errors.append({"field": "quantity", "message": "用量格式错误"})

            # 检查项目和物料是否存在
            if project_code:
                project = db.query(Project).filter(Project.project_code == project_code).first()
                if not project:
                    row_errors.append({"field": "project_code", "message": f"项目 {project_code} 不存在"})

            if material_code:
                from app.models.material import Material
                material = db.query(Material).filter(Material.material_code == material_code).first()
                if not material:
                    row_errors.append({"field": "material_code", "message": f"物料 {material_code} 不存在"})

        if row_errors:
            errors.append({
                "row": idx,
                "errors": row_errors
            })
        else:
            valid_count += 1

    return ImportValidateResponse(
        is_valid=len(errors) == 0,
        valid_count=valid_count,
        invalid_count=len(validate_in.data) - valid_count,
        errors=errors
    )


@router.post("/upload", response_model=ImportUploadResponse, status_code=status.HTTP_201_CREATED)
def upload_and_import_data(
    *,
    db: Session = Depends(deps.get_db),
    file: UploadFile = File(...),
    template_type: str = Query(..., description="模板类型"),
    update_existing: bool = Query(False, description="是否更新已存在的数据"),
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    上传并导入数据（执行导入）

    支持多种数据类型导入：
    - PROJECT: 项目导入
    - USER: 用户导入
    - TIMESHEET: 工时导入
    - TASK: 任务导入
    - MATERIAL: 物料导入
    - BOM: BOM导入
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas"
        )

    from app.services.unified_import_service import unified_import_service

    try:
        file_content = file.file.read()

        # 使用统一导入服务
        result = unified_import_service.import_data(
            db=db,
            file_content=file_content,
            filename=file.filename,
            template_type=template_type,
            current_user_id=current_user.id,
            update_existing=update_existing
        )

        db.commit()

        # 创建导入任务记录
        task_code = f"IMP-{datetime.now().strftime('%y%m%d%H%M%S')}"
        import_task = DataImportTask(
            task_code=task_code,
            template_type=template_type,
            file_url=file.filename,
            status="COMPLETED" if result.get("failed_count", 0) == 0 else "PARTIAL",
            requested_by=current_user.id,
            import_options={"update_existing": update_existing}
        )
        db.add(import_task)
        db.commit()

        # 生成消息
        imported = result.get("imported_count", 0)
        updated = result.get("updated_count", 0)
        failed = result.get("failed_count", 0)

        message = f"导入完成：成功导入 {imported} 条"
        if updated > 0:
            message += f"，更新 {updated} 条"
        if failed > 0:
            message += f"，失败 {failed} 条"

        return ImportUploadResponse(
            task_id=import_task.id,
            task_code=import_task.task_code,
            status="COMPLETED" if failed == 0 else "PARTIAL",
            message=message
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")


# ==================== 数据导出 ====================

@router.post("/export/project_list", response_class=StreamingResponse)
def export_project_list(
    *,
    db: Session = Depends(deps.get_db),
    export_in: ExportProjectListRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    导出项目列表（Excel）
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
        import openpyxl
        import io
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 构建查询
    from app.models.project import Project
    from app.services.data_scope_service import DataScopeService
    from sqlalchemy import desc, or_
    
    query = db.query(Project).filter(Project.is_active == True)
    
    # 应用过滤条件
    filters = export_in.filters or {}
    if filters.get('keyword'):
        keyword = filters['keyword']
        query = query.filter(
            or_(
                Project.project_name.contains(keyword),
                Project.project_code.contains(keyword),
                Project.contract_no.contains(keyword),
            )
        )
    if filters.get('customer_id'):
        query = query.filter(Project.customer_id == filters['customer_id'])
    if filters.get('stage'):
        query = query.filter(Project.stage == filters['stage'])
    if filters.get('status'):
        query = query.filter(Project.status == filters['status'])
    if filters.get('health'):
        query = query.filter(Project.health == filters['health'])
    if filters.get('pm_id'):
        query = query.filter(Project.pm_id == filters['pm_id'])
    if filters.get('project_type'):
        query = query.filter(Project.project_type == filters['project_type'])
    
    # 应用数据权限过滤
    query = DataScopeService.filter_projects_by_scope(db, query, current_user)
    
    # 获取项目列表
    projects = query.order_by(desc(Project.created_at)).all()
    
    # 阶段名称映射
    stage_names = {
        'S1': '需求进入', 'S2': '方案设计', 'S3': '采购备料', 'S4': '加工制造',
        'S5': '装配调试', 'S6': '出厂验收(FAT)', 'S7': '包装发运', 'S8': '现场安装(SAT)', 'S9': '质保结项'
    }
    
    # 健康度名称映射
    health_names = {
        'H1': '正常(绿色)', 'H2': '有风险(黄色)', 'H3': '阻塞(红色)', 'H4': '已完结(灰色)'
    }
    
    # 构建DataFrame
    data = []
    for project in projects:
        data.append({
            '项目编码': project.project_code or '',
            '项目名称': project.project_name or '',
            '客户名称': project.customer_name or '',
            '合同编号': project.contract_no or '',
            '合同金额': float(project.contract_amount or 0),
            '项目经理': project.pm_name or '',
            '项目类型': project.project_type or '',
            '阶段': project.stage or '',
            '阶段名称': stage_names.get(project.stage, project.stage or ''),
            '状态': project.status or '',
            '健康度': project.health or '',
            '健康度名称': health_names.get(project.health, project.health or '') if project.health else '',
            '进度(%)': float(project.progress_pct or 0),
            '计划开始日期': project.planned_start_date.strftime('%Y-%m-%d') if project.planned_start_date else '',
            '计划结束日期': project.planned_end_date.strftime('%Y-%m-%d') if project.planned_end_date else '',
            '实际开始日期': project.actual_start_date.strftime('%Y-%m-%d') if project.actual_start_date else '',
            '实际结束日期': project.actual_end_date.strftime('%Y-%m-%d') if project.actual_end_date else '',
            '创建时间': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
            '更新时间': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else '',
        })
    
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='项目列表', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['项目列表']
        column_widths = {
            'A': 15,  # 项目编码
            'B': 30,  # 项目名称
            'C': 20,  # 客户名称
            'D': 15,  # 合同编号
            'E': 12,  # 合同金额
            'F': 12,  # 项目经理
            'G': 12,  # 项目类型
            'H': 8,   # 阶段
            'I': 15,  # 阶段名称
            'J': 10,  # 状态
            'K': 8,   # 健康度
            'L': 15,  # 健康度名称
            'M': 10,  # 进度(%)
            'N': 12,  # 计划开始日期
            'O': 12,  # 计划结束日期
            'P': 12,  # 实际开始日期
            'Q': 12,  # 实际结束日期
            'R': 18,  # 创建时间
            'S': 18,  # 更新时间
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    output.seek(0)
    
    # 生成文件名
    from datetime import datetime
    filename = f"项目列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.post("/export/project_detail", response_model=ResponseModel, status_code=status.HTTP_200_OK)
def export_project_detail(
    *,
    db: Session = Depends(deps.get_db),
    export_in: ExportProjectDetailRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    导出项目详情（含任务/成本）
    """
    from app.utils.permission_helpers import check_project_access_or_raise
    from app.services.project_export_service import create_project_detail_excel
    
    # 检查项目访问权限
    project = check_project_access_or_raise(db, current_user, export_in.project_id)
    
    # 创建Excel文件
    try:
        output = create_project_detail_excel(
            db, project, export_in.include_tasks, export_in.include_costs
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装openpyxl"
        )
    
    # 生成文件名
    filename = f"项目详情_{project.project_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.post("/export/task_list", response_class=StreamingResponse)
def export_task_list(
    *,
    db: Session = Depends(deps.get_db),
    export_in: ExportTaskListRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    导出任务列表（Excel）
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
        import openpyxl
        import io
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 导入Task模型
    from app.models.progress import Task
    from app.services.data_scope_service import DataScopeService
    from sqlalchemy import desc, or_
    
    # 构建查询
    query = db.query(Task)
    
    # 应用过滤条件
    filters = export_in.filters or {}
    if filters.get('project_id'):
        query = query.filter(Task.project_id == filters['project_id'])
    if filters.get('machine_id'):
        query = query.filter(Task.machine_id == filters['machine_id'])
    if filters.get('stage'):
        query = query.filter(Task.stage == filters['stage'])
    if filters.get('status'):
        query = query.filter(Task.status == filters['status'])
    if filters.get('owner_id'):
        query = query.filter(Task.owner_id == filters['owner_id'])
    if filters.get('keyword'):
        keyword = filters['keyword']
        query = query.filter(Task.task_name.contains(keyword))
    
    # 应用数据权限过滤（通过项目范围限制任务）
    scoped_project_query = db.query(Project.id)
    if filters.get('project_id'):
        scoped_project_query = scoped_project_query.filter(Project.id == filters['project_id'])
    scoped_project_query = DataScopeService.filter_projects_by_scope(
        db,
        scoped_project_query,
        current_user
    )
    allowed_projects_subquery = scoped_project_query.subquery()
    query = query.filter(Task.project_id.in_(allowed_projects_subquery))
    
    # 获取任务列表
    tasks = query.order_by(desc(Task.created_at)).all()
    
    # 状态名称映射
    status_names = {
        'TODO': '待办',
        'IN_PROGRESS': '进行中',
        'BLOCKED': '阻塞',
        'DONE': '已完成',
        'CANCELLED': '已取消'
    }
    
    # 构建DataFrame
    data = []
    for task in tasks:
        project = db.query(Project).filter(Project.id == task.project_id).first()
        owner = db.query(User).filter(User.id == task.owner_id).first() if task.owner_id else None
        
        data.append({
            '任务ID': task.id,
            '任务名称': task.task_name or '',
            '项目编码': project.project_code if project else '',
            '项目名称': project.project_name if project else '',
            '阶段': task.stage or '',
            '状态': task.status or '',
            '状态名称': status_names.get(task.status, task.status or ''),
            '负责人': owner.real_name or owner.username if owner else '',
            '计划开始日期': task.plan_start.strftime('%Y-%m-%d') if task.plan_start else '',
            '计划结束日期': task.plan_end.strftime('%Y-%m-%d') if task.plan_end else '',
            '实际开始日期': task.actual_start.strftime('%Y-%m-%d') if task.actual_start else '',
            '实际结束日期': task.actual_end.strftime('%Y-%m-%d') if task.actual_end else '',
            '进度(%)': task.progress_percent or 0,
            '权重': float(task.weight or 0),
            '阻塞原因': task.block_reason or '',
            '创建时间': task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else '',
            '更新时间': task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else '',
        })
    
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='任务列表', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['任务列表']
        column_widths = {
            'A': 10,  # 任务ID
            'B': 30,  # 任务名称
            'C': 15,  # 项目编码
            'D': 30,  # 项目名称
            'E': 8,   # 阶段
            'F': 12,  # 状态
            'G': 12,  # 状态名称
            'H': 12,  # 负责人
            'I': 12,  # 计划开始日期
            'J': 12,  # 计划结束日期
            'K': 12,  # 实际开始日期
            'L': 12,  # 实际结束日期
            'M': 10,  # 进度(%)
            'N': 8,   # 权重
            'O': 40,  # 阻塞原因
            'P': 18,  # 创建时间
            'Q': 18,  # 更新时间
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    output.seek(0)
    
    # 生成文件名
    filename = f"任务列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.post("/export/timesheet", response_class=StreamingResponse)
def export_timesheet(
    *,
    db: Session = Depends(deps.get_db),
    export_in: ExportTimesheetRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    导出工时数据（按日期范围，Excel）
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
        import openpyxl
        import io
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 导入Timesheet模型
    from app.models.timesheet import Timesheet
    from sqlalchemy import desc, and_
    
    # 构建查询
    query = db.query(Timesheet).filter(
        and_(
            Timesheet.work_date >= export_in.start_date,
            Timesheet.work_date <= export_in.end_date
        )
    )
    
    # 应用过滤条件
    filters = export_in.filters or {}
    if filters.get('user_id'):
        query = query.filter(Timesheet.user_id == filters['user_id'])
    if filters.get('project_id'):
        query = query.filter(Timesheet.project_id == filters['project_id'])
    if filters.get('department_id'):
        query = query.filter(Timesheet.department_id == filters['department_id'])
    if filters.get('status'):
        query = query.filter(Timesheet.status == filters['status'])
    
    # 获取工时记录
    timesheets = query.order_by(Timesheet.work_date, Timesheet.user_id).all()
    
    # 状态名称映射
    status_names = {
        'DRAFT': '草稿',
        'SUBMITTED': '已提交',
        'APPROVED': '已通过',
        'REJECTED': '已驳回',
        'CANCELLED': '已取消'
    }
    
    overtime_type_names = {
        'NORMAL': '正常工时',
        'OVERTIME': '加班',
        'WEEKEND': '周末加班',
        'HOLIDAY': '节假日加班'
    }
    
    # 构建DataFrame
    data = []
    for ts in timesheets:
        data.append({
            '工作日期': ts.work_date.strftime('%Y-%m-%d') if ts.work_date else '',
            '人员姓名': ts.user_name or '',
            '部门': ts.department_name or '',
            '项目编码': ts.project_code or '',
            '项目名称': ts.project_name or '',
            '任务名称': ts.task_name or '',
            '工时(小时)': float(ts.hours or 0),
            '加班类型': ts.overtime_type or '',
            '加班类型名称': overtime_type_names.get(ts.overtime_type, ts.overtime_type or ''),
            '工作内容': ts.work_content or '',
            '工作成果': ts.work_result or '',
            '更新前进度(%)': ts.progress_before or 0,
            '更新后进度(%)': ts.progress_after or 0,
            '状态': ts.status or '',
            '状态名称': status_names.get(ts.status, ts.status or ''),
            '提交时间': ts.submit_time.strftime('%Y-%m-%d %H:%M:%S') if ts.submit_time else '',
            '审核人': ts.approver_name or '',
            '审核时间': ts.approve_time.strftime('%Y-%m-%d %H:%M:%S') if ts.approve_time else '',
            '审核意见': ts.approve_comment or '',
        })
    
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='工时数据', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['工时数据']
        column_widths = {
            'A': 12,  # 工作日期
            'B': 12,  # 人员姓名
            'C': 15,  # 部门
            'D': 15,  # 项目编码
            'E': 30,  # 项目名称
            'F': 30,  # 任务名称
            'G': 12,  # 工时(小时)
            'H': 12,  # 加班类型
            'I': 15,  # 加班类型名称
            'J': 40,  # 工作内容
            'K': 40,  # 工作成果
            'L': 12,  # 更新前进度(%)
            'M': 12,  # 更新后进度(%)
            'N': 12,  # 状态
            'O': 12,  # 状态名称
            'P': 18,  # 提交时间
            'Q': 12,  # 审核人
            'R': 18,  # 审核时间
            'S': 40,  # 审核意见
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    output.seek(0)
    
    # 生成文件名
    filename = f"工时数据_{export_in.start_date.strftime('%Y%m%d')}_{export_in.end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.post("/export/workload", response_class=StreamingResponse)
def export_workload(
    *,
    db: Session = Depends(deps.get_db),
    export_in: ExportWorkloadRequest,
    current_user: User = Depends(security.require_permission("data_import_export:manage")),
) -> Any:
    """
    导出负荷数据（Excel）
    基于资源分配和工时数据计算人员负荷
    """
    # 检查Excel库是否可用
    try:
        import pandas as pd
        import openpyxl
        import io
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel处理库未安装，请安装pandas和openpyxl"
        )
    
    # 导入相关模型
    from app.models.pmo import PmoResourceAllocation
    from app.models.timesheet import Timesheet
    from sqlalchemy import func, and_
    
    # 获取日期范围内的所有用户
    users = db.query(User).filter(User.is_active == True).all()
    
    # 构建负荷数据
    data = []
    for user in users:
        # 计算该用户在日期范围内的工时
        timesheets = db.query(Timesheet).filter(
            and_(
                Timesheet.user_id == user.id,
                Timesheet.work_date >= export_in.start_date,
                Timesheet.work_date <= export_in.end_date,
                Timesheet.status == 'APPROVED'
            )
        ).all()
        
        total_hours = sum(float(ts.hours or 0) for ts in timesheets)
        
        # 计算工作日数
        from datetime import timedelta
        work_days = 0
        current_date = export_in.start_date
        while current_date <= export_in.end_date:
            # 简单判断：周一到周五为工作日
            if current_date.weekday() < 5:
                work_days += 1
            current_date += timedelta(days=1)
        
        # 计算平均每日工时
        avg_daily_hours = total_hours / work_days if work_days > 0 else 0
        
        # 计算标准工时（假设每天8小时）
        standard_hours = work_days * 8
        utilization_rate = (total_hours / standard_hours * 100) if standard_hours > 0 else 0
        
        # 获取资源分配信息
        allocations = db.query(PmoResourceAllocation).filter(
            and_(
                PmoResourceAllocation.resource_id == user.id,
                PmoResourceAllocation.status.in_(['PLANNED', 'ACTIVE'])
            )
        ).all()
        
        project_count = len(set([a.project_id for a in allocations if a.project_id]))
        
        # 获取部门信息
        department_name = user.department if hasattr(user, 'department') else ''
        
        data.append({
            '人员姓名': user.real_name or user.username,
            '用户名': user.username,
            '部门': department_name,
            '总工时(小时)': round(total_hours, 2),
            '工作日数': work_days,
            '平均每日工时': round(avg_daily_hours, 2),
            '标准工时(小时)': standard_hours,
            '利用率(%)': round(utilization_rate, 2),
            '分配项目数': project_count,
            '负荷状态': '超负荷' if utilization_rate > 100 else ('正常' if utilization_rate > 80 else '空闲'),
        })
    
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='负荷数据', index=False)
        
        # 设置列宽
        worksheet = writer.sheets['负荷数据']
        column_widths = {
            'A': 12,  # 人员姓名
            'B': 15,  # 用户名
            'C': 15,  # 部门
            'D': 12,  # 总工时(小时)
            'E': 10,  # 工作日数
            'F': 12,  # 平均每日工时
            'G': 12,  # 标准工时(小时)
            'H': 12,  # 利用率(%)
            'I': 12,  # 分配项目数
            'J': 12,  # 负荷状态
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    output.seek(0)
    
    # 生成文件名
    filename = f"负荷数据_{export_in.start_date.strftime('%Y%m%d')}_{export_in.end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )

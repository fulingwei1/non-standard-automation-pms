# -*- coding: utf-8 -*-
"""
产能分析报告接口
"""
from datetime import date, timedelta
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from app.dependencies import get_db
from app.models.production import (
    Equipment,
    EquipmentOEERecord,
    Worker,
    WorkerEfficiencyRecord,
    Workshop,
)

router = APIRouter()


@router.get("/report")
async def get_capacity_report(
    workshop_id: Optional[int] = Query(None, description="车间ID"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    report_type: str = Query("summary", description="报告类型: summary/detailed/analysis"),
    format: str = Query("json", description="输出格式: json/excel"),
    db: Session = Depends(get_db),
):
    """
    产能分析报告
    
    支持生成:
    - 汇总报告: 整体产能概况
    - 详细报告: 包含所有明细数据
    - 分析报告: 包含趋势、瓶颈、建议
    """
    # 默认查询最近30天
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    # 构建过滤条件
    oee_filters = [
        EquipmentOEERecord.record_date >= start_date,
        EquipmentOEERecord.record_date <= end_date,
    ]
    
    efficiency_filters = [
        WorkerEfficiencyRecord.record_date >= start_date,
        WorkerEfficiencyRecord.record_date <= end_date,
    ]
    
    if workshop_id:
        oee_filters.append(EquipmentOEERecord.workshop_id == workshop_id)
        efficiency_filters.append(WorkerEfficiencyRecord.workshop_id == workshop_id)
    
    # 生成报告数据
    if report_type == "summary":
        report_data = _generate_summary_report(db, oee_filters, efficiency_filters)
    elif report_type == "detailed":
        report_data = _generate_detailed_report(db, oee_filters, efficiency_filters)
    elif report_type == "analysis":
        report_data = _generate_analysis_report(db, oee_filters, efficiency_filters)
    else:
        return {
            "code": 400,
            "message": "不支持的报告类型",
            "data": None,
        }
    
    # 添加报告头部信息
    report = {
        "report_info": {
            "title": f"产能分析{_get_report_type_name(report_type)}",
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
            "workshop_id": workshop_id,
            "generated_at": date.today().isoformat(),
        },
        "data": report_data,
    }
    
    # 如果需要Excel格式
    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "产能分析报告"
            
            # 写入报告头
            ws['A1'] = report['report_info']['title']
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"统计期间: {start_date.isoformat()} ~ {end_date.isoformat()}"
            
            # 写入数据(简化版)
            row = 4
            ws[f'A{row}'] = "产能分析数据"
            ws[f'A{row}'].font = Font(bold=True)
            
            # 这里可以根据实际需求填充更多Excel数据
            
            # 生成Excel文件
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return StreamingResponse(
                excel_buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=capacity_report_{start_date}_{end_date}.xlsx"
                }
            )
        except ImportError:
            return {
                "code": 500,
                "message": "Excel导出功能未启用,请安装openpyxl",
                "data": None,
            }
    
    # 返回JSON格式
    return {
        "code": 200,
        "message": "报告生成成功",
        "data": report,
    }


def _generate_summary_report(db: Session, oee_filters, efficiency_filters):
    """生成汇总报告"""
    # OEE汇总
    oee_summary = (
        db.query(
            func.count(EquipmentOEERecord.id).label('total_records'),
            func.avg(EquipmentOEERecord.oee).label('avg_oee'),
            func.sum(EquipmentOEERecord.actual_output).label('total_output'),
            func.sum(EquipmentOEERecord.qualified_qty).label('total_qualified'),
            func.sum(EquipmentOEERecord.defect_qty).label('total_defects'),
        )
        .filter(and_(*oee_filters))
        .first()
    )
    
    # 工人效率汇总
    efficiency_summary = (
        db.query(
            func.count(WorkerEfficiencyRecord.id).label('total_records'),
            func.avg(WorkerEfficiencyRecord.efficiency).label('avg_efficiency'),
            func.sum(WorkerEfficiencyRecord.completed_qty).label('total_completed'),
            func.sum(WorkerEfficiencyRecord.actual_hours).label('total_hours'),
        )
        .filter(and_(*efficiency_filters))
        .first()
    )
    
    return {
        "oee_summary": {
            "total_records": oee_summary.total_records or 0,
            "avg_oee": float(oee_summary.avg_oee) if oee_summary.avg_oee else 0,
            "total_output": oee_summary.total_output or 0,
            "total_qualified": oee_summary.total_qualified or 0,
            "total_defects": oee_summary.total_defects or 0,
            "quality_rate": round((oee_summary.total_qualified / oee_summary.total_output * 100) if oee_summary.total_output else 0, 2),
        },
        "efficiency_summary": {
            "total_records": efficiency_summary.total_records or 0,
            "avg_efficiency": float(efficiency_summary.avg_efficiency) if efficiency_summary.avg_efficiency else 0,
            "total_completed": efficiency_summary.total_completed or 0,
            "total_hours": float(efficiency_summary.total_hours) if efficiency_summary.total_hours else 0,
        },
    }


def _generate_detailed_report(db: Session, oee_filters, efficiency_filters):
    """生成详细报告"""
    # 详细的OEE数据
    oee_details = (
        db.query(
            Equipment.equipment_code,
            Equipment.equipment_name,
            func.avg(EquipmentOEERecord.oee).label('avg_oee'),
            func.avg(EquipmentOEERecord.availability).label('avg_availability'),
            func.avg(EquipmentOEERecord.performance).label('avg_performance'),
            func.avg(EquipmentOEERecord.quality).label('avg_quality'),
            func.sum(EquipmentOEERecord.actual_output).label('total_output'),
        )
        .join(Equipment, EquipmentOEERecord.equipment_id == Equipment.id)
        .filter(and_(*oee_filters))
        .group_by(Equipment.equipment_code, Equipment.equipment_name)
        .all()
    )
    
    # 详细的工人效率数据
    efficiency_details = (
        db.query(
            Worker.worker_code,
            Worker.worker_name,
            func.avg(WorkerEfficiencyRecord.efficiency).label('avg_efficiency'),
            func.avg(WorkerEfficiencyRecord.quality_rate).label('avg_quality_rate'),
            func.sum(WorkerEfficiencyRecord.completed_qty).label('total_completed'),
        )
        .join(Worker, WorkerEfficiencyRecord.worker_id == Worker.id)
        .filter(and_(*efficiency_filters))
        .group_by(Worker.worker_code, Worker.worker_name)
        .all()
    )
    
    return {
        "oee_details": [
            {
                "equipment_code": row.equipment_code,
                "equipment_name": row.equipment_name,
                "avg_oee": float(row.avg_oee) if row.avg_oee else 0,
                "avg_availability": float(row.avg_availability) if row.avg_availability else 0,
                "avg_performance": float(row.avg_performance) if row.avg_performance else 0,
                "avg_quality": float(row.avg_quality) if row.avg_quality else 0,
                "total_output": row.total_output or 0,
            }
            for row in oee_details
        ],
        "efficiency_details": [
            {
                "worker_code": row.worker_code,
                "worker_name": row.worker_name,
                "avg_efficiency": float(row.avg_efficiency) if row.avg_efficiency else 0,
                "avg_quality_rate": float(row.avg_quality_rate) if row.avg_quality_rate else 0,
                "total_completed": row.total_completed or 0,
            }
            for row in efficiency_details
        ],
    }


def _generate_analysis_report(db: Session, oee_filters, efficiency_filters):
    """生成分析报告"""
    # 包含趋势、瓶颈和改进建议
    summary = _generate_summary_report(db, oee_filters, efficiency_filters)
    
    # 识别瓶颈
    bottlenecks = (
        db.query(
            Equipment.equipment_code,
            Equipment.equipment_name,
            func.avg(EquipmentOEERecord.oee).label('avg_oee'),
            func.sum(EquipmentOEERecord.unplanned_downtime).label('total_downtime'),
        )
        .join(Equipment, EquipmentOEERecord.equipment_id == Equipment.id)
        .filter(and_(*oee_filters))
        .group_by(Equipment.equipment_code, Equipment.equipment_name)
        .having(func.avg(EquipmentOEERecord.oee) < 60)
        .all()
    )
    
    # 改进建议
    suggestions = []
    avg_oee = summary['oee_summary']['avg_oee']
    
    if avg_oee < 60:
        suggestions.append("整体OEE低于60%,需要重点改进设备可用率和性能")
    elif avg_oee < 85:
        suggestions.append("OEE处于良好水平,继续优化可达到世界级标准")
    
    if bottlenecks:
        suggestions.append(f"发现{len(bottlenecks)}个瓶颈设备,建议优先处理")
    
    return {
        **summary,
        "bottlenecks": [
            {
                "equipment_code": row.equipment_code,
                "equipment_name": row.equipment_name,
                "avg_oee": float(row.avg_oee) if row.avg_oee else 0,
                "total_downtime": row.total_downtime or 0,
            }
            for row in bottlenecks
        ],
        "improvement_suggestions": suggestions,
    }


def _get_report_type_name(report_type: str) -> str:
    """获取报告类型名称"""
    type_names = {
        "summary": "汇总报告",
        "detailed": "详细报告",
        "analysis": "分析报告",
    }
    return type_names.get(report_type, "报告")

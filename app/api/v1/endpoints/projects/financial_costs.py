# -*- coding: utf-8 -*-
"""财务项目成本兼容路由。

前端历史成本上传页使用 /projects/financial-costs；该静态路径必须在
/projects/{project_id} 动态路由之前注册。
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.api import deps
from app.common.pagination import PaginationParams, get_pagination_query
from app.common.query_filters import apply_pagination
from app.core import security
from app.models.project import FinancialProjectCost, Project
from app.models.user import User
from app.schemas.common import PaginatedResponse, ResponseModel
from app.schemas.project import FinancialProjectCostResponse

router = APIRouter()


def _build_cost_response(cost: FinancialProjectCost) -> FinancialProjectCostResponse:
    data = {column.name: getattr(cost, column.name) for column in cost.__table__.columns}
    data["tax_amount"] = data.get("tax_amount") or Decimal("0")
    data["currency"] = data.get("currency") or "CNY"
    data["source_type"] = data.get("source_type") or "FINANCIAL_UPLOAD"
    data["is_verified"] = bool(data.get("is_verified"))
    data["uploaded_by_name"] = cost.uploader.real_name if cost.uploader else None
    data["verified_by_name"] = cost.verifier.real_name if cost.verifier else None
    return FinancialProjectCostResponse(**data)


@router.get("/financial-costs", response_model=PaginatedResponse[FinancialProjectCostResponse])
def list_financial_project_costs(
    db: Session = Depends(deps.get_db),
    pagination: PaginationParams = Depends(get_pagination_query),
    project_id: Optional[int] = Query(None, description="项目ID"),
    cost_type: Optional[str] = Query(None, description="成本类型"),
    cost_category: Optional[str] = Query(None, description="成本分类"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取财务手工录入的项目成本。"""
    query = db.query(FinancialProjectCost)

    if project_id:
        query = query.filter(FinancialProjectCost.project_id == project_id)
    if cost_type:
        query = query.filter(FinancialProjectCost.cost_type == cost_type)
    if cost_category:
        query = query.filter(FinancialProjectCost.cost_category == cost_category)
    if start_date:
        query = query.filter(FinancialProjectCost.cost_date >= start_date)
    if end_date:
        query = query.filter(FinancialProjectCost.cost_date <= end_date)

    total = query.count()
    rows = apply_pagination(
        query.order_by(desc(FinancialProjectCost.cost_date), desc(FinancialProjectCost.id)),
        pagination.offset,
        pagination.limit,
    ).all()

    return PaginatedResponse(
        items=[_build_cost_response(row) for row in rows],
        total=total,
        page=pagination.page,
        page_size=pagination.page_size,
        pages=pagination.pages_for_total(total),
    )


@router.get("/financial-costs/template")
def download_financial_cost_template(
    current_user: User = Depends(security.get_current_active_user),
):
    """下载财务项目成本上传模板。"""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel模板依赖未安装") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "财务项目成本"
    sheet.append(
        [
            "project_id",
            "project_code",
            "project_name",
            "cost_type",
            "cost_category",
            "cost_item",
            "amount",
            "tax_amount",
            "currency",
            "cost_date",
            "cost_month",
            "description",
            "location",
            "participants",
            "purpose",
            "user_name",
            "hours",
            "hourly_rate",
            "source_no",
            "invoice_no",
        ]
    )
    sheet.append(
        [
            1,
            "PRJ-001",
            "示例项目",
            "TRAVEL",
            "差旅费",
            "交通费",
            1000,
            0,
            "CNY",
            date.today(),
            date.today().strftime("%Y-%m"),
            "示例说明",
            "上海",
            "张三",
            "项目出差",
            "",
            "",
            "",
            "BX-001",
            "FP-001",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="financial_cost_template.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _parse_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, ""):
        return default
    return Decimal(str(value))


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    raise ValueError("cost_date不能为空")


@router.post("/financial-costs/upload", response_model=ResponseModel)
async def upload_financial_project_costs(
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """上传财务项目成本 Excel。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel上传依赖未安装") from exc

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="只支持 .xlsx/.xlsm Excel 文件")

    workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    required = {"project_id", "cost_type", "cost_category", "amount", "cost_date"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"模板缺少字段: {', '.join(sorted(missing))}")

    created = 0
    errors = []
    batch_no = f"FC-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(headers, row))
        if not any(value not in (None, "") for value in values.values()):
            continue
        try:
            project_id = int(values["project_id"])
            project = db.query(Project).filter(Project.id == project_id).first()
            if not project:
                raise ValueError(f"项目不存在: {project_id}")

            cost_date = _parse_date(values.get("cost_date"))
            cost = FinancialProjectCost(
                project_id=project_id,
                project_code=values.get("project_code") or project.project_code,
                project_name=values.get("project_name") or project.project_name,
                cost_type=str(values.get("cost_type") or "").strip(),
                cost_category=str(values.get("cost_category") or "").strip(),
                cost_item=values.get("cost_item"),
                amount=_parse_decimal(values.get("amount")),
                tax_amount=_parse_decimal(values.get("tax_amount")),
                currency=values.get("currency") or "CNY",
                cost_date=cost_date,
                cost_month=values.get("cost_month") or cost_date.strftime("%Y-%m"),
                description=values.get("description"),
                location=values.get("location"),
                participants=values.get("participants"),
                purpose=values.get("purpose"),
                user_name=values.get("user_name"),
                hours=_parse_decimal(values.get("hours"), default=Decimal("0"))
                if values.get("hours") not in (None, "")
                else None,
                hourly_rate=_parse_decimal(values.get("hourly_rate"), default=Decimal("0"))
                if values.get("hourly_rate") not in (None, "")
                else None,
                source_no=values.get("source_no"),
                invoice_no=values.get("invoice_no"),
                upload_batch_no=batch_no,
                uploaded_by=current_user.id,
            )
            db.add(cost)
            created += 1
        except Exception as exc:  # noqa: BLE001 - 返回逐行导入错误给前端
            errors.append({"row": row_no, "error": str(exc)})

    if created:
        db.commit()
    else:
        db.rollback()

    return ResponseModel(
        code=200 if not errors else 207,
        message="上传完成" if not errors else "部分数据上传失败",
        data={"created": created, "errors": errors, "upload_batch_no": batch_no},
    )


@router.delete("/financial-costs/{cost_id}", response_model=ResponseModel)
def delete_financial_project_cost(
    cost_id: int,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    cost = db.query(FinancialProjectCost).filter(FinancialProjectCost.id == cost_id).first()
    if not cost:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="成本记录不存在")

    db.delete(cost)
    db.commit()
    return ResponseModel(code=200, message="删除成功")

# -*- coding: utf-8 -*-
"""
销售管理模块 API endpoints
"""

import csv
import io
import calendar
from collections import defaultdict
from typing import Any, List, Optional, Dict
from datetime import date, datetime, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, aliased
from sqlalchemy import desc, or_, and_, func

from app.api import deps
from app.core.config import settings
from app.core import security
from app.models.user import User, UserRole, Role
from app.models.project import Customer, Project, ProjectMilestone, ProjectPaymentPlan
from app.models.organization import Department
from app.schemas.project import ProjectPaymentPlanResponse
from app.models.sales import (
    Lead, LeadFollowUp, Opportunity, OpportunityRequirement, Quote, QuoteVersion, QuoteItem,
    QuoteCostTemplate, QuoteCostApproval, QuoteCostHistory, PurchaseMaterialCost,
    MaterialCostUpdateReminder,
    Contract, ContractDeliverable, ContractAmendment, Invoice, ReceivableDispute,
    QuoteApproval, ContractApproval, InvoiceApproval,
    CpqRuleSet, QuoteTemplate, QuoteTemplateVersion,
    ContractTemplate, ContractTemplateVersion,
    TechnicalAssessment, ScoringRule, FailureCase, OpenItem, LeadRequirementDetail,
    RequirementFreeze, AIClarification,
    ApprovalWorkflow, ApprovalWorkflowStep, ApprovalRecord, ApprovalHistory,
    SalesTarget
)
from app.schemas.sales import (
    LeadCreate, LeadUpdate, LeadResponse, LeadFollowUpResponse, LeadFollowUpCreate,
    OpportunityCreate, OpportunityUpdate, OpportunityResponse, OpportunityRequirementResponse,
    GateSubmitRequest,
    QuoteCreate, QuoteUpdate, QuoteResponse, QuoteVersionCreate, QuoteVersionResponse,
    QuoteItemCreate, QuoteItemUpdate, QuoteItemBatchUpdate, QuoteItemResponse, QuoteApproveRequest,
    QuoteCostTemplateCreate, QuoteCostTemplateUpdate, QuoteCostTemplateResponse,
    QuoteCostApprovalCreate, QuoteCostApprovalResponse, QuoteCostApprovalAction,
    CostBreakdownResponse, CostCheckResponse, CostComparisonResponse,
    PurchaseMaterialCostCreate, PurchaseMaterialCostUpdate, PurchaseMaterialCostResponse,
    MaterialCostMatchRequest, MaterialCostMatchResponse,
    CostMatchSuggestion, CostMatchSuggestionsResponse, ApplyCostSuggestionsRequest,
    MaterialCostUpdateReminderResponse, MaterialCostUpdateReminderUpdate,
    ContractCreate, ContractUpdate, ContractResponse, ContractDeliverableResponse,
    ContractAmendmentCreate, ContractAmendmentResponse,
    ContractSignRequest, ContractProjectCreateRequest,
    InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceIssueRequest,
    ReceivableDisputeCreate, ReceivableDisputeUpdate, ReceivableDisputeResponse,
    QuoteApprovalResponse, QuoteApprovalCreate,
    ContractApprovalResponse, ContractApprovalCreate,
    InvoiceApprovalResponse, InvoiceApprovalCreate,
    QuoteTemplateCreate, QuoteTemplateUpdate, QuoteTemplateResponse,
    QuoteTemplateVersionCreate, QuoteTemplateVersionResponse,
    QuoteTemplateApplyResponse,
    ContractTemplateCreate, ContractTemplateUpdate, ContractTemplateResponse,
    ContractTemplateVersionCreate, ContractTemplateVersionResponse,
    ContractTemplateApplyResponse,
    CpqRuleSetCreate, CpqRuleSetUpdate, CpqRuleSetResponse,
    CpqPricePreviewRequest, CpqPricePreviewResponse,
    TemplateVersionDiff, TemplateApprovalHistoryRecord,
    TechnicalAssessmentApplyRequest, TechnicalAssessmentEvaluateRequest, TechnicalAssessmentResponse,
    ScoringRuleCreate, ScoringRuleResponse,
    FailureCaseCreate, FailureCaseResponse,
    OpenItemCreate, OpenItemResponse,
    LeadRequirementDetailCreate, LeadRequirementDetailResponse,
    RequirementFreezeCreate, RequirementFreezeResponse,
    AIClarificationCreate, AIClarificationUpdate, AIClarificationResponse,
    ApprovalWorkflowCreate, ApprovalWorkflowUpdate, ApprovalWorkflowResponse,
    ApprovalWorkflowStepCreate, ApprovalWorkflowStepResponse,
    ApprovalRecordResponse, ApprovalHistoryResponse,
    ApprovalStartRequest, ApprovalActionRequest, ApprovalStatusResponse,
    SalesTargetCreate, SalesTargetUpdate, SalesTargetResponse
)
from app.schemas.common import PaginatedResponse, ResponseModel
from app.services.cpq_pricing_service import CpqPricingService
from app.services.technical_assessment_service import TechnicalAssessmentService
from app.services.ai_assessment_service import AIAssessmentService
from app.services.approval_workflow_service import ApprovalWorkflowService
from app.services.sales_team_service import SalesTeamService
from app.models.enums import (
    AssessmentSourceTypeEnum, AssessmentStatusEnum, AssessmentDecisionEnum,
    WorkflowTypeEnum, ApprovalRecordStatusEnum, ApprovalActionEnum,
    QuoteStatusEnum, ContractStatusEnum, InvoiceStatusEnum, InvoiceStatusEnum as InvoiceStatus
)

router = APIRouter()


def _get_entity_creator_id(entity) -> Optional[int]:
    """Safely fetch created_by if the ORM model defines it."""
    return getattr(entity, "created_by", None)


def _normalize_date_range(
    start_date_value: Optional[date],
    end_date_value: Optional[date],
) -> tuple[date, date]:
    """Normalize and validate date range."""
    today = date.today()
    normalized_start = start_date_value or date(today.year, today.month, 1)

    if end_date_value:
        normalized_end = end_date_value
    else:
        if normalized_start.month == 12:
            normalized_end = date(normalized_start.year, 12, 31)
        else:
            normalized_end = date(normalized_start.year, normalized_start.month + 1, 1) - timedelta(days=1)

    if normalized_start > normalized_end:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")

    return normalized_start, normalized_end


def _get_user_role_code(db: Session, user: User) -> str:
    """获取用户的角色代码（返回第一个角色的代码）"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == user.id).all()
    if user_roles and user_roles[0].role:
        return user_roles[0].role.role_code
    return "USER"


def _get_user_role_name(db: Session, user: User) -> str:
    """获取用户的角色名称（返回第一个角色的名称）"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == user.id).all()
    if user_roles and user_roles[0].role:
        return user_roles[0].role.role_name
    return "普通用户"


def _get_visible_sales_users(
    db: Session,
    current_user: User,
    department_id: Optional[int],
    region_keyword: Optional[str],
) -> List[User]:
    """根据角色、部门和区域过滤可见的销售用户"""
    # 直接查询当前用户的角色代码
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    user_role_codes = [ur.role.role_code for ur in user_roles if ur.role]
    user_role_codes_lower = [rc.lower() for rc in user_role_codes]

    # 检查是否是销售总监
    is_sales_director = 'SALES_DIR' in user_role_codes
    # 检查是否是销售经理
    is_sales_manager = 'SALES_MANAGER' in user_role_codes

    query = db.query(User).filter(User.is_active == True)

    if is_sales_director:
        # 销售总监可以看到所有销售人员
        # 通过子查询获取有销售角色的用户ID
        sales_role_codes = ['SALES', 'SALES_DIR', 'SALES_MANAGER', 'SA']
        sales_user_ids = (
            db.query(UserRole.user_id)
            .join(Role, Role.id == UserRole.role_id)
            .filter(Role.role_code.in_(sales_role_codes))
            .distinct()
            .all()
        )
        sales_user_ids = [uid for (uid,) in sales_user_ids]
        return query.filter(User.id.in_(sales_user_ids)).all()
    elif is_sales_manager:
        # 销售经理可以看到自己部门的所有用户
        dept_id = getattr(current_user, 'department_id', None)
        if dept_id:
            return query.filter(User.department_id == dept_id).all()
        else:
            return [current_user]
    else:
        # 其他用户只能看到自己
        return [current_user]


def _build_department_name_map(db: Session, users: List[User]) -> Dict[str, str]:
    """批量获取部门名称，减少数据库查询"""
    # User表使用department字符串字段，不是department_id外键
    dept_names = {user.department for user in users if user.department}
    return {name: name for name in dept_names}


def _collect_sales_team_members(
    db: Session,
    users: List[User],
    department_names: Dict[str, str],
    start_date_value: date,
    end_date_value: date,
) -> List[dict]:
    """构建销售团队成员的统计数据列表"""
    if not users:
        return []

    # 获取用户角色映射
    user_ids = [user.id for user in users]
    user_roles_map = {}
    for uid in user_ids:
        user_roles = db.query(UserRole).filter(UserRole.user_id == uid).all()
        role_names = [ur.role.role_name for ur in user_roles if ur.role]
        user_roles_map[uid] = role_names[0] if role_names else "销售专员"

    start_datetime = datetime.combine(start_date_value, datetime.min.time())
    end_datetime = datetime.combine(end_date_value, datetime.max.time())
    month_value = start_date_value.strftime("%Y-%m")
    year_value = str(start_date_value.year)

    team_service = SalesTeamService(db)
    personal_targets_map = team_service.build_personal_target_map(user_ids, month_value, year_value)
    recent_followups_map = team_service.get_recent_followups_map(user_ids, start_datetime, end_datetime)
    customer_distribution_map = team_service.get_customer_distribution_map(user_ids, start_date_value, end_date_value)

    team_members: List[dict] = []
    for user in users:
        lead_query = db.query(Lead).filter(Lead.owner_id == user.id)
        lead_query = lead_query.filter(Lead.created_at >= start_datetime)
        lead_query = lead_query.filter(Lead.created_at <= end_datetime)
        lead_count = lead_query.count()

        opp_query = db.query(Opportunity).filter(Opportunity.owner_id == user.id)
        opp_query = opp_query.filter(Opportunity.created_at >= start_datetime)
        opp_query = opp_query.filter(Opportunity.created_at <= end_datetime)
        opp_count = opp_query.count()

        contract_query = db.query(Contract).filter(Contract.owner_id == user.id)
        contract_query = contract_query.filter(Contract.created_at >= start_datetime)
        contract_query = contract_query.filter(Contract.created_at <= end_datetime)
        contracts = contract_query.all()
        contract_count = len(contracts)
        contract_amount = sum(float(c.contract_amount or 0) for c in contracts)

        invoice_query = db.query(Invoice).join(Contract).filter(Contract.owner_id == user.id)
        invoice_query = invoice_query.filter(Invoice.paid_date.isnot(None))
        invoice_query = invoice_query.filter(Invoice.paid_date >= start_date_value)
        invoice_query = invoice_query.filter(Invoice.paid_date <= end_date_value)
        invoices = invoice_query.filter(Invoice.payment_status.in_(["PAID", "PARTIAL"])).all()
        collection_amount = sum(float(inv.paid_amount or 0) for inv in invoices)

        # User表使用department字符串字段
        department_name = user.department or "未分配"
        region_name = department_name

        target_snapshot = personal_targets_map.get(user.id, {})
        monthly_target_info = target_snapshot.get("monthly", {})
        yearly_target_info = target_snapshot.get("yearly", {})

        customer_distribution = customer_distribution_map.get(user.id, {})
        recent_follow_up = recent_followups_map.get(user.id)

        monthly_target_value = monthly_target_info.get("target_value", 0.0)
        monthly_actual_value = monthly_target_info.get("actual_value", 0.0)
        monthly_completion_rate = monthly_target_info.get("completion_rate", 0.0)

        yearly_target_value = yearly_target_info.get("target_value", 0.0)
        yearly_actual_value = yearly_target_info.get("actual_value", 0.0)
        yearly_completion_rate = yearly_target_info.get("completion_rate", 0.0)

        team_members.append({
            "user_id": user.id,
            "user_name": user.real_name or user.username,
            "username": user.username,
            "role": user_roles_map.get(user.id, "销售专员"),
            "department_name": department_name,
            "email": user.email,
            "phone": user.phone,
            "lead_count": lead_count,
            "opportunity_count": opp_count,
            "contract_count": contract_count,
            "contract_amount": float(contract_amount),
            "collection_amount": float(collection_amount),
            "monthly_target": monthly_target_value,
            "monthly_actual": monthly_actual_value,
            "monthly_completion_rate": monthly_completion_rate,
            "year_target": yearly_target_value,
            "year_actual": yearly_actual_value,
            "year_completion_rate": yearly_completion_rate,
            "personal_targets": target_snapshot,
            "recent_follow_up": recent_follow_up,
            "customer_distribution": customer_distribution.get("categories", []),
            "customer_total": customer_distribution.get("total", 0),
            "new_customers": customer_distribution.get("new_customers", 0),
            "region": region_name,
        })

    return team_members


def _shift_month(year: int, month: int, delta: int) -> tuple[int, int]:
    """根据delta偏移月数"""
    total_months = year * 12 + (month - 1) + delta
    new_year = total_months // 12
    new_month = total_months % 12 + 1
    return new_year, new_month


def _generate_trend_buckets(period: str, count: int) -> List[dict]:
    """根据周期生成统计区间"""
    today = date.today()
    buckets: List[dict] = []
    period = period.upper()

    if period == "QUARTER":
        period = "QUARTER"
    elif period == "YEAR":
        period = "YEAR"
    else:
        period = "MONTH"

    if period == "MONTH":
        for offset in range(count - 1, -1, -1):
            target_year, target_month = _shift_month(today.year, today.month, -offset)
            start = date(target_year, target_month, 1)
            _, last_day = calendar.monthrange(target_year, target_month)
            end = date(target_year, target_month, last_day)
            label = f"{target_year}-{target_month:02d}"
            buckets.append({
                "label": label,
                "target_label": label,
                "start": start,
                "end": end,
            })
    elif period == "QUARTER":
        current_quarter = (today.month - 1) // 3 + 1
        for offset in range(count - 1, -1, -1):
            quarter_delta = -offset
            total_quarters = (today.year * 4 + (current_quarter - 1)) + quarter_delta
            target_year = total_quarters // 4
            target_quarter = total_quarters % 4 + 1
            start_month = (target_quarter - 1) * 3 + 1
            start = date(target_year, start_month, 1)
            end_month = start_month + 2
            _, last_day = calendar.monthrange(target_year, end_month)
            end = date(target_year, end_month, last_day)
            label = f"{target_year}-Q{target_quarter}"
            buckets.append({
                "label": label,
                "target_label": label,
                "start": start,
                "end": end,
            })
    else:  # YEAR
        for offset in range(count - 1, -1, -1):
            target_year = today.year - offset
            start = date(target_year, 1, 1)
            end = date(target_year, 12, 31)
            label = str(target_year)
            buckets.append({
                "label": label,
                "target_label": label,
                "start": start,
                "end": end,
            })

    return buckets


def _calculate_growth(current: float, previous: Optional[float]) -> float:
    """计算增长率"""
    if previous is None or previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 2)


def _get_previous_range(start_date_value: date, end_date_value: date) -> tuple[date, date]:
    """根据当前区间计算上一对等区间"""
    delta_days = (end_date_value - start_date_value).days + 1
    prev_end = start_date_value - timedelta(days=1)
    prev_start = prev_end - timedelta(days=delta_days - 1)
    return prev_start, prev_end


# ==================== 阶段门验证函数 ====================


def validate_g1_lead_to_opportunity(lead: Lead, requirement: Optional[OpportunityRequirement] = None, 
                                   db: Optional[Session] = None) -> tuple[bool, List[str]]:
    """
    G1：线索 → 商机 验证
    需求模板必填：行业/产品对象/节拍/接口/现场约束/验收依据
    客户基本信息与联系人齐全
    可选：检查技术评估状态（如果已申请）
    """
    errors = []
    warnings = []
    
    # 检查客户基本信息
    if not lead.customer_name:
        errors.append("客户名称不能为空")
    if not lead.contact_name:
        errors.append("联系人不能为空")
    if not lead.contact_phone:
        errors.append("联系电话不能为空")
    
    # 检查需求模板必填项
    if requirement:
        if not requirement.product_object:
            errors.append("产品对象不能为空")
        if not requirement.ct_seconds:
            errors.append("节拍(秒)不能为空")
        if not requirement.interface_desc:
            errors.append("接口/通信协议不能为空")
        if not requirement.site_constraints:
            errors.append("现场约束不能为空")
        if not requirement.acceptance_criteria:
            errors.append("验收依据不能为空")
    else:
        errors.append("需求信息不能为空")
    
    # 可选：检查技术评估状态（如果已申请）
    if db and lead.assessment_id:
        from app.models.sales import TechnicalAssessment
        assessment = db.query(TechnicalAssessment).filter(TechnicalAssessment.id == lead.assessment_id).first()
        if assessment:
            if assessment.status != AssessmentStatusEnum.COMPLETED.value:
                warnings.append(f"技术评估状态为{assessment.status}，建议等待评估完成")
            elif assessment.decision == AssessmentDecisionEnum.NOT_RECOMMEND.value:
                warnings.append("技术评估建议为'不建议立项'，请谨慎考虑")
    
    # 检查未决事项（阻塞报价的）
    if db:
        from app.models.sales import OpenItem
        blocking_items = db.query(OpenItem).filter(
            and_(
                OpenItem.source_type == AssessmentSourceTypeEnum.LEAD.value,
                OpenItem.source_id == lead.id,
                OpenItem.blocks_quotation == True,
                OpenItem.status != 'CLOSED'
            )
        ).count()
        if blocking_items > 0:
            warnings.append(f"存在{blocking_items}个阻塞报价的未决事项，建议先解决")
    
    # 返回验证结果（warnings不作为错误，但会记录）
    return len(errors) == 0, errors + warnings


def validate_g2_opportunity_to_quote(opportunity: Opportunity) -> tuple[bool, List[str]]:
    """
    G2：商机 → 报价 验证
    预算范围、决策链、交付窗口、验收标准明确
    技术可行性初评通过
    """
    errors = []
    
    if not opportunity.budget_range:
        errors.append("预算范围不能为空")
    if not opportunity.decision_chain:
        errors.append("决策链不能为空")
    if not opportunity.delivery_window:
        errors.append("交付窗口不能为空")
    if not opportunity.acceptance_basis:
        errors.append("验收标准不能为空")
    
    # 技术可行性初评（简化：检查是否有评分）
    if opportunity.score is None or opportunity.score < 60:
        errors.append("技术可行性初评未通过（评分需≥60分）")
    
    return len(errors) == 0, errors


def validate_g3_quote_to_contract(quote: Quote, version: QuoteVersion, items: List[QuoteItem], 
                                   db: Optional[Session] = None) -> tuple[bool, List[str], Optional[str]]:
    """
    G3：报价 → 合同 验证
    成本拆解齐备，毛利率低于阈值自动预警
    交期校验通过（关键物料交期 + 设计/装配/调试周期）
    风险条款与边界条款已补充
    """
    from app.core.config import settings
    
    errors = []
    warnings = []
    
    # 检查成本拆解
    if not items or len(items) == 0:
        errors.append("报价明细不能为空")
    else:
        # 检查每个明细项是否有成本
        items_without_cost = [item for item in items if not item.cost or float(item.cost or 0) == 0]
        if items_without_cost:
            errors.append(f"有{len(items_without_cost)}个报价明细项未填写成本")
        
        total_cost = sum(float(item.cost or 0) * float(item.qty or 0) for item in items)
        if total_cost == 0:
            errors.append("成本拆解不完整，总成本不能为0")
    
    # 检查毛利率（使用配置的阈值）
    total_price = float(version.total_price or 0)
    total_cost = float(version.cost_total or 0)
    if total_price > 0:
        gross_margin = (total_price - total_cost) / total_price * 100
        margin_threshold = settings.SALES_GROSS_MARGIN_THRESHOLD
        margin_warning = settings.SALES_GROSS_MARGIN_WARNING
        
        if gross_margin < margin_threshold:
            errors.append(f"毛利率过低（{gross_margin:.2f}%），低于最低阈值{margin_threshold}%，需要审批")
        elif gross_margin < margin_warning:
            warnings.append(f"毛利率较低（{gross_margin:.2f}%），低于警告阈值{margin_warning}%，建议重新评估")

    # 检查交期（使用配置的最小交期）
    if not version.lead_time_days or version.lead_time_days <= 0:
        errors.append("交期不能为空或必须大于0")
    else:
        min_lead_time = settings.SALES_MIN_LEAD_TIME_DAYS
        if version.lead_time_days < min_lead_time:
            warnings.append(f"交期较短（{version.lead_time_days}天），低于建议最小交期{min_lead_time}天，请确认关键物料交期和设计/装配/调试周期")

        # 详细的交期校验（如果有数据库会话）
        if db is not None:
            from app.services.delivery_validation_service import delivery_validation_service
            validation_result = delivery_validation_service.validate_delivery_date(
                db, quote, version, items
            )

            # 根据校验结果添加警告
            if validation_result["status"] == "WARNING":
                for warn in validation_result.get("warnings", []):
                    msg = warn.get("message", "")
                    if msg:
                        warnings.append(f"[交期校验] {msg}")

            # 添加优化建议
            for suggestion in validation_result.get("suggestions", []):
                warnings.append(f"[优化建议] {suggestion}")

    # 检查风险条款
    if not version.risk_terms:
        warnings.append("风险条款未补充，建议补充风险条款与边界条款")
    
    warning_msg = "; ".join(warnings) if warnings else None
    return len(errors) == 0, errors, warning_msg


def validate_g4_contract_to_project(contract: Contract, deliverables: List[ContractDeliverable], 
                                     db: Optional[Session] = None) -> tuple[bool, List[str]]:
    """
    G4：合同 → 项目 验证
    付款节点与可交付物绑定
    SOW/验收标准/BOM初版/里程碑基线冻结
    """
    errors = []
    
    # 检查交付物
    if not deliverables or len(deliverables) == 0:
        errors.append("合同交付物不能为空")
    else:
        # 检查交付物是否必需（required_for_payment字段）
        required_deliverables = [d for d in deliverables if d.required_for_payment]
        if len(required_deliverables) == 0:
            errors.append("至少需要一个付款必需的交付物")
        
        # 检查交付物名称是否完整
        incomplete_deliverables = [d for d in deliverables if not d.deliverable_name or len(d.deliverable_name.strip()) == 0]
        if incomplete_deliverables:
            errors.append(f"有{len(incomplete_deliverables)}个交付物名称不完整")
    
    # 检查合同关键信息
    if not contract.contract_amount or contract.contract_amount <= 0:
        errors.append("合同金额不能为空或必须大于0")
    
    # 检查验收标准（简化：检查是否有验收摘要）
    if not contract.acceptance_summary:
        errors.append("验收摘要不能为空，请补充验收标准")
    
    # 检查付款条款摘要
    if not contract.payment_terms_summary:
        errors.append("付款条款摘要不能为空，请补充付款节点信息")
    
    # 检查合同是否已签订
    if contract.status != "SIGNED":
        errors.append("只有已签订的合同才能生成项目")

    # 检查合同是否已关联项目（避免重复生成）
    if contract.project_id:
        errors.append("合同已关联项目，不能重复生成")

    # 检查 SOW/验收标准/BOM初版/里程碑基线是否已冻结
    if db is not None:
        freeze_checks = []

        # 1. 检查是否有 SOW 文档（通过合同附件或技术文档）
        from app.models.project import ProjectDocument
        sow_documents = db.query(ProjectDocument).filter(
            ProjectDocument.contract_id == contract.id,
            ProjectDocument.document_type == "SOW"
        ).count()

        if sow_documents == 0:
            freeze_checks.append("SOW文档未上传")
        else:
            freeze_checks.append(f"SOW文档已上传({sow_documents}份)")

        # 2. 检查验收标准是否已确认
        if contract.acceptance_summary:
            freeze_checks.append("验收标准已填写")
        else:
            freeze_checks.append("验收标准未确认")

        # 3. 检查 BOM 初版是否已冻结
        from app.models.material import BomHeader
        bom_count = db.query(BomHeader).filter(
            BomHeader.contract_id == contract.id,
            BomHeader.version == "1.0"  # 初版
        ).count()

        if bom_count == 0:
            freeze_checks.append("BOM初版未创建")
        else:
            freeze_checks.append(f"BOM初版已创建({bom_count}个)")

        # 4. 检查里程碑基线是否已确定
        milestone_count = 0
        if hasattr(contract, 'deliverables') and contract.deliverables:
            # 通过交付物推算里程碑
            from app.models.project import ProjectMilestone
            milestone_count = db.query(ProjectMilestone).filter(
                ProjectMilestone.contract_id == contract.id
            ).count()

        if milestone_count == 0:
            freeze_checks.append("里程碑基线未确定")
        else:
            freeze_checks.append(f"里程碑已设置({milestone_count}个)")

        # 汇总冻结检查结果
        missing_items = [item for item in freeze_checks if "未" in item]
        if missing_items:
            # 不阻止生成项目，但给出警告
            freeze_checks.append(f"警告：以下项目未完成 - {', '.join(missing_items)}")

    return len(errors) == 0, errors


# ==================== 编码生成函数 ====================


def generate_lead_code(db: Session) -> str:
    """生成线索编码：L2507-001"""
    today = datetime.now()
    month_str = today.strftime("%y%m")
    prefix = f"L{month_str}-"
    
    max_lead = (
        db.query(Lead)
        .filter(Lead.lead_code.like(f"{prefix}%"))
        .order_by(desc(Lead.lead_code))
        .first()
    )
    
    if max_lead:
        try:
            seq = int(max_lead.lead_code.split("-")[-1]) + 1
        except:
            seq = 1
    else:
        seq = 1
    
    return f"{prefix}{seq:03d}"


def generate_opportunity_code(db: Session) -> str:
    """生成商机编码：O2507-001"""
    today = datetime.now()
    month_str = today.strftime("%y%m")
    prefix = f"O{month_str}-"
    
    max_opp = (
        db.query(Opportunity)
        .filter(Opportunity.opp_code.like(f"{prefix}%"))
        .order_by(desc(Opportunity.opp_code))
        .first()
    )
    
    if max_opp:
        try:
            seq = int(max_opp.opp_code.split("-")[-1]) + 1
        except:
            seq = 1
    else:
        seq = 1
    
    return f"{prefix}{seq:03d}"


def generate_quote_code(db: Session) -> str:
    """生成报价编码：Q2507-001"""
    today = datetime.now()
    month_str = today.strftime("%y%m")
    prefix = f"Q{month_str}-"
    
    max_quote = (
        db.query(Quote)
        .filter(Quote.quote_code.like(f"{prefix}%"))
        .order_by(desc(Quote.quote_code))
        .first()
    )
    
    if max_quote:
        try:
            seq = int(max_quote.quote_code.split("-")[-1]) + 1
        except:
            seq = 1
    else:
        seq = 1
    
    return f"{prefix}{seq:03d}"


def generate_contract_code(db: Session) -> str:
    """生成合同编码：HT2507-001"""
    today = datetime.now()
    month_str = today.strftime("%y%m")
    prefix = f"HT{month_str}-"
    
    max_contract = (
        db.query(Contract)
        .filter(Contract.contract_code.like(f"{prefix}%"))
        .order_by(desc(Contract.contract_code))
        .first()
    )
    
    if max_contract:
        try:
            seq = int(max_contract.contract_code.split("-")[-1]) + 1
        except:
            seq = 1
    else:
        seq = 1
    
    return f"{prefix}{seq:03d}"


def generate_amendment_no(db: Session, contract_code: str) -> str:
    """
    生成合同变更编号
    格式：{合同编码}-BG{序号}
    """
    prefix = f"{contract_code}-BG"
    
    # 查询该合同已有的变更记录数量
    count = db.query(ContractAmendment).filter(
        ContractAmendment.amendment_no.like(f"{prefix}%")
    ).count()
    
    seq = count + 1
    return f"{prefix}{seq:03d}"


def generate_invoice_code(db: Session) -> str:
    """生成发票编码：INV-yymmdd-xxx"""
    today = datetime.now().strftime("%y%m%d")
    prefix = f"INV-{today}-"
    
    max_invoice = (
        db.query(Invoice)
        .filter(Invoice.invoice_code.like(f"{prefix}%"))
        .order_by(desc(Invoice.invoice_code))
        .first()
    )
    
    if max_invoice:
        try:
            seq = int(max_invoice.invoice_code.split("-")[-1]) + 1
        except:
            seq = 1
    else:
        seq = 1
    
    return f"{prefix}{seq:03d}"


# ==================== 线索 ====================


@router.get("/leads", response_model=PaginatedResponse[LeadResponse])
def read_leads(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取线索列表
    Issue 7.1: 已集成数据权限过滤
    """
    query = db.query(Lead)

    # Issue 7.1: 应用数据权限过滤
    query = security.filter_sales_data_by_scope(query, current_user, db, Lead, 'owner_id')

    if keyword:
        query = query.filter(
            or_(
                Lead.lead_code.contains(keyword),
                Lead.customer_name.contains(keyword),
                Lead.contact_name.contains(keyword),
            )
        )

    if status:
        query = query.filter(Lead.status == status)

    if owner_id:
        query = query.filter(Lead.owner_id == owner_id)

    total = query.count()
    offset = (page - 1) * page_size
    leads = query.order_by(desc(Lead.created_at)).offset(offset).limit(page_size).all()

    # 填充负责人名称
    lead_responses = []
    for lead in leads:
        lead_dict = {
            **{c.name: getattr(lead, c.name) for c in lead.__table__.columns},
            "owner_name": lead.owner.real_name if lead.owner else None,
        }
        lead_responses.append(LeadResponse(**lead_dict))

    return PaginatedResponse(
        items=lead_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/leads", response_model=LeadResponse, status_code=201)
def create_lead(
    *,
    db: Session = Depends(deps.get_db),
    lead_in: LeadCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建线索
    """
    # 如果没有提供编码，自动生成
    lead_data = lead_in.model_dump()
    if not lead_data.get("lead_code"):
        lead_data["lead_code"] = generate_lead_code(db)
    else:
        # 检查编码是否已存在
        existing = db.query(Lead).filter(Lead.lead_code == lead_data["lead_code"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="线索编码已存在")

    # 如果没有指定负责人，默认使用当前用户
    if not lead_data.get("owner_id"):
        lead_data["owner_id"] = current_user.id

    lead = Lead(**lead_data)
    db.add(lead)
    db.commit()
    db.refresh(lead)

    lead_dict = {
        **{c.name: getattr(lead, c.name) for c in lead.__table__.columns},
        "owner_name": lead.owner.real_name if lead.owner else None,
    }
    return LeadResponse(**lead_dict)


@router.get("/leads/{lead_id}", response_model=LeadResponse)
def read_lead(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取线索详情
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    lead_dict = {
        **{c.name: getattr(lead, c.name) for c in lead.__table__.columns},
        "owner_name": lead.owner.real_name if lead.owner else None,
    }
    return LeadResponse(**lead_dict)


@router.put("/leads/{lead_id}", response_model=LeadResponse)
def update_lead(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    lead_in: LeadUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新线索
    Issue 7.2: 已集成操作权限检查
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    # Issue 7.2: 检查编辑权限
    if not security.check_sales_edit_permission(
        current_user,
        db,
        _get_entity_creator_id(lead),
        lead.owner_id,
    ):
        raise HTTPException(status_code=403, detail="您没有权限编辑此线索")

    update_data = lead_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(lead, field, value)

    db.commit()
    db.refresh(lead)

    lead_dict = {
        **{c.name: getattr(lead, c.name) for c in lead.__table__.columns},
        "owner_name": lead.owner.real_name if lead.owner else None,
    }
    return LeadResponse(**lead_dict)


@router.delete("/leads/{lead_id}", response_model=ResponseModel)
def delete_lead(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除线索
    Issue 7.2: 已集成操作权限检查（仅创建人、销售总监、管理员可以删除）
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    # Issue 7.2: 检查删除权限
    if not security.check_sales_delete_permission(
        current_user,
        db,
        _get_entity_creator_id(lead),
    ):
        raise HTTPException(status_code=403, detail="您没有权限删除此线索")

    db.delete(lead)
    db.commit()

    return ResponseModel(code=200, message="线索已删除")


@router.get("/leads/{lead_id}/follow-ups", response_model=List[LeadFollowUpResponse])
def get_lead_follow_ups(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取线索跟进记录列表
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    follow_ups = db.query(LeadFollowUp).filter(
        LeadFollowUp.lead_id == lead_id
    ).order_by(desc(LeadFollowUp.created_at)).all()

    result = []
    for follow_up in follow_ups:
        result.append({
            "id": follow_up.id,
            "lead_id": follow_up.lead_id,
            "follow_up_type": follow_up.follow_up_type,
            "content": follow_up.content,
            "next_action": follow_up.next_action,
            "next_action_at": follow_up.next_action_at,
            "created_by": follow_up.created_by,
            "attachments": follow_up.attachments,
            "created_at": follow_up.created_at,
            "updated_at": follow_up.updated_at,
            "creator_name": follow_up.creator.real_name if follow_up.creator else None,
        })

    return result


@router.post("/leads/{lead_id}/follow-ups", response_model=LeadFollowUpResponse, status_code=201)
def create_lead_follow_up(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    follow_up_in: LeadFollowUpCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    添加线索跟进记录
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    follow_up = LeadFollowUp(
        lead_id=lead_id,
        follow_up_type=follow_up_in.follow_up_type,
        content=follow_up_in.content,
        next_action=follow_up_in.next_action,
        next_action_at=follow_up_in.next_action_at,
        created_by=current_user.id,
        attachments=follow_up_in.attachments,
    )

    db.add(follow_up)
    
    # 如果设置了下次行动时间，更新线索的next_action_at
    if follow_up_in.next_action_at:
        lead.next_action_at = follow_up_in.next_action_at
    
    db.commit()
    db.refresh(follow_up)

    return {
        "id": follow_up.id,
        "lead_id": follow_up.lead_id,
        "follow_up_type": follow_up.follow_up_type,
        "content": follow_up.content,
        "next_action": follow_up.next_action,
        "next_action_at": follow_up.next_action_at,
        "created_by": follow_up.created_by,
        "attachments": follow_up.attachments,
        "created_at": follow_up.created_at,
        "updated_at": follow_up.updated_at,
        "creator_name": follow_up.creator.real_name if follow_up.creator else None,
    }


@router.post("/leads/{lead_id}/convert", response_model=OpportunityResponse, status_code=201)
def convert_lead_to_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    customer_id: int = Query(..., description="客户ID"),
    requirement_data: Optional[dict] = None,  # 需求数据，可以通过请求体传入
    skip_validation: bool = Query(False, description="跳过G1验证"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    线索转商机（G1阶段门验证）
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    # G1验证
    requirement = None
    if requirement_data:
        requirement = OpportunityRequirement(**requirement_data)
    
    if not skip_validation:
        is_valid, messages = validate_g1_lead_to_opportunity(lead, requirement, db)
        # 区分错误和警告
        errors = [msg for msg in messages if not msg.startswith("技术评估") and not msg.startswith("存在")]
        warnings = [msg for msg in messages if msg.startswith("技术评估") or msg.startswith("存在")]
        
        if errors:
            raise HTTPException(
                status_code=400,
                detail=f"G1阶段门验证失败: {', '.join(errors)}"
            )
        # warnings 只记录，不阻止转换

    # 生成商机编码
    opp_code = generate_opportunity_code(db)

    opportunity = Opportunity(
        opp_code=opp_code,
        lead_id=lead_id,
        customer_id=customer_id,
        opp_name=f"{lead.customer_name} - {lead.demand_summary[:50] if lead.demand_summary else '商机'}",
        owner_id=lead.owner_id,
        stage="DISCOVERY",
        gate_status="PASS" if not skip_validation else "PENDING",
        gate_passed_at=datetime.now() if not skip_validation else None
    )

    db.add(opportunity)
    db.flush()

    # 创建需求信息
    if requirement:
        requirement.opportunity_id = opportunity.id
        db.add(requirement)

    # 更新线索状态
    lead.status = "CONVERTED"

    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": customer.customer_name,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})
    
    return OpportunityResponse(**opp_dict)


# ==================== 商机 ====================


@router.get("/opportunities", response_model=PaginatedResponse[OpportunityResponse])
def read_opportunities(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    stage: Optional[str] = Query(None, description="阶段筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取商机列表
    Issue 7.1: 已集成数据权限过滤
    """
    query = db.query(Opportunity).options(joinedload(Opportunity.customer))

    # Issue 7.1: 应用数据权限过滤
    query = security.filter_sales_data_by_scope(query, current_user, db, Opportunity, 'owner_id')

    if keyword:
        query = query.filter(
            or_(
                Opportunity.opp_code.contains(keyword),
                Opportunity.opp_name.contains(keyword),
            )
        )

    if stage:
        query = query.filter(Opportunity.stage == stage)

    if customer_id:
        query = query.filter(Opportunity.customer_id == customer_id)

    if owner_id:
        query = query.filter(Opportunity.owner_id == owner_id)

    total = query.count()
    offset = (page - 1) * page_size
    # 使用 eager loading 避免 N+1 查询
    opportunities = query.options(
        joinedload(Opportunity.customer),
        joinedload(Opportunity.owner),
        joinedload(Opportunity.requirements)
    ).order_by(desc(Opportunity.created_at)).offset(offset).limit(page_size).all()

    opp_responses = []
    for opp in opportunities:
        # 获取第一个 requirement（如果存在）
        req = opp.requirements[0] if opp.requirements else None
        opp_dict = {
            **{c.name: getattr(opp, c.name) for c in opp.__table__.columns},
            "customer_name": opp.customer.customer_name if opp.customer else None,
            "owner_name": opp.owner.real_name if opp.owner else None,
            "requirement": None,
        }
        if req:
            opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})
        opp_responses.append(OpportunityResponse(**opp_dict))

    return PaginatedResponse(
        items=opp_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/opportunities", response_model=OpportunityResponse, status_code=201)
def create_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_in: OpportunityCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建商机
    """
    opp_data = opp_in.model_dump(exclude={"requirement"})
    
    # 如果没有提供编码，自动生成
    if not opp_data.get("opp_code"):
        opp_data["opp_code"] = generate_opportunity_code(db)
    else:
        # 检查编码是否已存在
        existing = db.query(Opportunity).filter(Opportunity.opp_code == opp_data["opp_code"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="商机编码已存在")
    
    # 如果没有指定负责人，默认使用当前用户
    if not opp_data.get("owner_id"):
        opp_data["owner_id"] = current_user.id

    customer = db.query(Customer).filter(Customer.id == opp_data["customer_id"]).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    opportunity = Opportunity(**opp_data)
    db.add(opportunity)
    db.flush()

    # 创建需求信息
    if opp_in.requirement:
        req_data = opp_in.requirement.model_dump()
        req_data["opportunity_id"] = opportunity.id
        requirement = OpportunityRequirement(**req_data)
        db.add(requirement)

    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": customer.customer_name,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.get("/opportunities/{opp_id}", response_model=OpportunityResponse)
def read_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取商机详情
    """
    opportunity = db.query(Opportunity).options(joinedload(Opportunity.customer)).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.put("/opportunities/{opp_id}", response_model=OpportunityResponse)
def update_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    opp_in: OpportunityUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新商机
    Issue 7.2: 已集成操作权限检查
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    # Issue 7.2: 检查编辑权限
    if not security.check_sales_edit_permission(
        current_user,
        db,
        _get_entity_creator_id(opportunity),
        opportunity.owner_id,
    ):
        raise HTTPException(status_code=403, detail="您没有权限编辑此商机")

    update_data = opp_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(opportunity, field, value)

    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.post("/opportunities/{opp_id}/gate", response_model=ResponseModel)
def submit_opportunity_gate(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    gate_request: GateSubmitRequest,
    gate_type: str = Query("G2", description="阶段门类型: G1, G2, G3, G4"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    提交商机阶段门（带自动验证）
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    # 根据阶段门类型进行验证
    validation_errors = []
    if gate_type == "G2":
        is_valid, errors = validate_g2_opportunity_to_quote(opportunity)
        if not is_valid:
            validation_errors = errors
    # G1在转换时验证，G3在报价转合同时验证，G4在合同转项目时验证

    if validation_errors and gate_request.gate_status == "PASS":
        raise HTTPException(
            status_code=400,
            detail=f"{gate_type}阶段门验证失败: {', '.join(validation_errors)}"
        )

    opportunity.gate_status = gate_request.gate_status
    if gate_request.gate_status == "PASS":
        opportunity.gate_passed_at = datetime.now()

    db.commit()

    return ResponseModel(
        code=200,
        message=f"{gate_type}阶段门{'通过' if gate_request.gate_status == 'PASS' else '拒绝'}",
        data={"validation_errors": validation_errors} if validation_errors else None
    )


# ==================== 报价 ====================


@router.get("/quotes", response_model=PaginatedResponse[QuoteResponse])
def read_quotes(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    opportunity_id: Optional[int] = Query(None, description="商机ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价列表
    Issue 7.1: 已集成数据权限过滤
    """
    query = db.query(Quote).options(
        joinedload(Quote.opportunity),
        joinedload(Quote.customer),
        joinedload(Quote.owner)
    )

    # Issue 7.1: 应用数据权限过滤
    query = security.filter_sales_data_by_scope(query, current_user, db, Quote, 'owner_id')

    if keyword:
        query = query.filter(Quote.quote_code.contains(keyword))

    if status:
        query = query.filter(Quote.status == status)

    if opportunity_id:
        query = query.filter(Quote.opportunity_id == opportunity_id)

    total = query.count()
    offset = (page - 1) * page_size
    quotes = query.order_by(desc(Quote.created_at)).offset(offset).limit(page_size).all()

    quote_responses = []
    for quote in quotes:
        versions = db.query(QuoteVersion).options(
            joinedload(QuoteVersion.creator),
            joinedload(QuoteVersion.approver)
        ).filter(QuoteVersion.quote_id == quote.id).all()
        quote_dict = {
            **{c.name: getattr(quote, c.name) for c in quote.__table__.columns},
            "opportunity_code": quote.opportunity.opp_code if quote.opportunity else None,
            "customer_name": quote.customer.customer_name if quote.customer else None,
            "owner_name": quote.owner.real_name if quote.owner else None,
            "versions": [],
        }
        for v in versions:
            items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == v.id).all()
            v_dict = {
                **{c.name: getattr(v, c.name) for c in v.__table__.columns},
                "created_by_name": v.creator.real_name if v.creator else None,
                "approved_by_name": v.approver.real_name if v.approver else None,
                "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
            }
            quote_dict["versions"].append(QuoteVersionResponse(**v_dict))
        quote_responses.append(QuoteResponse(**quote_dict))

    return PaginatedResponse(
        items=quote_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/quotes", response_model=QuoteResponse, status_code=201)
def create_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_in: QuoteCreate,
    skip_g2_validation: bool = Query(False, description="跳过G2验证"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建报价（G2阶段门验证）
    """
    # 检查商机是否存在
    opportunity = db.query(Opportunity).filter(Opportunity.id == quote_in.opportunity_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    # G2验证
    if not skip_g2_validation:
        is_valid, errors = validate_g2_opportunity_to_quote(opportunity)
        if not is_valid:
            raise HTTPException(
                status_code=400,
                detail=f"G2阶段门验证失败: {', '.join(errors)}"
            )
    
    quote_data = quote_in.model_dump(exclude={"version"})
    
    # 如果没有提供编码，自动生成
    if not quote_data.get("quote_code"):
        quote_data["quote_code"] = generate_quote_code(db)
    else:
        existing = db.query(Quote).filter(Quote.quote_code == quote_data["quote_code"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="报价编码已存在")
    
    if not quote_data.get("owner_id"):
        quote_data["owner_id"] = current_user.id

    quote = Quote(**quote_data)
    db.add(quote)
    db.flush()

    # 创建报价版本
    if quote_in.version:
        version_data = quote_in.version.model_dump(exclude={"items"})
        version_data["quote_id"] = quote.id
        version_data["created_by"] = current_user.id
        version = QuoteVersion(**version_data)
        db.add(version)
        db.flush()

        quote.current_version_id = version.id

        # 创建报价明细
        if quote_in.version.items:
            for item_data in quote_in.version.items:
                item_dict = item_data.model_dump()
                item_dict["quote_version_id"] = version.id
                item = QuoteItem(**item_dict)
                db.add(item)

    db.commit()
    db.refresh(quote)

    version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first() if quote.current_version_id else None
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all() if version else []
    
    quote_dict = {
        **{c.name: getattr(quote, c.name) for c in quote.__table__.columns},
        "opportunity_code": opportunity.opp_code if opportunity else None,
        "customer_name": quote.customer.customer_name if quote.customer else None,
        "owner_name": quote.owner.real_name if quote.owner else None,
        "versions": [],
    }
    
    if version:
        version_dict = {
            **{c.name: getattr(version, c.name) for c in version.__table__.columns},
            "created_by_name": version.creator.real_name if version.creator else None,
            "approved_by_name": version.approver.real_name if version.approver else None,
            "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
        }
        quote_dict["versions"] = [QuoteVersionResponse(**version_dict)]
    
    return QuoteResponse(**quote_dict)


@router.post("/quotes/{quote_id}/versions", response_model=QuoteVersionResponse, status_code=201)
def create_quote_version(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_in: QuoteVersionCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建报价版本
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    version_data = version_in.model_dump()
    version_data["quote_id"] = quote_id
    version_data["created_by"] = current_user.id
    version = QuoteVersion(**version_data)
    db.add(version)
    db.flush()

    # 创建报价明细
    if version_in.items:
        for item_data in version_in.items:
            item = QuoteItem(quote_version_id=version.id, **item_data.model_dump())
            db.add(item)

    db.commit()
    db.refresh(version)

    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
    version_dict = {
        **{c.name: getattr(version, c.name) for c in version.__table__.columns},
        "created_by_name": version.creator.real_name if version.creator else None,
        "approved_by_name": version.approver.real_name if version.approver else None,
        "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
    }
    return QuoteVersionResponse(**version_dict)


@router.post("/quotes/{quote_id}/approve", response_model=ResponseModel)
def approve_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    approve_request: QuoteApproveRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批报价（单级审批，兼容旧接口）
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    if not quote.current_version_id:
        raise HTTPException(status_code=400, detail="报价没有当前版本")

    version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")

    if approve_request.approved:
        quote.status = "APPROVED"
        version.approved_by = current_user.id
        version.approved_at = datetime.now()
    else:
        quote.status = "REJECTED"

    db.commit()

    return ResponseModel(
        code=200,
        message="报价审批完成" if approve_request.approved else "报价已驳回"
    )


@router.get("/quotes/{quote_id}/approvals", response_model=List[QuoteApprovalResponse])
def get_quote_approvals(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价审批记录列表
    """
    approvals = db.query(QuoteApproval).filter(QuoteApproval.quote_id == quote_id).order_by(QuoteApproval.approval_level).all()
    
    result = []
    for approval in approvals:
        approver_name = None
        if approval.approver_id:
            approver = db.query(User).filter(User.id == approval.approver_id).first()
            approver_name = approver.real_name if approver else None
        
        result.append(QuoteApprovalResponse(
            id=approval.id,
            quote_id=approval.quote_id,
            approval_level=approval.approval_level,
            approval_role=approval.approval_role,
            approver_id=approval.approver_id,
            approver_name=approver_name,
            approval_result=approval.approval_result,
            approval_opinion=approval.approval_opinion,
            status=approval.status,
            approved_at=approval.approved_at,
            due_date=approval.due_date,
            is_overdue=approval.is_overdue or False,
            created_at=approval.created_at,
            updated_at=approval.updated_at
        ))
    
    return result


@router.put("/quote-approvals/{approval_id}/approve", response_model=QuoteApprovalResponse)
def approve_quote_approval(
    *,
    db: Session = Depends(deps.get_db),
    approval_id: int,
    approval_opinion: Optional[str] = Query(None, description="审批意见"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批通过（多级审批）
    """
    approval = db.query(QuoteApproval).filter(QuoteApproval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")

    if approval.status != "PENDING":
        raise HTTPException(status_code=400, detail="只能审批待审批状态的记录")

    # 检查审批权限
    if not security.check_sales_approval_permission(current_user, approval, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批此记录"
        )

    approval.approval_result = "APPROVED"
    approval.approval_opinion = approval_opinion
    approval.approved_at = datetime.now()
    approval.status = "COMPLETED"
    approval.approver_id = current_user.id
    approver = db.query(User).filter(User.id == current_user.id).first()
    if approver:
        approval.approver_name = approver.real_name

    # 检查是否所有审批都已完成
    quote = db.query(Quote).filter(Quote.id == approval.quote_id).first()
    if quote:
        pending_approvals = db.query(QuoteApproval).filter(
            QuoteApproval.quote_id == approval.quote_id,
            QuoteApproval.status == "PENDING"
        ).count()

        if pending_approvals == 0:
            # 所有审批都已完成，更新报价状态
            quote.status = "APPROVED"
            version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
            if version:
                version.approved_by = current_user.id
                version.approved_at = datetime.now()

    db.commit()
    db.refresh(approval)

    approver_name = approval.approver_name
    return QuoteApprovalResponse(
        id=approval.id,
        quote_id=approval.quote_id,
        approval_level=approval.approval_level,
        approval_role=approval.approval_role,
        approver_id=approval.approver_id,
        approver_name=approver_name,
        approval_result=approval.approval_result,
        approval_opinion=approval.approval_opinion,
        status=approval.status,
        approved_at=approval.approved_at,
        due_date=approval.due_date,
        is_overdue=approval.is_overdue or False,
        created_at=approval.created_at,
        updated_at=approval.updated_at
    )


@router.put("/quote-approvals/{approval_id}/reject", response_model=QuoteApprovalResponse)
def reject_quote_approval(
    *,
    db: Session = Depends(deps.get_db),
    approval_id: int,
    rejection_reason: str = Query(..., description="驳回原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批驳回（多级审批）
    """
    approval = db.query(QuoteApproval).filter(QuoteApproval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")

    if approval.status != "PENDING":
        raise HTTPException(status_code=400, detail="只能审批待审批状态的记录")

    # 检查审批权限
    if not security.check_sales_approval_permission(current_user, approval, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批此记录"
        )

    approval.approval_result = "REJECTED"
    approval.approval_opinion = rejection_reason
    approval.approved_at = datetime.now()
    approval.status = "COMPLETED"
    approval.approver_id = current_user.id
    approver = db.query(User).filter(User.id == current_user.id).first()
    if approver:
        approval.approver_name = approver.real_name

    # 驳回后，报价状态变为被拒
    quote = db.query(Quote).filter(Quote.id == approval.quote_id).first()
    if quote:
        quote.status = "REJECTED"

    db.commit()
    db.refresh(approval)

    approver_name = approval.approver_name
    return QuoteApprovalResponse(
        id=approval.id,
        quote_id=approval.quote_id,
        approval_level=approval.approval_level,
        approval_role=approval.approval_role,
        approver_id=approval.approver_id,
        approver_name=approver_name,
        approval_result=approval.approval_result,
        approval_opinion=approval.approval_opinion,
        status=approval.status,
        approved_at=approval.approved_at,
        due_date=approval.due_date,
        is_overdue=approval.is_overdue or False,
        created_at=approval.created_at,
        updated_at=approval.updated_at
    )


# ==================== 合同 ====================


@router.get("/contracts", response_model=PaginatedResponse[ContractResponse])
def read_contracts(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同列表
    Issue 7.1: 已集成数据权限过滤
    """
    query = db.query(Contract).options(
        joinedload(Contract.customer),
        joinedload(Contract.project),
        joinedload(Contract.opportunity),
        joinedload(Contract.owner)
    )

    # Issue 7.1: 应用数据权限过滤
    query = security.filter_sales_data_by_scope(query, current_user, db, Contract, 'owner_id')

    if keyword:
        query = query.filter(Contract.contract_code.contains(keyword))

    if status:
        query = query.filter(Contract.status == status)

    if customer_id:
        query = query.filter(Contract.customer_id == customer_id)

    total = query.count()
    offset = (page - 1) * page_size
    contracts = query.order_by(desc(Contract.created_at)).offset(offset).limit(page_size).all()

    contract_responses = []
    for contract in contracts:
        deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract.id).all()
        contract_dict = {
            **{c.name: getattr(contract, c.name) for c in contract.__table__.columns},
            "opportunity_code": contract.opportunity.opp_code if contract.opportunity else None,
            "customer_name": contract.customer.customer_name if contract.customer else None,
            "project_code": contract.project.project_code if contract.project else None,
            "owner_name": contract.owner.real_name if contract.owner else None,
            "deliverables": [ContractDeliverableResponse(**{c.name: getattr(d, c.name) for c in d.__table__.columns}) for d in deliverables],
        }
        contract_responses.append(ContractResponse(**contract_dict))

    return PaginatedResponse(
        items=contract_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/contracts", response_model=ContractResponse, status_code=201)
def create_contract(
    *,
    db: Session = Depends(deps.get_db),
    contract_in: ContractCreate,
    skip_g3_validation: bool = Query(False, description="跳过G3验证"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建合同（G3阶段门验证）
    """
    contract_data = contract_in.model_dump(exclude={"deliverables"})
    
    # 检查报价是否存在（如果提供了quote_version_id）
    quote = None
    version = None
    items: List[QuoteItem] = []
    if contract_data.get("quote_version_id"):
        version = db.query(QuoteVersion).filter(QuoteVersion.id == contract_data["quote_version_id"]).first()
        if not version:
            raise HTTPException(status_code=404, detail="报价版本不存在")

        quote = db.query(Quote).filter(Quote.id == version.quote_id).first()
        if not quote:
            raise HTTPException(status_code=404, detail="报价不存在")

        items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
        
        # G3验证
        if not skip_g3_validation:
            is_valid, errors, warning = validate_g3_quote_to_contract(quote, version, items, db)
            if not is_valid:
                raise HTTPException(
                    status_code=400,
                    detail=f"G3阶段门验证失败: {', '.join(errors)}"
                )
            if warning:
                # 警告信息可以通过响应返回，但不阻止创建
                pass
    
    # 如果没有提供编码，自动生成
    if not contract_data.get("contract_code"):
        contract_data["contract_code"] = generate_contract_code(db)
    else:
        existing = db.query(Contract).filter(Contract.contract_code == contract_data["contract_code"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="合同编码已存在")
    
    # 如果没有指定负责人，默认使用当前用户
    if not contract_data.get("owner_id"):
        contract_data["owner_id"] = current_user.id

    opportunity = db.query(Opportunity).filter(Opportunity.id == contract_data["opportunity_id"]).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    customer = db.query(Customer).filter(Customer.id == contract_data["customer_id"]).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    contract = Contract(**contract_data)
    db.add(contract)
    db.flush()

    # 创建交付物清单
    if contract_in.deliverables:
        for deliverable_data in contract_in.deliverables:
            deliverable = ContractDeliverable(contract_id=contract.id, **deliverable_data.model_dump())
            db.add(deliverable)

    db.commit()
    db.refresh(contract)

    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract.id).all()
    contract_dict = {
        **{c.name: getattr(contract, c.name) for c in contract.__table__.columns},
        "opportunity_code": opportunity.opp_code,
        "customer_name": customer.customer_name,
        "project_code": contract.project.project_code if contract.project else None,
        "owner_name": contract.owner.real_name if contract.owner else None,
        "deliverables": [ContractDeliverableResponse(**{c.name: getattr(d, c.name) for c in d.__table__.columns}) for d in deliverables],
    }
    return ContractResponse(**contract_dict)


@router.post("/contracts/{contract_id}/sign", response_model=ResponseModel)
def sign_contract(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    sign_request: ContractSignRequest,
    auto_generate_payment_plans: bool = Query(True, description="自动生成收款计划"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    合同签订（自动生成收款计划）
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    contract.signed_date = sign_request.signed_date
    contract.status = "SIGNED"
    
    # Sprint 2.1 + Issue 1.2: 合同签订自动创建项目并触发阶段流转
    auto_create_project = getattr(sign_request, 'auto_create_project', True) if hasattr(sign_request, 'auto_create_project') else True
    created_project = None
    
    if auto_create_project:
        try:
            from app.services.status_transition_service import StatusTransitionService
            transition_service = StatusTransitionService(db)
            created_project = transition_service.handle_contract_signed(contract_id, auto_create_project=True)
            
            if created_project:
                # 更新合同关联的项目ID（如果之前没有）
                if not contract.project_id:
                    contract.project_id = created_project.id
                
                # Issue 1.2: 合同签订后自动触发阶段流转检查（S3→S4）
                try:
                    auto_transition_result = transition_service.check_auto_stage_transition(
                        created_project.id,
                        auto_advance=True  # 自动推进
                    )
                    if auto_transition_result.get("auto_advanced"):
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.info(f"合同签订后自动推进项目 {created_project.id} 至 {auto_transition_result.get('target_stage')} 阶段")
                except Exception as e:
                    # 自动流转失败不影响合同签订，记录日志
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"合同签订后自动阶段流转失败：{str(e)}", exc_info=True)
        except Exception as e:
            # 自动创建项目失败不影响合同签订，记录日志
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"合同签订自动创建项目失败：{str(e)}", exc_info=True)
    
    # 自动生成收款计划
    if auto_generate_payment_plans and contract.project_id:
        _generate_payment_plans_from_contract(db, contract)
    
    db.commit()
    
    # 发送合同签订通知
    try:
        from app.services.sales_reminder_service import notify_contract_signed
        notify_contract_signed(db, contract.id)
        db.commit()
    except Exception as e:
        # 通知失败不影响主流程
        pass

    response_data = {"contract_id": contract.id}
    if created_project:
        response_data["project_id"] = created_project.id
        response_data["project_code"] = created_project.project_code
        return ResponseModel(code=200, message="合同签订成功，项目已自动创建", data=response_data)
    
    return ResponseModel(code=200, message="合同签订成功", data=response_data)


def _generate_payment_plans_from_contract(db: Session, contract: Contract) -> List[ProjectPaymentPlan]:
    """
    根据合同自动生成收款计划
    默认规则：
    - 预付款：30%（合同签订后）
    - 发货款：40%（发货里程碑）
    - 验收款：25%（验收里程碑）
    - 质保款：5%（质保结束）
    """
    if not contract.project_id:
        return []
    
    project = db.query(Project).filter(Project.id == contract.project_id).first()
    if not project:
        return []
    
    contract_amount = float(contract.contract_amount or 0)
    if contract_amount <= 0:
        return []
    
    # 检查是否已有收款计划
    existing_plans = db.query(ProjectPaymentPlan).filter(
        ProjectPaymentPlan.contract_id == contract.id
    ).count()
    if existing_plans > 0:
        return []  # 已有计划，不重复生成
    
    # 默认收款计划配置
    payment_configs = [
        {
            "payment_no": 1,
            "payment_name": "预付款",
            "payment_type": "ADVANCE",
            "payment_ratio": 30.0,
            "trigger_milestone": "合同签订",
            "trigger_condition": "合同签订后"
        },
        {
            "payment_no": 2,
            "payment_name": "发货款",
            "payment_type": "DELIVERY",
            "payment_ratio": 40.0,
            "trigger_milestone": "发货",
            "trigger_condition": "设备发货后"
        },
        {
            "payment_no": 3,
            "payment_name": "验收款",
            "payment_type": "ACCEPTANCE",
            "payment_ratio": 25.0,
            "trigger_milestone": "终验通过",
            "trigger_condition": "终验通过后"
        },
        {
            "payment_no": 4,
            "payment_name": "质保款",
            "payment_type": "WARRANTY",
            "payment_ratio": 5.0,
            "trigger_milestone": "质保结束",
            "trigger_condition": "质保期结束后"
        }
    ]
    
    # 如果合同有付款条款摘要，可以解析自定义的付款计划
    # 这里简化处理，使用默认配置
    
    plans = []
    from datetime import timedelta
    
    for config in payment_configs:
        planned_amount = contract_amount * config["payment_ratio"] / 100
        
        # 计算计划收款日期（基于合同签订日期和项目计划）
        planned_date = None
        if config["payment_no"] == 1:
            # 预付款：合同签订后7天
            planned_date = contract.signed_date + timedelta(days=7) if contract.signed_date else None
        elif config["payment_no"] == 2:
            # 发货款：预计项目中期
            if project.planned_end_date and project.planned_start_date:
                days = (project.planned_end_date - project.planned_start_date).days
                planned_date = project.planned_start_date + timedelta(days=int(days * 0.6)) if days > 0 else None
        elif config["payment_no"] == 3:
            # 验收款：项目结束前
            planned_date = project.planned_end_date
        elif config["payment_no"] == 4:
            # 质保款：项目结束后1年
            if project.planned_end_date:
                planned_date = project.planned_end_date + timedelta(days=365)
        
        # 尝试查找关联的里程碑
        milestone_id = None
        if config["payment_no"] == 2:  # 发货款
            # 查找发货相关的里程碑
            milestone = db.query(ProjectMilestone).filter(
                and_(
                    ProjectMilestone.project_id == contract.project_id,
                    or_(
                        ProjectMilestone.milestone_name.like("%发货%"),
                        ProjectMilestone.milestone_name.like("%发运%"),
                        ProjectMilestone.milestone_type == "DELIVERY"
                    )
                )
            ).first()
            if milestone:
                milestone_id = milestone.id
        elif config["payment_no"] == 3:  # 验收款
            # 查找验收相关的里程碑
            milestone = db.query(ProjectMilestone).filter(
                and_(
                    ProjectMilestone.project_id == contract.project_id,
                    or_(
                        ProjectMilestone.milestone_name.like("%验收%"),
                        ProjectMilestone.milestone_name.like("%终验%"),
                        ProjectMilestone.milestone_type == "GATE"
                    )
                )
            ).first()
            if milestone:
                milestone_id = milestone.id
        
        plan = ProjectPaymentPlan(
            project_id=contract.project_id,
            contract_id=contract.id,
            payment_no=config["payment_no"],
            payment_name=config["payment_name"],
            payment_type=config["payment_type"],
            payment_ratio=config["payment_ratio"],
            planned_amount=planned_amount,
            planned_date=planned_date,
            milestone_id=milestone_id,  # 关联里程碑
            trigger_milestone=config["trigger_milestone"],
            trigger_condition=config["trigger_condition"],
            status="PENDING"
        )
        db.add(plan)
        plans.append(plan)
    
    db.flush()
    return plans


@router.post("/contracts/{contract_id}/project", response_model=ResponseModel)
def create_contract_project(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    project_request: ContractProjectCreateRequest,
    skip_g4_validation: bool = Query(False, description="跳过G4验证"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    合同生成项目（G4阶段门验证）
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # 获取交付物清单
    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract_id).all()
    
    # G4验证
    if not skip_g4_validation:
        is_valid, errors = validate_g4_contract_to_project(contract, deliverables, db)
        if not is_valid:
            raise HTTPException(
                status_code=400,
                detail=f"G4阶段门验证失败: {', '.join(errors)}"
            )

    # 检查项目编码是否已存在
    existing = db.query(Project).filter(Project.project_code == project_request.project_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="项目编码已存在")

    # 创建项目
    project = Project(
        project_code=project_request.project_code,
        project_name=project_request.project_name,
        customer_id=contract.customer_id,
        contract_no=contract.contract_code,
        contract_amount=contract.contract_amount,
        contract_date=contract.signed_date,
        pm_id=project_request.pm_id,
        planned_start_date=project_request.planned_start_date,
        planned_end_date=project_request.planned_end_date,
    )

    # 填充客户信息
    customer = db.query(Customer).filter(Customer.id == contract.customer_id).first()
    if customer:
        project.customer_name = customer.customer_name
        project.customer_contact = customer.contact_person
        project.customer_phone = customer.contact_phone

    db.add(project)
    db.flush()

    # 关联合同和项目
    contract.project_id = project.id
    db.commit()

    return ResponseModel(code=200, message="项目创建成功", data={"project_id": project.id})


@router.get("/contracts/{contract_id}/payment-plans", response_model=List[ProjectPaymentPlanResponse])
def get_contract_payment_plans(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同的收款计划列表
    """
    
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    
    # 查询收款计划
    payment_plans = db.query(ProjectPaymentPlan).filter(
        ProjectPaymentPlan.contract_id == contract_id
    ).options(
        joinedload(ProjectPaymentPlan.milestone),
        joinedload(ProjectPaymentPlan.project)
    ).order_by(ProjectPaymentPlan.payment_no).all()
    
    result = []
    for plan in payment_plans:
        plan_dict = {
            **{c.name: getattr(plan, c.name) for c in plan.__table__.columns},
            "project_code": plan.project.project_code if plan.project else None,
            "project_name": plan.project.project_name if plan.project else None,
            "contract_code": contract.contract_code,
            "milestone_code": plan.milestone.milestone_code if plan.milestone else None,
            "milestone_name": plan.milestone.milestone_name if plan.milestone else None,
        }
        result.append(ProjectPaymentPlanResponse(**plan_dict))
    
    return result


# ==================== 审批工作流 ====================


@router.post("/quotes/{quote_id}/approval/start", response_model=ResponseModel)
def start_quote_approval(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    approval_request: ApprovalStartRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    启动报价审批流程
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    if quote.status != QuoteStatusEnum.IN_REVIEW:
        raise HTTPException(status_code=400, detail="只有待审批状态的报价才能启动审批流程")
    
    # 获取报价金额用于路由
    version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
    routing_params = {
        "amount": float(version.total_price or 0) if version else 0
    }
    
    # 启动审批流程
    workflow_service = ApprovalWorkflowService(db)
    try:
        record = workflow_service.start_approval(
            entity_type=WorkflowTypeEnum.QUOTE,
            entity_id=quote_id,
            initiator_id=current_user.id,
            workflow_id=approval_request.workflow_id,
            routing_params=routing_params,
            comment=approval_request.comment
        )
        
        # 更新报价状态
        quote.status = QuoteStatusEnum.IN_REVIEW
        
        db.commit()
        
        return ResponseModel(
            code=200,
            message="审批流程已启动",
            data={"approval_record_id": record.id}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/quotes/{quote_id}/approval-status", response_model=ApprovalStatusResponse)
def get_quote_approval_status(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价审批状态
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.QUOTE,
        entity_id=quote_id
    )
    
    if not record:
        return ApprovalStatusResponse(
            record=None,
            current_step_info=None,
            can_approve=False,
            can_reject=False,
            can_delegate=False,
            can_withdraw=False
        )
    
    # 获取当前步骤信息
    current_step_info = workflow_service.get_current_step(record.id)
    
    # 判断当前用户的操作权限
    can_approve = False
    can_reject = False
    can_delegate = False
    can_withdraw = False
    
    if record.status == ApprovalRecordStatusEnum.PENDING:
        if current_step_info:
            # 检查是否是当前审批人
            if current_step_info.get("approver_id") == current_user.id:
                can_approve = True
                can_reject = True
                if current_step_info.get("can_delegate"):
                    can_delegate = True
        
        # 检查是否可以撤回（只有发起人可以撤回）
        if record.initiator_id == current_user.id:
            can_withdraw = True
    
    # 构建响应
    record_dict = {
        **{c.name: getattr(record, c.name) for c in record.__table__.columns},
        "workflow_name": record.workflow.workflow_name if record.workflow else None,
        "initiator_name": record.initiator.real_name if record.initiator else None,
        "history": []
    }
    
    # 获取审批历史
    history_list = workflow_service.get_approval_history(record.id)
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        record_dict["history"].append(ApprovalHistoryResponse(**history_dict))
    
    return ApprovalStatusResponse(
        record=ApprovalRecordResponse(**record_dict),
        current_step_info=current_step_info,
        can_approve=can_approve,
        can_reject=can_reject,
        can_delegate=can_delegate,
        can_withdraw=can_withdraw
    )


@router.post("/quotes/{quote_id}/approval/action", response_model=ResponseModel)
def quote_approval_action(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    action_request: ApprovalActionRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    报价审批操作（通过/驳回/委托/撤回）
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.QUOTE,
        entity_id=quote_id
    )
    
    if not record:
        raise HTTPException(status_code=404, detail="审批记录不存在")
    
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    try:
        if action_request.action == ApprovalActionEnum.APPROVE:
            record = workflow_service.approve_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment
            )
            
            # 如果审批完成，更新报价状态
            if record.status == ApprovalRecordStatusEnum.APPROVED:
                quote.status = QuoteStatusEnum.APPROVED
                version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
                if version:
                    version.approved_by = current_user.id
                    version.approved_at = datetime.now()
            
            message = "审批通过"
            
        elif action_request.action == ApprovalActionEnum.REJECT:
            record = workflow_service.reject_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment or "审批驳回"
            )
            
            # 驳回后更新报价状态
            quote.status = QuoteStatusEnum.REJECTED
            message = "审批已驳回"
            
        elif action_request.action == ApprovalActionEnum.DELEGATE:
            if not action_request.delegate_to_id:
                raise HTTPException(status_code=400, detail="委托操作需要指定委托给的用户ID")
            
            record = workflow_service.delegate_step(
                record_id=record.id,
                approver_id=current_user.id,
                delegate_to_id=action_request.delegate_to_id,
                comment=action_request.comment
            )
            message = "审批已委托"
            
        elif action_request.action == ApprovalActionEnum.WITHDRAW:
            record = workflow_service.withdraw_approval(
                record_id=record.id,
                initiator_id=current_user.id,
                comment=action_request.comment
            )
            
            # 撤回后恢复报价状态
            quote.status = QuoteStatusEnum.DRAFT
            message = "审批已撤回"
            
        else:
            raise HTTPException(status_code=400, detail=f"不支持的审批操作: {action_request.action}")
        
        db.commit()
        
        return ResponseModel(
            code=200,
            message=message,
            data={"approval_record_id": record.id, "status": record.status}
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/quotes/{quote_id}/approval-history", response_model=List[ApprovalHistoryResponse])
def get_quote_approval_history(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价审批历史
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.QUOTE,
        entity_id=quote_id
    )
    
    if not record:
        return []
    
    history_list = workflow_service.get_approval_history(record.id)
    result = []
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        result.append(ApprovalHistoryResponse(**history_dict))
    
    return result


# ==================== 审批工作流管理 ====================


@router.get("/approval-workflows", response_model=PaginatedResponse[ApprovalWorkflowResponse])
def list_approval_workflows(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    workflow_type: Optional[str] = Query(None, description="工作流类型筛选"),
    is_active: Optional[bool] = Query(None, description="是否启用筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取审批工作流列表
    """
    query = db.query(ApprovalWorkflow).options(
        joinedload(ApprovalWorkflow.steps)
    )
    
    if workflow_type:
        query = query.filter(ApprovalWorkflow.workflow_type == workflow_type)
    
    if is_active is not None:
        query = query.filter(ApprovalWorkflow.is_active == is_active)
    
    total = query.count()
    offset = (page - 1) * page_size
    workflows = query.order_by(ApprovalWorkflow.created_at.desc()).offset(offset).limit(page_size).all()
    
    result = []
    for workflow in workflows:
        workflow_dict = {
            **{c.name: getattr(workflow, c.name) for c in workflow.__table__.columns},
            "steps": [
                ApprovalWorkflowStepResponse(
                    **{c.name: getattr(step, c.name) for c in step.__table__.columns},
                    approver_name=step.approver.real_name if step.approver else None
                )
                for step in sorted(workflow.steps, key=lambda x: x.step_order)
            ]
        }
        result.append(ApprovalWorkflowResponse(**workflow_dict))
    
    return PaginatedResponse(
        items=result,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/approval-workflows", response_model=ApprovalWorkflowResponse, status_code=201)
def create_approval_workflow(
    *,
    db: Session = Depends(deps.get_db),
    workflow_in: ApprovalWorkflowCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建审批工作流
    """
    workflow_data = workflow_in.model_dump(exclude={"steps"})
    workflow = ApprovalWorkflow(**workflow_data)
    db.add(workflow)
    db.flush()
    
    # 创建审批步骤
    for step_data in workflow_in.steps:
        step = ApprovalWorkflowStep(
            workflow_id=workflow.id,
            **step_data.model_dump()
        )
        db.add(step)
    
    db.commit()
    db.refresh(workflow)
    
    # 加载步骤
    workflow_dict = {
        **{c.name: getattr(workflow, c.name) for c in workflow.__table__.columns},
        "steps": [
            ApprovalWorkflowStepResponse(
                **{c.name: getattr(step, c.name) for c in step.__table__.columns},
                approver_name=step.approver.real_name if step.approver else None
            )
            for step in sorted(workflow.steps, key=lambda x: x.step_order)
        ]
    }
    
    return ApprovalWorkflowResponse(**workflow_dict)


@router.get("/approval-workflows/{workflow_id}", response_model=ApprovalWorkflowResponse)
def get_approval_workflow(
    *,
    db: Session = Depends(deps.get_db),
    workflow_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取审批工作流详情
    """
    workflow = db.query(ApprovalWorkflow).options(
        joinedload(ApprovalWorkflow.steps)
    ).filter(ApprovalWorkflow.id == workflow_id).first()
    
    if not workflow:
        raise HTTPException(status_code=404, detail="审批工作流不存在")
    
    workflow_dict = {
        **{c.name: getattr(workflow, c.name) for c in workflow.__table__.columns},
        "steps": [
            ApprovalWorkflowStepResponse(
                **{c.name: getattr(step, c.name) for c in step.__table__.columns},
                approver_name=step.approver.real_name if step.approver else None
            )
            for step in sorted(workflow.steps, key=lambda x: x.step_order)
        ]
    }
    
    return ApprovalWorkflowResponse(**workflow_dict)


@router.put("/approval-workflows/{workflow_id}", response_model=ApprovalWorkflowResponse)
def update_approval_workflow(
    *,
    db: Session = Depends(deps.get_db),
    workflow_id: int,
    workflow_in: ApprovalWorkflowUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新审批工作流
    """
    workflow = db.query(ApprovalWorkflow).filter(ApprovalWorkflow.id == workflow_id).first()
    if not workflow:
        raise HTTPException(status_code=404, detail="审批工作流不存在")
    
    update_data = workflow_in.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(workflow, key, value)
    
    db.commit()
    db.refresh(workflow)
    
    # 加载步骤
    workflow_dict = {
        **{c.name: getattr(workflow, c.name) for c in workflow.__table__.columns},
        "steps": [
            ApprovalWorkflowStepResponse(
                **{c.name: getattr(step, c.name) for c in step.__table__.columns},
                approver_name=step.approver.real_name if step.approver else None
            )
            for step in sorted(workflow.steps, key=lambda x: x.step_order)
        ]
    }
    
    return ApprovalWorkflowResponse(**workflow_dict)


@router.post("/contracts/{contract_id}/approval/start", response_model=ResponseModel)
def start_contract_approval(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    approval_request: ApprovalStartRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    启动合同审批流程
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    
    if contract.status != ContractStatusEnum.DRAFT and contract.status != ContractStatusEnum.IN_REVIEW:
        raise HTTPException(status_code=400, detail="只有草稿或待审批状态的合同才能启动审批流程")
    
    # 获取合同金额用于路由
    routing_params = {
        "amount": float(contract.contract_amount or 0)
    }
    
    # 启动审批流程
    workflow_service = ApprovalWorkflowService(db)
    try:
        record = workflow_service.start_approval(
            entity_type=WorkflowTypeEnum.CONTRACT,
            entity_id=contract_id,
            initiator_id=current_user.id,
            workflow_id=approval_request.workflow_id,
            routing_params=routing_params,
            comment=approval_request.comment
        )

        # 更新合同状态
        contract.status = ContractStatusEnum.IN_REVIEW

        db.commit()

        # 发送审批通知给当前审批人
        from app.services.notification_service import notification_service, NotificationType, NotificationPriority
        try:
            # 获取当前待审批的步骤
            current_step = workflow_service.get_current_step(record.id)
            if current_step and current_step.get("approver_id"):
                notification_service.send_notification(
                    db=db,
                    recipient_id=current_step["approver_id"],
                    notification_type=NotificationType.TASK_ASSIGNED,
                    title=f"待审批合同: {contract.contract_name}",
                    content=f"合同编码: {contract.contract_no}\n合同金额: ¥{float(contract.contract_amount or 0):,.2f}\n发起人: {current_user.real_name or current_user.username}",
                    priority=NotificationPriority.HIGH,
                    link=f"/sales/contracts/{contract.id}/approval"
                )
        except Exception:
            pass  # 通知失败不影响主流程

        return ResponseModel(
            code=200,
            message="审批流程已启动",
            data={"approval_record_id": record.id}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/contracts/{contract_id}/approval-status", response_model=ApprovalStatusResponse)
def get_contract_approval_status(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同审批状态
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.CONTRACT,
        entity_id=contract_id
    )
    
    if not record:
        return ApprovalStatusResponse(
            record=None,
            current_step_info=None,
            can_approve=False,
            can_reject=False,
            can_delegate=False,
            can_withdraw=False
        )
    
    current_step_info = workflow_service.get_current_step(record.id)
    
    can_approve = False
    can_reject = False
    can_delegate = False
    can_withdraw = False
    
    if record.status == ApprovalRecordStatusEnum.PENDING:
        if current_step_info:
            if current_step_info.get("approver_id") == current_user.id:
                can_approve = True
                can_reject = True
                if current_step_info.get("can_delegate"):
                    can_delegate = True
        
        if record.initiator_id == current_user.id:
            can_withdraw = True
    
    record_dict = {
        **{c.name: getattr(record, c.name) for c in record.__table__.columns},
        "workflow_name": record.workflow.workflow_name if record.workflow else None,
        "initiator_name": record.initiator.real_name if record.initiator else None,
        "history": []
    }
    
    history_list = workflow_service.get_approval_history(record.id)
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        record_dict["history"].append(ApprovalHistoryResponse(**history_dict))
    
    return ApprovalStatusResponse(
        record=ApprovalRecordResponse(**record_dict),
        current_step_info=current_step_info,
        can_approve=can_approve,
        can_reject=can_reject,
        can_delegate=can_delegate,
        can_withdraw=can_withdraw
    )


@router.post("/contracts/{contract_id}/approval/action", response_model=ResponseModel)
def contract_approval_action(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    action_request: ApprovalActionRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    合同审批操作（通过/驳回/委托/撤回）
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.CONTRACT,
        entity_id=contract_id
    )
    
    if not record:
        raise HTTPException(status_code=404, detail="审批记录不存在")
    
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    
    try:
        if action_request.action == ApprovalActionEnum.APPROVE:
            record = workflow_service.approve_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment
            )

            if record.status == ApprovalRecordStatusEnum.APPROVED:
                # 审批完成，允许合同签订
                contract.status = ContractStatusEnum.IN_REVIEW  # 保持待审批状态，等待签订

                # 发送审批完成通知
                from app.services.notification_service import notification_service, NotificationType, NotificationPriority
                try:
                    # 通知合同创建人审批已完成
                    notification_service.send_notification(
                        db=db,
                        recipient_id=contract.created_by,
                        notification_type=NotificationType.TASK_APPROVED,
                        title=f"合同审批已完成: {contract.contract_name}",
                        content=f"合同编号: {contract.contract_no}\n审批人: {current_user.real_name or current_user.username}",
                        priority=NotificationPriority.NORMAL,
                        link=f"/sales/contracts/{contract.id}"
                    )
                except Exception:
                    pass  # 通知失败不影响主流程
            message = "审批通过"

        elif action_request.action == ApprovalActionEnum.REJECT:
            record = workflow_service.reject_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment or "审批驳回"
            )
            contract.status = ContractStatusEnum.CANCELLED
            message = "审批已驳回"

            # 发送驳回通知
            from app.services.notification_service import notification_service, NotificationType, NotificationPriority
            try:
                notification_service.send_notification(
                    db=db,
                    recipient_id=contract.created_by,
                    notification_type=NotificationType.TASK_REJECTED,
                    title=f"合同审批已驳回: {contract.contract_name}",
                    content=f"合同编号: {contract.contract_no}\n驳回原因: {action_request.comment or '无'}",
                    priority=NotificationPriority.HIGH,
                    link=f"/sales/contracts/{contract.id}"
                )
            except Exception:
                pass

        elif action_request.action == ApprovalActionEnum.DELEGATE:
            if not action_request.delegate_to_id:
                raise HTTPException(status_code=400, detail="委托操作需要指定委托给的用户ID")

            record = workflow_service.delegate_step(
                record_id=record.id,
                approver_id=current_user.id,
                delegate_to_id=action_request.delegate_to_id,
                comment=action_request.comment
            )
            message = "审批已委托"

            # 发送委托通知
            from app.services.notification_service import notification_service, NotificationType
            try:
                notification_service.send_notification(
                    db=db,
                    recipient_id=action_request.delegate_to_id,
                    notification_type=NotificationType.TASK_ASSIGNED,
                    title=f"合同审批已委托给您: {contract.contract_name}",
                    content=f"原审批人: {current_user.real_name or current_user.username}\n合同编码: {contract.contract_no}",
                    priority=notification_service.NotificationPriority.NORMAL,
                    link=f"/sales/contracts/{contract.id}/approval"
                )
            except Exception:
                pass

        elif action_request.action == ApprovalActionEnum.WITHDRAW:
            record = workflow_service.withdraw_approval(
                record_id=record.id,
                initiator_id=current_user.id,
                comment=action_request.comment
            )
            contract.status = ContractStatusEnum.DRAFT
            message = "审批已撤回"

        else:
            raise HTTPException(status_code=400, detail=f"不支持的审批操作: {action_request.action}")

        db.commit()

        return ResponseModel(
            code=200,
            message=message,
            data={"approval_record_id": record.id, "status": record.status}
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/contracts/{contract_id}/approval-history", response_model=List[ApprovalHistoryResponse])
def get_contract_approval_history(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同审批历史
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.CONTRACT,
        entity_id=contract_id
    )
    
    if not record:
        return []
    
    history_list = workflow_service.get_approval_history(record.id)
    result = []
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        result.append(ApprovalHistoryResponse(**history_dict))
    
    return result


# ==================== 发票 ====================


@router.get("/invoices", response_model=PaginatedResponse[InvoiceResponse])
def read_invoices(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票列表
    Issue 7.1: 已集成数据权限过滤（财务可以看到所有发票）
    """
    query = db.query(Invoice).options(
        joinedload(Invoice.contract)
    )

    # Issue 7.1: 应用财务数据权限过滤（财务可以看到所有发票）
    # 注意：Invoice 模型没有 owner_id 字段，所以跳过此过滤
    # query = security.filter_sales_finance_data_by_scope(query, current_user, db, Invoice, 'owner_id')

    if keyword:
        query = query.filter(Invoice.invoice_code.contains(keyword))

    if status:
        query = query.filter(Invoice.status == status)

    if customer_id:
        # 通过 contract 关联过滤客户
        query = query.join(Contract).filter(Contract.customer_id == customer_id)

    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(desc(Invoice.created_at)).offset(offset).limit(page_size).all()

    invoice_responses = []
    for invoice in invoices:
        # 获取客户名称
        customer_name = None
        if invoice.contract and invoice.contract.customer:
            customer_name = invoice.contract.customer.customer_name

        invoice_dict = {
            **{c.name: getattr(invoice, c.name) for c in invoice.__table__.columns},
            "contract_code": invoice.contract.contract_code if invoice.contract else None,
            "customer_name": customer_name,
        }
        invoice_responses.append(InvoiceResponse(**invoice_dict))

    return PaginatedResponse(
        items=invoice_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/invoices_old", response_model=PaginatedResponse[InvoiceResponse])
def read_invoices(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票列表
    Issue 7.1: 已集成数据权限过滤（财务和销售总监可以看到所有发票）
    """
    query = db.query(Invoice).options(
        joinedload(Invoice.contract).joinedload(Contract.customer),
        joinedload(Invoice.project)
    )

    # Issue 7.1: 应用财务数据权限过滤（财务和销售总监可以看到所有发票）
    query = security.filter_sales_finance_data_by_scope(query, current_user, db, Invoice, 'owner_id')

    if keyword:
        query = query.filter(Invoice.invoice_code.contains(keyword))

    if status:
        query = query.filter(Invoice.status == status)

    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)

    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(desc(Invoice.created_at)).offset(offset).limit(page_size).all()

    invoice_responses = []
    for invoice in invoices:
        contract = invoice.contract
        project = invoice.project
        customer = contract.customer if contract else None
        invoice_dict = {
            **{c.name: getattr(invoice, c.name) for c in invoice.__table__.columns},
            "contract_code": contract.contract_code if contract else None,
            "project_code": project.project_code if project else None,
            "project_name": project.project_name if project else None,
            "customer_name": customer.customer_name if customer else None,
        }
        invoice_responses.append(InvoiceResponse(**invoice_dict))

    return PaginatedResponse(
        items=invoice_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/invoices", response_model=InvoiceResponse, status_code=201)
def create_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_in: InvoiceCreate,
    current_user: User = Depends(security.require_finance_access()),
) -> Any:
    """
    创建发票
    """
    invoice_data = invoice_in.model_dump()
    
    # 如果没有提供编码，自动生成
    if not invoice_data.get("invoice_code"):
        invoice_data["invoice_code"] = generate_invoice_code(db)
    else:
        existing = db.query(Invoice).filter(Invoice.invoice_code == invoice_data["invoice_code"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="发票编码已存在")

    contract = db.query(Contract).filter(Contract.id == invoice_data["contract_id"]).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    invoice = Invoice(**invoice_data)
    db.add(invoice)
    db.flush()
    
    # 如果发票状态是 APPLIED，自动启动审批流程
    if invoice.status == InvoiceStatus.APPLIED:
        try:
            workflow_service = ApprovalWorkflowService(db)
            routing_params = {"amount": float(invoice.amount or 0)}
            record = workflow_service.start_approval(
                entity_type=WorkflowTypeEnum.INVOICE,
                entity_id=invoice.id,
                initiator_id=current_user.id,
                workflow_id=None,  # 自动选择
                routing_params=routing_params,
                comment="发票申请"
            )
            invoice.status = InvoiceStatus.IN_REVIEW
        except Exception as e:
            # 如果启动审批失败，记录日志但不阻止发票创建
            # TODO: 添加日志记录
            pass
    
    db.commit()
    db.refresh(invoice)

    project = invoice.project
    customer = contract.customer if contract else None
    invoice_dict = {
        **{c.name: getattr(invoice, c.name) for c in invoice.__table__.columns},
        "contract_code": contract.contract_code,
        "project_code": project.project_code if project else None,
        "project_name": project.project_name if project else None,
        "customer_name": customer.customer_name if customer else None,
    }
    return InvoiceResponse(**invoice_dict)


@router.post("/invoices/{invoice_id}/issue", response_model=ResponseModel)
def issue_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    issue_request: InvoiceIssueRequest,
    current_user: User = Depends(security.require_finance_access()),
) -> Any:
    """
    开票
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 检查是否已通过审批（如果启用了审批工作流）
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.INVOICE,
        entity_id=invoice_id
    )
    
    if record and record.status != ApprovalRecordStatusEnum.APPROVED:
        raise HTTPException(status_code=400, detail="发票尚未通过审批，无法开票")
    
    invoice.issue_date = issue_request.issue_date
    invoice.status = InvoiceStatus.ISSUED
    invoice.payment_status = "PENDING"
    
    # 如果没有设置到期日期，默认设置为开票日期后30天
    if not invoice.due_date and invoice.issue_date:
        from datetime import timedelta
        invoice.due_date = invoice.issue_date + timedelta(days=30)
    
    db.commit()
    
    # 发送发票开具通知
    try:
        from app.services.sales_reminder_service import notify_invoice_issued
        notify_invoice_issued(db, invoice.id)
        db.commit()
    except Exception as e:
        # 通知失败不影响主流程
        pass

    return ResponseModel(code=200, message="发票开票成功")


@router.delete("/invoices/{invoice_id}", status_code=200)
def delete_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.require_finance_access()),
) -> Any:
    """
    删除发票（仅限草稿状态）
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 只有草稿状态才能删除
    if invoice.status != "DRAFT":
        raise HTTPException(status_code=400, detail="只有草稿状态的发票才能删除")

    db.delete(invoice)
    db.commit()

    return ResponseModel(code=200, message="发票已删除")


@router.post("/invoices/{invoice_id}/receive-payment", response_model=ResponseModel)
def receive_payment(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    paid_amount: str = Query(..., description="收款金额"),
    paid_date: date = Query(..., description="收款日期"),
    remark: Optional[str] = Query(None, description="备注"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    记录发票收款
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if invoice.status != "ISSUED":
        raise HTTPException(status_code=400, detail="只有已开票的发票才能记录收款")

    # 更新收款信息
    current_paid = invoice.paid_amount or Decimal("0")
    paid_amount_decimal = Decimal(str(paid_amount))
    new_paid = current_paid + paid_amount_decimal
    invoice.paid_amount = new_paid
    invoice.paid_date = paid_date

    # 更新收款状态
    total = invoice.total_amount or invoice.amount or Decimal("0")
    if new_paid >= total:
        invoice.payment_status = "PAID"
    elif new_paid > Decimal("0"):
        invoice.payment_status = "PARTIAL"
    else:
        invoice.payment_status = "PENDING"

    if remark:
        invoice.remark = (invoice.remark or "") + f"\n收款备注: {remark}"

    db.commit()

    return ResponseModel(code=200, message="收款记录成功", data={
        "paid_amount": float(new_paid),
        "payment_status": invoice.payment_status
    })


@router.post("/invoices/{invoice_id}/approval/start", response_model=ResponseModel)
def start_invoice_approval(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    approval_request: ApprovalStartRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    启动发票审批流程
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    if invoice.status != InvoiceStatus.APPLIED:
        raise HTTPException(status_code=400, detail="只有已申请状态的发票才能启动审批流程")
    
    # 获取发票金额用于路由
    routing_params = {
        "amount": float(invoice.amount or 0)
    }
    
    # 启动审批流程
    workflow_service = ApprovalWorkflowService(db)
    try:
        record = workflow_service.start_approval(
            entity_type=WorkflowTypeEnum.INVOICE,
            entity_id=invoice_id,
            initiator_id=current_user.id,
            workflow_id=approval_request.workflow_id,
            routing_params=routing_params,
            comment=approval_request.comment
        )
        
        # 更新发票状态
        invoice.status = InvoiceStatus.IN_REVIEW
        
        db.commit()
        
        return ResponseModel(
            code=200,
            message="审批流程已启动",
            data={"approval_record_id": record.id}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/invoices/{invoice_id}/approval-status", response_model=ApprovalStatusResponse)
def get_invoice_approval_status(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票审批状态
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.INVOICE,
        entity_id=invoice_id
    )
    
    if not record:
        return ApprovalStatusResponse(
            record=None,
            current_step_info=None,
            can_approve=False,
            can_reject=False,
            can_delegate=False,
            can_withdraw=False
        )
    
    current_step_info = workflow_service.get_current_step(record.id)
    
    can_approve = False
    can_reject = False
    can_delegate = False
    can_withdraw = False
    
    if record.status == ApprovalRecordStatusEnum.PENDING:
        if current_step_info:
            if current_step_info.get("approver_id") == current_user.id:
                can_approve = True
                can_reject = True
                if current_step_info.get("can_delegate"):
                    can_delegate = True
        
        if record.initiator_id == current_user.id:
            can_withdraw = True
    
    record_dict = {
        **{c.name: getattr(record, c.name) for c in record.__table__.columns},
        "workflow_name": record.workflow.workflow_name if record.workflow else None,
        "initiator_name": record.initiator.real_name if record.initiator else None,
        "history": []
    }
    
    history_list = workflow_service.get_approval_history(record.id)
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        record_dict["history"].append(ApprovalHistoryResponse(**history_dict))
    
    return ApprovalStatusResponse(
        record=ApprovalRecordResponse(**record_dict),
        current_step_info=current_step_info,
        can_approve=can_approve,
        can_reject=can_reject,
        can_delegate=can_delegate,
        can_withdraw=can_withdraw
    )


@router.post("/invoices/{invoice_id}/approval/action", response_model=ResponseModel)
def invoice_approval_action(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    action_request: ApprovalActionRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    发票审批操作（通过/驳回/委托/撤回）
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.INVOICE,
        entity_id=invoice_id
    )
    
    if not record:
        raise HTTPException(status_code=404, detail="审批记录不存在")
    
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    try:
        if action_request.action == ApprovalActionEnum.APPROVE:
            record = workflow_service.approve_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment
            )
            
            if record.status == ApprovalRecordStatusEnum.APPROVED:
                # 审批完成，允许开票
                invoice.status = InvoiceStatus.APPROVED
            message = "审批通过"
            
        elif action_request.action == ApprovalActionEnum.REJECT:
            record = workflow_service.reject_step(
                record_id=record.id,
                approver_id=current_user.id,
                comment=action_request.comment or "审批驳回"
            )
            invoice.status = InvoiceStatus.REJECTED
            message = "审批已驳回"
            
        elif action_request.action == ApprovalActionEnum.DELEGATE:
            if not action_request.delegate_to_id:
                raise HTTPException(status_code=400, detail="委托操作需要指定委托给的用户ID")
            
            record = workflow_service.delegate_step(
                record_id=record.id,
                approver_id=current_user.id,
                delegate_to_id=action_request.delegate_to_id,
                comment=action_request.comment
            )
            message = "审批已委托"
            
        elif action_request.action == ApprovalActionEnum.WITHDRAW:
            record = workflow_service.withdraw_approval(
                record_id=record.id,
                initiator_id=current_user.id,
                comment=action_request.comment
            )
            invoice.status = InvoiceStatus.APPLIED
            message = "审批已撤回"
            
        else:
            raise HTTPException(status_code=400, detail=f"不支持的审批操作: {action_request.action}")
        
        db.commit()
        
        return ResponseModel(
            code=200,
            message=message,
            data={"approval_record_id": record.id, "status": record.status}
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/invoices/{invoice_id}/approval-history", response_model=List[ApprovalHistoryResponse])
def get_invoice_approval_history(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票审批历史
    """
    workflow_service = ApprovalWorkflowService(db)
    record = workflow_service.get_approval_record(
        entity_type=WorkflowTypeEnum.INVOICE,
        entity_id=invoice_id
    )
    
    if not record:
        return []
    
    history_list = workflow_service.get_approval_history(record.id)
    result = []
    for h in history_list:
        history_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "approver_name": h.approver.real_name if h.approver else None,
            "delegate_to_name": h.delegate_to.real_name if h.delegate_to else None
        }
        result.append(ApprovalHistoryResponse(**history_dict))
    
    return result


# ==================== 回款争议 ====================


@router.get("/disputes", response_model=PaginatedResponse[ReceivableDisputeResponse])
def read_disputes(
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款争议列表
    """
    query = db.query(ReceivableDispute)

    if status:
        query = query.filter(ReceivableDispute.status == status)

    total = query.count()
    offset = (page - 1) * page_size
    disputes = query.order_by(desc(ReceivableDispute.created_at)).offset(offset).limit(page_size).all()

    dispute_responses = []
    for dispute in disputes:
        dispute_dict = {
            **{c.name: getattr(dispute, c.name) for c in dispute.__table__.columns},
            "responsible_name": dispute.responsible.real_name if dispute.responsible else None,
        }
        dispute_responses.append(ReceivableDisputeResponse(**dispute_dict))

    return PaginatedResponse(
        items=dispute_responses,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/disputes", response_model=ReceivableDisputeResponse, status_code=201)
def create_dispute(
    *,
    db: Session = Depends(deps.get_db),
    dispute_in: ReceivableDisputeCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建回款争议
    """
    dispute = ReceivableDispute(**dispute_in.model_dump())
    db.add(dispute)
    db.commit()
    db.refresh(dispute)

    dispute_dict = {
        **{c.name: getattr(dispute, c.name) for c in dispute.__table__.columns},
        "responsible_name": dispute.responsible.real_name if dispute.responsible else None,
    }
    return ReceivableDisputeResponse(**dispute_dict)


# ==================== 统计报表 ====================


@router.get("/statistics/funnel", response_model=ResponseModel)
def get_sales_funnel(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取销售漏斗统计
    """
    query_leads = db.query(Lead)
    query_opps = db.query(Opportunity)
    query_quotes = db.query(Quote)
    query_contracts = db.query(Contract)

    if start_date:
        query_leads = query_leads.filter(Lead.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_opps = query_opps.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_quotes = query_quotes.filter(Quote.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_contracts = query_contracts.filter(Contract.created_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query_leads = query_leads.filter(Lead.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_opps = query_opps.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_quotes = query_quotes.filter(Quote.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_contracts = query_contracts.filter(Contract.created_at <= datetime.combine(end_date, datetime.max.time()))

    # 统计各阶段数量
    leads_count = query_leads.count()
    opps_count = query_opps.count()
    quotes_count = query_quotes.count()
    contracts_count = query_contracts.count()
    
    # 统计金额
    won_opps = query_opps.filter(Opportunity.stage == "WON").all()
    total_opp_amount = sum([float(opp.est_amount or 0) for opp in won_opps])
    
    signed_contracts = query_contracts.filter(Contract.status == "SIGNED").all()
    total_contract_amount = sum([float(contract.contract_amount or 0) for contract in signed_contracts])

    return ResponseModel(
        code=200,
        message="success",
        data={
            "leads": leads_count,
            "opportunities": opps_count,
            "quotes": quotes_count,
            "contracts": contracts_count,
            "total_opportunity_amount": total_opp_amount,
            "total_contract_amount": total_contract_amount,
            "conversion_rates": {
                "lead_to_opp": round(opps_count / leads_count * 100, 2) if leads_count > 0 else 0,
                "opp_to_quote": round(quotes_count / opps_count * 100, 2) if opps_count > 0 else 0,
                "quote_to_contract": round(contracts_count / quotes_count * 100, 2) if quotes_count > 0 else 0,
            }
        }
    )


@router.get("/statistics/opportunities-by-stage", response_model=ResponseModel)
def get_opportunities_by_stage(
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    按阶段统计商机
    """
    stages = ["DISCOVERY", "QUALIFIED", "PROPOSAL", "NEGOTIATION", "WON", "LOST", "ON_HOLD"]
    result = {}
    
    for stage in stages:
        count = db.query(Opportunity).filter(Opportunity.stage == stage).count()
        total_amount = db.query(func.sum(Opportunity.est_amount)).filter(Opportunity.stage == stage).scalar() or 0
        result[stage] = {
            "count": count,
            "total_amount": float(total_amount)
        }
    
    return ResponseModel(code=200, message="success", data=result)


@router.get("/statistics/revenue-forecast", response_model=ResponseModel)
def get_revenue_forecast(
    db: Session = Depends(deps.get_db),
    months: int = Query(3, ge=1, le=12, description="预测月数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    收入预测（基于已签订合同和进行中的商机）
    """
    from datetime import timedelta
    from calendar import monthrange
    
    today = date.today()
    forecast = []
    
    for i in range(months):
        forecast_date = today + timedelta(days=30 * (i + 1))
        month_start = forecast_date.replace(day=1)
        _, last_day = monthrange(forecast_date.year, forecast_date.month)
        month_end = forecast_date.replace(day=last_day)
        
        # 统计该月预计签约的合同金额（基于商机预计金额）
        opps_in_month = (
            db.query(Opportunity)
            .filter(Opportunity.stage.in_(["PROPOSAL", "NEGOTIATION"]))
            .all()
        )
        
        # 简化处理：假设进行中的商机在接下来几个月平均分布
        estimated_revenue = sum([float(opp.est_amount or 0) for opp in opps_in_month]) / months
        
        forecast.append({
            "month": forecast_date.strftime("%Y-%m"),
            "estimated_revenue": round(estimated_revenue, 2)
        })
    
    return ResponseModel(code=200, message="success", data={"forecast": forecast})


@router.get("/statistics/prediction", response_model=ResponseModel)
def get_sales_prediction(
    *,
    db: Session = Depends(deps.get_db),
    days: int = Query(90, ge=30, le=365, description="预测天数（30/60/90）"),
    method: str = Query("moving_average", description="预测方法：moving_average/exponential_smoothing"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.3: 销售预测增强
    使用移动平均法或指数平滑法进行收入预测
    """
    from app.services.sales_prediction_service import SalesPredictionService
    
    service = SalesPredictionService(db)
    prediction = service.predict_revenue(
        days=days,
        method=method,
        customer_id=customer_id,
        owner_id=owner_id,
    )
    
    return ResponseModel(
        code=200,
        message="success",
        data=prediction
    )


@router.get("/opportunities/{opp_id}/win-probability", response_model=ResponseModel)
def get_opportunity_win_probability(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.3: 商机赢单概率预测
    基于商机阶段、金额、历史赢单率
    """
    from app.services.sales_prediction_service import SalesPredictionService
    
    service = SalesPredictionService(db)
    probability = service.predict_win_probability(opportunity_id=opp_id)
    
    return ResponseModel(
        code=200,
        message="success",
        data=probability
    )


@router.get("/statistics/prediction/accuracy", response_model=ResponseModel)
def get_prediction_accuracy(
    *,
    db: Session = Depends(deps.get_db),
    days_back: int = Query(90, ge=30, le=365, description="评估时间段（天数）"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.3: 预测准确度评估
    对比历史预测值和实际值
    """
    from app.services.sales_prediction_service import SalesPredictionService
    
    service = SalesPredictionService(db)
    accuracy = service.evaluate_prediction_accuracy(days_back=days_back)
    
    return ResponseModel(
        code=200,
        message="success",
        data=accuracy
    )


@router.get("/contracts/{contract_id}/deliverables", response_model=List[ContractDeliverableResponse])
def get_contract_deliverables(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同交付物清单
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract_id).all()
    return [ContractDeliverableResponse(**{c.name: getattr(d, c.name) for c in d.__table__.columns}) for d in deliverables]


@router.get("/contracts/{contract_id}/amendments", response_model=List[ContractAmendmentResponse])
def get_contract_amendments(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同变更记录列表
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    query = db.query(ContractAmendment).filter(ContractAmendment.contract_id == contract_id)
    
    if status:
        query = query.filter(ContractAmendment.status == status)
    
    amendments = query.order_by(desc(ContractAmendment.request_date), desc(ContractAmendment.created_at)).all()

    result = []
    for amendment in amendments:
        result.append({
            "id": amendment.id,
            "contract_id": amendment.contract_id,
            "amendment_no": amendment.amendment_no,
            "amendment_type": amendment.amendment_type,
            "title": amendment.title,
            "description": amendment.description,
            "reason": amendment.reason,
            "old_value": amendment.old_value,
            "new_value": amendment.new_value,
            "amount_change": amendment.amount_change,
            "schedule_impact": amendment.schedule_impact,
            "other_impact": amendment.other_impact,
            "requestor_id": amendment.requestor_id,
            "requestor_name": amendment.requestor.real_name if amendment.requestor else None,
            "request_date": amendment.request_date,
            "status": amendment.status,
            "approver_id": amendment.approver_id,
            "approver_name": amendment.approver.real_name if amendment.approver else None,
            "approval_date": amendment.approval_date,
            "approval_comment": amendment.approval_comment,
            "attachments": amendment.attachments,
            "created_at": amendment.created_at,
            "updated_at": amendment.updated_at,
        })

    return result


@router.post("/contracts/{contract_id}/amendments", response_model=ContractAmendmentResponse, status_code=201)
def create_contract_amendment(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    amendment_in: ContractAmendmentCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建合同变更记录
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # 生成变更编号
    amendment_no = generate_amendment_no(db, contract.contract_code)

    amendment = ContractAmendment(
        contract_id=contract_id,
        amendment_no=amendment_no,
        amendment_type=amendment_in.amendment_type,
        title=amendment_in.title,
        description=amendment_in.description,
        reason=amendment_in.reason,
        old_value=amendment_in.old_value,
        new_value=amendment_in.new_value,
        amount_change=amendment_in.amount_change,
        schedule_impact=amendment_in.schedule_impact,
        other_impact=amendment_in.other_impact,
        requestor_id=current_user.id,
        request_date=amendment_in.request_date,
        status="PENDING",
        attachments=amendment_in.attachments,
    )

    db.add(amendment)
    db.commit()
    db.refresh(amendment)

    return {
        "id": amendment.id,
        "contract_id": amendment.contract_id,
        "amendment_no": amendment.amendment_no,
        "amendment_type": amendment.amendment_type,
        "title": amendment.title,
        "description": amendment.description,
        "reason": amendment.reason,
        "old_value": amendment.old_value,
        "new_value": amendment.new_value,
        "amount_change": amendment.amount_change,
        "schedule_impact": amendment.schedule_impact,
        "other_impact": amendment.other_impact,
        "requestor_id": amendment.requestor_id,
        "requestor_name": amendment.requestor.real_name if amendment.requestor else None,
        "request_date": amendment.request_date,
        "status": amendment.status,
        "approver_id": amendment.approver_id,
        "approver_name": amendment.approver.real_name if amendment.approver else None,
        "approval_date": amendment.approval_date,
        "approval_comment": amendment.approval_comment,
        "attachments": amendment.attachments,
        "created_at": amendment.created_at,
        "updated_at": amendment.updated_at,
    }


@router.get("/quotes/{quote_id}/versions", response_model=List[QuoteVersionResponse])
def get_quote_versions(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价的所有版本
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    versions = db.query(QuoteVersion).filter(QuoteVersion.quote_id == quote_id).order_by(desc(QuoteVersion.created_at)).all()
    
    version_responses = []
    for v in versions:
        items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == v.id).all()
        v_dict = {
            **{c.name: getattr(v, c.name) for c in v.__table__.columns},
            "created_by_name": v.creator.real_name if v.creator else None,
            "approved_by_name": v.approver.real_name if v.approver else None,
            "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
        }
        version_responses.append(QuoteVersionResponse(**v_dict))
    
    return version_responses


@router.get("/statistics/summary", response_model=ResponseModel)
def get_sales_summary(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取销售汇总统计
    """
    query_leads = db.query(Lead)
    query_opps = db.query(Opportunity)
    query_contracts = db.query(Contract)
    query_invoices = db.query(Invoice)

    if start_date:
        query_leads = query_leads.filter(Lead.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_opps = query_opps.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_contracts = query_contracts.filter(Contract.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_invoices = query_invoices.filter(Invoice.created_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query_leads = query_leads.filter(Lead.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_opps = query_opps.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_contracts = query_contracts.filter(Contract.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_invoices = query_invoices.filter(Invoice.created_at <= datetime.combine(end_date, datetime.max.time()))

    # 线索统计
    total_leads = query_leads.count()
    converted_leads = query_leads.filter(Lead.status == "CONVERTED").count()

    # 商机统计
    total_opportunities = query_opps.count()
    won_opportunities = query_opps.filter(Opportunity.stage == "WON").count()

    # 合同统计
    signed_contracts = query_contracts.filter(Contract.status == "SIGNED").all()
    total_contract_amount = sum([float(c.contract_amount or 0) for c in signed_contracts])

    # 发票统计
    paid_invoices = query_invoices.filter(Invoice.payment_status == "PAID").all()
    paid_amount = sum([float(inv.paid_amount or 0) for inv in paid_invoices])

    # 计算转化率
    conversion_rate = round((converted_leads / total_leads * 100), 2) if total_leads > 0 else 0
    win_rate = round((won_opportunities / total_opportunities * 100), 2) if total_opportunities > 0 else 0

    return ResponseModel(
        code=200,
        message="success",
        data={
            "total_leads": total_leads,
            "converted_leads": converted_leads,
            "total_opportunities": total_opportunities,
            "won_opportunities": won_opportunities,
            "total_contract_amount": total_contract_amount,
            "paid_amount": paid_amount,
            "conversion_rate": conversion_rate,
            "win_rate": win_rate,
        }
    )


# ==================== 线索管理补充 ====================


@router.put("/leads/{lead_id}/invalid", response_model=ResponseModel)
def mark_lead_invalid(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    reason: Optional[str] = Query(None, description="无效原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    标记线索无效
    """
    from app.models.enums import LeadStatusEnum
    
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    if lead.status == LeadStatusEnum.CONVERTED:
        raise HTTPException(status_code=400, detail="已转商机的线索不能标记为无效")

    lead.status = LeadStatusEnum.INVALID
    db.commit()

    # 可选：记录一条跟进记录说明无效原因
    if reason:
        follow_up = LeadFollowUp(
            lead_id=lead_id,
            follow_up_type="OTHER",
            content=f"标记为无效：{reason}",
            created_by=current_user.id,
        )
        db.add(follow_up)
        db.commit()

    return ResponseModel(code=200, message="线索已标记为无效")


# ==================== 商机管理补充 ====================


@router.put("/opportunities/{opp_id}/stage", response_model=OpportunityResponse)
def update_opportunity_stage(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    stage: str = Query(..., description="新阶段"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新商机阶段
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    valid_stages = ["DISCOVERY", "QUALIFIED", "PROPOSAL", "NEGOTIATION", "WON", "LOST", "ON_HOLD"]
    if stage not in valid_stages:
        raise HTTPException(status_code=400, detail=f"无效的阶段，必须是: {', '.join(valid_stages)}")

    opportunity.stage = stage
    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.put("/opportunities/{opp_id}/score", response_model=OpportunityResponse)
def update_opportunity_score(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    score: int = Query(..., ge=0, le=100, description="评分"),
    score_remark: Optional[str] = Query(None, description="评分说明"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    商机评分（准入评估）
    评分范围：0-100分
    评分标准：
    - 90-100分：优秀，优先跟进
    - 70-89分：良好，正常跟进
    - 60-69分：一般，谨慎跟进
    - 0-59分：较差，需要评估是否继续
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    old_score = opportunity.score
    opportunity.score = score
    
    # 根据评分自动更新风险等级
    if score >= 80:
        opportunity.risk_level = "LOW"
    elif score >= 60:
        opportunity.risk_level = "MEDIUM"
    else:
        opportunity.risk_level = "HIGH"
    
    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.put("/opportunities/{opp_id}/win", response_model=OpportunityResponse)
def win_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    赢单
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    opportunity.stage = "WON"
    opportunity.gate_status = "PASS"
    opportunity.gate_passed_at = datetime.now()
    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.put("/opportunities/{opp_id}/lose", response_model=OpportunityResponse)
def lose_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    reason: Optional[str] = Query(None, description="丢单原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    丢单
    """
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")

    opportunity.stage = "LOST"
    db.commit()
    db.refresh(opportunity)

    req = db.query(OpportunityRequirement).filter(OpportunityRequirement.opportunity_id == opportunity.id).first()
    opp_dict = {
        **{c.name: getattr(opportunity, c.name) for c in opportunity.__table__.columns},
        "customer_name": opportunity.customer.customer_name if opportunity.customer else None,
        "owner_name": opportunity.owner.real_name if opportunity.owner else None,
        "requirement": None,
    }
    if req:
        from app.schemas.sales import OpportunityRequirementResponse
        opp_dict["requirement"] = OpportunityRequirementResponse(**{c.name: getattr(req, c.name) for c in req.__table__.columns})

    return OpportunityResponse(**opp_dict)


@router.get("/opportunities/funnel", response_model=ResponseModel)
def get_opportunity_funnel(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    商机漏斗统计
    按阶段统计商机数量和金额，计算转化率
    """
    from app.models.enums import OpportunityStageEnum
    
    query = db.query(Opportunity)
    
    if start_date:
        query = query.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
    if owner_id:
        query = query.filter(Opportunity.owner_id == owner_id)
    
    stages = [stage.value for stage in OpportunityStageEnum]
    funnel_data = {}
    total_count = 0
    total_amount = 0
    
    for stage in stages:
        stage_query = query.filter(Opportunity.stage == stage)
        count = stage_query.count()
        amount_query = db.query(func.sum(Opportunity.est_amount)).filter(Opportunity.stage == stage)
        if start_date:
            amount_query = amount_query.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            amount_query = amount_query.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
        if owner_id:
            amount_query = amount_query.filter(Opportunity.owner_id == owner_id)
        amount_result = amount_query.scalar() or 0
        
        funnel_data[stage] = {
            "count": count,
            "total_amount": float(amount_result),
            "avg_amount": float(amount_result / count) if count > 0 else 0
        }
        total_count += count
        total_amount += float(amount_result)
    
    # 计算转化率（从前一阶段到当前阶段）
    conversion_rates = {}
    prev_count = None
    for stage in stages:
        current_count = funnel_data[stage]["count"]
        if prev_count is not None and prev_count > 0:
            conversion_rates[stage] = round((current_count / prev_count) * 100, 2)
        else:
            conversion_rates[stage] = 100.0 if current_count > 0 else 0.0
        prev_count = current_count
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "funnel": funnel_data,
            "conversion_rates": conversion_rates,
            "total_count": total_count,
            "total_amount": total_amount
        }
    )


# ==================== 报价管理补充 ====================


@router.get("/quotes/{quote_id}", response_model=QuoteResponse)
def read_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价详情
    """
    quote = db.query(Quote).options(joinedload(Quote.opportunity), joinedload(Quote.customer)).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    versions = db.query(QuoteVersion).filter(QuoteVersion.quote_id == quote.id).all()
    quote_dict = {
        **{c.name: getattr(quote, c.name) for c in quote.__table__.columns},
        "opportunity_code": quote.opportunity.opp_code if quote.opportunity else None,
        "customer_name": quote.customer.customer_name if quote.customer else None,
        "owner_name": quote.owner.real_name if quote.owner else None,
        "versions": [],
    }
    for v in versions:
        items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == v.id).all()
        v_dict = {
            **{c.name: getattr(v, c.name) for c in v.__table__.columns},
            "created_by_name": v.creator.real_name if v.creator else None,
            "approved_by_name": v.approver.real_name if v.approver else None,
            "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
        }
        quote_dict["versions"].append(QuoteVersionResponse(**v_dict))
        if v.id == quote.current_version_id:
            quote_dict["current_version"] = QuoteVersionResponse(**v_dict)

    return QuoteResponse(**quote_dict)


@router.put("/quotes/{quote_id}", response_model=QuoteResponse)
def update_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    quote_in: QuoteUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新报价
    Issue 7.2: 已集成操作权限检查
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    # Issue 7.2: 检查编辑权限
    if not security.check_sales_edit_permission(
        current_user,
        db,
        _get_entity_creator_id(quote),
        quote.owner_id,
    ):
        raise HTTPException(status_code=403, detail="您没有权限编辑此报价")

    update_data = quote_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(quote, field, value)

    db.commit()
    db.refresh(quote)

    versions = db.query(QuoteVersion).filter(QuoteVersion.quote_id == quote.id).all()
    quote_dict = {
        **{c.name: getattr(quote, c.name) for c in quote.__table__.columns},
        "opportunity_code": quote.opportunity.opp_code if quote.opportunity else None,
        "customer_name": quote.customer.customer_name if quote.customer else None,
        "owner_name": quote.owner.real_name if quote.owner else None,
        "versions": [],
    }
    for v in versions:
        items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == v.id).all()
        v_dict = {
            **{c.name: getattr(v, c.name) for c in v.__table__.columns},
            "created_by_name": v.creator.real_name if v.creator else None,
            "approved_by_name": v.approver.real_name if v.approver else None,
            "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
        }
        quote_dict["versions"].append(QuoteVersionResponse(**v_dict))
        if v.id == quote.current_version_id:
            quote_dict["current_version"] = QuoteVersionResponse(**v_dict)

    return QuoteResponse(**quote_dict)


@router.get("/quotes/{quote_id}/items", response_model=List[QuoteItemResponse])
def get_quote_items(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取报价明细列表
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")

    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    return [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items]


@router.post("/quotes/{quote_id}/items", response_model=QuoteItemResponse, status_code=201)
def create_quote_item(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    item_in: QuoteItemCreate,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    添加报价明细
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")

    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")

    item = QuoteItem(quote_version_id=target_version_id, **item_in.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)

    return QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns})


@router.put("/quotes/{quote_id}/items/{item_id}", response_model=QuoteItemResponse)
def update_quote_item(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    item_id: int,
    item_in: QuoteItemUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新报价明细
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    item = db.query(QuoteItem).filter(
        QuoteItem.id == item_id,
        QuoteItem.quote_version_id.in_([v.id for v in quote.versions])
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="报价明细不存在")
    
    # 更新字段
    update_data = item_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(item, field):
            setattr(item, field, value)
    
    db.add(item)
    db.commit()
    db.refresh(item)
    
    return QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns})


@router.delete("/quotes/{quote_id}/items/{item_id}", status_code=200)
def delete_quote_item(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    item_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除报价明细
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    item = db.query(QuoteItem).filter(
        QuoteItem.id == item_id,
        QuoteItem.quote_version_id.in_([v.id for v in quote.versions])
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="报价明细不存在")
    
    db.delete(item)
    db.commit()
    
    return ResponseModel(code=200, message="删除成功")


@router.put("/quotes/{quote_id}/items/batch", response_model=ResponseModel)
def batch_update_quote_items(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    batch_data: QuoteItemBatchUpdate,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    批量更新报价明细
    支持更新、新增、删除操作
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")
    
    updated_count = 0
    created_count = 0
    
    # 处理批量更新
    for item_data in batch_data.items:
        item_dict = item_data.model_dump(exclude_unset=True)
        item_id = item_dict.pop('id', None)
        
        if item_id:
            # 更新现有明细
            item = db.query(QuoteItem).filter(
                QuoteItem.id == item_id,
                QuoteItem.quote_version_id == target_version_id
            ).first()
            
            if item:
                for field, value in item_dict.items():
                    if hasattr(item, field) and field != 'id':
                        setattr(item, field, value)
                db.add(item)
                updated_count += 1
        else:
            # 创建新明细
            item = QuoteItem(quote_version_id=target_version_id, **item_dict)
            db.add(item)
            created_count += 1
    
    db.commit()
    
    # 重新计算成本和毛利率
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    total_price = sum([float(item.qty or 0) * float(item.unit_price or 0) for item in items])
    total_cost = sum([float(item.cost or 0) * float(item.qty or 0) for item in items])
    gross_margin = ((total_price - total_cost) / total_price * 100) if total_price > 0 else 0
    
    version.total_price = total_price
    version.cost_total = total_cost
    version.gross_margin = gross_margin
    db.add(version)
    db.commit()
    
    return ResponseModel(
        code=200,
        message=f"批量更新完成，新增 {created_count} 条，更新 {updated_count} 条",
        data={
            "created": created_count,
            "updated": updated_count,
            "total_price": total_price,
            "total_cost": total_cost,
            "gross_margin": gross_margin
        }
    )


@router.get("/quotes/{quote_id}/cost-breakdown", response_model=ResponseModel)
def get_quote_cost_breakdown(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    报价成本拆解
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")

    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")

    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()

    total_price = float(version.total_price or 0)
    total_cost = float(version.cost_total or 0)
    gross_margin = float(version.gross_margin or 0) if version.gross_margin else (total_price - total_cost) / total_price * 100 if total_price > 0 else 0

    cost_breakdown = []
    for item in items:
        item_price = float(item.qty or 0) * float(item.unit_price or 0)
        item_cost = float(item.cost or 0) * float(item.qty or 0)
        item_margin = (item_price - item_cost) / item_price * 100 if item_price > 0 else 0
        cost_breakdown.append({
            "item_name": item.item_name,
            "item_type": item.item_type,
            "qty": float(item.qty or 0),
            "unit_price": float(item.unit_price or 0),
            "total_price": item_price,
            "unit_cost": float(item.cost or 0),
            "total_cost": item_cost,
            "margin": round(item_margin, 2)
        })

    return ResponseModel(
        code=200,
        message="success",
        data={
            "total_price": total_price,
            "total_cost": total_cost,
            "gross_margin": round(gross_margin, 2),
            "breakdown": cost_breakdown
        }
    )


@router.put("/quotes/{quote_id}/submit", response_model=ResponseModel)
def submit_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    提交审批
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    if not quote.current_version_id:
        raise HTTPException(status_code=400, detail="报价没有当前版本")

    quote.status = "PENDING_APPROVAL"
    db.commit()

    return ResponseModel(code=200, message="报价已提交审批")


@router.put("/quotes/{quote_id}/reject", response_model=ResponseModel)
def reject_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    reason: Optional[str] = Query(None, description="驳回原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批驳回
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    quote.status = "REJECTED"
    db.commit()

    return ResponseModel(code=200, message="报价已驳回")


@router.put("/quotes/{quote_id}/send", response_model=ResponseModel)
def send_quote_to_customer(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    send_method: Optional[str] = Query("EMAIL", description="发送方式：EMAIL/PRINT/OTHER"),
    send_to: Optional[str] = Query(None, description="发送对象（邮箱/联系人等）"),
    remark: Optional[str] = Query(None, description="发送备注"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    发送报价给客户
    只有已审批通过的报价才能发送给客户
    """
    from app.models.enums import QuoteStatusEnum
    
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    if quote.status != QuoteStatusEnum.APPROVED:
        raise HTTPException(status_code=400, detail="只有已审批通过的报价才能发送给客户")

    # 更新报价状态为已发送（如果有SENT状态，否则保持APPROVED）
    # 这里简化处理，不改变状态，只记录发送操作
    
    # 可选：记录发送日志或通知
    
    return ResponseModel(
        code=200,
        message="报价已发送给客户",
        data={
            "quote_id": quote_id,
            "send_method": send_method,
            "send_to": send_to,
            "sent_at": datetime.now().isoformat()
        }
    )


# ==================== 合同管理补充 ====================


@router.get("/contracts/{contract_id}", response_model=ContractResponse)
def read_contract(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取合同详情
    """
    contract = db.query(Contract).options(
        joinedload(Contract.customer),
        joinedload(Contract.project),
        joinedload(Contract.opportunity),
        joinedload(Contract.owner)
    ).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract.id).all()
    contract_dict = {
        **{c.name: getattr(contract, c.name) for c in contract.__table__.columns},
        "opportunity_code": contract.opportunity.opp_code if contract.opportunity else None,
        "customer_name": contract.customer.customer_name if contract.customer else None,
        "project_code": contract.project.project_code if contract.project else None,
        "owner_name": contract.owner.real_name if contract.owner else None,
        "deliverables": [ContractDeliverableResponse(**{c.name: getattr(d, c.name) for c in d.__table__.columns}) for d in deliverables],
    }
    return ContractResponse(**contract_dict)


@router.put("/contracts/{contract_id}", response_model=ContractResponse)
def update_contract(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    contract_in: ContractUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新合同
    Issue 7.2: 已集成操作权限检查
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # Issue 7.2: 检查编辑权限
    if not security.check_sales_edit_permission(
        current_user,
        db,
        _get_entity_creator_id(contract),
        contract.owner_id,
    ):
        raise HTTPException(status_code=403, detail="您没有权限编辑此合同")

    update_data = contract_in.model_dump(exclude_unset=True)
    
    # 记录需要同步的字段
    need_sync = any(field in update_data for field in ["contract_amount", "signed_date", "delivery_deadline"])
    
    for field, value in update_data.items():
        setattr(contract, field, value)

    # Sprint 2.4: 合同变更时自动同步到项目
    if need_sync and contract.project_id:
        try:
            from app.services.data_sync_service import DataSyncService
            sync_service = DataSyncService(db)
            sync_result = sync_service.sync_contract_to_project(contract_id)
            if sync_result.get("success"):
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"合同变更已同步到项目：{sync_result.get('message')}")
        except Exception as e:
            # 同步失败不影响合同更新，记录日志
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"合同变更同步到项目失败：{str(e)}", exc_info=True)
    
    db.commit()
    db.refresh(contract)

    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract.id).all()
    contract_dict = {
        **{c.name: getattr(contract, c.name) for c in contract.__table__.columns},
        "opportunity_code": contract.opportunity.opp_code if contract.opportunity else None,
        "customer_name": contract.customer.customer_name if contract.customer else None,
        "project_code": contract.project.project_code if contract.project else None,
        "owner_name": contract.owner.real_name if contract.owner else None,
        "deliverables": [ContractDeliverableResponse(**{c.name: getattr(d, c.name) for c in d.__table__.columns}) for d in deliverables],
    }
    return ContractResponse(**contract_dict)


@router.put("/contracts/{contract_id}/approve", response_model=ResponseModel)
def approve_contract(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    approved: bool = Query(..., description="是否批准"),
    remark: Optional[str] = Query(None, description="审批意见"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    合同审批
    """
    # 检查审批权限
    if not security.has_sales_approval_access(current_user, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批合同"
        )

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    if approved:
        contract.status = "APPROVED"
    else:
        contract.status = "REJECTED"

    db.commit()

    return ResponseModel(code=200, message="合同审批完成" if approved else "合同已驳回")


# ==================== 开票管理补充 ====================


@router.get("/invoices/{invoice_id}", response_model=InvoiceResponse)
def read_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票详情
    """
    invoice = db.query(Invoice).options(
        joinedload(Invoice.contract).joinedload(Contract.customer),
        joinedload(Invoice.project)
    ).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    contract = invoice.contract
    project = invoice.project
    customer = contract.customer if contract else None
    invoice_dict = {
        **{c.name: getattr(invoice, c.name) for c in invoice.__table__.columns},
        "contract_code": contract.contract_code if contract else None,
        "project_code": project.project_code if project else None,
        "project_name": project.project_name if project else None,
        "customer_name": customer.customer_name if customer else None,
    }
    return InvoiceResponse(**invoice_dict)


@router.put("/invoices/{invoice_id}/approve", response_model=ResponseModel)
def approve_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    approved: bool = Query(..., description="是否批准"),
    remark: Optional[str] = Query(None, description="审批意见"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    开票审批（单级审批，兼容旧接口）
    """
    # 检查审批权限
    if not security.has_sales_approval_access(current_user, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批发票"
        )

    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if approved:
        invoice.status = "APPROVED"
    else:
        invoice.status = "REJECTED"

    db.commit()

    return ResponseModel(code=200, message="发票审批完成" if approved else "发票已驳回")


@router.put("/invoices/{invoice_id}/submit-approval", response_model=ResponseModel)
def submit_invoice_for_approval(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    提交发票审批（创建多级审批记录）
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 检查是否已有审批记录
    existing_approvals = db.query(InvoiceApproval).filter(InvoiceApproval.invoice_id == invoice_id).count()
    if existing_approvals > 0:
        raise HTTPException(status_code=400, detail="发票已提交审批，请勿重复提交")

    # 根据发票金额确定审批流程
    invoice_amount = float(invoice.total_amount or invoice.amount or 0)
    
    # 审批流程：根据金额确定审批层级
    # 小于10万：财务（1级）
    # 10-50万：财务（1级）+ 财务经理（2级）
    # 大于50万：财务（1级）+ 财务经理（2级）+ 财务总监（3级）
    
    approval_levels = []
    if invoice_amount < 100000:
        approval_levels = [1]  # 财务
    elif invoice_amount < 500000:
        approval_levels = [1, 2]  # 财务 + 财务经理
    else:
        approval_levels = [1, 2, 3]  # 财务 + 财务经理 + 财务总监

    # 创建审批记录
    from datetime import timedelta
    role_map = {1: "财务", 2: "财务经理", 3: "财务总监"}
    for level in approval_levels:
        approval = InvoiceApproval(
            invoice_id=invoice_id,
            approval_level=level,
            approval_role=role_map.get(level, "审批人"),
            status="PENDING",
            due_date=datetime.now() + timedelta(days=2)  # 默认2天审批期限
        )
        db.add(approval)

    invoice.status = "APPLIED"
    db.commit()

    return ResponseModel(code=200, message="发票已提交审批", data={"approval_levels": len(approval_levels)})


@router.get("/invoices/{invoice_id}/approvals", response_model=List[InvoiceApprovalResponse])
def get_invoice_approvals(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取发票审批记录列表
    """
    approvals = db.query(InvoiceApproval).filter(InvoiceApproval.invoice_id == invoice_id).order_by(InvoiceApproval.approval_level).all()
    
    result = []
    for approval in approvals:
        approver_name = None
        if approval.approver_id:
            approver = db.query(User).filter(User.id == approval.approver_id).first()
            approver_name = approver.real_name if approver else None
        
        result.append(InvoiceApprovalResponse(
            id=approval.id,
            invoice_id=approval.invoice_id,
            approval_level=approval.approval_level,
            approval_role=approval.approval_role,
            approver_id=approval.approver_id,
            approver_name=approver_name,
            approval_result=approval.approval_result,
            approval_opinion=approval.approval_opinion,
            status=approval.status,
            approved_at=approval.approved_at,
            due_date=approval.due_date,
            is_overdue=approval.is_overdue or False,
            created_at=approval.created_at,
            updated_at=approval.updated_at
        ))
    
    return result


@router.put("/invoice-approvals/{approval_id}/approve", response_model=InvoiceApprovalResponse)
def approve_invoice_approval(
    *,
    db: Session = Depends(deps.get_db),
    approval_id: int,
    approval_opinion: Optional[str] = Query(None, description="审批意见"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批通过（多级审批）
    """
    approval = db.query(InvoiceApproval).filter(InvoiceApproval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")

    if approval.status != "PENDING":
        raise HTTPException(status_code=400, detail="只能审批待审批状态的记录")

    # 检查审批权限
    if not security.check_sales_approval_permission(current_user, approval, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批此记录"
        )

    approval.approval_result = "APPROVED"
    approval.approval_opinion = approval_opinion
    approval.approved_at = datetime.now()
    approval.status = "COMPLETED"
    approval.approver_id = current_user.id
    approver = db.query(User).filter(User.id == current_user.id).first()
    if approver:
        approval.approver_name = approver.real_name

    # 检查是否所有审批都已完成
    invoice = db.query(Invoice).filter(Invoice.id == approval.invoice_id).first()
    if invoice:
        pending_approvals = db.query(InvoiceApproval).filter(
            InvoiceApproval.invoice_id == approval.invoice_id,
            InvoiceApproval.status == "PENDING"
        ).count()

        if pending_approvals == 0:
            # 所有审批都已完成，更新发票状态
            invoice.status = "APPROVED"

    db.commit()
    db.refresh(approval)

    approver_name = approval.approver_name
    return InvoiceApprovalResponse(
        id=approval.id,
        invoice_id=approval.invoice_id,
        approval_level=approval.approval_level,
        approval_role=approval.approval_role,
        approver_id=approval.approver_id,
        approver_name=approver_name,
        approval_result=approval.approval_result,
        approval_opinion=approval.approval_opinion,
        status=approval.status,
        approved_at=approval.approved_at,
        due_date=approval.due_date,
        is_overdue=approval.is_overdue or False,
        created_at=approval.created_at,
        updated_at=approval.updated_at
    )


@router.put("/invoice-approvals/{approval_id}/reject", response_model=InvoiceApprovalResponse)
def reject_invoice_approval(
    *,
    db: Session = Depends(deps.get_db),
    approval_id: int,
    rejection_reason: str = Query(..., description="驳回原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批驳回（多级审批）
    """
    approval = db.query(InvoiceApproval).filter(InvoiceApproval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")

    if approval.status != "PENDING":
        raise HTTPException(status_code=400, detail="只能审批待审批状态的记录")

    # 检查审批权限
    if not security.check_sales_approval_permission(current_user, approval, db):
        raise HTTPException(
            status_code=403,
            detail="您没有权限审批此记录"
        )

    approval.approval_result = "REJECTED"
    approval.approval_opinion = rejection_reason
    approval.approved_at = datetime.now()
    approval.status = "COMPLETED"
    approval.approver_id = current_user.id
    approver = db.query(User).filter(User.id == current_user.id).first()
    if approver:
        approval.approver_name = approver.real_name

    # 驳回后，发票状态变为被拒
    invoice = db.query(Invoice).filter(Invoice.id == approval.invoice_id).first()
    if invoice:
        invoice.status = "REJECTED"

    db.commit()
    db.refresh(approval)

    approver_name = approval.approver_name
    return InvoiceApprovalResponse(
        id=approval.id,
        invoice_id=approval.invoice_id,
        approval_level=approval.approval_level,
        approval_role=approval.approval_role,
        approver_id=approval.approver_id,
        approver_name=approver_name,
        approval_result=approval.approval_result,
        approval_opinion=approval.approval_opinion,
        status=approval.status,
        approved_at=approval.approved_at,
        due_date=approval.due_date,
        is_overdue=approval.is_overdue or False,
        created_at=approval.created_at,
        updated_at=approval.updated_at
    )


@router.put("/invoices/{invoice_id}/void", response_model=ResponseModel)
def void_invoice(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    reason: Optional[str] = Query(None, description="作废原因"),
    current_user: User = Depends(security.require_finance_access()),
) -> Any:
    """
    作废发票
    """
    from app.models.enums import InvoiceStatusEnum
    
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 只有已开票或已审批的发票才能作废
    if invoice.status not in [InvoiceStatusEnum.ISSUED, InvoiceStatusEnum.APPROVED]:
        raise HTTPException(status_code=400, detail="只有已开票或已审批的发票才能作废")

    # 如果已收款，不能作废
    if invoice.paid_amount and invoice.paid_amount > 0:
        raise HTTPException(status_code=400, detail="已收款的发票不能作废，请先处理收款")

    invoice.status = InvoiceStatusEnum.VOIDED
    if reason:
        invoice.remark = (invoice.remark or "") + f"\n作废原因: {reason}"
    
    db.commit()

    return ResponseModel(code=200, message="发票已作废")


# ==================== 销售报表补充 ====================


@router.get("/reports/sales-funnel", response_model=ResponseModel)
def get_sales_funnel_report(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    销售漏斗报表
    """
    # 复用已有的统计逻辑
    return get_sales_funnel(db, start_date, end_date, current_user)


@router.get("/reports/win-loss", response_model=ResponseModel)
def get_win_loss_analysis(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    赢单/丢单分析
    """
    query = db.query(Opportunity)

    if start_date:
        query = query.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query = query.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))

    won_opps = query.filter(Opportunity.stage == "WON").all()
    lost_opps = query.filter(Opportunity.stage == "LOST").all()

    won_count = len(won_opps)
    lost_count = len(lost_opps)
    total_count = won_count + lost_count
    win_rate = round(won_count / total_count * 100, 2) if total_count > 0 else 0

    won_amount = sum([float(opp.est_amount or 0) for opp in won_opps])
    lost_amount = sum([float(opp.est_amount or 0) for opp in lost_opps])

    return ResponseModel(
        code=200,
        message="success",
        data={
            "won": {
                "count": won_count,
                "amount": won_amount
            },
            "lost": {
                "count": lost_count,
                "amount": lost_amount
            },
            "win_rate": win_rate,
            "total_count": total_count
        }
    )


@router.get("/reports/sales-performance", response_model=ResponseModel)
def get_sales_performance(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    owner_id: Optional[int] = Query(None, description="负责人ID"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    销售业绩统计
    """
    query_opps = db.query(Opportunity)
    query_contracts = db.query(Contract)
    query_invoices = db.query(Invoice)

    if start_date:
        query_opps = query_opps.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_contracts = query_contracts.filter(Contract.created_at >= datetime.combine(start_date, datetime.min.time()))
        query_invoices = query_invoices.filter(Invoice.created_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query_opps = query_opps.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_contracts = query_contracts.filter(Contract.created_at <= datetime.combine(end_date, datetime.max.time()))
        query_invoices = query_invoices.filter(Invoice.created_at <= datetime.combine(end_date, datetime.max.time()))

    if owner_id:
        query_opps = query_opps.filter(Opportunity.owner_id == owner_id)
        query_contracts = query_contracts.filter(Contract.owner_id == owner_id)

    won_opps = query_opps.filter(Opportunity.stage == "WON").all()
    signed_contracts = query_contracts.filter(Contract.status == "SIGNED").all()
    issued_invoices = query_invoices.filter(Invoice.status == "ISSUED").all()

    total_opp_amount = sum([float(opp.est_amount or 0) for opp in won_opps])
    total_contract_amount = sum([float(c.contract_amount or 0) for c in signed_contracts])
    total_invoice_amount = sum([float(inv.amount or 0) for inv in issued_invoices])

    return ResponseModel(
        code=200,
        message="success",
        data={
            "won_opportunities": len(won_opps),
            "total_opportunity_amount": total_opp_amount,
            "signed_contracts": len(signed_contracts),
            "total_contract_amount": total_contract_amount,
            "issued_invoices": len(issued_invoices),
            "total_invoice_amount": total_invoice_amount
        }
    )


@router.get("/reports/customer-contribution", response_model=ResponseModel)
def get_customer_contribution(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    top_n: int = Query(10, ge=1, le=50, description="返回前N名"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    客户贡献分析
    """
    query_contracts = db.query(Contract).filter(Contract.status == "SIGNED")

    if start_date:
        query_contracts = query_contracts.filter(Contract.created_at >= datetime.combine(start_date, datetime.min.time()))

    if end_date:
        query_contracts = query_contracts.filter(Contract.created_at <= datetime.combine(end_date, datetime.max.time()))

    contracts = query_contracts.all()

    # 按客户统计
    customer_stats = {}
    for contract in contracts:
        customer_id = contract.customer_id
        if customer_id not in customer_stats:
            customer = contract.customer
            customer_stats[customer_id] = {
                "customer_id": customer_id,
                "customer_name": customer.customer_name if customer else None,
                "contract_count": 0,
                "total_amount": 0
            }
        customer_stats[customer_id]["contract_count"] += 1
        customer_stats[customer_id]["total_amount"] += float(contract.contract_amount or 0)

    # 排序并取前N名
    sorted_customers = sorted(customer_stats.values(), key=lambda x: x["total_amount"], reverse=True)[:top_n]

    return ResponseModel(
        code=200,
        message="success",
        data={
            "customers": sorted_customers,
            "total_customers": len(customer_stats)
        }
    )


@router.get("/reports/o2c-pipeline", response_model=ResponseModel)
def get_o2c_pipeline(
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    O2C流程全链路统计
    """
    from datetime import date as date_type
    today = date_type.today()

    # 线索统计
    query_leads = db.query(Lead)
    if start_date:
        query_leads = query_leads.filter(Lead.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_leads = query_leads.filter(Lead.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_leads = query_leads.count()
    converted_leads = query_leads.filter(Lead.status == "CONVERTED").count()

    # 商机统计
    query_opps = db.query(Opportunity)
    if start_date:
        query_opps = query_opps.filter(Opportunity.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_opps = query_opps.filter(Opportunity.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_opps = query_opps.count()
    won_opps = query_opps.filter(Opportunity.stage == "WON").all()
    won_count = len(won_opps)
    won_amount = sum([float(opp.est_amount or 0) for opp in won_opps])

    # 报价统计
    query_quotes = db.query(Quote)
    if start_date:
        query_quotes = query_quotes.filter(Quote.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_quotes = query_quotes.filter(Quote.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_quotes = query_quotes.count()
    approved_quotes = query_quotes.filter(Quote.status == "APPROVED").all()
    approved_amount = sum([float(q.total_price or 0) for q in approved_quotes])

    # 合同统计
    query_contracts = db.query(Contract)
    if start_date:
        query_contracts = query_contracts.filter(Contract.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_contracts = query_contracts.filter(Contract.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_contracts = query_contracts.count()
    signed_contracts = query_contracts.filter(Contract.status == "SIGNED").all()
    signed_amount = sum([float(c.contract_amount or 0) for c in signed_contracts])

    # 发票统计
    query_invoices = db.query(Invoice)
    if start_date:
        query_invoices = query_invoices.filter(Invoice.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_invoices = query_invoices.filter(Invoice.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    total_invoices = query_invoices.filter(Invoice.status == "ISSUED").count()
    issued_invoices = query_invoices.filter(Invoice.status == "ISSUED").all()
    issued_amount = sum([float(inv.total_amount or inv.amount or 0) for inv in issued_invoices])

    # 收款统计
    paid_invoices = query_invoices.filter(Invoice.payment_status == "PAID").all()
    paid_amount = sum([float(inv.paid_amount or 0) for inv in paid_invoices])
    
    partial_invoices = query_invoices.filter(Invoice.payment_status == "PARTIAL").all()
    partial_amount = sum([float(inv.paid_amount or 0) for inv in partial_invoices])
    
    pending_invoices = query_invoices.filter(Invoice.payment_status == "PENDING").all()
    pending_amount = sum([float(inv.total_amount or inv.amount or 0) - float(inv.paid_amount or 0) for inv in pending_invoices])
    
    # 逾期统计
    overdue_invoices = query_invoices.filter(
        Invoice.status == "ISSUED",
        Invoice.due_date < today,
        Invoice.payment_status.in_(["PENDING", "PARTIAL"])
    ).all()
    overdue_amount = sum([float(inv.total_amount or inv.amount or 0) - float(inv.paid_amount or 0) for inv in overdue_invoices])

    # 计算转化率
    conversion_rate = round(converted_leads / total_leads * 100, 2) if total_leads > 0 else 0
    win_rate = round(won_count / total_opps * 100, 2) if total_opps > 0 else 0
    quote_to_contract_rate = round(total_contracts / total_quotes * 100, 2) if total_quotes > 0 else 0
    contract_to_invoice_rate = round(total_invoices / len(signed_contracts) * 100, 2) if signed_contracts else 0
    collection_rate = round(paid_amount / issued_amount * 100, 2) if issued_amount > 0 else 0

    return ResponseModel(
        code=200,
        message="success",
        data={
            "leads": {
                "total": total_leads,
                "converted": converted_leads,
                "conversion_rate": conversion_rate
            },
            "opportunities": {
                "total": total_opps,
                "won": won_count,
                "won_amount": won_amount,
                "win_rate": win_rate
            },
            "quotes": {
                "total": total_quotes,
                "approved": len(approved_quotes),
                "approved_amount": approved_amount
            },
            "contracts": {
                "total": total_contracts,
                "signed": len(signed_contracts),
                "signed_amount": signed_amount,
                "quote_to_contract_rate": quote_to_contract_rate
            },
            "invoices": {
                "total": total_invoices,
                "issued_amount": issued_amount,
                "contract_to_invoice_rate": contract_to_invoice_rate
            },
            "receivables": {
                "paid_amount": paid_amount,
                "partial_amount": partial_amount,
                "pending_amount": pending_amount,
                "overdue_count": len(overdue_invoices),
                "overdue_amount": overdue_amount,
                "collection_rate": collection_rate
            },
            "pipeline_health": {
                "lead_to_opp_rate": round(total_opps / total_leads * 100, 2) if total_leads > 0 else 0,
                "opp_to_quote_rate": round(total_quotes / total_opps * 100, 2) if total_opps > 0 else 0,
                "quote_to_contract_rate": quote_to_contract_rate,
                "contract_to_invoice_rate": contract_to_invoice_rate,
                "collection_rate": collection_rate
            }
        }
    )


# ==================== 回款管理 ====================


@router.post("/payment-plans/{plan_id}/adjust", response_model=ResponseModel)
def adjust_payment_plan(
    *,
    db: Session = Depends(deps.get_db),
    plan_id: int,
    new_date: date = Query(..., description="新的收款日期"),
    reason: str = Query(..., description="调整原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 7.3: 手动调整收款计划
    记录调整历史并发送通知
    """
    from app.services.payment_adjustment_service import PaymentAdjustmentService
    
    service = PaymentAdjustmentService(db)
    result = service.manual_adjust_payment_plan(
        plan_id=plan_id,
        new_date=new_date,
        reason=reason,
        adjusted_by=current_user.id,
    )
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "调整失败"))
    
    return ResponseModel(
        code=200,
        message=result.get("message", "收款计划已调整"),
        data=result
    )


@router.get("/payment-plans/{plan_id}/adjustment-history", response_model=ResponseModel)
def get_payment_adjustment_history(
    *,
    db: Session = Depends(deps.get_db),
    plan_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 7.3: 获取收款计划调整历史
    """
    from app.services.payment_adjustment_service import PaymentAdjustmentService
    
    service = PaymentAdjustmentService(db)
    history = service.get_adjustment_history(plan_id)
    
    return ResponseModel(
        code=200,
        message="success",
        data={"history": history}
    )


@router.get("/payments", response_model=PaginatedResponse)
def get_payment_records(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    payment_status: Optional[str] = Query(None, description="收款状态筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款记录列表（基于发票）
    Issue 7.1: 已集成数据权限过滤（财务和销售总监可以看到所有收款数据）
    """
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")
    
    # Issue 7.1: 应用财务数据权限过滤（财务和销售总监可以看到所有收款数据）
    query = security.filter_sales_finance_data_by_scope(query, current_user, db, Invoice, 'owner_id')
    
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    
    if project_id:
        query = query.filter(Invoice.project_id == project_id)
    
    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)
    
    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)
    
    if start_date:
        query = query.filter(Invoice.paid_date >= start_date)
    
    if end_date:
        query = query.filter(Invoice.paid_date <= end_date)
    
    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(desc(Invoice.paid_date)).offset(offset).limit(page_size).all()
    
    items = []
    for invoice in invoices:
        contract = invoice.contract
        items.append({
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": invoice.project.project_code if invoice.project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "invoice_amount": float(invoice.total_amount or invoice.amount or 0),
            "paid_amount": float(invoice.paid_amount or 0),
            "unpaid_amount": float((invoice.total_amount or invoice.amount or 0) - (invoice.paid_amount or 0)),
            "payment_status": invoice.payment_status,
            "issue_date": invoice.issue_date,
            "due_date": invoice.due_date,
            "paid_date": invoice.paid_date,
            "overdue_days": (date.today() - invoice.due_date).days if invoice.due_date and invoice.due_date < date.today() and invoice.payment_status in ["PENDING", "PARTIAL"] else None,
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/payments", response_model=ResponseModel)
def create_payment_record(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int = Query(..., description="发票ID"),
    paid_amount: Decimal = Query(..., description="收款金额"),
    paid_date: date = Query(..., description="收款日期"),
    payment_method: Optional[str] = Query(None, description="收款方式"),
    bank_account: Optional[str] = Query(None, description="收款账户"),
    remark: Optional[str] = Query(None, description="备注"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    登记回款
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    if invoice.status != "ISSUED":
        raise HTTPException(status_code=400, detail="只有已开票的发票才能登记回款")
    
    # 更新收款信息
    current_paid = invoice.paid_amount or Decimal("0")
    new_paid = current_paid + paid_amount
    invoice.paid_amount = new_paid
    invoice.paid_date = paid_date
    
    # 更新收款状态
    total = invoice.total_amount or invoice.amount or Decimal("0")
    if new_paid >= total:
        invoice.payment_status = "PAID"
    elif new_paid > Decimal("0"):
        invoice.payment_status = "PARTIAL"
    else:
        invoice.payment_status = "PENDING"
    
    # 更新备注
    payment_note = f"收款记录: {paid_date}, 金额: {paid_amount}"
    if payment_method:
        payment_note += f", 方式: {payment_method}"
    if bank_account:
        payment_note += f", 账户: {bank_account}"
    if remark:
        payment_note += f", 备注: {remark}"
    
    invoice.remark = (invoice.remark or "") + f"\n{payment_note}"
    
    db.commit()
    
    return ResponseModel(
        code=200,
        message="回款登记成功",
        data={
            "invoice_id": invoice.id,
            "paid_amount": float(new_paid),
            "payment_status": invoice.payment_status,
            "unpaid_amount": float(total - new_paid)
        }
    )


@router.get("/payments/statistics", response_model=ResponseModel)
def get_payment_statistics(
    *,
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    回款统计分析

    注意：此路由必须定义在 /payments/{payment_id} 之前，否则 FastAPI 会将 "statistics" 解析为 payment_id
    """
    from app.services.payment_statistics_service import (
        build_invoice_query,
        calculate_monthly_statistics,
        calculate_customer_statistics,
        calculate_status_statistics,
        calculate_overdue_amount,
        build_monthly_list,
        build_customer_list
    )
    
    # 构建查询
    query = build_invoice_query(db, customer_id, start_date, end_date)
    invoices = query.all()
    
    today = date.today()
    
    # 计算各项统计
    monthly_stats = calculate_monthly_statistics(invoices)
    customer_stats = calculate_customer_statistics(invoices)
    status_stats = calculate_status_statistics(invoices)
    
    # 计算汇总
    total_invoiced = sum([invoice.total_amount or invoice.amount or Decimal("0") for invoice in invoices])
    total_paid = sum([invoice.paid_amount or Decimal("0") for invoice in invoices])
    total_unpaid = total_invoiced - total_paid
    total_overdue = calculate_overdue_amount(invoices, today)
    
    collection_rate = (total_paid / total_invoiced * 100) if total_invoiced > 0 else Decimal("0")
    
    # 构建列表
    monthly_list = build_monthly_list(monthly_stats)
    customer_list = build_customer_list(customer_stats, limit=10)
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "summary": {
                "total_invoiced": float(total_invoiced),
                "total_paid": float(total_paid),
                "total_unpaid": float(total_unpaid),
                "total_overdue": float(total_overdue),
                "collection_rate": float(collection_rate),
                "invoice_count": len(invoices),
            },
            "monthly_statistics": monthly_list,
            "customer_statistics": customer_list,
            "status_statistics": {
                "PAID": {
                    "count": status_stats["PAID"]["count"],
                    "amount": float(status_stats["PAID"]["amount"])
                },
                "PARTIAL": {
                    "count": status_stats["PARTIAL"]["count"],
                    "amount": float(status_stats["PARTIAL"]["amount"])
                },
                "PENDING": {
                    "count": status_stats["PENDING"]["count"],
                    "amount": float(status_stats["PENDING"]["amount"])
                },
            }
        }
    )


@router.get("/payments/reminders", response_model=PaginatedResponse)
def get_payment_reminders(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    days_before: int = Query(7, ge=0, description="提前提醒天数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款提醒列表（即将到期和已逾期的回款）

    注意：此路由必须定义在 /payments/{payment_id} 之前，否则 FastAPI 会将 "reminders" 解析为 payment_id
    """
    today = date.today()
    reminder_date = today + timedelta(days=days_before)

    query = db.query(Invoice).filter(
        Invoice.status == "ISSUED",
        Invoice.payment_status.in_(["PENDING", "PARTIAL"]),
        Invoice.due_date.isnot(None),
        Invoice.due_date <= reminder_date
    )

    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(Invoice.due_date).offset(offset).limit(page_size).all()

    items = []
    for invoice in invoices:
        contract = invoice.contract
        unpaid = (invoice.total_amount or invoice.amount or Decimal("0")) - (invoice.paid_amount or Decimal("0"))
        days_until_due = (invoice.due_date - today).days if invoice.due_date else None
        is_overdue = days_until_due is not None and days_until_due < 0

        items.append({
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": invoice.project.project_code if invoice.project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "unpaid_amount": float(unpaid),
            "due_date": invoice.due_date,
            "days_until_due": days_until_due,
            "is_overdue": is_overdue,
            "overdue_days": abs(days_until_due) if is_overdue else None,
            "payment_status": invoice.payment_status,
            "reminder_level": "urgent" if is_overdue else ("warning" if days_until_due is not None and days_until_due <= 3 else "normal"),
        })

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/payments/{payment_id}", response_model=ResponseModel)
def get_payment_detail(
    *,
    db: Session = Depends(deps.get_db),
    payment_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款详情（基于发票ID）
    """
    invoice = db.query(Invoice).options(
        joinedload(Invoice.contract),
        joinedload(Invoice.project)
    ).filter(Invoice.id == payment_id).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    contract = invoice.contract
    project = invoice.project
    
    total = invoice.total_amount or invoice.amount or Decimal("0")
    paid = invoice.paid_amount or Decimal("0")
    unpaid = total - paid
    
    overdue_days = None
    if invoice.due_date and invoice.due_date < date.today() and invoice.payment_status in ["PENDING", "PARTIAL"]:
        overdue_days = (date.today() - invoice.due_date).days
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": project.project_code if project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "invoice_amount": float(total),
            "paid_amount": float(paid),
            "unpaid_amount": float(unpaid),
            "payment_status": invoice.payment_status,
            "issue_date": invoice.issue_date,
            "due_date": invoice.due_date,
            "paid_date": invoice.paid_date,
            "overdue_days": overdue_days,
            "remark": invoice.remark,
        }
    )


@router.put("/payments/{payment_id}/match-invoice", response_model=ResponseModel)
def match_payment_to_invoice(
    *,
    db: Session = Depends(deps.get_db),
    payment_id: int,
    invoice_id: int = Query(..., description="发票ID"),
    match_amount: Optional[Decimal] = Query(None, description="核销金额，不指定则核销全部未收金额"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    核销发票（将回款记录与发票关联）
    注意：这里 payment_id 实际上是发票ID，用于保持API路径一致性
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    if invoice.status != "ISSUED":
        raise HTTPException(status_code=400, detail="只有已开票的发票才能核销")
    
    total = invoice.total_amount or invoice.amount or Decimal("0")
    current_paid = invoice.paid_amount or Decimal("0")
    unpaid = total - current_paid
    
    if match_amount:
        if match_amount > unpaid:
            raise HTTPException(status_code=400, detail=f"核销金额不能超过未收金额 {unpaid}")
        new_paid = current_paid + match_amount
    else:
        # 核销全部未收金额
        new_paid = total
    
    invoice.paid_amount = new_paid
    invoice.paid_date = date.today()
    
    # 更新收款状态
    if new_paid >= total:
        invoice.payment_status = "PAID"
    elif new_paid > Decimal("0"):
        invoice.payment_status = "PARTIAL"
    
    db.commit()
    
    return ResponseModel(
        code=200,
        message="发票核销成功",
        data={
            "invoice_id": invoice.id,
            "matched_amount": float(match_amount or unpaid),
            "paid_amount": float(new_paid),
            "payment_status": invoice.payment_status
        }
    )


@router.get("/receivables/aging", response_model=ResponseModel)
def get_receivables_aging(
    *,
    db: Session = Depends(deps.get_db),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    应收账龄分析
    """
    query = db.query(Invoice).filter(
        Invoice.status == "ISSUED",
        Invoice.payment_status.in_(["PENDING", "PARTIAL"])
    )
    
    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)
    
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    
    invoices = query.all()
    
    today = date.today()
    aging_buckets = {
        "0-30": {"count": 0, "amount": Decimal("0")},
        "31-60": {"count": 0, "amount": Decimal("0")},
        "61-90": {"count": 0, "amount": Decimal("0")},
        "90+": {"count": 0, "amount": Decimal("0")},
    }
    
    total_unpaid = Decimal("0")
    
    for invoice in invoices:
        if not invoice.due_date:
            continue
        
        unpaid = (invoice.total_amount or invoice.amount or Decimal("0")) - (invoice.paid_amount or Decimal("0"))
        if unpaid <= 0:
            continue
        
        total_unpaid += unpaid
        days_overdue = (today - invoice.due_date).days
        
        if days_overdue <= 30:
            aging_buckets["0-30"]["count"] += 1
            aging_buckets["0-30"]["amount"] += unpaid
        elif days_overdue <= 60:
            aging_buckets["31-60"]["count"] += 1
            aging_buckets["31-60"]["amount"] += unpaid
        elif days_overdue <= 90:
            aging_buckets["61-90"]["count"] += 1
            aging_buckets["61-90"]["amount"] += unpaid
        else:
            aging_buckets["90+"]["count"] += 1
            aging_buckets["90+"]["amount"] += unpaid
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "total_unpaid": float(total_unpaid),
            "aging_buckets": {
                "0-30": {
                    "count": aging_buckets["0-30"]["count"],
                    "amount": float(aging_buckets["0-30"]["amount"])
                },
                "31-60": {
                    "count": aging_buckets["31-60"]["count"],
                    "amount": float(aging_buckets["31-60"]["amount"])
                },
                "61-90": {
                    "count": aging_buckets["61-90"]["count"],
                    "amount": float(aging_buckets["61-90"]["amount"])
                },
                "90+": {
                    "count": aging_buckets["90+"]["count"],
                    "amount": float(aging_buckets["90+"]["amount"])
                }
            }
        }
    )


@router.get("/receivables/overdue", response_model=PaginatedResponse)
def get_overdue_receivables(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    min_overdue_days: Optional[int] = Query(None, description="最小逾期天数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    逾期应收列表
    """
    today = date.today()
    
    query = db.query(Invoice).filter(
        Invoice.status == "ISSUED",
        Invoice.payment_status.in_(["PENDING", "PARTIAL"]),
        Invoice.due_date.isnot(None),
        Invoice.due_date < today
    )
    
    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)
    
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    
    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(Invoice.due_date).offset(offset).limit(page_size).all()
    
    items = []
    for invoice in invoices:
        contract = invoice.contract
        unpaid = (invoice.total_amount or invoice.amount or Decimal("0")) - (invoice.paid_amount or Decimal("0"))
        overdue_days = (today - invoice.due_date).days
        
        if min_overdue_days and overdue_days < min_overdue_days:
            continue
        
        items.append({
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "invoice_amount": float(invoice.total_amount or invoice.amount or 0),
            "paid_amount": float(invoice.paid_amount or 0),
            "unpaid_amount": float(unpaid),
            "due_date": invoice.due_date,
            "overdue_days": overdue_days,
            "payment_status": invoice.payment_status,
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/receivables/summary", response_model=ResponseModel)
def get_receivables_summary(
    *,
    db: Session = Depends(deps.get_db),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    应收账款统计
    """
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")
    
    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)
    
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    
    invoices = query.all()
    
    total_amount = Decimal("0")
    paid_amount = Decimal("0")
    unpaid_amount = Decimal("0")
    partial_amount = Decimal("0")
    overdue_amount = Decimal("0")
    overdue_count = 0
    
    today = date.today()
    
    for invoice in invoices:
        total = invoice.total_amount or invoice.amount or Decimal("0")
        paid = invoice.paid_amount or Decimal("0")
        unpaid = total - paid
        
        total_amount += total
        paid_amount += paid
        unpaid_amount += unpaid
        
        if invoice.payment_status == "PARTIAL":
            partial_amount += unpaid
        
        if invoice.due_date and invoice.due_date < today and invoice.payment_status in ["PENDING", "PARTIAL"]:
            overdue_amount += unpaid
            overdue_count += 1
    
    collection_rate = (paid_amount / total_amount * 100) if total_amount > 0 else Decimal("0")
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "total_amount": float(total_amount),
            "paid_amount": float(paid_amount),
            "unpaid_amount": float(unpaid_amount),
            "partial_amount": float(partial_amount),
            "overdue_amount": float(overdue_amount),
            "overdue_count": overdue_count,
            "collection_rate": float(collection_rate),
            "invoice_count": len(invoices),
            "paid_count": len([inv for inv in invoices if inv.payment_status == "PAID"]),
            "partial_count": len([inv for inv in invoices if inv.payment_status == "PARTIAL"]),
            "pending_count": len([inv for inv in invoices if inv.payment_status == "PENDING"]),
        }
    )


@router.get("/payments/invoices/export")
def export_payment_invoices(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    payment_status: Optional[str] = Query(None, description="收款状态筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出回款记录（基于发票，Excel格式）
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from fastapi.responses import StreamingResponse
    
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")
    
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    
    if project_id:
        query = query.filter(Invoice.project_id == project_id)
    
    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)
    
    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)
    
    if start_date:
        query = query.filter(Invoice.paid_date >= start_date)
    
    if end_date:
        query = query.filter(Invoice.paid_date <= end_date)
    
    invoices = query.order_by(desc(Invoice.paid_date)).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "回款记录"
    
    # 设置表头
    headers = [
        "发票号", "合同编号", "项目编号", "客户名称", "发票金额", 
        "已收金额", "未收金额", "收款状态", "开票日期", "到期日期", 
        "实际收款日期", "逾期天数", "备注"
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充数据
    today = date.today()
    for row, invoice in enumerate(invoices, 2):
        contract = invoice.contract
        total = invoice.total_amount or invoice.amount or Decimal("0")
        paid = invoice.paid_amount or Decimal("0")
        unpaid = total - paid
        
        overdue_days = None
        if invoice.due_date and invoice.due_date < today and invoice.payment_status in ["PENDING", "PARTIAL"]:
            overdue_days = (today - invoice.due_date).days
        
        ws.cell(row=row, column=1, value=invoice.invoice_code or "")
        ws.cell(row=row, column=2, value=contract.contract_code if contract else "")
        ws.cell(row=row, column=3, value=invoice.project.project_code if invoice.project else "")
        ws.cell(row=row, column=4, value=contract.customer.customer_name if contract and contract.customer else "")
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=6, value=float(paid))
        ws.cell(row=row, column=7, value=float(unpaid))
        ws.cell(row=row, column=8, value=invoice.payment_status or "")
        ws.cell(row=row, column=9, value=invoice.issue_date.strftime("%Y-%m-%d") if invoice.issue_date else "")
        ws.cell(row=row, column=10, value=invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "")
        ws.cell(row=row, column=11, value=invoice.paid_date.strftime("%Y-%m-%d") if invoice.paid_date else "")
        ws.cell(row=row, column=12, value=overdue_days if overdue_days else "")
        ws.cell(row=row, column=13, value=invoice.remark or "")
    
    # 调整列宽
    column_widths = [15, 15, 15, 20, 12, 12, 12, 12, 12, 12, 12, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"回款记录_{date.today().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/payment-plans", response_model=PaginatedResponse)
def get_payment_plans(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取收款计划列表
    """
    query = db.query(ProjectPaymentPlan)
    
    if project_id:
        query = query.filter(ProjectPaymentPlan.project_id == project_id)
    
    if contract_id:
        query = query.filter(ProjectPaymentPlan.contract_id == contract_id)
    
    if status:
        query = query.filter(ProjectPaymentPlan.status == status)
    
    total = query.count()
    offset = (page - 1) * page_size
    plans = query.order_by(ProjectPaymentPlan.planned_date).offset(offset).limit(page_size).all()
    
    items = []
    for plan in plans:
        items.append({
            "id": plan.id,
            "payment_no": plan.payment_no,
            "project_id": plan.project_id,
            "project_code": plan.project.project_code if plan.project else None,
            "contract_id": plan.contract_id,
            "contract_code": plan.contract.contract_code if plan.contract else None,
            "payment_stage": plan.payment_stage,
            "payment_ratio": float(plan.payment_ratio or 0),
            "planned_amount": float(plan.planned_amount or 0),
            "actual_amount": float(plan.actual_amount or 0),
            "planned_date": plan.planned_date,
            "actual_date": plan.actual_date,
            "milestone_id": plan.milestone_id,
            "milestone_name": plan.milestone.milestone_name if plan.milestone else None,
            "trigger_milestone": plan.trigger_milestone,
            "status": plan.status,
            "invoice_id": plan.invoice_id,
            "invoice_no": plan.invoice_no,
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


# ==================== 销售团队管理与业绩排名 ====================


@router.get("/team", response_model=ResponseModel)
def get_sales_team(
    *,
    db: Session = Depends(deps.get_db),
    department_id: Optional[int] = Query(None, description="部门ID筛选"),
    region: Optional[str] = Query(None, description="区域关键字筛选"),
    start_date: Optional[date] = Query(None, description="统计开始日期"),
    end_date: Optional[date] = Query(None, description="统计结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.4: 获取销售团队列表
    返回销售团队成员信息，包括角色、负责区域等
    """
    normalized_start, normalized_end = _normalize_date_range(start_date, end_date)
    users = _get_visible_sales_users(db, current_user, department_id, region)
    department_names = _build_department_name_map(db, users)
    team_members = _collect_sales_team_members(db, users, department_names, normalized_start, normalized_end)
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "team_members": team_members,
            "total_count": len(team_members),
            "filters": {
                "start_date": normalized_start.isoformat(),
                "end_date": normalized_end.isoformat(),
                "department_id": department_id,
                "region": region,
            },
        }
    )


@router.get("/team/export")
def export_sales_team(
    *,
    db: Session = Depends(deps.get_db),
    department_id: Optional[int] = Query(None, description="部门ID筛选"),
    region: Optional[str] = Query(None, description="区域关键字筛选"),
    start_date: Optional[date] = Query(None, description="统计开始日期"),
    end_date: Optional[date] = Query(None, description="统计结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """导出销售团队数据为CSV"""
    normalized_start, normalized_end = _normalize_date_range(start_date, end_date)
    users = _get_visible_sales_users(db, current_user, department_id, region)
    department_names = _build_department_name_map(db, users)
    team_members = _collect_sales_team_members(db, users, department_names, normalized_start, normalized_end)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "成员ID", "姓名", "角色", "部门", "区域", "邮箱", "电话",
        "线索数量", "商机数量", "合同数量", "合同金额", "回款金额",
        "月度目标", "月度完成", "月度完成率(%)",
        "年度目标", "年度完成", "年度完成率(%)",
        "客户总数", "本期新增客户",
    ])

    for member in team_members:
        writer.writerow([
            member.get("user_id"),
            member.get("user_name"),
            member.get("role"),
            member.get("department_name") or "",
            member.get("region") or "",
            member.get("email") or "",
            member.get("phone") or "",
            member.get("lead_count") or 0,
            member.get("opportunity_count") or 0,
            member.get("contract_count") or 0,
            member.get("contract_amount") or 0,
            member.get("collection_amount") or 0,
            member.get("monthly_target") or 0,
            member.get("monthly_actual") or 0,
            member.get("monthly_completion_rate") or 0,
            member.get("year_target") or 0,
            member.get("year_actual") or 0,
            member.get("year_completion_rate") or 0,
            member.get("customer_total") or 0,
            member.get("new_customers") or 0,
        ])

    output.seek(0)
    filename = f"sales-team-{normalized_start.strftime('%Y%m%d')}-{normalized_end.strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


@router.get("/team/ranking", response_model=ResponseModel)
def get_sales_team_ranking(
    *,
    db: Session = Depends(deps.get_db),
    ranking_type: str = Query("contract_amount", description="排名类型：lead_count/opportunity_count/contract_amount/collection_amount"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    department_id: Optional[int] = Query(None, description="部门ID筛选"),
    region: Optional[str] = Query(None, description="区域关键字筛选"),
    limit: int = Query(20, ge=1, le=100, description="返回数量"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.4: 销售业绩排名
    支持按线索、商机、合同、回款排名
    """
    normalized_start, normalized_end = _normalize_date_range(start_date, end_date)
    start_datetime = datetime.combine(normalized_start, datetime.min.time())
    end_datetime = datetime.combine(normalized_end, datetime.max.time())

    users = _get_visible_sales_users(db, current_user, department_id, region)
    department_names = _build_department_name_map(db, users)

    # 获取用户角色映射
    user_ids = [user.id for user in users]
    user_roles_map = {}
    for uid in user_ids:
        user_roles_map[uid] = _get_user_role_name(db, db.query(User).filter(User.id == uid).first())

    rankings = []
    for user in users:
        stats = {
            "user_id": user.id,
            "user_name": user.real_name or user.username,
            "username": user.username,
            "role": user_roles_map.get(user.id, "销售专员"),
            "department_name": user.department or "未分配",
        }

        stats["region"] = stats["department_name"]
        
        lead_query = db.query(Lead).filter(Lead.owner_id == user.id)
        lead_query = lead_query.filter(Lead.created_at >= start_datetime)
        lead_query = lead_query.filter(Lead.created_at <= end_datetime)
        stats["lead_count"] = lead_query.count()
        
        opp_query = db.query(Opportunity).filter(Opportunity.owner_id == user.id)
        opp_query = opp_query.filter(Opportunity.created_at >= start_datetime)
        opp_query = opp_query.filter(Opportunity.created_at <= end_datetime)
        stats["opportunity_count"] = opp_query.count()
        
        contract_query = db.query(Contract).filter(Contract.owner_id == user.id)
        contract_query = contract_query.filter(Contract.created_at >= start_datetime)
        contract_query = contract_query.filter(Contract.created_at <= end_datetime)
        contracts = contract_query.all()
        stats["contract_count"] = len(contracts)
        stats["contract_amount"] = float(sum([c.contract_amount or 0 for c in contracts]))
        
        invoice_query = db.query(Invoice).join(Contract).filter(Contract.owner_id == user.id)
        invoice_query = invoice_query.filter(Invoice.paid_date.isnot(None))
        invoice_query = invoice_query.filter(Invoice.paid_date >= normalized_start)
        invoice_query = invoice_query.filter(Invoice.paid_date <= normalized_end)
        invoices = invoice_query.filter(Invoice.payment_status.in_(["PAID", "PARTIAL"])).all()
        stats["collection_amount"] = float(sum([inv.paid_amount or 0 for inv in invoices]))
        
        rankings.append(stats)
    
    # 根据排名类型排序
    valid_ranking_types = ["lead_count", "opportunity_count", "contract_amount", "collection_amount"]
    if ranking_type not in valid_ranking_types:
        ranking_type = "contract_amount"
    
    rankings.sort(key=lambda x: x.get(ranking_type, 0), reverse=True)
    
    # 添加排名
    for idx, ranking in enumerate(rankings[:limit], 1):
        ranking["rank"] = idx
        ranking["value"] = ranking.get(ranking_type, 0)
    
    return ResponseModel(
        code=200,
        message="success",
        data={
            "ranking_type": ranking_type,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "rankings": rankings[:limit],
            "total_count": len(rankings)
        }
    )


# ==================== 销售目标管理 ====================


@router.get("/targets", response_model=PaginatedResponse)
def get_sales_targets(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    target_scope: Optional[str] = Query(None, description="目标范围筛选：PERSONAL/TEAM/DEPARTMENT"),
    target_type: Optional[str] = Query(None, description="目标类型筛选"),
    target_period: Optional[str] = Query(None, description="目标周期筛选：MONTHLY/QUARTERLY/YEARLY"),
    period_value: Optional[str] = Query(None, description="周期值筛选：2025-01/2025-Q1/2025"),
    user_id: Optional[int] = Query(None, description="用户ID筛选"),
    department_id: Optional[int] = Query(None, description="部门ID筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.5: 获取销售目标列表
    支持多种筛选条件，并根据用户角色返回可见的目标
    """
    from app.models.organization import Department
    
    query = db.query(SalesTarget)

    # 根据用户角色确定可见范围
    user_role_code = _get_user_role_code(db, current_user)

    if user_role_code == 'SALES_DIR':
        # 销售总监可以看到所有目标
        pass
    elif user_role_code == 'SALES_MANAGER':
        # 销售经理可以看到自己部门的目标
        # 注意：User表没有department_id字段，需要根据department字符串匹配
        dept_name = getattr(current_user, 'department', None)
        if dept_name:
            # 查找对应的部门ID
            dept = db.query(Department).filter(Department.dept_name == dept_name).first()
            if dept:
                query = query.filter(
                    or_(
                        SalesTarget.department_id == dept.id,
                        SalesTarget.user_id == current_user.id
                    )
                )
            else:
                query = query.filter(SalesTarget.user_id == current_user.id)
        else:
            query = query.filter(SalesTarget.user_id == current_user.id)
    else:
        # 其他角色只能看到自己的目标
        query = query.filter(SalesTarget.user_id == current_user.id)
    
    # 应用筛选条件
    if target_scope:
        query = query.filter(SalesTarget.target_scope == target_scope)
    if target_type:
        query = query.filter(SalesTarget.target_type == target_type)
    if target_period:
        query = query.filter(SalesTarget.target_period == target_period)
    if period_value:
        query = query.filter(SalesTarget.period_value == period_value)
    if user_id:
        query = query.filter(SalesTarget.user_id == user_id)
    if department_id:
        query = query.filter(SalesTarget.department_id == department_id)
    if status:
        query = query.filter(SalesTarget.status == status)
    
    total = query.count()
    offset = (page - 1) * page_size
    targets = query.order_by(desc(SalesTarget.created_at)).offset(offset).limit(page_size).all()

    team_service = SalesTeamService(db)
    
    # 计算实际完成值和完成率
    items = []
    for target in targets:
        actual_value, completion_rate = team_service.calculate_target_performance(target)
        
        # 获取用户/部门名称
        user_name = None
        if target.user_id:
            user = db.query(User).filter(User.id == target.user_id).first()
            user_name = user.real_name or user.username if user else None
        
        department_name = None
        if target.department_id:
            dept = db.query(Department).filter(Department.id == target.department_id).first()
            department_name = dept.dept_name if dept else None
        
        items.append({
            "id": target.id,
            "target_scope": target.target_scope,
            "user_id": target.user_id,
            "department_id": target.department_id,
            "team_id": target.team_id,
            "target_type": target.target_type,
            "target_period": target.target_period,
            "period_value": target.period_value,
            "target_value": float(target.target_value),
            "description": target.description,
            "status": target.status,
            "created_by": target.created_by,
            "actual_value": float(actual_value),
            "completion_rate": completion_rate,
            "user_name": user_name,
            "department_name": department_name,
            "created_at": target.created_at,
            "updated_at": target.updated_at,
        })
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )
@router.post("/targets", response_model=SalesTargetResponse, status_code=201)
def create_sales_target(
    *,
    db: Session = Depends(deps.get_db),
    target_data: SalesTargetCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.5: 创建销售目标
    """
    # 验证目标范围和数据
    if target_data.target_scope == "PERSONAL" and not target_data.user_id:
        raise HTTPException(status_code=400, detail="个人目标必须指定用户ID")
    if target_data.target_scope == "DEPARTMENT" and not target_data.department_id:
        raise HTTPException(status_code=400, detail="部门目标必须指定部门ID")
    
    # 创建目标
    target = SalesTarget(
        target_scope=target_data.target_scope,
        user_id=target_data.user_id,
        department_id=target_data.department_id,
        team_id=target_data.team_id,
        target_type=target_data.target_type,
        target_period=target_data.target_period,
        period_value=target_data.period_value,
        target_value=target_data.target_value,
        description=target_data.description,
        status=target_data.status or "ACTIVE",
        created_by=current_user.id,
    )
    
    db.add(target)
    db.commit()
    db.refresh(target)
    
    # 获取用户/部门名称
    user_name = None
    if target.user_id:
        user = db.query(User).filter(User.id == target.user_id).first()
        user_name = user.real_name or user.username if user else None
    
    department_name = None
    if target.department_id:
        from app.models.organization import Department
        dept = db.query(Department).filter(Department.id == target.department_id).first()
        department_name = dept.dept_name if dept else None
    
    return SalesTargetResponse(
        id=target.id,
        target_scope=target.target_scope,
        user_id=target.user_id,
        department_id=target.department_id,
        team_id=target.team_id,
        target_type=target.target_type,
        target_period=target.target_period,
        period_value=target.period_value,
        target_value=target.target_value,
        description=target.description,
        status=target.status,
        created_by=target.created_by,
        actual_value=None,
        completion_rate=None,
        user_name=user_name,
        department_name=department_name,
        created_at=target.created_at,
        updated_at=target.updated_at,
    )


@router.put("/targets/{target_id}", response_model=SalesTargetResponse)
def update_sales_target(
    *,
    db: Session = Depends(deps.get_db),
    target_id: int,
    target_data: SalesTargetUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 6.5: 更新销售目标
    """
    target = db.query(SalesTarget).filter(SalesTarget.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="目标不存在")
    
    # 权限检查：只能修改自己创建的目标或自己部门的目标
    if target.created_by != current_user.id:
        user_role_code = _get_user_role_code(db, current_user)
        if user_role_code != 'SALES_DIR':
            # User表没有department_id，需要通过department字符串匹配
            dept_name = getattr(current_user, 'department', None)
            if dept_name:
                dept = db.query(Department).filter(Department.dept_name == dept_name).first()
                if dept and target.department_id != dept.id:
                    raise HTTPException(status_code=403, detail="无权修改此目标")
            else:
                raise HTTPException(status_code=403, detail="无权修改此目标")
    
    # 更新字段
    if target_data.target_value is not None:
        target.target_value = target_data.target_value
    if target_data.description is not None:
        target.description = target_data.description
    if target_data.status is not None:
        target.status = target_data.status
    
    db.commit()
    db.refresh(target)
    
    # 获取用户/部门名称
    user_name = None
    if target.user_id:
        user = db.query(User).filter(User.id == target.user_id).first()
        user_name = user.real_name or user.username if user else None
    
    department_name = None
    if target.department_id:
        from app.models.organization import Department
        dept = db.query(Department).filter(Department.id == target.department_id).first()
        department_name = dept.dept_name if dept else None
    
    return SalesTargetResponse(
        id=target.id,
        target_scope=target.target_scope,
        user_id=target.user_id,
        department_id=target.department_id,
        team_id=target.team_id,
        target_type=target.target_type,
        target_period=target.target_period,
        period_value=target.period_value,
        target_value=target.target_value,
        description=target.description,
        status=target.status,
        created_by=target.created_by,
        actual_value=None,
        completion_rate=None,
        user_name=user_name,
        department_name=department_name,
        created_at=target.created_at,
        updated_at=target.updated_at,
    )


# ==================== 模板 & CPQ ====================


def _flatten_structure(data: Any, prefix: str = "") -> Dict[str, Any]:
    """将嵌套结构展开便于比较"""
    entries: Dict[str, Any] = {}
    if data is None:
        return entries
    if isinstance(data, dict):
        for key, value in data.items():
            sub_prefix = f"{prefix}.{key}" if prefix else str(key)
            entries.update(_flatten_structure(value, sub_prefix))
    elif isinstance(data, list):
        for index, value in enumerate(data):
            sub_prefix = f"{prefix}[{index}]" if prefix else f"[{index}]"
            entries.update(_flatten_structure(value, sub_prefix))
    else:
        entries[prefix or "value"] = data
    return entries


def _build_diff_section(current: Any, previous: Any) -> Dict[str, Any]:
    current_flat = _flatten_structure(current)
    previous_flat = _flatten_structure(previous)

    added = sorted(list(set(current_flat.keys()) - set(previous_flat.keys())))
    removed = sorted(list(set(previous_flat.keys()) - set(current_flat.keys())))
    changed = []
    for key in current_flat.keys() & previous_flat.keys():
        if current_flat[key] != previous_flat[key]:
            changed.append(
                {
                    "path": key,
                    "old": previous_flat[key],
                    "new": current_flat[key],
                }
            )
    return {
        "added": added,
        "removed": removed,
        "changed": changed,
    }


def _get_previous_version(template, current_version):
    if not template or not template.versions:
        return None
    versions = sorted(
        template.versions,
        key=lambda v: v.created_at or datetime.min,
        reverse=True,
    )
    for idx, version in enumerate(versions):
        if version.id == current_version.id and idx + 1 < len(versions):
            return versions[idx + 1]
    return None


def _build_template_version_diff(current_version, previous_version) -> TemplateVersionDiff:
    sections_diff = _build_diff_section(
        getattr(current_version, "sections", None),
        getattr(previous_version, "sections", None) if previous_version else None,
    )
    pricing_diff = _build_diff_section(
        getattr(current_version, "pricing_rules", None),
        getattr(previous_version, "pricing_rules", None) if previous_version else None,
    )
    clause_diff = _build_diff_section(
        getattr(current_version, "clause_sections", None),
        getattr(previous_version, "clause_sections", None) if previous_version else None,
    )
    return TemplateVersionDiff(
        sections=sections_diff,
        pricing_rules=pricing_diff,
        clause_sections=clause_diff,
    )


def _build_template_history(template) -> List[TemplateApprovalHistoryRecord]:
    if not template or not template.versions:
        return []
    versions = sorted(
        template.versions,
        key=lambda v: v.created_at or datetime.min,
        reverse=True,
    )
    history = []
    for version in versions[:10]:
        history.append(
            TemplateApprovalHistoryRecord(
                version_id=version.id,
                version_no=version.version_no,
                status=version.status,
                published_by=version.published_by,
                published_at=version.published_at,
                release_notes=version.release_notes,
            )
        )
    return history


def _filter_template_visibility(query, model, user: User):
    """根据可见范围过滤模板"""
    if user.is_superuser:
        return query
    owner_alias = aliased(User)
    query = query.outerjoin(owner_alias, model.owner)
    conditions = [
        model.visibility_scope == "ALL",
        model.owner_id == user.id,
    ]
    if user.department:
        conditions.append(
            and_(
                model.visibility_scope.in_(["TEAM", "DEPT"]),
                owner_alias.department == user.department,
            )
        )
    return query.filter(or_(*conditions))


def _serialize_quote_template(template: QuoteTemplate) -> QuoteTemplateResponse:
    versions = sorted(
        template.versions or [],
        key=lambda v: v.created_at or datetime.min,
        reverse=True,
    )
    version_responses = [
        QuoteTemplateVersionResponse(
            id=v.id,
            template_id=v.template_id,
            version_no=v.version_no,
            status=v.status,
            sections=v.sections,
            pricing_rules=v.pricing_rules,
            config_schema=v.config_schema,
            discount_rules=v.discount_rules,
            release_notes=v.release_notes,
            rule_set_id=v.rule_set_id,
            created_by=v.created_by,
            published_by=v.published_by,
            published_at=v.published_at,
            created_at=v.created_at,
            updated_at=v.updated_at,
        )
        for v in versions
    ]
    return QuoteTemplateResponse(
        id=template.id,
        template_code=template.template_code,
        template_name=template.template_name,
        category=template.category,
        description=template.description,
        status=template.status,
        visibility_scope=template.visibility_scope,
        is_default=template.is_default,
        current_version_id=template.current_version_id,
        owner_id=template.owner_id,
        versions=version_responses,
        created_at=template.created_at,
        updated_at=template.updated_at,
    )


def _serialize_contract_template(template: ContractTemplate) -> ContractTemplateResponse:
    versions = sorted(
        template.versions or [],
        key=lambda v: v.created_at or datetime.min,
        reverse=True,
    )
    version_responses = [
        ContractTemplateVersionResponse(
            id=v.id,
            template_id=v.template_id,
            version_no=v.version_no,
            status=v.status,
            clause_sections=v.clause_sections,
            clause_library=v.clause_library,
            attachment_refs=v.attachment_refs,
            approval_flow=v.approval_flow,
            release_notes=v.release_notes,
            created_by=v.created_by,
            published_by=v.published_by,
            published_at=v.published_at,
            created_at=v.created_at,
            updated_at=v.updated_at,
        )
        for v in versions
    ]
    return ContractTemplateResponse(
        id=template.id,
        template_code=template.template_code,
        template_name=template.template_name,
        contract_type=template.contract_type,
        description=template.description,
        status=template.status,
        visibility_scope=template.visibility_scope,
        is_default=template.is_default,
        current_version_id=template.current_version_id,
        owner_id=template.owner_id,
        versions=version_responses,
        created_at=template.created_at,
        updated_at=template.updated_at,
    )


def _serialize_rule_set(rule_set: CpqRuleSet) -> CpqRuleSetResponse:
    return CpqRuleSetResponse(
        id=rule_set.id,
        rule_code=rule_set.rule_code,
        rule_name=rule_set.rule_name,
        description=rule_set.description,
        status=rule_set.status,
        base_price=rule_set.base_price or Decimal("0"),
        currency=rule_set.currency or "CNY",
        config_schema=rule_set.config_schema,
        pricing_matrix=rule_set.pricing_matrix,
        approval_threshold=rule_set.approval_threshold,
        visibility_scope=rule_set.visibility_scope or "ALL",
        is_default=rule_set.is_default or False,
        owner_role=rule_set.owner_role,
        created_at=rule_set.created_at,
        updated_at=rule_set.updated_at,
    )


@router.get("/quote-templates", response_model=PaginatedResponse[QuoteTemplateResponse])
def list_quote_templates(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    query = db.query(QuoteTemplate).options(joinedload(QuoteTemplate.versions))
    query = _filter_template_visibility(query, QuoteTemplate, current_user)

    if keyword:
        query = query.filter(
            or_(
                QuoteTemplate.template_name.contains(keyword),
                QuoteTemplate.template_code.contains(keyword),
            )
        )
    if status:
        query = query.filter(QuoteTemplate.status == status)

    total = query.count()
    offset = (page - 1) * page_size
    templates = (
        query.order_by(desc(QuoteTemplate.created_at))
        .offset(offset)
        .limit(page_size)
        .all()
    )
    items = [_serialize_quote_template(t) for t in templates]
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.post("/quote-templates", response_model=QuoteTemplateResponse)
def create_quote_template(
    *,
    db: Session = Depends(deps.get_db),
    template_in: QuoteTemplateCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    existing = (
        db.query(QuoteTemplate)
        .filter(QuoteTemplate.template_code == template_in.template_code)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="模板编码已存在")

    template = QuoteTemplate(
        template_code=template_in.template_code,
        template_name=template_in.template_name,
        category=template_in.category,
        description=template_in.description,
        visibility_scope=template_in.visibility_scope or "TEAM",
        is_default=template_in.is_default or False,
        owner_id=template_in.owner_id or current_user.id,
        status="DRAFT",
    )
    db.add(template)
    db.flush()

    if template_in.initial_version:
        version_data = template_in.initial_version
        version = QuoteTemplateVersion(
            template_id=template.id,
            version_no=version_data.version_no,
            sections=version_data.sections,
            pricing_rules=version_data.pricing_rules,
            config_schema=version_data.config_schema,
            discount_rules=version_data.discount_rules,
            release_notes=version_data.release_notes,
            rule_set_id=version_data.rule_set_id,
            status="DRAFT",
            created_by=current_user.id,
        )
        db.add(version)
        db.flush()
        template.current_version_id = version.id

    db.commit()
    db.refresh(template)
    template = (
        db.query(QuoteTemplate)
        .options(joinedload(QuoteTemplate.versions))
        .filter(QuoteTemplate.id == template.id)
        .first()
    )
    return _serialize_quote_template(template)


@router.put("/quote-templates/{template_id}", response_model=QuoteTemplateResponse)
def update_quote_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    template_in: QuoteTemplateUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(QuoteTemplate).options(joinedload(QuoteTemplate.versions)),
            QuoteTemplate,
            current_user,
        )
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    update_data = template_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(template, field, value)
    db.commit()
    db.refresh(template)
    template = (
        db.query(QuoteTemplate)
        .options(joinedload(QuoteTemplate.versions))
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    return _serialize_quote_template(template)


@router.post("/quote-templates/{template_id}/versions", response_model=QuoteTemplateVersionResponse)
def create_quote_template_version(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    version_in: QuoteTemplateVersionCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(QuoteTemplate),
            QuoteTemplate,
            current_user,
        )
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    version = QuoteTemplateVersion(
        template_id=template_id,
        version_no=version_in.version_no,
        sections=version_in.sections,
        pricing_rules=version_in.pricing_rules,
        config_schema=version_in.config_schema,
        discount_rules=version_in.discount_rules,
        release_notes=version_in.release_notes,
        rule_set_id=version_in.rule_set_id,
        status="DRAFT",
        created_by=current_user.id,
    )
    db.add(version)
    db.commit()
    db.refresh(version)
    return QuoteTemplateVersionResponse(
        id=version.id,
        template_id=version.template_id,
        version_no=version.version_no,
        status=version.status,
        sections=version.sections,
        pricing_rules=version.pricing_rules,
        config_schema=version.config_schema,
        discount_rules=version.discount_rules,
        release_notes=version.release_notes,
        rule_set_id=version.rule_set_id,
        created_by=version.created_by,
        published_by=version.published_by,
        published_at=version.published_at,
        created_at=version.created_at,
        updated_at=version.updated_at,
    )


@router.post("/quote-templates/{template_id}/versions/{version_id}/publish", response_model=QuoteTemplateResponse)
def publish_quote_template_version(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    version_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        db.query(QuoteTemplate)
        .options(joinedload(QuoteTemplate.versions))
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    version = next((v for v in template.versions if v.id == version_id), None)
    if not version:
        raise HTTPException(status_code=404, detail="模板版本不存在")

    version.status = "PUBLISHED"
    version.published_by = current_user.id
    version.published_at = datetime.utcnow()
    template.current_version_id = version.id
    template.status = "ACTIVE"
    db.query(QuoteTemplateVersion).filter(
        QuoteTemplateVersion.template_id == template_id,
        QuoteTemplateVersion.id != version_id,
        QuoteTemplateVersion.status == "PUBLISHED",
    ).update({"status": "ARCHIVED"}, synchronize_session=False)

    db.commit()
    template = (
        db.query(QuoteTemplate)
        .options(joinedload(QuoteTemplate.versions))
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    return _serialize_quote_template(template)


@router.post("/quote-templates/{template_id}/apply", response_model=QuoteTemplateApplyResponse)
def apply_quote_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    preview_request: Optional[CpqPricePreviewRequest] = None,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(QuoteTemplate).options(joinedload(QuoteTemplate.versions)),
            QuoteTemplate,
            current_user,
        )
        .filter(QuoteTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    version = None
    if preview_request and preview_request.template_version_id:
        version = next(
            (v for v in template.versions if v.id == preview_request.template_version_id),
            None,
        )
    if not version:
        version = template.current_version or (template.versions[0] if template.versions else None)

    if not version:
        raise HTTPException(status_code=400, detail="模板尚未创建版本")

    preview = preview_request or CpqPricePreviewRequest()
    service = CpqPricingService(db)
    preview_data = service.preview_price(
        rule_set_id=preview.rule_set_id or version.rule_set_id,
        template_version_id=version.id,
        selections=preview.selections,
        manual_discount_pct=preview.manual_discount_pct,
        manual_markup_pct=preview.manual_markup_pct,
    )

    previous_version = _get_previous_version(template, version)
    version_diff = _build_template_version_diff(version, previous_version)
    history = _build_template_history(template)

    return QuoteTemplateApplyResponse(
        template=_serialize_quote_template(template),
        version=QuoteTemplateVersionResponse(
            id=version.id,
            template_id=version.template_id,
            version_no=version.version_no,
            status=version.status,
            sections=version.sections,
            pricing_rules=version.pricing_rules,
            config_schema=version.config_schema,
            discount_rules=version.discount_rules,
            release_notes=version.release_notes,
            rule_set_id=version.rule_set_id,
            created_by=version.created_by,
            published_by=version.published_by,
            published_at=version.published_at,
            created_at=version.created_at,
            updated_at=version.updated_at,
        ),
        cpq_preview=CpqPricePreviewResponse(**preview_data),
        version_diff=version_diff,
        approval_history=history,
    )


@router.get("/contract-templates", response_model=PaginatedResponse[ContractTemplateResponse])
def list_contract_templates(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    keyword: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    query = db.query(ContractTemplate).options(joinedload(ContractTemplate.versions))
    query = _filter_template_visibility(query, ContractTemplate, current_user)

    if keyword:
        query = query.filter(
            or_(
                ContractTemplate.template_name.contains(keyword),
                ContractTemplate.template_code.contains(keyword),
            )
        )
    if status:
        query = query.filter(ContractTemplate.status == status)

    total = query.count()
    offset = (page - 1) * page_size
    templates = (
        query.order_by(desc(ContractTemplate.created_at))
        .offset(offset)
        .limit(page_size)
        .all()
    )
    return PaginatedResponse(
        items=[_serialize_contract_template(t) for t in templates],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.post("/contract-templates", response_model=ContractTemplateResponse)
def create_contract_template(
    *,
    db: Session = Depends(deps.get_db),
    template_in: ContractTemplateCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    existing = (
        db.query(ContractTemplate)
        .filter(ContractTemplate.template_code == template_in.template_code)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="模板编码已存在")

    template = ContractTemplate(
        template_code=template_in.template_code,
        template_name=template_in.template_name,
        contract_type=template_in.contract_type,
        description=template_in.description,
        visibility_scope=template_in.visibility_scope or "TEAM",
        is_default=template_in.is_default or False,
        owner_id=template_in.owner_id or current_user.id,
        status="DRAFT",
    )
    db.add(template)
    db.flush()

    if template_in.initial_version:
        version_data = template_in.initial_version
        version = ContractTemplateVersion(
            template_id=template.id,
            version_no=version_data.version_no,
            clause_sections=version_data.clause_sections,
            clause_library=version_data.clause_library,
            attachment_refs=version_data.attachment_refs,
            approval_flow=version_data.approval_flow,
            release_notes=version_data.release_notes,
            status="DRAFT",
            created_by=current_user.id,
        )
        db.add(version)
        db.flush()
        template.current_version_id = version.id

    db.commit()
    template = (
        db.query(ContractTemplate)
        .options(joinedload(ContractTemplate.versions))
        .filter(ContractTemplate.id == template.id)
        .first()
    )
    return _serialize_contract_template(template)


@router.put("/contract-templates/{template_id}", response_model=ContractTemplateResponse)
def update_contract_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    template_in: ContractTemplateUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(ContractTemplate).options(joinedload(ContractTemplate.versions)),
            ContractTemplate,
            current_user,
        )
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    update_data = template_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(template, field, value)

    db.commit()
    template = (
        db.query(ContractTemplate)
        .options(joinedload(ContractTemplate.versions))
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    return _serialize_contract_template(template)


@router.post("/contract-templates/{template_id}/versions", response_model=ContractTemplateVersionResponse)
def create_contract_template_version(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    version_in: ContractTemplateVersionCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(ContractTemplate),
            ContractTemplate,
            current_user,
        )
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    version = ContractTemplateVersion(
        template_id=template_id,
        version_no=version_in.version_no,
        clause_sections=version_in.clause_sections,
        clause_library=version_in.clause_library,
        attachment_refs=version_in.attachment_refs,
        approval_flow=version_in.approval_flow,
        release_notes=version_in.release_notes,
        status="DRAFT",
        created_by=current_user.id,
    )
    db.add(version)
    db.commit()
    db.refresh(version)
    return ContractTemplateVersionResponse(
        id=version.id,
        template_id=version.template_id,
        version_no=version.version_no,
        status=version.status,
        clause_sections=version.clause_sections,
        clause_library=version.clause_library,
        attachment_refs=version.attachment_refs,
        approval_flow=version.approval_flow,
        release_notes=version.release_notes,
        created_by=version.created_by,
        published_by=version.published_by,
        published_at=version.published_at,
        created_at=version.created_at,
        updated_at=version.updated_at,
    )


@router.post("/contract-templates/{template_id}/versions/{version_id}/publish", response_model=ContractTemplateResponse)
def publish_contract_template_version(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    version_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        db.query(ContractTemplate)
        .options(joinedload(ContractTemplate.versions))
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    version = next((v for v in template.versions if v.id == version_id), None)
    if not version:
        raise HTTPException(status_code=404, detail="模板版本不存在")

    version.status = "PUBLISHED"
    version.published_by = current_user.id
    version.published_at = datetime.utcnow()
    template.current_version_id = version.id
    template.status = "ACTIVE"
    db.query(ContractTemplateVersion).filter(
        ContractTemplateVersion.template_id == template_id,
        ContractTemplateVersion.id != version_id,
        ContractTemplateVersion.status == "PUBLISHED",
    ).update({"status": "ARCHIVED"}, synchronize_session=False)

    db.commit()
    template = (
        db.query(ContractTemplate)
        .options(joinedload(ContractTemplate.versions))
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    return _serialize_contract_template(template)


@router.get("/contract-templates/{template_id}/apply", response_model=ContractTemplateApplyResponse)
def apply_contract_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    template_version_id: Optional[int] = Query(None, description="指定模板版本ID"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    template = (
        _filter_template_visibility(
            db.query(ContractTemplate).options(joinedload(ContractTemplate.versions)),
            ContractTemplate,
            current_user,
        )
        .filter(ContractTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在或无权访问")

    version = None
    if template_version_id:
        version = next((v for v in template.versions if v.id == template_version_id), None)
    if not version:
        version = template.current_version or (template.versions[0] if template.versions else None)
    if not version:
        raise HTTPException(status_code=400, detail="模板尚未创建版本")

    previous_version = _get_previous_version(template, version)
    version_diff = _build_template_version_diff(version, previous_version)
    history = _build_template_history(template)

    return ContractTemplateApplyResponse(
        template=_serialize_contract_template(template),
        version=ContractTemplateVersionResponse(
            id=version.id,
            template_id=version.template_id,
            version_no=version.version_no,
            status=version.status,
            clause_sections=version.clause_sections,
            clause_library=version.clause_library,
            attachment_refs=version.attachment_refs,
            approval_flow=version.approval_flow,
            release_notes=version.release_notes,
            created_by=version.created_by,
            published_by=version.published_by,
            published_at=version.published_at,
            created_at=version.created_at,
            updated_at=version.updated_at,
        ),
        version_diff=version_diff,
        approval_history=history,
    )


@router.get("/cpq/rule-sets", response_model=PaginatedResponse[CpqRuleSetResponse])
def list_cpq_rule_sets(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    keyword: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    query = db.query(CpqRuleSet)
    if keyword:
        query = query.filter(
            or_(
                CpqRuleSet.rule_name.contains(keyword),
                CpqRuleSet.rule_code.contains(keyword),
            )
        )
    if status:
        query = query.filter(CpqRuleSet.status == status)

    total = query.count()
    offset = (page - 1) * page_size
    rule_sets = (
        query.order_by(desc(CpqRuleSet.created_at))
        .offset(offset)
        .limit(page_size)
        .all()
    )
    return PaginatedResponse(
        items=[_serialize_rule_set(r) for r in rule_sets],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.post("/cpq/rule-sets", response_model=CpqRuleSetResponse)
def create_cpq_rule_set(
    *,
    db: Session = Depends(deps.get_db),
    rule_set_in: CpqRuleSetCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    existing = (
        db.query(CpqRuleSet)
        .filter(CpqRuleSet.rule_code == rule_set_in.rule_code)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="规则集编码已存在")

    rule_set = CpqRuleSet(
        rule_code=rule_set_in.rule_code,
        rule_name=rule_set_in.rule_name,
        description=rule_set_in.description,
        status=rule_set_in.status or "ACTIVE",
        base_price=rule_set_in.base_price or Decimal("0"),
        currency=rule_set_in.currency or "CNY",
        config_schema=rule_set_in.config_schema,
        pricing_matrix=rule_set_in.pricing_matrix,
        approval_threshold=rule_set_in.approval_threshold,
        visibility_scope=rule_set_in.visibility_scope or "ALL",
        is_default=rule_set_in.is_default or False,
        owner_role=rule_set_in.owner_role or (current_user.department or "SALES"),
    )
    db.add(rule_set)
    db.commit()
    db.refresh(rule_set)
    return _serialize_rule_set(rule_set)


@router.put("/cpq/rule-sets/{rule_set_id}", response_model=CpqRuleSetResponse)
def update_cpq_rule_set(
    *,
    db: Session = Depends(deps.get_db),
    rule_set_id: int,
    rule_set_in: CpqRuleSetUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    rule_set = db.query(CpqRuleSet).filter(CpqRuleSet.id == rule_set_id).first()
    if not rule_set:
        raise HTTPException(status_code=404, detail="规则集不存在")

    update_data = rule_set_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule_set, field, value)
    db.commit()
    db.refresh(rule_set)
    return _serialize_rule_set(rule_set)


@router.post("/cpq/price-preview", response_model=CpqPricePreviewResponse)
def preview_cpq_price(
    *,
    db: Session = Depends(deps.get_db),
    preview_request: CpqPricePreviewRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    service = CpqPricingService(db)
    preview_data = service.preview_price(
        rule_set_id=preview_request.rule_set_id,
        template_version_id=preview_request.template_version_id,
        selections=preview_request.selections,
        manual_discount_pct=preview_request.manual_discount_pct,
        manual_markup_pct=preview_request.manual_markup_pct,
    )
    return CpqPricePreviewResponse(**preview_data)


# ==================== 报价成本管理 ====================


@router.get("/cost-templates", response_model=PaginatedResponse[QuoteCostTemplateResponse])
def get_cost_templates(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    template_type: Optional[str] = Query(None, description="模板类型筛选"),
    equipment_type: Optional[str] = Query(None, description="设备类型筛选"),
    industry: Optional[str] = Query(None, description="行业筛选"),
    is_active: Optional[bool] = Query(None, description="是否启用"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取成本模板列表
    """
    query = db.query(QuoteCostTemplate)
    
    if template_type:
        query = query.filter(QuoteCostTemplate.template_type == template_type)
    if equipment_type:
        query = query.filter(QuoteCostTemplate.equipment_type == equipment_type)
    if industry:
        query = query.filter(QuoteCostTemplate.industry == industry)
    if is_active is not None:
        query = query.filter(QuoteCostTemplate.is_active == is_active)
    
    total = query.count()
    offset = (page - 1) * page_size
    templates = query.order_by(desc(QuoteCostTemplate.created_at)).offset(offset).limit(page_size).all()
    
    items = []
    for template in templates:
        template_dict = {
            **{c.name: getattr(template, c.name) for c in template.__table__.columns},
            "creator_name": template.creator.real_name if template.creator else None
        }
        items.append(QuoteCostTemplateResponse(**template_dict))
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/cost-templates/{template_id}", response_model=QuoteCostTemplateResponse)
def get_cost_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取成本模板详情
    """
    template = db.query(QuoteCostTemplate).filter(QuoteCostTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="成本模板不存在")
    
    template_dict = {
        **{c.name: getattr(template, c.name) for c in template.__table__.columns},
        "creator_name": template.creator.real_name if template.creator else None
    }
    return QuoteCostTemplateResponse(**template_dict)


@router.post("/cost-templates", response_model=QuoteCostTemplateResponse, status_code=201)
def create_cost_template(
    *,
    db: Session = Depends(deps.get_db),
    template_in: QuoteCostTemplateCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建成本模板
    """
    # 检查模板编码是否已存在
    existing = db.query(QuoteCostTemplate).filter(QuoteCostTemplate.template_code == template_in.template_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="模板编码已存在")
    
    # 计算总成本和分类
    total_cost = Decimal('0')
    categories = set()
    if template_in.cost_structure:
        import json
        cost_structure = template_in.cost_structure
        for category in cost_structure.get('categories', []):
            categories.add(category.get('category', ''))
            for item in category.get('items', []):
                qty = Decimal(str(item.get('default_qty', 0)))
                cost = Decimal(str(item.get('default_cost', 0)))
                total_cost += qty * cost
    
    template_data = template_in.model_dump()
    template_data['total_cost'] = float(total_cost)
    template_data['cost_categories'] = ','.join(categories) if categories else None
    template_data['created_by'] = current_user.id
    
    template = QuoteCostTemplate(**template_data)
    db.add(template)
    db.commit()
    db.refresh(template)
    
    template_dict = {
        **{c.name: getattr(template, c.name) for c in template.__table__.columns},
        "creator_name": template.creator.real_name if template.creator else None
    }
    return QuoteCostTemplateResponse(**template_dict)


@router.put("/cost-templates/{template_id}", response_model=QuoteCostTemplateResponse)
def update_cost_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    template_in: QuoteCostTemplateUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新成本模板
    """
    template = db.query(QuoteCostTemplate).filter(QuoteCostTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="成本模板不存在")
    
    update_data = template_in.model_dump(exclude_unset=True)
    
    # 如果更新了成本结构，重新计算总成本和分类
    if 'cost_structure' in update_data and update_data['cost_structure']:
        total_cost = Decimal('0')
        categories = set()
        cost_structure = update_data['cost_structure']
        for category in cost_structure.get('categories', []):
            categories.add(category.get('category', ''))
            for item in category.get('items', []):
                qty = Decimal(str(item.get('default_qty', 0)))
                cost = Decimal(str(item.get('default_cost', 0)))
                total_cost += qty * cost
        update_data['total_cost'] = float(total_cost)
        update_data['cost_categories'] = ','.join(categories) if categories else None
    
    for field, value in update_data.items():
        if hasattr(template, field):
            setattr(template, field, value)
    
    db.commit()
    db.refresh(template)
    
    template_dict = {
        **{c.name: getattr(template, c.name) for c in template.__table__.columns},
        "creator_name": template.creator.real_name if template.creator else None
    }
    return QuoteCostTemplateResponse(**template_dict)


@router.delete("/cost-templates/{template_id}", status_code=200)
def delete_cost_template(
    *,
    db: Session = Depends(deps.get_db),
    template_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除成本模板
    """
    template = db.query(QuoteCostTemplate).filter(QuoteCostTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="成本模板不存在")
    
    db.delete(template)
    db.commit()
    
    return ResponseModel(code=200, message="成本模板已删除")


@router.post("/quotes/{quote_id}/apply-template", response_model=QuoteVersionResponse)
def apply_cost_template_to_quote(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    template_id: int = Query(..., description="成本模板ID"),
    version_id: Optional[int] = Query(None, description="报价版本ID，不指定则使用当前版本或创建新版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    应用成本模板到报价
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    template = db.query(QuoteCostTemplate).filter(QuoteCostTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="成本模板不存在")
    
    if not template.is_active:
        raise HTTPException(status_code=400, detail="成本模板未启用")
    
    # 获取或创建报价版本
    if version_id:
        version = db.query(QuoteVersion).filter(QuoteVersion.id == version_id).first()
        if not version or version.quote_id != quote_id:
            raise HTTPException(status_code=404, detail="报价版本不存在")
    elif quote.current_version_id:
        version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
    else:
        # 创建新版本
        version_no = f"V{len(quote.versions) + 1}"
        version = QuoteVersion(
            quote_id=quote_id,
            version_no=version_no,
            created_by=current_user.id
        )
        db.add(version)
        db.flush()
        quote.current_version_id = version.id
    
    # 解析模板成本结构
    import json
    cost_structure = template.cost_structure if isinstance(template.cost_structure, dict) else json.loads(template.cost_structure) if template.cost_structure else {}
    
    # 解析调整项
    adj_dict = json.loads(adjustments) if isinstance(adjustments, str) else adjustments or {}
    
    # 清空现有成本明细（可选，这里保留现有明细）
    # db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).delete()
    
    # 应用模板成本项
    total_cost = Decimal('0')
    total_price = Decimal('0')
    
    for category in cost_structure.get('categories', []):
        for item_template in category.get('items', []):
            item_name = item_template.get('item_name', '')
            
            # 检查是否有调整
            if item_name in adj_dict:
                adj = adj_dict[item_name]
                qty = Decimal(str(adj.get('qty', item_template.get('default_qty', 0))))
                unit_price = Decimal(str(adj.get('unit_price', item_template.get('default_unit_price', 0))))
                cost = Decimal(str(adj.get('cost', item_template.get('default_cost', 0))))
            else:
                qty = Decimal(str(item_template.get('default_qty', 0)))
                unit_price = Decimal(str(item_template.get('default_unit_price', 0)))
                cost = Decimal(str(item_template.get('default_cost', 0)))
            
            # 创建报价明细
            item = QuoteItem(
                quote_version_id=version.id,
                item_type=item_template.get('item_type', category.get('category', '')),
                item_name=item_name,
                specification=item_template.get('specification'),
                unit=item_template.get('unit'),
                qty=float(qty),
                unit_price=float(unit_price),
                cost=float(cost),
                cost_category=category.get('category', ''),
                cost_source='TEMPLATE',
                lead_time_days=item_template.get('lead_time_days')
            )
            db.add(item)
            
            total_cost += cost * qty
            total_price += unit_price * qty
    
    # 更新版本信息
    version.cost_template_id = template_id
    version.total_price = float(total_price)
    version.cost_total = float(total_cost)
    if total_price > 0:
        version.gross_margin = float((total_price - total_cost) / total_price * 100)
    
    # 更新模板使用次数
    template.usage_count = (template.usage_count or 0) + 1
    
    db.commit()
    db.refresh(version)
    
    # 返回版本信息
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
    version_dict = {
        **{c.name: getattr(version, c.name) for c in version.__table__.columns},
        "created_by_name": version.creator.real_name if version.creator else None,
        "approved_by_name": version.approver.real_name if version.approver else None,
        "items": [QuoteItemResponse(**{c.name: getattr(item, c.name) for c in item.__table__.columns}) for item in items],
    }
    return QuoteVersionResponse(**version_dict)


@router.post("/quotes/{quote_id}/calculate-cost", response_model=ResponseModel)
def calculate_quote_cost(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="报价版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    自动计算报价成本
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")
    
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    
    total_cost = Decimal('0')
    total_price = Decimal(str(version.total_price or 0))
    
    for item in items:
        item_cost = Decimal(str(item.cost or 0)) * Decimal(str(item.qty or 0))
        total_cost += item_cost
    
    # 更新版本成本
    version.cost_total = float(total_cost)
    
    # 计算毛利率
    if total_price > 0:
        gross_margin = ((total_price - total_cost) / total_price * 100)
        version.gross_margin = float(gross_margin)
    else:
        version.gross_margin = None
    
    db.commit()
    db.refresh(version)
    
    return ResponseModel(
        code=200,
        message="成本计算完成",
        data={
            "total_price": float(total_price),
            "total_cost": float(total_cost),
            "gross_margin": float(version.gross_margin) if version.gross_margin else None
        }
    )


@router.get("/quotes/{quote_id}/cost-check", response_model=CostCheckResponse)
def check_quote_cost(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="报价版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    成本完整性检查
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")
    
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    
    checks = []
    
    # 检查1：是否有成本明细
    if not items:
        checks.append({
            "check_item": "成本明细",
            "status": "FAIL",
            "message": "未添加任何成本明细"
        })
    else:
        checks.append({
            "check_item": "成本明细",
            "status": "PASS",
            "message": f"已添加{len(items)}项成本明细"
        })
    
    # 检查2：成本项是否完整
    incomplete_items = []
    for item in items:
        if not item.cost or item.cost == 0:
            incomplete_items.append(item.item_name or f"项目{item.id}")
    
    if incomplete_items:
        checks.append({
            "check_item": "成本项完整性",
            "status": "FAIL",
            "message": f"以下成本项未填写成本：{', '.join(incomplete_items[:5])}{'...' if len(incomplete_items) > 5 else ''}"
        })
    else:
        checks.append({
            "check_item": "成本项完整性",
            "status": "PASS",
            "message": "所有成本项已填写"
        })
    
    # 检查3：毛利率检查
    margin_threshold = Decimal('20.0')  # 默认阈值20%
    current_margin = version.gross_margin or 0
    
    if current_margin < margin_threshold:
        checks.append({
            "check_item": "毛利率检查",
            "status": "WARNING" if current_margin >= 15 else "FAIL",
            "message": f"毛利率{current_margin:.2f}%，低于阈值{margin_threshold}%",
            "current_margin": float(current_margin),
            "threshold": float(margin_threshold)
        })
    else:
        checks.append({
            "check_item": "毛利率检查",
            "status": "PASS",
            "message": f"毛利率{current_margin:.2f}%，符合要求"
        })
    
    # 检查4：交期检查
    items_without_leadtime = []
    for item in items:
        if not item.lead_time_days and item.item_type in ['硬件', '外购件', '标准件']:
            items_without_leadtime.append(item.item_name or f"项目{item.id}")
    
    if items_without_leadtime:
        checks.append({
            "check_item": "交期校验",
            "status": "WARNING",
            "message": f"以下关键物料未填写交期：{', '.join(items_without_leadtime[:5])}{'...' if len(items_without_leadtime) > 5 else ''}"
        })
    else:
        checks.append({
            "check_item": "交期校验",
            "status": "PASS",
            "message": "关键物料交期已填写"
        })
    
    is_complete = all(check["status"] == "PASS" for check in checks)
    
    return CostCheckResponse(
        is_complete=is_complete,
        checks=checks,
        total_price=Decimal(str(version.total_price or 0)),
        total_cost=Decimal(str(version.cost_total or 0)),
        gross_margin=Decimal(str(current_margin))
    )


@router.get("/quotes/{quote_id}/delivery-validation", response_model=ResponseModel)
def validate_quote_delivery(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="报价版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    交期校验API

    验证报价交期的合理性，包括：
    - 物料交期查询
    - 项目周期估算
    - 交期合理性分析
    - 优化建议
    """
    from app.services.delivery_validation_service import delivery_validation_service

    # 获取报价单
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")

    # 获取版本
    if version_id:
        version = db.query(QuoteVersion).filter(
            QuoteVersion.id == version_id,
            QuoteVersion.quote_id == quote_id
        ).first()
    else:
        version = db.query(QuoteVersion).filter(
            QuoteVersion.quote_id == quote_id,
            QuoteVersion.is_current == True
        ).first()

    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")

    # 获取报价明细
    items = db.query(QuoteItem).filter(
        QuoteItem.quote_version_id == version.id
    ).all()

    # 执行交期校验
    validation_result = delivery_validation_service.validate_delivery_date(
        db, quote, version, items
    )

    return ResponseModel(
        code=200,
        message="交期校验完成",
        data=validation_result
    )


@router.get("/quotes/project-cycle-estimate", response_model=ResponseModel)
def estimate_project_cycle(
    *,
    contract_amount: Optional[float] = Query(None, description="合同金额（用于估算）"),
    project_type: Optional[str] = Query(None, description="项目类型"),
    complexity_level: str = Query("MEDIUM", description="复杂度：SIMPLE/MEDIUM/COMPLEX"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    项目周期估算API

    根据项目类型、金额、复杂度估算项目周期
    返回各阶段工期建议
    """
    from app.services.delivery_validation_service import DeliveryValidationService

    cycle_estimate = DeliveryValidationService.estimate_project_cycle(
        db=None,  # 不需要数据库
        contract_amount=contract_amount,
        project_type=project_type,
        complexity_level=complexity_level
    )

    return ResponseModel(
        code=200,
        message="项目周期估算完成",
        data=cycle_estimate
    )


@router.post("/quotes/{quote_id}/cost-approval/submit", response_model=QuoteCostApprovalResponse, status_code=201)
def submit_cost_approval(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    approval_in: QuoteCostApprovalCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    提交成本审批
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == approval_in.quote_version_id).first()
    if not version or version.quote_id != quote_id:
        raise HTTPException(status_code=404, detail="报价版本不存在")
    
    # 执行成本检查
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
    total_cost = sum([float(item.cost or 0) * float(item.qty or 0) for item in items])
    total_price = float(version.total_price or 0)
    gross_margin = float(version.gross_margin or 0) if version.gross_margin else ((total_price - total_cost) / total_price * 100 if total_price > 0 else 0)
    
    # 判断毛利率状态
    margin_threshold = 20.0 if approval_in.approval_level == 1 else 15.0
    if gross_margin >= margin_threshold:
        margin_status = "PASS"
    elif gross_margin >= 15.0:
        margin_status = "WARNING"
    else:
        margin_status = "FAIL"
    
    # 创建审批记录
    approval = QuoteCostApproval(
        quote_id=quote_id,
        quote_version_id=approval_in.quote_version_id,
        approval_status="PENDING",
        approval_level=approval_in.approval_level,
        current_approver_id=None,  # 根据审批层级确定审批人
        total_price=total_price,
        total_cost=total_cost,
        gross_margin=gross_margin,
        margin_threshold=margin_threshold,
        margin_status=margin_status,
        cost_complete=len(items) > 0 and all(item.cost and item.cost > 0 for item in items),
        delivery_check=all(item.lead_time_days for item in items if item.item_type in ['硬件', '外购件']),
        risk_terms_check=bool(version.risk_terms),
        approval_comment=approval_in.comment
    )
    db.add(approval)
    db.commit()
    db.refresh(approval)
    
    approval_dict = {
        **{c.name: getattr(approval, c.name) for c in approval.__table__.columns},
        "current_approver_name": approval.current_approver.real_name if approval.current_approver else None,
        "approved_by_name": approval.approver.real_name if approval.approver else None
    }
    return QuoteCostApprovalResponse(**approval_dict)


@router.post("/quotes/{quote_id}/cost-approval/{approval_id}/approve", response_model=QuoteCostApprovalResponse)
def approve_cost(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    approval_id: int,
    action: QuoteCostApprovalAction,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批通过
    """
    approval = db.query(QuoteCostApproval).filter(
        QuoteCostApproval.id == approval_id,
        QuoteCostApproval.quote_id == quote_id
    ).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")
    
    if approval.approval_status != "PENDING":
        raise HTTPException(status_code=400, detail="审批记录不是待审批状态")
    
    approval.approval_status = "APPROVED"
    approval.approved_by = current_user.id
    approval.approved_at = datetime.now()
    approval.approval_comment = action.comment
    
    # 如果是最低层级审批通过，更新报价版本状态
    if approval.approval_level == 1:
        version = db.query(QuoteVersion).filter(QuoteVersion.id == approval.quote_version_id).first()
        if version:
            version.cost_breakdown_complete = approval.cost_complete
            version.margin_warning = approval.margin_status in ["WARNING", "FAIL"]
    
    db.commit()
    db.refresh(approval)
    
    approval_dict = {
        **{c.name: getattr(approval, c.name) for c in approval.__table__.columns},
        "current_approver_name": approval.current_approver.real_name if approval.current_approver else None,
        "approved_by_name": approval.approver.real_name if approval.approver else None
    }
    return QuoteCostApprovalResponse(**approval_dict)


@router.post("/quotes/{quote_id}/cost-approval/{approval_id}/reject", response_model=QuoteCostApprovalResponse)
def reject_cost(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    approval_id: int,
    action: QuoteCostApprovalAction,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    审批驳回
    """
    if not action.reason:
        raise HTTPException(status_code=400, detail="驳回原因不能为空")
    
    approval = db.query(QuoteCostApproval).filter(
        QuoteCostApproval.id == approval_id,
        QuoteCostApproval.quote_id == quote_id
    ).first()
    if not approval:
        raise HTTPException(status_code=404, detail="审批记录不存在")
    
    if approval.approval_status != "PENDING":
        raise HTTPException(status_code=400, detail="审批记录不是待审批状态")
    
    approval.approval_status = "REJECTED"
    approval.approved_by = current_user.id
    approval.approved_at = datetime.now()
    approval.rejected_reason = action.reason
    approval.approval_comment = action.comment
    
    db.commit()
    db.refresh(approval)
    
    approval_dict = {
        **{c.name: getattr(approval, c.name) for c in approval.__table__.columns},
        "current_approver_name": approval.current_approver.real_name if approval.current_approver else None,
        "approved_by_name": approval.approver.real_name if approval.approver else None
    }
    return QuoteCostApprovalResponse(**approval_dict)


@router.get("/quotes/{quote_id}/cost-approval/history", response_model=List[QuoteCostApprovalResponse])
def get_cost_approval_history(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取成本审批历史
    """
    approvals = db.query(QuoteCostApproval).filter(
        QuoteCostApproval.quote_id == quote_id
    ).order_by(desc(QuoteCostApproval.created_at)).all()
    
    result = []
    for approval in approvals:
        approval_dict = {
            **{c.name: getattr(approval, c.name) for c in approval.__table__.columns},
            "current_approver_name": approval.current_approver.real_name if approval.current_approver else None,
            "approved_by_name": approval.approver.real_name if approval.approver else None
        }
        result.append(QuoteCostApprovalResponse(**approval_dict))
    
    return result


@router.get("/quotes/{quote_id}/cost-comparison", response_model=CostComparisonResponse)
def compare_quote_costs(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_ids: Optional[str] = Query(None, description="版本ID列表（逗号分隔），对比多个版本"),
    compare_quote_id: Optional[int] = Query(None, description="对比报价ID（与其他报价对比）"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    报价成本对比分析
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    # 获取当前版本
    current_version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
    if not current_version:
        raise HTTPException(status_code=400, detail="报价没有当前版本")
    
    current_items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == current_version.id).all()
    current_total_price = float(current_version.total_price or 0)
    current_total_cost = float(current_version.cost_total or 0)
    current_margin = float(current_version.gross_margin or 0)
    
    current_version_data = {
        "version_no": current_version.version_no,
        "total_price": current_total_price,
        "total_cost": current_total_cost,
        "gross_margin": current_margin
    }
    
    # 对比数据
    previous_version_data = None
    comparison = None
    breakdown_comparison = []
    
    # 如果指定了版本ID列表
    if version_ids:
        version_id_list = [int(vid) for vid in version_ids.split(',') if vid.strip()]
        if len(version_id_list) > 0:
            prev_version = db.query(QuoteVersion).filter(QuoteVersion.id == version_id_list[0]).first()
            if prev_version:
                prev_items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == prev_version.id).all()
                prev_total_price = float(prev_version.total_price or 0)
                prev_total_cost = float(prev_version.cost_total or 0)
                prev_margin = float(prev_version.gross_margin or 0)
                
                previous_version_data = {
                    "version_no": prev_version.version_no,
                    "total_price": prev_total_price,
                    "total_cost": prev_total_cost,
                    "gross_margin": prev_margin
                }
                
                # 计算对比
                price_change = current_total_price - prev_total_price
                price_change_pct = (price_change / prev_total_price * 100) if prev_total_price > 0 else 0
                cost_change = current_total_cost - prev_total_cost
                cost_change_pct = (cost_change / prev_total_cost * 100) if prev_total_cost > 0 else 0
                margin_change = current_margin - prev_margin
                margin_change_pct = (margin_change / prev_margin * 100) if prev_margin > 0 else 0
                
                comparison = {
                    "price_change": round(price_change, 2),
                    "price_change_pct": round(price_change_pct, 2),
                    "cost_change": round(cost_change, 2),
                    "cost_change_pct": round(cost_change_pct, 2),
                    "margin_change": round(margin_change, 2),
                    "margin_change_pct": round(margin_change_pct, 2)
                }
                
                # 按分类对比
                current_by_category = {}
                for item in current_items:
                    category = item.cost_category or "其他"
                    if category not in current_by_category:
                        current_by_category[category] = 0
                    current_by_category[category] += float(item.cost or 0) * float(item.qty or 0)
                
                prev_by_category = {}
                for item in prev_items:
                    category = item.cost_category or "其他"
                    if category not in prev_by_category:
                        prev_by_category[category] = 0
                    prev_by_category[category] += float(item.cost or 0) * float(item.qty or 0)
                
                all_categories = set(list(current_by_category.keys()) + list(prev_by_category.keys()))
                for category in all_categories:
                    v1_amount = prev_by_category.get(category, 0)
                    v2_amount = current_by_category.get(category, 0)
                    change = v2_amount - v1_amount
                    change_pct = (change / v1_amount * 100) if v1_amount > 0 else 0
                    breakdown_comparison.append({
                        "category": category,
                        "v1_amount": round(v1_amount, 2),
                        "v2_amount": round(v2_amount, 2),
                        "change": round(change, 2),
                        "change_pct": round(change_pct, 2)
                    })
    
    return CostComparisonResponse(
        current_version=current_version_data,
        previous_version=previous_version_data,
        comparison=comparison,
        breakdown_comparison=breakdown_comparison if breakdown_comparison else None
    )


# ==================== 采购物料成本清单管理 ====================


@router.get("/purchase-material-costs", response_model=PaginatedResponse[PurchaseMaterialCostResponse])
def get_purchase_material_costs(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    material_name: Optional[str] = Query(None, description="物料名称搜索"),
    material_type: Optional[str] = Query(None, description="物料类型筛选"),
    is_standard_part: Optional[bool] = Query(None, description="是否标准件"),
    is_active: Optional[bool] = Query(None, description="是否启用"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取采购物料成本清单列表（采购部维护的标准件成本信息）
    """
    query = db.query(PurchaseMaterialCost)
    
    if material_name:
        query = query.filter(PurchaseMaterialCost.material_name.like(f"%{material_name}%"))
    if material_type:
        query = query.filter(PurchaseMaterialCost.material_type == material_type)
    if is_standard_part is not None:
        query = query.filter(PurchaseMaterialCost.is_standard_part == is_standard_part)
    if is_active is not None:
        query = query.filter(PurchaseMaterialCost.is_active == is_active)
    
    total = query.count()
    offset = (page - 1) * page_size
    costs = query.order_by(desc(PurchaseMaterialCost.match_priority), desc(PurchaseMaterialCost.created_at)).offset(offset).limit(page_size).all()
    
    items = []
    for cost in costs:
        cost_dict = {
            **{c.name: getattr(cost, c.name) for c in cost.__table__.columns},
            "submitter_name": cost.submitter.real_name if cost.submitter else None
        }
        items.append(PurchaseMaterialCostResponse(**cost_dict))
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/purchase-material-costs/{cost_id}", response_model=PurchaseMaterialCostResponse)
def get_purchase_material_cost(
    *,
    db: Session = Depends(deps.get_db),
    cost_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取采购物料成本详情
    """
    cost = db.query(PurchaseMaterialCost).filter(PurchaseMaterialCost.id == cost_id).first()
    if not cost:
        raise HTTPException(status_code=404, detail="采购物料成本不存在")
    
    cost_dict = {
        **{c.name: getattr(cost, c.name) for c in cost.__table__.columns},
        "submitter_name": cost.submitter.real_name if cost.submitter else None
    }
    return PurchaseMaterialCostResponse(**cost_dict)


@router.post("/purchase-material-costs", response_model=PurchaseMaterialCostResponse, status_code=201)
def create_purchase_material_cost(
    *,
    db: Session = Depends(deps.get_db),
    cost_in: PurchaseMaterialCostCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    创建采购物料成本（采购部提交）
    """
    cost = PurchaseMaterialCost(
        **cost_in.model_dump(),
        submitted_by=current_user.id
    )
    db.add(cost)
    db.commit()
    db.refresh(cost)
    
    cost_dict = {
        **{c.name: getattr(cost, c.name) for c in cost.__table__.columns},
        "submitter_name": cost.submitter.real_name if cost.submitter else None
    }
    return PurchaseMaterialCostResponse(**cost_dict)


@router.put("/purchase-material-costs/{cost_id}", response_model=PurchaseMaterialCostResponse)
def update_purchase_material_cost(
    *,
    db: Session = Depends(deps.get_db),
    cost_id: int,
    cost_in: PurchaseMaterialCostUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新采购物料成本
    """
    cost = db.query(PurchaseMaterialCost).filter(PurchaseMaterialCost.id == cost_id).first()
    if not cost:
        raise HTTPException(status_code=404, detail="采购物料成本不存在")
    
    update_data = cost_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(cost, field):
            setattr(cost, field, value)
    
    db.add(cost)
    db.commit()
    db.refresh(cost)
    
    cost_dict = {
        **{c.name: getattr(cost, c.name) for c in cost.__table__.columns},
        "submitter_name": cost.submitter.real_name if cost.submitter else None
    }
    return PurchaseMaterialCostResponse(**cost_dict)


@router.delete("/purchase-material-costs/{cost_id}", status_code=200)
def delete_purchase_material_cost(
    *,
    db: Session = Depends(deps.get_db),
    cost_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    删除采购物料成本
    """
    cost = db.query(PurchaseMaterialCost).filter(PurchaseMaterialCost.id == cost_id).first()
    if not cost:
        raise HTTPException(status_code=404, detail="采购物料成本不存在")
    
    db.delete(cost)
    db.commit()
    
    return ResponseModel(code=200, message="删除成功")


@router.post("/purchase-material-costs/match", response_model=MaterialCostMatchResponse)
def match_material_cost(
    *,
    db: Session = Depends(deps.get_db),
    match_request: MaterialCostMatchRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    自动匹配物料成本（根据物料名称、规格等匹配历史采购价格）
    """
    # 查询启用的标准件成本清单
    query = db.query(PurchaseMaterialCost).filter(
        PurchaseMaterialCost.is_active == True,
        PurchaseMaterialCost.is_standard_part == True
    )
    
    # 匹配逻辑：优先精确匹配，其次模糊匹配
    matched_cost = None
    suggestions = []
    match_score = 0
    
    # 1. 精确匹配物料名称
    exact_match = query.filter(
        PurchaseMaterialCost.material_name == match_request.item_name
    ).order_by(desc(PurchaseMaterialCost.match_priority), desc(PurchaseMaterialCost.purchase_date)).first()
    
    if exact_match:
        matched_cost = exact_match
        match_score = 100
    else:
        # 2. 模糊匹配物料名称
        name_matches = query.filter(
            PurchaseMaterialCost.material_name.like(f"%{match_request.item_name}%")
        ).order_by(desc(PurchaseMaterialCost.match_priority), desc(PurchaseMaterialCost.purchase_date)).limit(5).all()
        
        if name_matches:
            matched_cost = name_matches[0]
            match_score = 80
            suggestions = name_matches[1:5] if len(name_matches) > 1 else []
        else:
            # 3. 关键词匹配
            if match_request.item_name:
                keywords = match_request.item_name.split()
                for keyword in keywords:
                    if len(keyword) > 2:  # 只匹配长度大于2的关键词
                        keyword_matches = query.filter(
                            or_(
                                PurchaseMaterialCost.material_name.like(f"%{keyword}%"),
                                PurchaseMaterialCost.match_keywords.like(f"%{keyword}%")
                            )
                        ).order_by(desc(PurchaseMaterialCost.match_priority), desc(PurchaseMaterialCost.usage_count)).limit(5).all()
                        
                        if keyword_matches:
                            matched_cost = keyword_matches[0]
                            match_score = 60
                            suggestions = keyword_matches[1:5] if len(keyword_matches) > 1 else []
                            break
    
    # 如果匹配成功，更新使用次数
    if matched_cost:
        matched_cost.usage_count = (matched_cost.usage_count or 0) + 1
        matched_cost.last_used_at = datetime.now()
        db.add(matched_cost)
        db.commit()
        db.refresh(matched_cost)
    
    # 构建响应
    matched_cost_dict = None
    if matched_cost:
        matched_cost_dict = {
            **{c.name: getattr(matched_cost, c.name) for c in matched_cost.__table__.columns},
            "submitter_name": matched_cost.submitter.real_name if matched_cost.submitter else None
        }
        matched_cost_dict = PurchaseMaterialCostResponse(**matched_cost_dict)
    
    suggestions_list = []
    for sug in suggestions:
        sug_dict = {
            **{c.name: getattr(sug, c.name) for c in sug.__table__.columns},
            "submitter_name": sug.submitter.real_name if sug.submitter else None
        }
        suggestions_list.append(PurchaseMaterialCostResponse(**sug_dict))
    
    return MaterialCostMatchResponse(
        matched=matched_cost is not None,
        match_score=match_score if matched_cost else None,
        matched_cost=matched_cost_dict,
        suggestions=suggestions_list
    )


@router.post("/quotes/{quote_id}/items/auto-match-cost-suggestions", response_model=CostMatchSuggestionsResponse)
def get_cost_match_suggestions(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取成本匹配建议（AI生成建议，不直接更新）
    根据物料名称和规格，从采购物料成本清单中生成匹配建议，包含异常检查
    """
    from app.services.cost_match_suggestion_service import (
        process_cost_match_suggestions,
        check_overall_anomalies,
        calculate_summary
    )
    
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="报价版本不存在")
    
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    
    # 获取成本清单查询
    cost_query = db.query(PurchaseMaterialCost).filter(
        PurchaseMaterialCost.is_active == True,
        PurchaseMaterialCost.is_standard_part == True
    )
    
    # 处理成本匹配建议
    suggestions, matched_count, unmatched_count, _, current_total_cost = process_cost_match_suggestions(
        db, items, cost_query
    )
    
    # 计算当前总价格
    current_total_price = float(version.total_price or 0)
    
    # 计算建议总成本
    suggested_total_cost = sum([
        float(s.suggested_cost or s.current_cost or 0) * 
        float(next((item.qty for item in items if item.id == s.item_id), 0) or 0)
        for s in suggestions
    ])
    
    # 整体异常检查
    warnings = check_overall_anomalies(
        current_total_price, current_total_cost, suggested_total_cost, items, suggestions
    )
    
    # 计算汇总
    summary = calculate_summary(current_total_cost, current_total_price, items, suggestions)
    
    return CostMatchSuggestionsResponse(
        suggestions=suggestions,
        total_items=len(items),
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        warnings=warnings if warnings else None,
        summary=summary
    )


@router.post("/quotes/{quote_id}/items/apply-cost-suggestions", response_model=ResponseModel)
def apply_cost_suggestions(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    version_id: Optional[int] = Query(None, description="版本ID，不指定则使用当前版本"),
    request: ApplyCostSuggestionsRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    应用成本匹配建议（人工确认后应用）
    将用户确认（可能修改过）的建议应用到报价明细中
    """
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    target_version_id = version_id or quote.current_version_id
    if not target_version_id:
        raise HTTPException(status_code=400, detail="报价没有指定版本")
    
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    item_dict = {item.id: item for item in items}
    
    applied_count = 0
    updated_cost_records = set()  # 记录已更新的成本记录，避免重复更新使用次数
    
    for suggestion_data in request.suggestions:
        item_id = suggestion_data.get("item_id")
        if not item_id:
            continue
        
        item = item_dict.get(item_id)
        if not item:
            continue
        
        # 应用建议（用户可能已修改）
        if "cost" in suggestion_data:
            item.cost = Decimal(str(suggestion_data["cost"]))
            item.cost_source = "HISTORY"
        
        if "specification" in suggestion_data:
            item.specification = suggestion_data["specification"]
        
        if "unit" in suggestion_data:
            item.unit = suggestion_data["unit"]
        
        if "lead_time_days" in suggestion_data:
            item.lead_time_days = suggestion_data["lead_time_days"]
        
        if "cost_category" in suggestion_data:
            item.cost_category = suggestion_data["cost_category"]
        
        db.add(item)
        applied_count += 1
        
        # 如果应用了成本建议，更新对应的成本记录使用次数
        # 注意：这里需要根据item_name匹配，因为suggestion_data中可能没有cost_record_id
        if "cost" in suggestion_data and item.item_name:
            matched_cost = db.query(PurchaseMaterialCost).filter(
                PurchaseMaterialCost.is_active == True,
                PurchaseMaterialCost.material_name.like(f"%{item.item_name}%")
            ).order_by(desc(PurchaseMaterialCost.match_priority), desc(PurchaseMaterialCost.purchase_date)).first()
            
            if matched_cost and matched_cost.id not in updated_cost_records:
                matched_cost.usage_count = (matched_cost.usage_count or 0) + 1
                matched_cost.last_used_at = datetime.now()
                db.add(matched_cost)
                updated_cost_records.add(matched_cost.id)
    
    db.commit()
    
    # 重新计算总成本
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == target_version_id).all()
    total_cost = sum([float(item.cost or 0) * float(item.qty or 0) for item in items])
    version = db.query(QuoteVersion).filter(QuoteVersion.id == target_version_id).first()
    if version:
        version.cost_total = total_cost
        total_price = float(version.total_price or 0)
        if total_price > 0:
            version.gross_margin = ((total_price - total_cost) / total_price * 100)
        db.add(version)
        db.commit()
    
    return ResponseModel(
        code=200,
        message=f"已应用 {applied_count} 项成本建议",
        data={
            "applied_count": applied_count,
            "total_cost": total_cost
        }
    )


# ==================== 物料成本更新提醒 ====================


@router.get("/purchase-material-costs/reminder", response_model=MaterialCostUpdateReminderResponse)
def get_cost_update_reminder(
    *,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取物料成本更新提醒配置和状态
    """
    reminder = db.query(MaterialCostUpdateReminder).filter(
        MaterialCostUpdateReminder.is_enabled == True
    ).order_by(desc(MaterialCostUpdateReminder.created_at)).first()
    
    if not reminder:
        # 创建默认提醒配置
        from datetime import date, timedelta
        reminder = MaterialCostUpdateReminder(
            reminder_type="PERIODIC",
            reminder_interval_days=30,
            next_reminder_date=date.today() + timedelta(days=30),
            is_enabled=True,
            include_standard=True,
            include_non_standard=True,
            notify_roles=["procurement", "procurement_manager", "采购工程师", "采购专员", "采购部经理"],
            reminder_count=0
        )
        db.add(reminder)
        db.commit()
        db.refresh(reminder)
    
    # 计算距离下次提醒的天数
    from datetime import date
    days_until_next = None
    is_due = False
    
    if reminder.next_reminder_date:
        today = date.today()
        delta = (reminder.next_reminder_date - today).days
        days_until_next = delta
        is_due = delta <= 0
    
    reminder_dict = {
        **{c.name: getattr(reminder, c.name) for c in reminder.__table__.columns},
        "days_until_next": days_until_next,
        "is_due": is_due
    }
    
    return MaterialCostUpdateReminderResponse(**reminder_dict)


@router.put("/purchase-material-costs/reminder", response_model=MaterialCostUpdateReminderResponse)
def update_cost_update_reminder(
    *,
    db: Session = Depends(deps.get_db),
    reminder_in: MaterialCostUpdateReminderUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    更新物料成本更新提醒配置
    """
    reminder = db.query(MaterialCostUpdateReminder).filter(
        MaterialCostUpdateReminder.is_enabled == True
    ).order_by(desc(MaterialCostUpdateReminder.created_at)).first()
    
    if not reminder:
        from datetime import date, timedelta
        reminder = MaterialCostUpdateReminder(
            reminder_type="PERIODIC",
            reminder_interval_days=30,
            next_reminder_date=date.today() + timedelta(days=30),
            is_enabled=True,
            include_standard=True,
            include_non_standard=True,
            notify_roles=["procurement", "procurement_manager", "采购工程师", "采购专员", "采购部经理"],
            reminder_count=0
        )
        db.add(reminder)
    
    update_data = reminder_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(reminder, field):
            setattr(reminder, field, value)
    
    reminder.last_updated_by = current_user.id
    reminder.last_updated_at = datetime.now()
    
    db.add(reminder)
    db.commit()
    db.refresh(reminder)
    
    # 计算距离下次提醒的天数
    from datetime import date
    days_until_next = None
    is_due = False
    
    if reminder.next_reminder_date:
        today = date.today()
        delta = (reminder.next_reminder_date - today).days
        days_until_next = delta
        is_due = delta <= 0
    
    reminder_dict = {
        **{c.name: getattr(reminder, c.name) for c in reminder.__table__.columns},
        "days_until_next": days_until_next,
        "is_due": is_due
    }
    
    return MaterialCostUpdateReminderResponse(**reminder_dict)


@router.post("/purchase-material-costs/reminder/acknowledge", response_model=ResponseModel)
def acknowledge_cost_update_reminder(
    *,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    确认物料成本更新提醒（标记为已处理，更新下次提醒日期）
    """
    reminder = db.query(MaterialCostUpdateReminder).filter(
        MaterialCostUpdateReminder.is_enabled == True
    ).order_by(desc(MaterialCostUpdateReminder.created_at)).first()
    
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒配置不存在")
    
    from datetime import date, timedelta
    
    # 更新提醒日期
    reminder.last_reminder_date = date.today()
    reminder.next_reminder_date = date.today() + timedelta(days=reminder.reminder_interval_days)
    reminder.reminder_count = (reminder.reminder_count or 0) + 1
    reminder.last_updated_by = current_user.id
    reminder.last_updated_at = datetime.now()
    
    db.add(reminder)
    db.commit()
    
    return ResponseModel(
        code=200,
        message="提醒已确认",
        data={
            "next_reminder_date": reminder.next_reminder_date.isoformat(),
            "reminder_count": reminder.reminder_count
        }
    )


# ==================== 技术评估 ====================


@router.post("/leads/{lead_id}/assessments/apply", response_model=ResponseModel, status_code=201)
def apply_lead_assessment(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: TechnicalAssessmentApplyRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """申请技术评估（线索）"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    # 创建评估申请
    assessment = TechnicalAssessment(
        source_type=AssessmentSourceTypeEnum.LEAD.value,
        source_id=lead_id,
        evaluator_id=request.evaluator_id or current_user.id,
        status=AssessmentStatusEnum.PENDING.value
    )
    
    db.add(assessment)
    db.flush()
    
    # 更新线索
    lead.assessment_id = assessment.id
    lead.assessment_status = AssessmentStatusEnum.PENDING.value
    
    db.commit()
    
    return ResponseModel(message="技术评估申请已提交", data={"assessment_id": assessment.id})


@router.post("/opportunities/{opp_id}/assessments/apply", response_model=ResponseModel, status_code=201)
def apply_opportunity_assessment(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    request: TechnicalAssessmentApplyRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """申请技术评估（商机）"""
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    # 创建评估申请
    assessment = TechnicalAssessment(
        source_type=AssessmentSourceTypeEnum.OPPORTUNITY.value,
        source_id=opp_id,
        evaluator_id=request.evaluator_id or current_user.id,
        status=AssessmentStatusEnum.PENDING.value
    )
    
    db.add(assessment)
    db.flush()
    
    # 更新商机
    opportunity.assessment_id = assessment.id
    opportunity.assessment_status = AssessmentStatusEnum.PENDING.value
    
    db.commit()
    
    return ResponseModel(message="技术评估申请已提交", data={"assessment_id": assessment.id})


@router.post("/assessments/{assessment_id}/evaluate", response_model=TechnicalAssessmentResponse, status_code=200)
async def evaluate_assessment(
    *,
    db: Session = Depends(deps.get_db),
    assessment_id: int,
    request: TechnicalAssessmentEvaluateRequest,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """执行技术评估"""
    assessment = db.query(TechnicalAssessment).filter(TechnicalAssessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="技术评估不存在")
    
    if assessment.status != AssessmentStatusEnum.PENDING.value:
        raise HTTPException(status_code=400, detail="评估状态不正确")
    
    # 可选：AI分析
    ai_analysis = None
    if request.enable_ai:
        ai_service = AIAssessmentService()
        if ai_service.is_available():
            ai_analysis = await ai_service.analyze_requirement(request.requirement_data)
    
    # 执行评估
    service = TechnicalAssessmentService(db)
    assessment = service.evaluate(
        assessment.source_type,
        assessment.source_id,
        current_user.id,
        request.requirement_data,
        ai_analysis=ai_analysis
    )
    
    db.commit()
    
    db.refresh(assessment)
    
    # 构建响应
    evaluator_name = None
    if assessment.evaluator_id:
        evaluator = db.query(User).filter(User.id == assessment.evaluator_id).first()
        evaluator_name = evaluator.real_name if evaluator else None
    
    return TechnicalAssessmentResponse(
        id=assessment.id,
        source_type=assessment.source_type,
        source_id=assessment.source_id,
        evaluator_id=assessment.evaluator_id,
        status=assessment.status,
        total_score=assessment.total_score,
        dimension_scores=assessment.dimension_scores,
        veto_triggered=assessment.veto_triggered,
        veto_rules=assessment.veto_rules,
        decision=assessment.decision,
        risks=assessment.risks,
        similar_cases=assessment.similar_cases,
        ai_analysis=assessment.ai_analysis,
        conditions=assessment.conditions,
        evaluated_at=assessment.evaluated_at,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        evaluator_name=evaluator_name
    )


@router.get("/leads/{lead_id}/assessments", response_model=List[TechnicalAssessmentResponse])
def get_lead_assessments(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取线索的技术评估列表"""
    assessments = db.query(TechnicalAssessment).filter(
        and_(
            TechnicalAssessment.source_type == AssessmentSourceTypeEnum.LEAD.value,
            TechnicalAssessment.source_id == lead_id
        )
    ).order_by(desc(TechnicalAssessment.created_at)).all()
    
    result = []
    for assessment in assessments:
        evaluator_name = None
        if assessment.evaluator_id:
            evaluator = db.query(User).filter(User.id == assessment.evaluator_id).first()
            evaluator_name = evaluator.real_name if evaluator else None
        
        result.append(TechnicalAssessmentResponse(
            id=assessment.id,
            source_type=assessment.source_type,
            source_id=assessment.source_id,
            evaluator_id=assessment.evaluator_id,
            status=assessment.status,
            total_score=assessment.total_score,
            dimension_scores=assessment.dimension_scores,
            veto_triggered=assessment.veto_triggered,
            veto_rules=assessment.veto_rules,
            decision=assessment.decision,
            risks=assessment.risks,
            similar_cases=assessment.similar_cases,
            ai_analysis=assessment.ai_analysis,
            conditions=assessment.conditions,
            evaluated_at=assessment.evaluated_at,
            created_at=assessment.created_at,
            updated_at=assessment.updated_at,
            evaluator_name=evaluator_name
        ))
    
    return result


@router.get("/opportunities/{opp_id}/assessments", response_model=List[TechnicalAssessmentResponse])
def get_opportunity_assessments(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取商机的技术评估列表"""
    assessments = db.query(TechnicalAssessment).filter(
        and_(
            TechnicalAssessment.source_type == AssessmentSourceTypeEnum.OPPORTUNITY.value,
            TechnicalAssessment.source_id == opp_id
        )
    ).order_by(desc(TechnicalAssessment.created_at)).all()
    
    result = []
    for assessment in assessments:
        evaluator_name = None
        if assessment.evaluator_id:
            evaluator = db.query(User).filter(User.id == assessment.evaluator_id).first()
            evaluator_name = evaluator.real_name if evaluator else None
        
        result.append(TechnicalAssessmentResponse(
            id=assessment.id,
            source_type=assessment.source_type,
            source_id=assessment.source_id,
            evaluator_id=assessment.evaluator_id,
            status=assessment.status,
            total_score=assessment.total_score,
            dimension_scores=assessment.dimension_scores,
            veto_triggered=assessment.veto_triggered,
            veto_rules=assessment.veto_rules,
            decision=assessment.decision,
            risks=assessment.risks,
            similar_cases=assessment.similar_cases,
            ai_analysis=assessment.ai_analysis,
            conditions=assessment.conditions,
            evaluated_at=assessment.evaluated_at,
            created_at=assessment.created_at,
            updated_at=assessment.updated_at,
            evaluator_name=evaluator_name
        ))
    
    return result


@router.get("/assessments/{assessment_id}", response_model=TechnicalAssessmentResponse)
def get_assessment(
    *,
    db: Session = Depends(deps.get_db),
    assessment_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取技术评估详情"""
    assessment = db.query(TechnicalAssessment).filter(TechnicalAssessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="技术评估不存在")
    
    evaluator_name = None
    if assessment.evaluator_id:
        evaluator = db.query(User).filter(User.id == assessment.evaluator_id).first()
        evaluator_name = evaluator.real_name if evaluator else None
    
    return TechnicalAssessmentResponse(
        id=assessment.id,
        source_type=assessment.source_type,
        source_id=assessment.source_id,
        evaluator_id=assessment.evaluator_id,
        status=assessment.status,
        total_score=assessment.total_score,
        dimension_scores=assessment.dimension_scores,
        veto_triggered=assessment.veto_triggered,
        veto_rules=assessment.veto_rules,
        decision=assessment.decision,
        risks=assessment.risks,
        similar_cases=assessment.similar_cases,
        ai_analysis=assessment.ai_analysis,
        conditions=assessment.conditions,
        evaluated_at=assessment.evaluated_at,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        evaluator_name=evaluator_name
    )


@router.get("/failure-cases/similar", response_model=List[FailureCaseResponse])
def find_similar_cases(
    *,
    db: Session = Depends(deps.get_db),
    industry: Optional[str] = Query(None, description="行业"),
    product_types: Optional[str] = Query(None, description="产品类型(JSON Array)"),
    takt_time_s: Optional[int] = Query(None, description="节拍时间(秒)"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """查找相似失败案例"""
    query = db.query(FailureCase)
    
    if industry:
        query = query.filter(FailureCase.industry == industry)
    
    cases = query.limit(10).all()
    
    result = []
    for case in cases:
        creator_name = None
        if case.created_by:
            creator = db.query(User).filter(User.id == case.created_by).first()
            creator_name = creator.real_name if creator else None
        
        result.append(FailureCaseResponse(
            **{c.name: getattr(case, c.name) for c in case.__table__.columns},
            creator_name=creator_name
        ))
    
    return result


@router.get("/open-items", response_model=PaginatedResponse[OpenItemResponse])
def list_open_items(
    *,
    db: Session = Depends(deps.get_db),
    source_type: Optional[str] = Query(None, description="来源类型"),
    source_id: Optional[int] = Query(None, description="来源ID"),
    status: Optional[str] = Query(None, description="状态"),
    blocks_quotation: Optional[bool] = Query(None, description="是否阻塞报价"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取未决事项列表"""
    query = db.query(OpenItem)
    
    if source_type:
        query = query.filter(OpenItem.source_type == source_type)
    if source_id:
        query = query.filter(OpenItem.source_id == source_id)
    if status:
        query = query.filter(OpenItem.status == status)
    if blocks_quotation is not None:
        query = query.filter(OpenItem.blocks_quotation == blocks_quotation)
    
    total = query.count()
    items = query.order_by(desc(OpenItem.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    result = []
    for item in items:
        responsible_person_name = None
        if item.responsible_person_id:
            person = db.query(User).filter(User.id == item.responsible_person_id).first()
            responsible_person_name = person.real_name if person else None
        
        result.append(OpenItemResponse(
            id=item.id,
            source_type=item.source_type,
            source_id=item.source_id,
            item_code=item.item_code,
            item_type=item.item_type,
            description=item.description,
            responsible_party=item.responsible_party,
            responsible_person_id=item.responsible_person_id,
            due_date=item.due_date,
            status=item.status,
            close_evidence=item.close_evidence,
            blocks_quotation=item.blocks_quotation,
            closed_at=item.closed_at,
            created_at=item.created_at,
            updated_at=item.updated_at,
            responsible_person_name=responsible_person_name
        ))
    
    return PaginatedResponse(
        items=result,
        total=total,
        page=page,
        page_size=page_size
    )


@router.post("/leads/{lead_id}/open-items", response_model=OpenItemResponse, status_code=201)
def create_open_item(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: OpenItemCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建未决事项（线索）"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    # 生成编号
    from datetime import datetime
    item_code = f"OI-{datetime.now().strftime('%y%m%d')}-{lead_id:03d}"
    
    open_item = OpenItem(
        source_type=AssessmentSourceTypeEnum.LEAD.value,
        source_id=lead_id,
        item_code=item_code,
        item_type=request.item_type,
        description=request.description,
        responsible_party=request.responsible_party,
        responsible_person_id=request.responsible_person_id,
        due_date=request.due_date,
        blocks_quotation=request.blocks_quotation
    )
    
    db.add(open_item)
    db.commit()
    db.refresh(open_item)
    
    responsible_person_name = None
    if open_item.responsible_person_id:
        person = db.query(User).filter(User.id == open_item.responsible_person_id).first()
        responsible_person_name = person.real_name if person else None
    
    return OpenItemResponse(
        id=open_item.id,
        source_type=open_item.source_type,
        source_id=open_item.source_id,
        item_code=open_item.item_code,
        item_type=open_item.item_type,
        description=open_item.description,
        responsible_party=open_item.responsible_party,
        responsible_person_id=open_item.responsible_person_id,
        due_date=open_item.due_date,
        status=open_item.status,
        close_evidence=open_item.close_evidence,
        blocks_quotation=open_item.blocks_quotation,
        closed_at=open_item.closed_at,
        created_at=open_item.created_at,
        updated_at=open_item.updated_at,
        responsible_person_name=responsible_person_name
    )


@router.post("/opportunities/{opp_id}/open-items", response_model=OpenItemResponse, status_code=201)
def create_open_item_for_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    request: OpenItemCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建未决事项（商机）"""
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    # 生成编号
    from datetime import datetime
    item_code = f"OI-{datetime.now().strftime('%y%m%d')}-{opp_id:03d}"
    
    open_item = OpenItem(
        source_type=AssessmentSourceTypeEnum.OPPORTUNITY.value,
        source_id=opp_id,
        item_code=item_code,
        item_type=request.item_type,
        description=request.description,
        responsible_party=request.responsible_party,
        responsible_person_id=request.responsible_person_id,
        due_date=request.due_date,
        blocks_quotation=request.blocks_quotation
    )
    
    db.add(open_item)
    db.commit()
    db.refresh(open_item)
    
    responsible_person_name = None
    if open_item.responsible_person_id:
        person = db.query(User).filter(User.id == open_item.responsible_person_id).first()
        responsible_person_name = person.real_name if person else None
    
    return OpenItemResponse(
        id=open_item.id,
        source_type=open_item.source_type,
        source_id=open_item.source_id,
        item_code=open_item.item_code,
        item_type=open_item.item_type,
        description=open_item.description,
        responsible_party=open_item.responsible_party,
        responsible_person_id=open_item.responsible_person_id,
        due_date=open_item.due_date,
        status=open_item.status,
        close_evidence=open_item.close_evidence,
        blocks_quotation=open_item.blocks_quotation,
        closed_at=open_item.closed_at,
        created_at=open_item.created_at,
        updated_at=open_item.updated_at,
        responsible_person_name=responsible_person_name
    )


@router.put("/open-items/{item_id}", response_model=OpenItemResponse)
def update_open_item(
    *,
    db: Session = Depends(deps.get_db),
    item_id: int,
    request: OpenItemCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """更新未决事项"""
    open_item = db.query(OpenItem).filter(OpenItem.id == item_id).first()
    if not open_item:
        raise HTTPException(status_code=404, detail="未决事项不存在")
    
    open_item.item_type = request.item_type
    open_item.description = request.description
    open_item.responsible_party = request.responsible_party
    open_item.responsible_person_id = request.responsible_person_id
    open_item.due_date = request.due_date
    open_item.blocks_quotation = request.blocks_quotation
    
    db.commit()
    db.refresh(open_item)
    
    responsible_person_name = None
    if open_item.responsible_person_id:
        person = db.query(User).filter(User.id == open_item.responsible_person_id).first()
        responsible_person_name = person.real_name if person else None
    
    return OpenItemResponse(
        id=open_item.id,
        source_type=open_item.source_type,
        source_id=open_item.source_id,
        item_code=open_item.item_code,
        item_type=open_item.item_type,
        description=open_item.description,
        responsible_party=open_item.responsible_party,
        responsible_person_id=open_item.responsible_person_id,
        due_date=open_item.due_date,
        status=open_item.status,
        close_evidence=open_item.close_evidence,
        blocks_quotation=open_item.blocks_quotation,
        closed_at=open_item.closed_at,
        created_at=open_item.created_at,
        updated_at=open_item.updated_at,
        responsible_person_name=responsible_person_name
    )


@router.post("/open-items/{item_id}/close", response_model=ResponseModel)
def close_open_item(
    *,
    db: Session = Depends(deps.get_db),
    item_id: int,
    close_evidence: Optional[str] = Query(None, description="关闭证据"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """关闭未决事项"""
    open_item = db.query(OpenItem).filter(OpenItem.id == item_id).first()
    if not open_item:
        raise HTTPException(status_code=404, detail="未决事项不存在")
    
    from app.models.enums import OpenItemStatusEnum
    from datetime import datetime
    
    open_item.status = OpenItemStatusEnum.CLOSED.value
    open_item.close_evidence = close_evidence
    open_item.closed_at = datetime.now()
    
    db.commit()
    
    return ResponseModel(message="未决事项已关闭")


@router.get("/scoring-rules", response_model=List[ScoringRuleResponse])
def list_scoring_rules(
    *,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取评分规则列表"""
    rules = db.query(ScoringRule).order_by(desc(ScoringRule.created_at)).all()
    
    result = []
    for rule in rules:
        creator_name = None
        if rule.created_by:
            creator = db.query(User).filter(User.id == rule.created_by).first()
            creator_name = creator.real_name if creator else None
        
        result.append(ScoringRuleResponse(
            id=rule.id,
            version=rule.version,
            is_active=rule.is_active,
            description=rule.description,
            created_by=rule.created_by,
            created_at=rule.created_at,
            updated_at=rule.updated_at,
            creator_name=creator_name
        ))
    
    return result


@router.post("/scoring-rules", response_model=ScoringRuleResponse, status_code=201)
def create_scoring_rule(
    *,
    db: Session = Depends(deps.get_db),
    request: ScoringRuleCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建评分规则"""
    # 检查版本号是否已存在
    existing = db.query(ScoringRule).filter(ScoringRule.version == request.version).first()
    if existing:
        raise HTTPException(status_code=400, detail="版本号已存在")
    
    rule = ScoringRule(
        version=request.version,
        rules_json=request.rules_json,
        description=request.description,
        created_by=current_user.id
    )
    
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    return ScoringRuleResponse(
        id=rule.id,
        version=rule.version,
        is_active=rule.is_active,
        description=rule.description,
        created_by=rule.created_by,
        created_at=rule.created_at,
        updated_at=rule.updated_at,
        creator_name=current_user.real_name
    )


@router.put("/scoring-rules/{rule_id}/activate", response_model=ResponseModel)
def activate_scoring_rule(
    *,
    db: Session = Depends(deps.get_db),
    rule_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """激活评分规则版本"""
    rule = db.query(ScoringRule).filter(ScoringRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="评分规则不存在")
    
    # 取消其他规则的激活状态
    db.query(ScoringRule).update({ScoringRule.is_active: False})
    
    # 激活当前规则
    rule.is_active = True
    db.commit()
    
    return ResponseModel(message="评分规则已激活")


@router.get("/failure-cases", response_model=PaginatedResponse[FailureCaseResponse])
def list_failure_cases(
    *,
    db: Session = Depends(deps.get_db),
    industry: Optional[str] = Query(None, description="行业"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取失败案例列表"""
    query = db.query(FailureCase)
    
    if industry:
        query = query.filter(FailureCase.industry == industry)
    
    if keyword:
        query = query.filter(
            or_(
                FailureCase.project_name.like(f"%{keyword}%"),
                FailureCase.core_failure_reason.like(f"%{keyword}%")
            )
        )
    
    total = query.count()
    cases = query.order_by(desc(FailureCase.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    result = []
    for case in cases:
        creator_name = None
        if case.created_by:
            creator = db.query(User).filter(User.id == case.created_by).first()
            creator_name = creator.real_name if creator else None
        
        result.append(FailureCaseResponse(
            **{c.name: getattr(case, c.name) for c in case.__table__.columns},
            creator_name=creator_name
        ))
    
    return PaginatedResponse(
        items=result,
        total=total,
        page=page,
        page_size=page_size
    )


@router.post("/failure-cases", response_model=FailureCaseResponse, status_code=201)
def create_failure_case(
    *,
    db: Session = Depends(deps.get_db),
    request: FailureCaseCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建失败案例"""
    # 检查案例编号是否已存在
    existing = db.query(FailureCase).filter(FailureCase.case_code == request.case_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="案例编号已存在")
    
    case = FailureCase(
        case_code=request.case_code,
        project_name=request.project_name,
        industry=request.industry,
        product_types=request.product_types,
        processes=request.processes,
        takt_time_s=request.takt_time_s,
        annual_volume=request.annual_volume,
        budget_status=request.budget_status,
        customer_project_status=request.customer_project_status,
        spec_status=request.spec_status,
        price_sensitivity=request.price_sensitivity,
        delivery_months=request.delivery_months,
        failure_tags=request.failure_tags,
        core_failure_reason=request.core_failure_reason,
        early_warning_signals=request.early_warning_signals,
        final_result=request.final_result,
        lesson_learned=request.lesson_learned,
        keywords=request.keywords,
        created_by=current_user.id
    )
    
    db.add(case)
    db.commit()
    db.refresh(case)
    
    return FailureCaseResponse(
        id=case.id,
        case_code=case.case_code,
        project_name=case.project_name,
        industry=case.industry,
        product_types=case.product_types,
        processes=case.processes,
        takt_time_s=case.takt_time_s,
        annual_volume=case.annual_volume,
        budget_status=case.budget_status,
        customer_project_status=case.customer_project_status,
        spec_status=case.spec_status,
        price_sensitivity=case.price_sensitivity,
        delivery_months=case.delivery_months,
        failure_tags=case.failure_tags,
        core_failure_reason=case.core_failure_reason,
        early_warning_signals=case.early_warning_signals,
        final_result=case.final_result,
        lesson_learned=case.lesson_learned,
        keywords=case.keywords,
        created_by=case.created_by,
        created_at=case.created_at,
        updated_at=case.updated_at,
        creator_name=current_user.real_name
    )


# ==================== 需求详情管理 ====================

@router.get("/leads/{lead_id}/requirement-detail", response_model=LeadRequirementDetailResponse)
def get_lead_requirement_detail(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取线索的需求详情"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    requirement_detail = db.query(LeadRequirementDetail).filter(
        LeadRequirementDetail.lead_id == lead_id
    ).first()
    
    if not requirement_detail:
        raise HTTPException(status_code=404, detail="需求详情不存在")
    
    frozen_by_name = None
    if requirement_detail.frozen_by:
        user = db.query(User).filter(User.id == requirement_detail.frozen_by).first()
        frozen_by_name = user.real_name if user else None
    
    return LeadRequirementDetailResponse(
        id=requirement_detail.id,
        lead_id=requirement_detail.lead_id,
        customer_factory_location=requirement_detail.customer_factory_location,
        target_object_type=requirement_detail.target_object_type,
        application_scenario=requirement_detail.application_scenario,
        delivery_mode=requirement_detail.delivery_mode,
        expected_delivery_date=requirement_detail.expected_delivery_date,
        requirement_source=requirement_detail.requirement_source,
        participant_ids=requirement_detail.participant_ids,
        requirement_maturity=requirement_detail.requirement_maturity,
        has_sow=requirement_detail.has_sow,
        has_interface_doc=requirement_detail.has_interface_doc,
        has_drawing_doc=requirement_detail.has_drawing_doc,
        sample_availability=requirement_detail.sample_availability,
        customer_support_resources=requirement_detail.customer_support_resources,
        key_risk_factors=requirement_detail.key_risk_factors,
        veto_triggered=requirement_detail.veto_triggered,
        veto_reason=requirement_detail.veto_reason,
        target_capacity_uph=float(requirement_detail.target_capacity_uph) if requirement_detail.target_capacity_uph else None,
        target_capacity_daily=float(requirement_detail.target_capacity_daily) if requirement_detail.target_capacity_daily else None,
        target_capacity_shift=float(requirement_detail.target_capacity_shift) if requirement_detail.target_capacity_shift else None,
        cycle_time_seconds=float(requirement_detail.cycle_time_seconds) if requirement_detail.cycle_time_seconds else None,
        workstation_count=requirement_detail.workstation_count,
        changeover_method=requirement_detail.changeover_method,
        yield_target=float(requirement_detail.yield_target) if requirement_detail.yield_target else None,
        retest_allowed=requirement_detail.retest_allowed,
        retest_max_count=requirement_detail.retest_max_count,
        traceability_type=requirement_detail.traceability_type,
        data_retention_period=requirement_detail.data_retention_period,
        data_format=requirement_detail.data_format,
        test_scope=requirement_detail.test_scope,
        key_metrics_spec=requirement_detail.key_metrics_spec,
        coverage_boundary=requirement_detail.coverage_boundary,
        exception_handling=requirement_detail.exception_handling,
        acceptance_method=requirement_detail.acceptance_method,
        acceptance_basis=requirement_detail.acceptance_basis,
        delivery_checklist=requirement_detail.delivery_checklist,
        interface_types=requirement_detail.interface_types,
        io_point_estimate=requirement_detail.io_point_estimate,
        communication_protocols=requirement_detail.communication_protocols,
        upper_system_integration=requirement_detail.upper_system_integration,
        data_field_list=requirement_detail.data_field_list,
        it_security_restrictions=requirement_detail.it_security_restrictions,
        power_supply=requirement_detail.power_supply,
        air_supply=requirement_detail.air_supply,
        environment=requirement_detail.environment,
        safety_requirements=requirement_detail.safety_requirements,
        space_and_logistics=requirement_detail.space_and_logistics,
        customer_site_standards=requirement_detail.customer_site_standards,
        customer_supplied_materials=requirement_detail.customer_supplied_materials,
        restricted_brands=requirement_detail.restricted_brands,
        specified_brands=requirement_detail.specified_brands,
        long_lead_items=requirement_detail.long_lead_items,
        spare_parts_requirement=requirement_detail.spare_parts_requirement,
        after_sales_support=requirement_detail.after_sales_support,
        requirement_version=requirement_detail.requirement_version,
        is_frozen=requirement_detail.is_frozen,
        frozen_at=requirement_detail.frozen_at,
        frozen_by=requirement_detail.frozen_by,
        frozen_by_name=frozen_by_name,
        created_at=requirement_detail.created_at,
        updated_at=requirement_detail.updated_at
    )


@router.post("/leads/{lead_id}/requirement-detail", response_model=LeadRequirementDetailResponse, status_code=201)
def create_lead_requirement_detail(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: LeadRequirementDetailCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建线索的需求详情"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    # 检查是否已存在
    existing = db.query(LeadRequirementDetail).filter(
        LeadRequirementDetail.lead_id == lead_id
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="需求详情已存在，请使用PUT方法更新")
    
    # 创建需求详情
    requirement_detail = LeadRequirementDetail(
        lead_id=lead_id,
        **request.dict(exclude_unset=True)
    )
    
    db.add(requirement_detail)
    db.commit()
    db.refresh(requirement_detail)
    
    # 更新线索的requirement_detail_id
    lead.requirement_detail_id = requirement_detail.id
    db.commit()
    
    return get_lead_requirement_detail(db=db, lead_id=lead_id, current_user=current_user)


@router.put("/leads/{lead_id}/requirement-detail", response_model=LeadRequirementDetailResponse)
def update_lead_requirement_detail(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: LeadRequirementDetailCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """更新线索的需求详情"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    requirement_detail = db.query(LeadRequirementDetail).filter(
        LeadRequirementDetail.lead_id == lead_id
    ).first()
    
    if not requirement_detail:
        raise HTTPException(status_code=404, detail="需求详情不存在")
    
    # 检查是否已冻结
    if requirement_detail.is_frozen:
        raise HTTPException(status_code=400, detail="需求已冻结，无法修改。如需修改，请先解冻或创建ECR/ECN")
    
    # 更新字段
    update_data = request.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(requirement_detail, field, value)
    
    db.commit()
    db.refresh(requirement_detail)
    
    return get_lead_requirement_detail(db=db, lead_id=lead_id, current_user=current_user)


# ==================== 需求冻结管理 ====================

@router.get("/leads/{lead_id}/requirement-freezes", response_model=List[RequirementFreezeResponse])
def list_lead_requirement_freezes(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取线索的需求冻结记录列表"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    freezes = db.query(RequirementFreeze).filter(
        and_(
            RequirementFreeze.source_type == AssessmentSourceTypeEnum.LEAD.value,
            RequirementFreeze.source_id == lead_id
        )
    ).order_by(desc(RequirementFreeze.freeze_time)).all()
    
    result = []
    for freeze in freezes:
        frozen_by_name = None
        if freeze.frozen_by:
            user = db.query(User).filter(User.id == freeze.frozen_by).first()
            frozen_by_name = user.real_name if user else None
        
        result.append(RequirementFreezeResponse(
            id=freeze.id,
            source_type=freeze.source_type,
            source_id=freeze.source_id,
            freeze_type=freeze.freeze_type,
            freeze_time=freeze.freeze_time,
            frozen_by=freeze.frozen_by,
            version_number=freeze.version_number,
            requires_ecr=freeze.requires_ecr,
            description=freeze.description,
            created_at=freeze.created_at,
            updated_at=freeze.updated_at,
            frozen_by_name=frozen_by_name
        ))
    
    return result


@router.post("/leads/{lead_id}/requirement-freezes", response_model=RequirementFreezeResponse, status_code=201)
def create_lead_requirement_freeze(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: RequirementFreezeCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建线索的需求冻结记录"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    # 检查需求详情是否存在
    requirement_detail = db.query(LeadRequirementDetail).filter(
        LeadRequirementDetail.lead_id == lead_id
    ).first()
    
    if not requirement_detail:
        raise HTTPException(status_code=400, detail="需求详情不存在，请先创建需求详情")
    
    # 创建冻结记录
    freeze = RequirementFreeze(
        source_type=AssessmentSourceTypeEnum.LEAD.value,
        source_id=lead_id,
        freeze_type=request.freeze_type,
        version_number=request.version_number,
        requires_ecr=request.requires_ecr,
        description=request.description,
        frozen_by=current_user.id
    )
    
    db.add(freeze)
    
    # 更新需求详情为冻结状态
    from datetime import datetime
    requirement_detail.is_frozen = True
    requirement_detail.frozen_at = datetime.now()
    requirement_detail.frozen_by = current_user.id
    requirement_detail.requirement_version = request.version_number
    
    db.commit()
    db.refresh(freeze)
    
    frozen_by_name = current_user.real_name
    
    return RequirementFreezeResponse(
        id=freeze.id,
        source_type=freeze.source_type,
        source_id=freeze.source_id,
        freeze_type=freeze.freeze_type,
        freeze_time=freeze.freeze_time,
        frozen_by=freeze.frozen_by,
        version_number=freeze.version_number,
        requires_ecr=freeze.requires_ecr,
        description=freeze.description,
        created_at=freeze.created_at,
        updated_at=freeze.updated_at,
        frozen_by_name=frozen_by_name
    )


@router.get("/opportunities/{opp_id}/requirement-freezes", response_model=List[RequirementFreezeResponse])
def list_opportunity_requirement_freezes(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取商机的需求冻结记录列表"""
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    freezes = db.query(RequirementFreeze).filter(
        and_(
            RequirementFreeze.source_type == AssessmentSourceTypeEnum.OPPORTUNITY.value,
            RequirementFreeze.source_id == opp_id
        )
    ).order_by(desc(RequirementFreeze.freeze_time)).all()
    
    result = []
    for freeze in freezes:
        frozen_by_name = None
        if freeze.frozen_by:
            user = db.query(User).filter(User.id == freeze.frozen_by).first()
            frozen_by_name = user.real_name if user else None
        
        result.append(RequirementFreezeResponse(
            id=freeze.id,
            source_type=freeze.source_type,
            source_id=freeze.source_id,
            freeze_type=freeze.freeze_type,
            freeze_time=freeze.freeze_time,
            frozen_by=freeze.frozen_by,
            version_number=freeze.version_number,
            requires_ecr=freeze.requires_ecr,
            description=freeze.description,
            created_at=freeze.created_at,
            updated_at=freeze.updated_at,
            frozen_by_name=frozen_by_name
        ))
    
    return result


@router.post("/opportunities/{opp_id}/requirement-freezes", response_model=RequirementFreezeResponse, status_code=201)
def create_opportunity_requirement_freeze(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    request: RequirementFreezeCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建商机的需求冻结记录"""
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    # 创建冻结记录
    freeze = RequirementFreeze(
        source_type=AssessmentSourceTypeEnum.OPPORTUNITY.value,
        source_id=opp_id,
        freeze_type=request.freeze_type,
        version_number=request.version_number,
        requires_ecr=request.requires_ecr,
        description=request.description,
        frozen_by=current_user.id
    )
    
    db.add(freeze)
    db.commit()
    db.refresh(freeze)
    
    frozen_by_name = current_user.real_name
    
    return RequirementFreezeResponse(
        id=freeze.id,
        source_type=freeze.source_type,
        source_id=freeze.source_id,
        freeze_type=freeze.freeze_type,
        freeze_time=freeze.freeze_time,
        frozen_by=freeze.frozen_by,
        version_number=freeze.version_number,
        requires_ecr=freeze.requires_ecr,
        description=freeze.description,
        created_at=freeze.created_at,
        updated_at=freeze.updated_at,
        frozen_by_name=frozen_by_name
    )


# ==================== AI澄清管理 ====================

@router.get("/ai-clarifications", response_model=PaginatedResponse[AIClarificationResponse])
def list_ai_clarifications(
    *,
    db: Session = Depends(deps.get_db),
    source_type: Optional[str] = Query(None, description="来源类型"),
    source_id: Optional[int] = Query(None, description="来源ID"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取AI澄清记录列表"""
    query = db.query(AIClarification)
    
    if source_type:
        query = query.filter(AIClarification.source_type == source_type)
    if source_id:
        query = query.filter(AIClarification.source_id == source_id)
    
    total = query.count()
    clarifications = query.order_by(
        desc(AIClarification.source_type),
        desc(AIClarification.source_id),
        desc(AIClarification.round)
    ).offset((page - 1) * page_size).limit(page_size).all()
    
    items = []
    for clarification in clarifications:
        items.append(AIClarificationResponse(
            id=clarification.id,
            source_type=clarification.source_type,
            source_id=clarification.source_id,
            round=clarification.round,
            questions=clarification.questions,
            answers=clarification.answers,
            created_at=clarification.created_at,
            updated_at=clarification.updated_at
        ))
    
    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/leads/{lead_id}/ai-clarifications", response_model=AIClarificationResponse, status_code=201)
def create_ai_clarification_for_lead(
    *,
    db: Session = Depends(deps.get_db),
    lead_id: int,
    request: AIClarificationCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建AI澄清记录（线索）"""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    
    # 获取当前最大轮次
    max_round = db.query(func.max(AIClarification.round)).filter(
        and_(
            AIClarification.source_type == AssessmentSourceTypeEnum.LEAD.value,
            AIClarification.source_id == lead_id
        )
    ).scalar() or 0
    
    clarification = AIClarification(
        source_type=AssessmentSourceTypeEnum.LEAD.value,
        source_id=lead_id,
        round=max_round + 1,
        questions=request.questions,
        answers=request.answers
    )
    
    db.add(clarification)
    db.commit()
    db.refresh(clarification)
    
    return AIClarificationResponse(
        id=clarification.id,
        source_type=clarification.source_type,
        source_id=clarification.source_id,
        round=clarification.round,
        questions=clarification.questions,
        answers=clarification.answers,
        created_at=clarification.created_at,
        updated_at=clarification.updated_at
    )


@router.post("/opportunities/{opp_id}/ai-clarifications", response_model=AIClarificationResponse, status_code=201)
def create_ai_clarification_for_opportunity(
    *,
    db: Session = Depends(deps.get_db),
    opp_id: int,
    request: AIClarificationCreate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """创建AI澄清记录（商机）"""
    opportunity = db.query(Opportunity).filter(Opportunity.id == opp_id).first()
    if not opportunity:
        raise HTTPException(status_code=404, detail="商机不存在")
    
    # 获取当前最大轮次
    max_round = db.query(func.max(AIClarification.round)).filter(
        and_(
            AIClarification.source_type == AssessmentSourceTypeEnum.OPPORTUNITY.value,
            AIClarification.source_id == opp_id
        )
    ).scalar() or 0
    
    clarification = AIClarification(
        source_type=AssessmentSourceTypeEnum.OPPORTUNITY.value,
        source_id=opp_id,
        round=max_round + 1,
        questions=request.questions,
        answers=request.answers
    )
    
    db.add(clarification)
    db.commit()
    db.refresh(clarification)
    
    return AIClarificationResponse(
        id=clarification.id,
        source_type=clarification.source_type,
        source_id=clarification.source_id,
        round=clarification.round,
        questions=clarification.questions,
        answers=clarification.answers,
        created_at=clarification.created_at,
        updated_at=clarification.updated_at
    )


@router.put("/ai-clarifications/{clarification_id}", response_model=AIClarificationResponse)
def update_ai_clarification(
    *,
    db: Session = Depends(deps.get_db),
    clarification_id: int,
    request: AIClarificationUpdate,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """更新AI澄清记录（回答）"""
    clarification = db.query(AIClarification).filter(
        AIClarification.id == clarification_id
    ).first()
    
    if not clarification:
        raise HTTPException(status_code=404, detail="AI澄清记录不存在")
    
    clarification.answers = request.answers
    
    db.commit()
    db.refresh(clarification)
    
    return AIClarificationResponse(
        id=clarification.id,
        source_type=clarification.source_type,
        source_id=clarification.source_id,
        round=clarification.round,
        questions=clarification.questions,
        answers=clarification.answers,
        created_at=clarification.created_at,
        updated_at=clarification.updated_at
    )


@router.get("/ai-clarifications/{clarification_id}", response_model=AIClarificationResponse)
def get_ai_clarification(
    *,
    db: Session = Depends(deps.get_db),
    clarification_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """获取AI澄清记录详情"""
    clarification = db.query(AIClarification).filter(
        AIClarification.id == clarification_id
    ).first()
    
    if not clarification:
        raise HTTPException(status_code=404, detail="AI澄清记录不存在")
    
    return AIClarificationResponse(
        id=clarification.id,
        source_type=clarification.source_type,
        source_id=clarification.source_id,
        round=clarification.round,
        questions=clarification.questions,
        answers=clarification.answers,
        created_at=clarification.created_at,
        updated_at=clarification.updated_at
    )

# ==================== 数据导出 ====================


# ==================== 数据导出 ====================


@router.get("/leads/export")
def export_leads(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.2: 导出线索列表（Excel）
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    
    query = db.query(Lead)
    if keyword:
        query = query.filter(or_(Lead.lead_code.contains(keyword), Lead.customer_name.contains(keyword), Lead.contact_name.contains(keyword)))
    if status:
        query = query.filter(Lead.status == status)
    if owner_id:
        query = query.filter(Lead.owner_id == owner_id)
    
    leads = query.order_by(Lead.created_at.desc()).all()
    export_service = ExcelExportService()
    columns = [
        {"key": "lead_code", "label": "线索编码", "width": 15},
        {"key": "source", "label": "来源", "width": 15},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "industry", "label": "行业", "width": 15},
        {"key": "contact_name", "label": "联系人", "width": 15},
        {"key": "contact_phone", "label": "联系电话", "width": 15},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "owner_name", "label": "负责人", "width": 12},
        {"key": "next_action_at", "label": "下次行动时间", "width": 18, "format": export_service.format_date},
        {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
    ]
    
    data = [{
        "lead_code": lead.lead_code,
        "source": lead.source or '',
        "customer_name": lead.customer_name or '',
        "industry": lead.industry or '',
        "contact_name": lead.contact_name or '',
        "contact_phone": lead.contact_phone or '',
        "status": lead.status,
        "owner_name": lead.owner.real_name if lead.owner else '',
        "next_action_at": lead.next_action_at,
        "created_at": lead.created_at,
    } for lead in leads]
    
    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="线索列表", title="线索列表")
    filename = f"线索列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


@router.get("/opportunities/export")
def export_opportunities(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    stage: Optional[str] = Query(None, description="阶段筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.2: 导出商机列表（Excel）
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    
    query = db.query(Opportunity)
    if keyword:
        query = query.filter(or_(Opportunity.opp_code.contains(keyword), Opportunity.opp_name.contains(keyword), Opportunity.customer.has(Customer.customer_name.contains(keyword))))
    if stage:
        query = query.filter(Opportunity.stage == stage)
    if status:
        query = query.filter(Opportunity.status == status)
    if owner_id:
        query = query.filter(Opportunity.owner_id == owner_id)
    
    opportunities = query.order_by(Opportunity.created_at.desc()).all()
    export_service = ExcelExportService()
    columns = [
        {"key": "opp_code", "label": "商机编码", "width": 15},
        {"key": "opp_name", "label": "商机名称", "width": 30},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "stage", "label": "阶段", "width": 15},
        {"key": "est_amount", "label": "预估金额", "width": 15, "format": export_service.format_currency},
        {"key": "est_margin", "label": "预估毛利率", "width": 12, "format": export_service.format_percentage},
        {"key": "score", "label": "评分", "width": 8},
        {"key": "risk_level", "label": "风险等级", "width": 10},
        {"key": "owner_name", "label": "负责人", "width": 12},
        {"key": "gate_status", "label": "阶段门状态", "width": 15},
        {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
    ]
    
    data = [{
        "opp_code": opp.opp_code,
        "opp_name": opp.opp_name,
        "customer_name": opp.customer.customer_name if opp.customer else '',
        "stage": opp.stage,
        "est_amount": float(opp.est_amount) if opp.est_amount else 0,
        "est_margin": float(opp.est_margin) if opp.est_margin else 0,
        "score": opp.score or 0,
        "risk_level": opp.risk_level or '',
        "owner_name": opp.owner.real_name if opp.owner else '',
        "gate_status": opp.gate_status,
        "created_at": opp.created_at,
    } for opp in opportunities]
    
    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="商机列表", title="商机列表")
    filename = f"商机列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


@router.get("/quotes/export")
def export_quotes(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    include_items: bool = Query(False, description="是否包含明细"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.3: 导出报价列表（Excel）
    支持导出报价主表和明细（多 Sheet）
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    
    query = db.query(Quote)
    if keyword:
        query = query.filter(or_(Quote.quote_code.contains(keyword), Quote.opportunity.has(Opportunity.opp_name.contains(keyword))))
    if status:
        query = query.filter(Quote.status == status)
    if customer_id:
        query = query.filter(Quote.customer_id == customer_id)
    if owner_id:
        query = query.filter(Quote.owner_id == owner_id)
    
    quotes = query.order_by(Quote.created_at.desc()).all()
    export_service = ExcelExportService()
    
    if include_items:
        sheets = []
        quote_data = []
        for quote in quotes:
            version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
            quote_data.append({
                "quote_code": quote.quote_code,
                "opp_code": quote.opportunity.opp_code if quote.opportunity else '',
                "customer_name": quote.customer.customer_name if quote.customer else '',
                "status": quote.status,
                "total_price": float(version.total_price) if version and version.total_price else 0,
                "total_cost": float(version.total_cost) if version and version.total_cost else 0,
                "gross_margin": float(version.gross_margin) if version and version.gross_margin else 0,
                "valid_until": version.valid_until if version and version.valid_until else None,
                "owner_name": quote.owner.real_name if quote.owner else '',
                "created_at": quote.created_at,
            })
        sheets.append({
            "name": "报价列表",
            "data": quote_data,
            "columns": [
                {"key": "quote_code", "label": "报价编码", "width": 15},
                {"key": "opp_code", "label": "商机编码", "width": 15},
                {"key": "customer_name", "label": "客户名称", "width": 25},
                {"key": "status", "label": "状态", "width": 12},
                {"key": "total_price", "label": "报价金额", "width": 15, "format": export_service.format_currency},
                {"key": "total_cost", "label": "成本金额", "width": 15, "format": export_service.format_currency},
                {"key": "gross_margin", "label": "毛利率", "width": 12, "format": export_service.format_percentage},
                {"key": "valid_until", "label": "有效期至", "width": 12, "format": export_service.format_date},
                {"key": "owner_name", "label": "负责人", "width": 12},
                {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
            ],
            "title": "报价列表"
        })
        item_data = []
        for quote in quotes:
            version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
            if version:
                items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
                for item in items:
                    item_data.append({
                        "quote_code": quote.quote_code,
                        "item_name": item.item_name or '',
                        "specification": item.specification or '',
                        "qty": float(item.qty) if item.qty else 0,
                        "unit": item.unit or '',
                        "unit_price": float(item.unit_price) if item.unit_price else 0,
                        "total_price": float(item.total_price) if item.total_price else 0,
                        "cost": float(item.cost) if item.cost else 0,
                        "item_type": item.item_type or '',
                    })
        sheets.append({
            "name": "报价明细",
            "data": item_data,
            "columns": [
                {"key": "quote_code", "label": "报价编码", "width": 15},
                {"key": "item_name", "label": "物料名称", "width": 30},
                {"key": "specification", "label": "规格型号", "width": 25},
                {"key": "qty", "label": "数量", "width": 10},
                {"key": "unit", "label": "单位", "width": 8},
                {"key": "unit_price", "label": "单价", "width": 12, "format": export_service.format_currency},
                {"key": "total_price", "label": "总价", "width": 12, "format": export_service.format_currency},
                {"key": "cost", "label": "成本", "width": 12, "format": export_service.format_currency},
                {"key": "item_type", "label": "类型", "width": 12},
            ],
            "title": "报价明细"
        })
        excel_data = export_service.export_multisheet(sheets)
    else:
        quote_data = []
        for quote in quotes:
            version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
            quote_data.append({
                "quote_code": quote.quote_code,
                "opp_code": quote.opportunity.opp_code if quote.opportunity else '',
                "customer_name": quote.customer.customer_name if quote.customer else '',
                "status": quote.status,
                "total_price": float(version.total_price) if version and version.total_price else 0,
                "total_cost": float(version.total_cost) if version and version.total_cost else 0,
                "gross_margin": float(version.gross_margin) if version and version.gross_margin else 0,
                "valid_until": version.valid_until if version and version.valid_until else None,
                "owner_name": quote.owner.real_name if quote.owner else '',
                "created_at": quote.created_at,
            })
        columns = [
            {"key": "quote_code", "label": "报价编码", "width": 15},
            {"key": "opp_code", "label": "商机编码", "width": 15},
            {"key": "customer_name", "label": "客户名称", "width": 25},
            {"key": "status", "label": "状态", "width": 12},
            {"key": "total_price", "label": "报价金额", "width": 15, "format": export_service.format_currency},
            {"key": "total_cost", "label": "成本金额", "width": 15, "format": export_service.format_currency},
            {"key": "gross_margin", "label": "毛利率", "width": 12, "format": export_service.format_percentage},
            {"key": "valid_until", "label": "有效期至", "width": 12, "format": export_service.format_date},
            {"key": "owner_name", "label": "负责人", "width": 12},
            {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
        ]
        excel_data = export_service.export_to_excel(data=quote_data, columns=columns, sheet_name="报价列表", title="报价列表")
    
    filename = f"报价列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


@router.get("/contracts/export")
def export_contracts(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    owner_id: Optional[int] = Query(None, description="负责人ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.3: 导出合同列表（Excel）
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    
    query = db.query(Contract)
    if keyword:
        query = query.filter(or_(Contract.contract_code.contains(keyword), Contract.contract_name.contains(keyword), Contract.opportunity.has(Opportunity.opp_name.contains(keyword))))
    if status:
        query = query.filter(Contract.status == status)
    if customer_id:
        query = query.filter(Contract.customer_id == customer_id)
    if owner_id:
        query = query.filter(Contract.owner_id == owner_id)
    
    contracts = query.order_by(Contract.created_at.desc()).all()
    export_service = ExcelExportService()
    columns = [
        {"key": "contract_code", "label": "合同编码", "width": 15},
        {"key": "contract_name", "label": "合同名称", "width": 30},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "contract_amount", "label": "合同金额", "width": 15, "format": export_service.format_currency},
        {"key": "signed_date", "label": "签订日期", "width": 12, "format": export_service.format_date},
        {"key": "delivery_deadline", "label": "交期", "width": 12, "format": export_service.format_date},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "project_code", "label": "项目编码", "width": 15},
        {"key": "owner_name", "label": "负责人", "width": 12},
        {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
    ]
    
    data = [{
        "contract_code": contract.contract_code,
        "contract_name": contract.contract_name or '',
        "customer_name": contract.customer.customer_name if contract.customer else '',
        "contract_amount": float(contract.contract_amount) if contract.contract_amount else 0,
        "signed_date": contract.signed_date,
        "delivery_deadline": contract.delivery_deadline,
        "status": contract.status,
        "project_code": contract.project.project_code if contract.project else '',
        "owner_name": contract.owner.real_name if contract.owner else '',
        "created_at": contract.created_at,
    } for contract in contracts]
    
    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="合同列表", title="合同列表")
    filename = f"合同列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


@router.get("/invoices/export")
def export_invoices(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.4: 导出发票列表（Excel）
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    
    query = db.query(Invoice)
    if keyword:
        query = query.filter(or_(Invoice.invoice_code.contains(keyword), Invoice.contract.has(Contract.contract_code.contains(keyword))))
    if status:
        query = query.filter(Invoice.status == status)
    if customer_id:
        query = query.filter(Invoice.contract.has(Contract.customer_id == customer_id))
    
    invoices = query.order_by(Invoice.created_at.desc()).all()
    export_service = ExcelExportService()
    columns = [
        {"key": "invoice_code", "label": "发票编码", "width": 15},
        {"key": "contract_code", "label": "合同编码", "width": 15},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "invoice_type", "label": "发票类型", "width": 12},
        {"key": "amount", "label": "发票金额", "width": 15, "format": export_service.format_currency},
        {"key": "paid_amount", "label": "已收金额", "width": 15, "format": export_service.format_currency},
        {"key": "unpaid_amount", "label": "未收金额", "width": 15, "format": export_service.format_currency},
        {"key": "issue_date", "label": "开票日期", "width": 12, "format": export_service.format_date},
        {"key": "due_date", "label": "到期日期", "width": 12, "format": export_service.format_date},
        {"key": "payment_status", "label": "收款状态", "width": 12},
        {"key": "status", "label": "发票状态", "width": 12},
        {"key": "created_at", "label": "创建时间", "width": 18, "format": export_service.format_date},
    ]
    
    data = []
    for invoice in invoices:
        total_amount = float(invoice.total_amount or invoice.amount or 0)
        paid_amount = float(invoice.paid_amount or 0)
        unpaid_amount = total_amount - paid_amount
        data.append({
            "invoice_code": invoice.invoice_code,
            "contract_code": invoice.contract.contract_code if invoice.contract else '',
            "customer_name": invoice.contract.customer.customer_name if invoice.contract and invoice.contract.customer else '',
            "invoice_type": invoice.invoice_type or '',
            "amount": total_amount,
            "paid_amount": paid_amount,
            "unpaid_amount": unpaid_amount,
            "issue_date": invoice.issue_date,
            "due_date": invoice.due_date,
            "payment_status": invoice.payment_status or '',
            "status": invoice.status,
            "created_at": invoice.created_at,
        })
    
    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="发票列表", title="发票列表")
    filename = f"发票列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


@router.get("/payments/export")
def export_payments(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    include_aging: bool = Query(True, description="是否包含账龄分析"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.4: 导出应收账款列表（Excel）
    包含账龄分析
    """
    from app.services.excel_export_service import ExcelExportService, create_excel_response
    from app.models.project import ProjectPaymentPlan
    
    query = db.query(ProjectPaymentPlan).join(Contract).filter(ProjectPaymentPlan.status.in_(["PENDING", "INVOICED", "PARTIAL"]))
    if keyword:
        query = query.filter(or_(ProjectPaymentPlan.payment_name.contains(keyword), ProjectPaymentPlan.contract.has(Contract.contract_code.contains(keyword))))
    if status:
        query = query.filter(ProjectPaymentPlan.status == status)
    if customer_id:
        query = query.filter(ProjectPaymentPlan.contract.has(Contract.customer_id == customer_id))
    
    payment_plans = query.order_by(ProjectPaymentPlan.planned_date).all()
    export_service = ExcelExportService()
    today = date.today()
    
    columns = [
        {"key": "payment_name", "label": "收款计划名称", "width": 25},
        {"key": "contract_code", "label": "合同编码", "width": 15},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "project_code", "label": "项目编码", "width": 15},
        {"key": "planned_amount", "label": "计划金额", "width": 15, "format": export_service.format_currency},
        {"key": "actual_amount", "label": "已收金额", "width": 15, "format": export_service.format_currency},
        {"key": "unpaid_amount", "label": "未收金额", "width": 15, "format": export_service.format_currency},
        {"key": "planned_date", "label": "计划日期", "width": 12, "format": export_service.format_date},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "overdue_days", "label": "逾期天数", "width": 10},
    ]
    
    if include_aging:
        columns.extend([
            {"key": "aging_0_30", "label": "0-30天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_31_60", "label": "31-60天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_61_90", "label": "61-90天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_over_90", "label": "90天以上", "width": 12, "format": export_service.format_currency},
        ])
    
    data = []
    for plan in payment_plans:
        planned_amount = float(plan.planned_amount or 0)
        actual_amount = float(plan.actual_amount or 0)
        unpaid_amount = planned_amount - actual_amount
        overdue_days = 0
        if plan.planned_date and plan.planned_date < today:
            overdue_days = (today - plan.planned_date).days
        
        row_data = {
            "payment_name": plan.payment_name or '',
            "contract_code": plan.contract.contract_code if plan.contract else '',
            "customer_name": plan.contract.customer.customer_name if plan.contract and plan.contract.customer else '',
            "project_code": plan.project.project_code if plan.project else '',
            "planned_amount": planned_amount,
            "actual_amount": actual_amount,
            "unpaid_amount": unpaid_amount,
            "planned_date": plan.planned_date,
            "status": plan.status or '',
            "overdue_days": overdue_days,
        }
        
        if include_aging:
            aging_0_30 = aging_31_60 = aging_61_90 = aging_over_90 = 0
            if unpaid_amount > 0 and plan.planned_date:
                if overdue_days <= 30:
                    aging_0_30 = unpaid_amount
                elif overdue_days <= 60:
                    aging_31_60 = unpaid_amount
                elif overdue_days <= 90:
                    aging_61_90 = unpaid_amount
                else:
                    aging_over_90 = unpaid_amount
            row_data.update({"aging_0_30": aging_0_30, "aging_31_60": aging_31_60, "aging_61_90": aging_61_90, "aging_over_90": aging_over_90})
        
        data.append(row_data)
    
    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="应收账款列表", title="应收账款列表（含账龄分析）" if include_aging else "应收账款列表")
    filename = f"应收账款列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)


# ==================== PDF 导出 ====================


@router.get("/quotes/{quote_id}/pdf")
def export_quote_pdf(
    *,
    db: Session = Depends(deps.get_db),
    quote_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.5: 导出报价单 PDF
    """
    from app.services.pdf_export_service import PDFExportService, create_pdf_response
    
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="报价不存在")
    
    version = db.query(QuoteVersion).filter(QuoteVersion.id == quote.current_version_id).first()
    if not version:
        raise HTTPException(status_code=400, detail="报价没有当前版本")
    
    items = db.query(QuoteItem).filter(QuoteItem.quote_version_id == version.id).all()
    
    # 准备数据
    quote_data = {
        "quote_code": quote.quote_code,
        "customer_name": quote.customer.customer_name if quote.customer else '',
        "created_at": quote.created_at,
        "valid_until": version.valid_until,
        "total_price": float(version.total_price) if version.total_price else 0,
        "status": quote.status,
    }
    
    quote_items = [{
        "item_name": item.item_name or '',
        "specification": item.specification or '',
        "qty": float(item.qty) if item.qty else 0,
        "unit": item.unit or '',
        "unit_price": float(item.unit_price) if item.unit_price else 0,
        "total_price": float(item.total_price) if item.total_price else 0,
        "remark": item.remark or '',
    } for item in items]
    
    pdf_service = PDFExportService()
    pdf_data = pdf_service.export_quote_to_pdf(quote_data, quote_items)
    
    filename = f"报价单_{quote.quote_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return create_pdf_response(pdf_data, filename)


@router.get("/contracts/{contract_id}/pdf")
def export_contract_pdf(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.5: 导出合同 PDF
    """
    from app.services.pdf_export_service import PDFExportService, create_pdf_response
    
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    
    deliverables = db.query(ContractDeliverable).filter(ContractDeliverable.contract_id == contract_id).all()
    
    # 准备数据
    contract_data = {
        "contract_code": contract.contract_code,
        "contract_name": contract.contract_name or '',
        "customer_name": contract.customer.customer_name if contract.customer else '',
        "contract_amount": float(contract.contract_amount) if contract.contract_amount else 0,
        "signed_date": contract.signed_date,
        "delivery_deadline": contract.delivery_deadline,
        "status": contract.status,
    }
    
    deliverable_list = [{
        "deliverable_name": d.deliverable_name or '',
        "quantity": float(d.quantity) if d.quantity else 0,
        "unit": d.unit or '',
        "remark": d.remark or '',
    } for d in deliverables]
    
    pdf_service = PDFExportService()
    pdf_data = pdf_service.export_contract_to_pdf(contract_data, deliverable_list)
    
    filename = f"合同_{contract.contract_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return create_pdf_response(pdf_data, filename)


@router.get("/invoices/{invoice_id}/pdf")
def export_invoice_pdf(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.5: 导出发票 PDF
    """
    from app.services.pdf_export_service import PDFExportService, create_pdf_response
    
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    
    # 准备数据
    total_amount = float(invoice.total_amount or invoice.amount or 0)
    paid_amount = float(invoice.paid_amount or 0)
    
    invoice_data = {
        "invoice_code": invoice.invoice_code,
        "contract_code": invoice.contract.contract_code if invoice.contract else '',
        "customer_name": invoice.contract.customer.customer_name if invoice.contract and invoice.contract.customer else '',
        "invoice_type": invoice.invoice_type or '',
        "total_amount": total_amount,
        "amount": total_amount,
        "paid_amount": paid_amount,
        "issue_date": invoice.issue_date,
        "due_date": invoice.due_date,
        "payment_status": invoice.payment_status or '',
        "status": invoice.status,
    }
    
    pdf_service = PDFExportService()
    pdf_data = pdf_service.export_invoice_to_pdf(invoice_data)
    
    filename = f"发票_{invoice.invoice_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return create_pdf_response(pdf_data, filename)

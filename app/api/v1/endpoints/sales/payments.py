# -*- coding: utf-8 -*-
"""
回款管理 API endpoints

包含：
- 收款计划调整
- 收款计划调整历史
- 回款记录列表
- 回款登记
- 回款统计分析
- 回款提醒
- 回款详情
- 发票核销
- 收款计划列表
- 回款记录导出
- 应收账款导出
"""

from io import BytesIO
from typing import Any, Optional
from datetime import date, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, or_

from app.api import deps
from app.core.config import settings
from app.core import security
from app.models.user import User
from app.models.project import ProjectPaymentPlan
from app.models.sales import Contract, Invoice
from app.schemas.common import PaginatedResponse, ResponseModel

router = APIRouter()


# ==================== 收款计划调整 ====================


@router.post("/payment-plans/{plan_id}/adjust", response_model=ResponseModel)
def adjust_payment_plan(
    *,
    db: Session = Depends(deps.get_db),
    plan_id: int,
    new_date: date = Query(..., description="新的收款日期"),
    reason: str = Query(..., description="调整原因"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 7.3: 手动调整收款计划
    记录调整历史并发送通知
    """
    from app.services.payment_adjustment_service import PaymentAdjustmentService

    service = PaymentAdjustmentService(db)
    result = service.manual_adjust_payment_plan(
        plan_id=plan_id,
        new_date=new_date,
        reason=reason,
        adjusted_by=current_user.id,
    )

    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "调整失败"))

    return ResponseModel(
        code=200,
        message=result.get("message", "收款计划已调整"),
        data=result
    )


@router.get("/payment-plans/{plan_id}/adjustment-history", response_model=ResponseModel)
def get_payment_adjustment_history(
    *,
    db: Session = Depends(deps.get_db),
    plan_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 7.3: 获取收款计划调整历史
    """
    from app.services.payment_adjustment_service import PaymentAdjustmentService

    service = PaymentAdjustmentService(db)
    history = service.get_adjustment_history(plan_id)

    return ResponseModel(
        code=200,
        message="success",
        data={"history": history}
    )


# ==================== 回款记录 ====================


@router.get("/payments", response_model=PaginatedResponse)
def get_payment_records(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    payment_status: Optional[str] = Query(None, description="收款状态筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款记录列表（基于发票）
    Issue 7.1: 已集成数据权限过滤（财务和销售总监可以看到所有收款数据）
    """
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")

    # Issue 7.1: 应用财务数据权限过滤（财务和销售总监可以看到所有收款数据）
    query = security.filter_sales_finance_data_by_scope(query, current_user, db, Invoice, 'owner_id')

    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)

    if project_id:
        query = query.filter(Invoice.project_id == project_id)

    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)

    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)

    if start_date:
        query = query.filter(Invoice.paid_date >= start_date)

    if end_date:
        query = query.filter(Invoice.paid_date <= end_date)

    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(desc(Invoice.paid_date)).offset(offset).limit(page_size).all()

    items = []
    for invoice in invoices:
        contract = invoice.contract
        items.append({
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": invoice.project.project_code if invoice.project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "invoice_amount": float(invoice.total_amount or invoice.amount or 0),
            "paid_amount": float(invoice.paid_amount or 0),
            "unpaid_amount": float((invoice.total_amount or invoice.amount or 0) - (invoice.paid_amount or 0)),
            "payment_status": invoice.payment_status,
            "issue_date": invoice.issue_date,
            "due_date": invoice.due_date,
            "paid_date": invoice.paid_date,
            "overdue_days": (date.today() - invoice.due_date).days if invoice.due_date and invoice.due_date < date.today() and invoice.payment_status in ["PENDING", "PARTIAL"] else None,
        })

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.post("/payments", response_model=ResponseModel)
def create_payment_record(
    *,
    db: Session = Depends(deps.get_db),
    invoice_id: int = Query(..., description="发票ID"),
    paid_amount: Decimal = Query(..., description="收款金额"),
    paid_date: date = Query(..., description="收款日期"),
    payment_method: Optional[str] = Query(None, description="收款方式"),
    bank_account: Optional[str] = Query(None, description="收款账户"),
    remark: Optional[str] = Query(None, description="备注"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    登记回款
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if invoice.status != "ISSUED":
        raise HTTPException(status_code=400, detail="只有已开票的发票才能登记回款")

    # 更新收款信息
    current_paid = invoice.paid_amount or Decimal("0")
    new_paid = current_paid + paid_amount
    invoice.paid_amount = new_paid
    invoice.paid_date = paid_date

    # 更新收款状态
    total = invoice.total_amount or invoice.amount or Decimal("0")
    if new_paid >= total:
        invoice.payment_status = "PAID"
    elif new_paid > Decimal("0"):
        invoice.payment_status = "PARTIAL"
    else:
        invoice.payment_status = "PENDING"

    # 更新备注
    payment_note = f"收款记录: {paid_date}, 金额: {paid_amount}"
    if payment_method:
        payment_note += f", 方式: {payment_method}"
    if bank_account:
        payment_note += f", 账户: {bank_account}"
    if remark:
        payment_note += f", 备注: {remark}"

    invoice.remark = (invoice.remark or "") + f"\n{payment_note}"

    db.commit()

    return ResponseModel(
        code=200,
        message="回款登记成功",
        data={
            "invoice_id": invoice.id,
            "paid_amount": float(new_paid),
            "payment_status": invoice.payment_status,
            "unpaid_amount": float(total - new_paid)
        }
    )


@router.get("/payments/statistics", response_model=ResponseModel)
def get_payment_statistics(
    *,
    db: Session = Depends(deps.get_db),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    回款统计分析

    注意：此路由必须定义在 /payments/{payment_id} 之前，否则 FastAPI 会将 "statistics" 解析为 payment_id
    """
    from app.services.payment_statistics_service import (
        build_invoice_query,
        calculate_monthly_statistics,
        calculate_customer_statistics,
        calculate_status_statistics,
        calculate_overdue_amount,
        build_monthly_list,
        build_customer_list
    )

    # 构建查询
    query = build_invoice_query(db, customer_id, start_date, end_date)
    invoices = query.all()

    today = date.today()

    # 计算各项统计
    monthly_stats = calculate_monthly_statistics(invoices)
    customer_stats = calculate_customer_statistics(invoices)
    status_stats = calculate_status_statistics(invoices)

    # 计算汇总
    total_invoiced = sum([invoice.total_amount or invoice.amount or Decimal("0") for invoice in invoices])
    total_paid = sum([invoice.paid_amount or Decimal("0") for invoice in invoices])
    total_unpaid = total_invoiced - total_paid
    total_overdue = calculate_overdue_amount(invoices, today)

    collection_rate = (total_paid / total_invoiced * 100) if total_invoiced > 0 else Decimal("0")

    # 构建列表
    monthly_list = build_monthly_list(monthly_stats)
    customer_list = build_customer_list(customer_stats, limit=10)

    return ResponseModel(
        code=200,
        message="success",
        data={
            "summary": {
                "total_invoiced": float(total_invoiced),
                "total_paid": float(total_paid),
                "total_unpaid": float(total_unpaid),
                "total_overdue": float(total_overdue),
                "collection_rate": float(collection_rate),
                "invoice_count": len(invoices),
            },
            "monthly_statistics": monthly_list,
            "customer_statistics": customer_list,
            "status_statistics": {
                "PAID": {
                    "count": status_stats["PAID"]["count"],
                    "amount": float(status_stats["PAID"]["amount"])
                },
                "PARTIAL": {
                    "count": status_stats["PARTIAL"]["count"],
                    "amount": float(status_stats["PARTIAL"]["amount"])
                },
                "PENDING": {
                    "count": status_stats["PENDING"]["count"],
                    "amount": float(status_stats["PENDING"]["amount"])
                },
            }
        }
    )


@router.get("/payments/reminders", response_model=PaginatedResponse)
def get_payment_reminders(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    days_before: int = Query(7, ge=0, description="提前提醒天数"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款提醒列表（即将到期和已逾期的回款）

    注意：此路由必须定义在 /payments/{payment_id} 之前，否则 FastAPI 会将 "reminders" 解析为 payment_id
    """
    today = date.today()
    reminder_date = today + timedelta(days=days_before)

    query = db.query(Invoice).filter(
        Invoice.status == "ISSUED",
        Invoice.payment_status.in_(["PENDING", "PARTIAL"]),
        Invoice.due_date.isnot(None),
        Invoice.due_date <= reminder_date
    )

    total = query.count()
    offset = (page - 1) * page_size
    invoices = query.order_by(Invoice.due_date).offset(offset).limit(page_size).all()

    items = []
    for invoice in invoices:
        contract = invoice.contract
        unpaid = (invoice.total_amount or invoice.amount or Decimal("0")) - (invoice.paid_amount or Decimal("0"))
        days_until_due = (invoice.due_date - today).days if invoice.due_date else None
        is_overdue = days_until_due is not None and days_until_due < 0

        items.append({
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": invoice.project.project_code if invoice.project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "unpaid_amount": float(unpaid),
            "due_date": invoice.due_date,
            "days_until_due": days_until_due,
            "is_overdue": is_overdue,
            "overdue_days": abs(days_until_due) if is_overdue else None,
            "payment_status": invoice.payment_status,
            "reminder_level": "urgent" if is_overdue else ("warning" if days_until_due is not None and days_until_due <= 3 else "normal"),
        })

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


@router.get("/payments/{payment_id}", response_model=ResponseModel)
def get_payment_detail(
    *,
    db: Session = Depends(deps.get_db),
    payment_id: int,
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取回款详情（基于发票ID）
    """
    invoice = db.query(Invoice).options(
        joinedload(Invoice.contract),
        joinedload(Invoice.project)
    ).filter(Invoice.id == payment_id).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    contract = invoice.contract
    project = invoice.project

    total = invoice.total_amount or invoice.amount or Decimal("0")
    paid = invoice.paid_amount or Decimal("0")
    unpaid = total - paid

    overdue_days = None
    if invoice.due_date and invoice.due_date < date.today() and invoice.payment_status in ["PENDING", "PARTIAL"]:
        overdue_days = (date.today() - invoice.due_date).days

    return ResponseModel(
        code=200,
        message="success",
        data={
            "id": invoice.id,
            "invoice_code": invoice.invoice_code,
            "contract_id": invoice.contract_id,
            "contract_code": contract.contract_code if contract else None,
            "project_id": invoice.project_id,
            "project_code": project.project_code if project else None,
            "customer_id": contract.customer_id if contract else None,
            "customer_name": contract.customer.customer_name if contract and contract.customer else None,
            "invoice_amount": float(total),
            "paid_amount": float(paid),
            "unpaid_amount": float(unpaid),
            "payment_status": invoice.payment_status,
            "issue_date": invoice.issue_date,
            "due_date": invoice.due_date,
            "paid_date": invoice.paid_date,
            "overdue_days": overdue_days,
            "remark": invoice.remark,
        }
    )


@router.put("/payments/{payment_id}/match-invoice", response_model=ResponseModel)
def match_payment_to_invoice(
    *,
    db: Session = Depends(deps.get_db),
    payment_id: int,
    invoice_id: int = Query(..., description="发票ID"),
    match_amount: Optional[Decimal] = Query(None, description="核销金额，不指定则核销全部未收金额"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    核销发票（将回款记录与发票关联）
    注意：这里 payment_id 实际上是发票ID，用于保持API路径一致性
    """
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if invoice.status != "ISSUED":
        raise HTTPException(status_code=400, detail="只有已开票的发票才能核销")

    total = invoice.total_amount or invoice.amount or Decimal("0")
    current_paid = invoice.paid_amount or Decimal("0")
    unpaid = total - current_paid

    if match_amount:
        if match_amount > unpaid:
            raise HTTPException(status_code=400, detail=f"核销金额不能超过未收金额 {unpaid}")
        new_paid = current_paid + match_amount
    else:
        # 核销全部未收金额
        new_paid = total

    invoice.paid_amount = new_paid
    invoice.paid_date = date.today()

    # 更新收款状态
    if new_paid >= total:
        invoice.payment_status = "PAID"
    elif new_paid > Decimal("0"):
        invoice.payment_status = "PARTIAL"

    db.commit()

    return ResponseModel(
        code=200,
        message="发票核销成功",
        data={
            "invoice_id": invoice.id,
            "matched_amount": float(match_amount or unpaid),
            "paid_amount": float(new_paid),
            "payment_status": invoice.payment_status
        }
    )


# ==================== 收款计划列表 ====================


@router.get("/payment-plans", response_model=PaginatedResponse)
def get_payment_plans(
    *,
    db: Session = Depends(deps.get_db),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE, description="每页数量"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    获取收款计划列表
    """
    query = db.query(ProjectPaymentPlan)

    if project_id:
        query = query.filter(ProjectPaymentPlan.project_id == project_id)

    if contract_id:
        query = query.filter(ProjectPaymentPlan.contract_id == contract_id)

    if status:
        query = query.filter(ProjectPaymentPlan.status == status)

    total = query.count()
    offset = (page - 1) * page_size
    plans = query.order_by(ProjectPaymentPlan.planned_date).offset(offset).limit(page_size).all()

    items = []
    for plan in plans:
        items.append({
            "id": plan.id,
            "payment_no": plan.payment_no,
            "project_id": plan.project_id,
            "project_code": plan.project.project_code if plan.project else None,
            "contract_id": plan.contract_id,
            "contract_code": plan.contract.contract_code if plan.contract else None,
            "payment_stage": plan.payment_stage,
            "payment_ratio": float(plan.payment_ratio or 0),
            "planned_amount": float(plan.planned_amount or 0),
            "actual_amount": float(plan.actual_amount or 0),
            "planned_date": plan.planned_date,
            "actual_date": plan.actual_date,
            "milestone_id": plan.milestone_id,
            "milestone_name": plan.milestone.milestone_name if plan.milestone else None,
            "trigger_milestone": plan.trigger_milestone,
            "status": plan.status,
            "invoice_id": plan.invoice_id,
            "invoice_no": plan.invoice_no,
        })

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size
    )


# ==================== 导出功能 ====================


@router.get("/payments/invoices/export")
def export_payment_invoices(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    payment_status: Optional[str] = Query(None, description="收款状态筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出回款记录（基于发票，Excel格式）
    """
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")

    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)

    if project_id:
        query = query.filter(Invoice.project_id == project_id)

    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)

    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)

    if start_date:
        query = query.filter(Invoice.paid_date >= start_date)

    if end_date:
        query = query.filter(Invoice.paid_date <= end_date)

    invoices = query.order_by(desc(Invoice.paid_date)).all()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "回款记录"

    # 设置表头
    headers = [
        "发票号", "合同编号", "项目编号", "客户名称", "发票金额",
        "已收金额", "未收金额", "收款状态", "开票日期", "到期日期",
        "实际收款日期", "逾期天数", "备注"
    ]

    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    today = date.today()
    for row, invoice in enumerate(invoices, 2):
        contract = invoice.contract
        total = invoice.total_amount or invoice.amount or Decimal("0")
        paid = invoice.paid_amount or Decimal("0")
        unpaid = total - paid

        overdue_days = None
        if invoice.due_date and invoice.due_date < today and invoice.payment_status in ["PENDING", "PARTIAL"]:
            overdue_days = (today - invoice.due_date).days

        ws.cell(row=row, column=1, value=invoice.invoice_code or "")
        ws.cell(row=row, column=2, value=contract.contract_code if contract else "")
        ws.cell(row=row, column=3, value=invoice.project.project_code if invoice.project else "")
        ws.cell(row=row, column=4, value=contract.customer.customer_name if contract and contract.customer else "")
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=6, value=float(paid))
        ws.cell(row=row, column=7, value=float(unpaid))
        ws.cell(row=row, column=8, value=invoice.payment_status or "")
        ws.cell(row=row, column=9, value=invoice.issue_date.strftime("%Y-%m-%d") if invoice.issue_date else "")
        ws.cell(row=row, column=10, value=invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "")
        ws.cell(row=row, column=11, value=invoice.paid_date.strftime("%Y-%m-%d") if invoice.paid_date else "")
        ws.cell(row=row, column=12, value=overdue_days if overdue_days else "")
        ws.cell(row=row, column=13, value=invoice.remark or "")

    # 调整列宽
    column_widths = [15, 15, 15, 20, 12, 12, 12, 12, 12, 12, 12, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f"回款记录_{date.today().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/payments/export")
def export_payments(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    include_aging: bool = Query(True, description="是否包含账龄分析"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.4: 导出应收账款列表（Excel）
    包含账龄分析
    """
    from datetime import datetime
    from app.services.excel_export_service import ExcelExportService, create_excel_response

    query = db.query(ProjectPaymentPlan).join(Contract).filter(ProjectPaymentPlan.status.in_(["PENDING", "INVOICED", "PARTIAL"]))
    if keyword:
        query = query.filter(or_(ProjectPaymentPlan.payment_name.contains(keyword), ProjectPaymentPlan.contract.has(Contract.contract_code.contains(keyword))))
    if status:
        query = query.filter(ProjectPaymentPlan.status == status)
    if customer_id:
        query = query.filter(ProjectPaymentPlan.contract.has(Contract.customer_id == customer_id))

    payment_plans = query.order_by(ProjectPaymentPlan.planned_date).all()
    export_service = ExcelExportService()
    today = date.today()

    columns = [
        {"key": "payment_name", "label": "收款计划名称", "width": 25},
        {"key": "contract_code", "label": "合同编码", "width": 15},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "project_code", "label": "项目编码", "width": 15},
        {"key": "planned_amount", "label": "计划金额", "width": 15, "format": export_service.format_currency},
        {"key": "actual_amount", "label": "已收金额", "width": 15, "format": export_service.format_currency},
        {"key": "unpaid_amount", "label": "未收金额", "width": 15, "format": export_service.format_currency},
        {"key": "planned_date", "label": "计划日期", "width": 12, "format": export_service.format_date},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "overdue_days", "label": "逾期天数", "width": 10},
    ]

    if include_aging:
        columns.extend([
            {"key": "aging_0_30", "label": "0-30天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_31_60", "label": "31-60天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_61_90", "label": "61-90天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_over_90", "label": "90天以上", "width": 12, "format": export_service.format_currency},
        ])

    data = []
    for plan in payment_plans:
        planned_amount = float(plan.planned_amount or 0)
        actual_amount = float(plan.actual_amount or 0)
        unpaid_amount = planned_amount - actual_amount
        overdue_days = 0
        if plan.planned_date and plan.planned_date < today:
            overdue_days = (today - plan.planned_date).days

        row_data = {
            "payment_name": plan.payment_name or '',
            "contract_code": plan.contract.contract_code if plan.contract else '',
            "customer_name": plan.contract.customer.customer_name if plan.contract and plan.contract.customer else '',
            "project_code": plan.project.project_code if plan.project else '',
            "planned_amount": planned_amount,
            "actual_amount": actual_amount,
            "unpaid_amount": unpaid_amount,
            "planned_date": plan.planned_date,
            "status": plan.status or '',
            "overdue_days": overdue_days,
        }

        if include_aging:
            aging_0_30 = aging_31_60 = aging_61_90 = aging_over_90 = 0
            if unpaid_amount > 0 and plan.planned_date:
                if overdue_days <= 30:
                    aging_0_30 = unpaid_amount
                elif overdue_days <= 60:
                    aging_31_60 = unpaid_amount
                elif overdue_days <= 90:
                    aging_61_90 = unpaid_amount
                else:
                    aging_over_90 = unpaid_amount
            row_data.update({"aging_0_30": aging_0_30, "aging_31_60": aging_31_60, "aging_61_90": aging_61_90, "aging_over_90": aging_over_90})

        data.append(row_data)

    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="应收账款列表", title="应收账款列表（含账龄分析）" if include_aging else "应收账款列表")
    filename = f"应收账款列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)

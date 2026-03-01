# -*- coding: utf-8 -*-
"""
回款导出功能 endpoints
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import desc, or_
from sqlalchemy.orm import Session

from app.api import deps
from app.core import security
from app.models.project import ProjectPaymentPlan
from app.models.sales import Contract, Invoice
from app.models.user import User

router = APIRouter()


@router.get("/payments/invoices/export")
def export_payment_invoices(
    *,
    db: Session = Depends(deps.get_db),
    contract_id: Optional[int] = Query(None, description="合同ID筛选"),
    project_id: Optional[int] = Query(None, description="项目ID筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    payment_status: Optional[str] = Query(None, description="收款状态筛选"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    导出回款记录（基于发票，Excel格式）
    """
    query = db.query(Invoice).filter(Invoice.status == "ISSUED")

    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)

    if project_id:
        query = query.filter(Invoice.project_id == project_id)

    if customer_id:
        query = query.join(Contract).filter(Contract.customer_id == customer_id)

    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)

    if start_date:
        query = query.filter(Invoice.paid_date >= start_date)

    if end_date:
        query = query.filter(Invoice.paid_date <= end_date)

    invoices = query.order_by(desc(Invoice.paid_date)).all()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "回款记录"

    # 设置表头
    headers = [
        "发票号", "合同编号", "项目编号", "客户名称", "发票金额",
        "已收金额", "未收金额", "收款状态", "开票日期", "到期日期",
        "实际收款日期", "逾期天数", "备注"
    ]

    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    today = date.today()
    for row, invoice in enumerate(invoices, 2):
        contract = invoice.contract
        total = invoice.total_amount or invoice.amount or Decimal("0")
        paid = invoice.paid_amount or Decimal("0")
        unpaid = total - paid

        overdue_days = None
        if invoice.due_date and invoice.due_date < today and invoice.payment_status in ["PENDING", "PARTIAL"]:
            overdue_days = (today - invoice.due_date).days

        ws.cell(row=row, column=1, value=invoice.invoice_code or "")
        ws.cell(row=row, column=2, value=contract.contract_code if contract else "")
        ws.cell(row=row, column=3, value=invoice.project.project_code if invoice.project else "")
        ws.cell(row=row, column=4, value=contract.customer.customer_name if contract and contract.customer else "")
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=6, value=float(paid))
        ws.cell(row=row, column=7, value=float(unpaid))
        ws.cell(row=row, column=8, value=invoice.payment_status or "")
        ws.cell(row=row, column=9, value=invoice.issue_date.strftime("%Y-%m-%d") if invoice.issue_date else "")
        ws.cell(row=row, column=10, value=invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "")
        ws.cell(row=row, column=11, value=invoice.paid_date.strftime("%Y-%m-%d") if invoice.paid_date else "")
        ws.cell(row=row, column=12, value=overdue_days if overdue_days else "")
        ws.cell(row=row, column=13, value=invoice.remark or "")

    # 调整列宽
    column_widths = [15, 15, 15, 20, 12, 12, 12, 12, 12, 12, 12, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f"回款记录_{date.today().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)
    ascii_filename = f"payment_invoices_{date.today().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename={ascii_filename}; "
                f"filename*=UTF-8''{encoded_filename}"
            )
        },
    )


@router.get("/payments/export")
def export_payments(
    *,
    db: Session = Depends(deps.get_db),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_id: Optional[int] = Query(None, description="客户ID筛选"),
    include_aging: bool = Query(True, description="是否包含账龄分析"),
    current_user: User = Depends(security.get_current_active_user),
) -> Any:
    """
    Issue 4.4: 导出应收账款列表（Excel）
    包含账龄分析
    """
    from app.services.excel_export_service import (
        ExcelExportService,
        create_excel_response,
    )

    query = db.query(ProjectPaymentPlan).join(Contract).filter(ProjectPaymentPlan.status.in_(["PENDING", "INVOICED", "PARTIAL"]))
    if keyword:
        query = query.filter(or_(ProjectPaymentPlan.payment_name.contains(keyword), ProjectPaymentPlan.contract.has(Contract.contract_code.contains(keyword))))
    if status:
        query = query.filter(ProjectPaymentPlan.status == status)
    if customer_id:
        query = query.filter(ProjectPaymentPlan.contract.has(Contract.customer_id == customer_id))

    payment_plans = query.order_by(ProjectPaymentPlan.planned_date).all()
    export_service = ExcelExportService()
    today = date.today()

    columns = [
        {"key": "payment_name", "label": "收款计划名称", "width": 25},
        {"key": "contract_code", "label": "合同编码", "width": 15},
        {"key": "customer_name", "label": "客户名称", "width": 25},
        {"key": "project_code", "label": "项目编码", "width": 15},
        {"key": "planned_amount", "label": "计划金额", "width": 15, "format": export_service.format_currency},
        {"key": "actual_amount", "label": "已收金额", "width": 15, "format": export_service.format_currency},
        {"key": "unpaid_amount", "label": "未收金额", "width": 15, "format": export_service.format_currency},
        {"key": "planned_date", "label": "计划日期", "width": 12, "format": export_service.format_date},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "overdue_days", "label": "逾期天数", "width": 10},
    ]

    if include_aging:
        columns.extend([
            {"key": "aging_0_30", "label": "0-30天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_31_60", "label": "31-60天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_61_90", "label": "61-90天", "width": 12, "format": export_service.format_currency},
            {"key": "aging_over_90", "label": "90天以上", "width": 12, "format": export_service.format_currency},
        ])

    data = []
    for plan in payment_plans:
        planned_amount = float(plan.planned_amount or 0)
        actual_amount = float(plan.actual_amount or 0)
        unpaid_amount = planned_amount - actual_amount
        overdue_days = 0
        if plan.planned_date and plan.planned_date < today:
            overdue_days = (today - plan.planned_date).days

        row_data = {
            "payment_name": plan.payment_name or '',
            "contract_code": plan.contract.contract_code if plan.contract else '',
            "customer_name": plan.contract.customer.customer_name if plan.contract and plan.contract.customer else '',
            "project_code": plan.project.project_code if plan.project else '',
            "planned_amount": planned_amount,
            "actual_amount": actual_amount,
            "unpaid_amount": unpaid_amount,
            "planned_date": plan.planned_date,
            "status": plan.status or '',
            "overdue_days": overdue_days,
        }

        if include_aging:
            aging_0_30 = aging_31_60 = aging_61_90 = aging_over_90 = 0
            if unpaid_amount > 0 and plan.planned_date:
                if overdue_days <= 30:
                    aging_0_30 = unpaid_amount
                elif overdue_days <= 60:
                    aging_31_60 = unpaid_amount
                elif overdue_days <= 90:
                    aging_61_90 = unpaid_amount
                else:
                    aging_over_90 = unpaid_amount
            row_data.update({"aging_0_30": aging_0_30, "aging_31_60": aging_31_60, "aging_61_90": aging_61_90, "aging_over_90": aging_over_90})

        data.append(row_data)

    excel_data = export_service.export_to_excel(data=data, columns=columns, sheet_name="应收账款列表", title="应收账款列表（含账龄分析）" if include_aging else "应收账款列表")
    filename = f"应收账款列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(excel_data, filename)

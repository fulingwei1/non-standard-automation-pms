# -*- coding: utf-8 -*-
"""
Excel 导出服务
提供通用的 Excel 导出功能，支持自定义列、排序、筛选、样式设置
"""

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable, Dict, List, Optional

try:
    import openpyxl
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelExportService:
    """Excel 导出服务类"""

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel处理库未安装，请安装pandas和openpyxl: pip install pandas openpyxl")

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[Dict[str, Any]]] = None,
        sheet_name: str = "Sheet1",
        filename: Optional[str] = None,
        title: Optional[str] = None,
        apply_styles: bool = True
    ) -> io.BytesIO:
        """
        导出数据到 Excel

        Args:
            data: 数据列表，每个元素是一个字典
            columns: 列配置列表，每个元素包含：
                - key: 数据字段名
                - label: 显示名称
                - width: 列宽（可选）
                - format: 格式化函数（可选）
            sheet_name: Sheet 名称
            filename: 文件名（可选，不指定则自动生成）
            title: 标题（可选）
            apply_styles: 是否应用样式

        Returns:
            io.BytesIO: Excel 文件的内存流
        """
        if not data:
            # 空数据，创建空的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            if title:
                ws['A1'] = title
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # 如果没有指定列配置，使用数据的所有键
        if not columns:
            columns = [
                {"key": key, "label": key}
                for key in data[0].keys()
            ]

        # 构建 DataFrame
        df_data = []
        for row in data:
            df_row = {}
            for col in columns:
                key = col["key"]
                value = row.get(key)

                # 应用格式化函数
                if "format" in col and callable(col["format"]):
                    value = col["format"](value)

                # 处理特殊类型
                if isinstance(value, (date, datetime)):
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                elif isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ''

                df_row[col["label"]] = value
            df_data.append(df_row)

        df = pd.DataFrame(df_data)

        # 创建 Excel 文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表
            worksheet = writer.sheets[sheet_name]

            # 设置列宽
            if apply_styles:
                for idx, col in enumerate(columns, start=1):
                    col_letter = get_column_letter(idx)
                    width = col.get("width", 15)
                    worksheet.column_dimensions[col_letter].width = width

                # 设置表头样式
                self._format_headers(worksheet, len(columns))

                # 如果有标题，添加标题行
                if title:
                    worksheet.insert_rows(1)
                    worksheet.merge_cells(f'A1:{get_column_letter(len(columns))}1')
                    title_cell = worksheet['A1']
                    title_cell.value = title
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")

        output.seek(0)
        return output

    def export_multisheet(
        self,
        sheets: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> io.BytesIO:
        """
        导出多 Sheet Excel 文件

        Args:
            sheets: Sheet 配置列表，每个元素包含：
                - name: Sheet 名称
                - data: 数据列表
                - columns: 列配置（可选）
                - title: 标题（可选）
            filename: 文件名（可选）

        Returns:
            io.BytesIO: Excel 文件的内存流
        """
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认的 Sheet

        for sheet_config in sheets:
            sheet_name = sheet_config["name"]
            data = sheet_config.get("data", [])
            columns = sheet_config.get("columns")
            title = sheet_config.get("title")

            # 创建 Sheet
            ws = wb.create_sheet(title=sheet_name)

            if not data:
                if title:
                    ws['A1'] = title
                continue

            # 如果没有指定列配置，使用数据的所有键
            if not columns:
                columns = [
                    {"key": key, "label": key}
                    for key in data[0].keys()
                ]

            # 写入标题（如果有）
            row = 1
            if title:
                ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
                title_cell = ws[f'A{row}']
                title_cell.value = title
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                row += 1

            # 写入表头
            header_row = row
            for idx, col in enumerate(columns, start=1):
                col_letter = get_column_letter(idx)
                cell = ws[f'{col_letter}{header_row}']
                cell.value = col["label"]
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                width = col.get("width", 15)
                ws.column_dimensions[col_letter].width = width

            # 写入数据
            for data_row in data:
                row += 1
                for idx, col in enumerate(columns, start=1):
                    col_letter = get_column_letter(idx)
                    key = col["key"]
                    value = data_row.get(key)

                    # 应用格式化函数
                    if "format" in col and callable(col["format"]):
                        value = col["format"](value)

                    # 处理特殊类型
                    if isinstance(value, (date, datetime)):
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
                    elif isinstance(value, Decimal):
                        value = float(value)
                    elif value is None:
                        value = ''

                    cell = ws[f'{col_letter}{row}']
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _format_headers(self, worksheet, num_columns: int):
        """
        格式化表头

        Args:
            worksheet: openpyxl 工作表对象
            num_columns: 列数
        """
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 找到标题行（如果有标题，表头在第2行，否则在第1行）
        header_row = 2 if worksheet['A1'].value and len(str(worksheet['A1'].value)) > 0 and worksheet['A1'].value != worksheet[f'{get_column_letter(1)}2'].value else 1

        for col_idx in range(1, num_columns + 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f'{col_letter}{header_row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def format_currency(self, value: Any) -> str:
        """
        格式化货币值

        Args:
            value: 数值

        Returns:
            str: 格式化后的字符串
        """
        if value is None or value == '':
            return '0.00'
        try:
            if isinstance(value, str):
                value = float(value)
            return f"{value:,.2f}"
        except (ValueError, TypeError):
            return str(value)

    def format_percentage(self, value: Any) -> str:
        """
        格式化百分比

        Args:
            value: 数值

        Returns:
            str: 格式化后的字符串
        """
        if value is None or value == '':
            return '0.00%'
        try:
            if isinstance(value, str):
                value = float(value)
            return f"{value:.2f}%"
        except (ValueError, TypeError):
            return str(value)

    def format_date(self, value: Any) -> str:
        """
        格式化日期

        Args:
            value: 日期值

        Returns:
            str: 格式化后的字符串
        """
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        return str(value)


def create_excel_response(
    excel_data: io.BytesIO,
    filename: str,
    media_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
) -> Any:
    """
    创建 Excel 下载响应

    Args:
        excel_data: Excel 文件的内存流
        filename: 文件名
        media_type: MIME 类型

    Returns:
        StreamingResponse: FastAPI 流式响应
    """
    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        excel_data,
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )

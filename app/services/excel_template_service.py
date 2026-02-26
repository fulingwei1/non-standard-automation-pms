# -*- coding: utf-8 -*-
"""
Excel模板生成服务

提供统一的Excel导入模板生成功能。
共享样式已迁移到 report_framework/renderers/excel_styles.py (#39)。
"""

import io
from datetime import datetime
from typing import Dict, List, Optional

from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill


def create_template_excel(
    template_data: Dict[str, List],
    sheet_name: str,
    column_widths: Dict[str, int],
    instructions: str,
    filename_prefix: str
) -> StreamingResponse:
    """
    创建标准化的Excel导入模板

    Args:
        template_data: 模板数据字典，键为列名，值为示例数据列表
        sheet_name: 工作表名称
        column_widths: 列宽配置，如 {'A': 18, 'B': 30}
        instructions: 说明文字
        filename_prefix: 文件名前缀

    Returns:
        StreamingResponse: Excel文件流
    """
    import pandas as pd

    # 创建DataFrame
    df = pd.DataFrame(template_data)

    # 创建Excel文件
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 获取工作表
        worksheet = writer.sheets[sheet_name]

        # 应用标准样式
        apply_template_styles(worksheet, column_widths, instructions)

    output.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


def apply_template_styles(
    worksheet,
    column_widths: Dict[str, int],
    instructions: str
) -> None:
    """
    应用Excel模板的标准样式

    Args:
        worksheet: openpyxl工作表对象
        column_widths: 列宽配置
        instructions: 说明文字
    """
    # 表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # 应用表头样式（第2行，因为第1行是说明）
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # 添加说明行
    if instructions:
        worksheet.insert_rows(1)
        # 计算合并范围（从A1到最后一列）
        last_col = chr(ord('A') + len(column_widths) - 1)
        merge_range = f'A1:{last_col}1'
        worksheet.merge_cells(merge_range)
        worksheet['A1'] = instructions
        worksheet['A1'].font = Font(size=10, italic=True)
        worksheet['A1'].alignment = Alignment(horizontal="left")


# 模板配置常量
TEMPLATE_CONFIGS = {
    "PROJECT": {
        "template_data": {
            '项目编码*': ['示例：PRJ-2025-001'],
            '项目名称*': ['示例：XX公司ICT测试设备'],
            '客户名称': ['示例：XX科技有限公司'],
            '合同编号': ['示例：HT2507-001'],
            '项目类型': ['单机类/线体类/改造类'],
            '合同日期': ['2025-01-15'],
            '合同金额': [1000000.00],
            '预算金额': [950000.00],
            '计划开始日期': ['2025-02-01'],
            '计划结束日期': ['2025-08-31'],
            '项目经理': ['请输入项目经理员工号或姓名'],
            '项目描述': ['项目简要描述'],
        },
        "sheet_name": "项目导入模板",
        "column_widths": {
            'A': 18, 'B': 30, 'C': 20, 'D': 15, 'E': 12,
            'F': 12, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
            'K': 15, 'L': 40
        },
        "instructions": "说明：1. 带*的列为必填项；2. 日期格式：YYYY-MM-DD；3. 金额为数字格式",
        "filename_prefix": "项目导入模板"
    },
    "TASK": {
        "template_data": {
            '任务名称*': ['示例：机械设计'],
            '项目编码*': ['示例：PRJ-2025-001'],
            '阶段': ['S1/S2/S3/S4/S5/S6/S7/S8/S9'],
            '负责人*': ['请输入项目负责人员工号或姓名'],
            '计划开始日期': ['2025-02-01'],
            '计划结束日期': ['2025-02-15'],
            '权重(%)': [10],
            '任务描述': ['任务详细描述'],
        },
        "sheet_name": "任务导入模板",
        "column_widths": {'A': 30, 'B': 18, 'C': 12, 'D': 12, 'E': 15, 'F': 15, 'G': 10, 'H': 40},
        "instructions": "说明：1. 带*的列为必填项；2. 日期格式：YYYY-MM-DD；3. 阶段：S1-需求进入到S9-质保结项",
        "filename_prefix": "任务导入模板"
    },
    "USER": {
        "template_data": {
            '用户名*': ['请输入员工用户名'],
            '真实姓名*': ['请输入员工真实姓名'],
            '邮箱': ['zhangsan@example.com'],
            '手机号': ['13800138000'],
            '部门': ['技术部'],
            '岗位': ['机械工程师'],
            '工号': ['EMP001'],
        },
        "sheet_name": "人员导入模板",
        "column_widths": {'A': 15, 'B': 15, 'C': 25, 'D': 15, 'E': 15, 'F': 15, 'G': 12},
        "instructions": "说明：1. 带*的列为必填项；2. 用户名必须唯一；3. 初始密码默认为123456",
        "filename_prefix": "人员导入模板"
    },
    "TIMESHEET": {
        "template_data": {
            '工作日期*': ['2025-01-15'],
            '人员姓名*': ['请输入员工姓名'],
            '项目编码*': ['PRJ-2025-001'],
            '任务名称': ['机械设计'],
            '工时(小时)*': [8],
            '工作内容': ['完成设备结构设计'],
            '工作成果': ['产出图纸10张'],
            '加班类型': ['NORMAL/OVERTIME/WEEKEND/HOLIDAY'],
        },
        "sheet_name": "工时导入模板",
        "column_widths": {'A': 12, 'B': 12, 'C': 18, 'D': 30, 'E': 12, 'F': 40, 'G': 40, 'H': 25},
        "instructions": "说明：1. 带*的列为必填项；2. 日期格式：YYYY-MM-DD；3. 加班类型：NORMAL-正常/OVERTIME-平时加班/WEEKEND-周末/HOLIDAY-节假日",
        "filename_prefix": "工时导入模板"
    },
    "MATERIAL": {
        "template_data": {
            '物料编码*': ['示例：MAT-001'],
            '物料名称*': ['示例：气缸'],
            '规格型号': ['例：CDJ2B10-45'],
            '单位': ['个'],
            '物料类型': ['标准件/机械件/电气件'],
            '默认供应商': ['某某公司'],
            '参考价格': [100.00],
            '安全库存': [10],
        },
        "sheet_name": "物料导入模板",
        "column_widths": {'A': 15, 'B': 25, 'C': 20, 'D': 8, 'E': 15, 'F': 15, 'G': 10, 'H': 10},
        "instructions": "说明：1. 带*的列为必填项；2. 物料编码必须唯一；3. 价格单位：元",
        "filename_prefix": "物料导入模板"
    },
    "BOM": {
        "template_data": {
            'BOM编码*': ['示例：BOM-PRJ-001'],
            '项目编码*': ['PRJ-2025-001'],
            '机台编号': ['PN001'],
            '物料编码*': ['MAT-001'],
            '用量*': [2],
            '单位': ['个'],
            '备注': ['说明'],
        },
        "sheet_name": "BOM导入模板",
        "column_widths": {'A': 18, 'B': 18, 'C': 10, 'D': 12, 'E': 8, 'F': 8, 'G': 30},
        "instructions": "说明：1. 带*的列为必填项；2. BOM编码格式：BOM-{项目编码}；3. 物料编码必须已存在",
        "filename_prefix": "BOM导入模板"
    }
}


def get_template_config(template_type: str) -> Optional[Dict]:
    """
    获取模板配置

    Args:
        template_type: 模板类型（PROJECT/TASK/USER等）

    Returns:
        模板配置字典，如果类型不存在则返回None
    """
    return TEMPLATE_CONFIGS.get(template_type.upper())

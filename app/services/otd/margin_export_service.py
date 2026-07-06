# -*- coding: utf-8 -*-
"""
毛利率 Dashboard 导出服务
"""

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from app.services.import_export_engine import ExcelExportEngine


class MarginExportService:
    """毛利率导出服务"""

    @classmethod
    def export_dashboard_to_excel(
        cls,
        dashboard_data: Dict[str, Any],
    ) -> io.BytesIO:
        """
        导出毛利率 Dashboard 到 Excel

        Args:
            dashboard_data: Dashboard 数据

        Returns:
            Excel 文件流
        """
        # Sheet 1: 汇总指标
        summary_labels = ["指标", "数值"]
        summary_widths = [30, 20]
        summary_columns = ExcelExportEngine.build_columns(summary_labels, widths=summary_widths)

        summary = dashboard_data.get("summary", {})
        summary_data = [
            {"指标": "在管项目数", "数值": summary.get("total_projects", 0)},
            {"指标": "平均毛利率(%)", "数值": summary.get("avg_margin_rate", 0)},
            {"指标": "目标毛利率(%)", "数值": summary.get("target_margin", 0)},
            {"指标": "健康项目数", "数值": summary.get("healthy_count", 0)},
            {"指标": "预警项目数", "数值": summary.get("warning_count", 0)},
            {"指标": "危险项目数", "数值": summary.get("critical_count", 0)},
            {"指标": "低于目标项目数", "数值": summary.get("below_target_count", 0)},
            {"指标": "严重低于目标项目数", "数值": summary.get("seriously_below_count", 0)},
        ]

        # Sheet 2: 项目详情
        projects = dashboard_data.get("projects", [])
        if projects:
            project_labels = [
                "项目编码",
                "项目名称",
                "合同金额(万)",
                "实际成本(万)",
                "毛利率(%)",
                "目标毛利率(%)",
                "偏差(%)",
                "健康度",
                "风险点",
            ]
            project_widths = [15, 30, 15, 15, 12, 12, 12, 10, 40]
            project_columns = ExcelExportEngine.build_columns(project_labels, widths=project_widths)

            project_data = []
            for p in projects:
                # 计算偏差
                current_rate = p.get("current_margin_rate", 0)
                target_rate = p.get("target_margin_rate", 25)
                deviation = round(current_rate - target_rate, 2)

                project_data.append({
                    "项目编码": p.get("project_code", ""),
                    "项目名称": p.get("project_name", ""),
                    "合同金额(万)": round(p.get("contract_amount", 0) / 10000, 2),
                    "实际成本(万)": round(p.get("actual_cost", 0) / 10000, 2),
                    "毛利率(%)": current_rate,
                    "目标毛利率(%)": target_rate,
                    "偏差(%)": deviation,
                    "健康度": p.get("health", ""),
                    "风险点": p.get("risk_point", ""),
                })

            # 按健康度排序
            health_order = {"critical": 0, "warning": 1, "healthy": 2}
            project_data.sort(key=lambda x: health_order.get(x["健康度"], 99))

            sheets = [
                {"name": "汇总指标", "data": summary_data, "columns": summary_columns},
                {"name": "项目详情", "data": project_data, "columns": project_columns},
            ]
        else:
            sheets = [
                {"name": "汇总指标", "data": summary_data, "columns": summary_columns},
            ]

        def post_process(worksheet, sheet_config):
            from openpyxl.styles import PatternFill

            # 给健康度列上色
            fills = {
                "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                "warning": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
                "healthy": PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid"),
            }

            # 给偏差列上色（负数红色）
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if worksheet.title == "项目详情":
                    # 健康度列（第8列）
                    health = row[7].value
                    if health in fills:
                        row[7].fill = fills[health]

                    # 偏差列（第7列）
                    deviation = row[6].value
                    if deviation is not None and deviation < 0:
                        row[6].font = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        return ExcelExportEngine.export_multi_sheet(sheets, sheet_post_process=post_process)

    @classmethod
    def export_metrics_to_excel(
        cls,
        metrics_data: Dict[str, Any],
    ) -> io.BytesIO:
        """
        导出 OTD 7 核心指标到 Excel

        Args:
            metrics_data: 指标数据

        Returns:
            Excel 文件流
        """
        # Sheet 1: 指标概览
        overview_labels = ["指标", "数值", "说明"]
        overview_widths = [30, 20, 50]
        overview_columns = ExcelExportEngine.build_columns(overview_labels, widths=overview_widths)

        metrics = metrics_data.get("metrics", {})
        overview_data = []

        # 准时交付率
        otd = metrics.get("on_time_delivery_rate", {})
        overview_data.append({
            "指标": "准时交付率(%)",
            "数值": otd.get("rate_pct", 0),
            "说明": f"按时交付 {otd.get('on_time', 0)} / 已完成 {otd.get('total_completed', 0)}",
        })

        # 延期天数
        delay = metrics.get("delay_days", {})
        overview_data.append({
            "指标": "平均延期天数",
            "数值": delay.get("avg_delay_days", 0),
            "说明": f"在途超期 {delay.get('in_progress_overdue_count', 0)} 个，已完成超期 {delay.get('completed_overdue_count', 0)} 个",
        })

        # 返工次数
        rework = metrics.get("rework_count", {})
        overview_data.append({
            "指标": "返工次数",
            "数值": rework.get("total_retry_count", 0),
            "说明": f"有返工的项目 {rework.get('projects_with_rework', 0)} 个",
        })

        # 变更次数
        change = metrics.get("change_count", {})
        overview_data.append({
            "指标": "变更次数",
            "数值": change.get("total_changes", 0),
            "说明": f"客户变更 {change.get('customer_changes', 0)} 个，内部变更 {change.get('internal_changes', 0)} 个",
        })

        # 毛利偏差
        margin = metrics.get("margin_deviation", {})
        overview_data.append({
            "指标": "平均毛利偏差(%)",
            "数值": margin.get("avg_margin_gap_pct", 0),
            "说明": f"低于目标 {margin.get('below_target_count', 0)} 个，严重低于目标 {margin.get('seriously_below_count', 0)} 个",
        })

        # 验收周期
        acceptance = metrics.get("acceptance_cycle_days", {})
        overview_data.append({
            "指标": "平均验收周期(天)",
            "数值": acceptance.get("avg_cycle_days", 0),
            "说明": f"已完成验收 {acceptance.get('completed_count', 0)} 个",
        })

        # 投诉率
        complaint = metrics.get("complaint_rate", {})
        overview_data.append({
            "指标": "客户投诉率(%)",
            "数值": complaint.get("complaint_rate_pct", 0),
            "说明": f"投诉 {complaint.get('complaint_count', 0)} / 反馈 {complaint.get('total_feedback', 0)}",
        })

        # Sheet 2: 下钻数据（top_offenders）
        offenders_data = []
        offenders_labels = ["指标", "项目编码", "项目名称", "数值", "详情"]
        offenders_widths = [20, 15, 30, 15, 50]
        offenders_columns = ExcelExportEngine.build_columns(offenders_labels, widths=offenders_widths)

        # 收集所有 top_offenders
        for metric_name, metric_data in metrics.items():
            top_offenders = metric_data.get("top_offenders", [])
            for offender in top_offenders:
                offenders_data.append({
                    "指标": metric_name,
                    "项目编码": offender.get("project_code", ""),
                    "项目名称": offender.get("project_name", ""),
                    "数值": offender.get("value", ""),
                    "详情": offender.get("detail", ""),
                })

        if offenders_data:
            sheets = [
                {"name": "指标概览", "data": overview_data, "columns": overview_columns},
                {"name": "下钻数据", "data": offenders_data, "columns": offenders_columns},
            ]
        else:
            sheets = [
                {"name": "指标概览", "data": overview_data, "columns": overview_columns},
            ]

        return ExcelExportEngine.export_multi_sheet(sheets)

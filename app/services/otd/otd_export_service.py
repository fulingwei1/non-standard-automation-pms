# -*- coding: utf-8 -*-
"""
OTD 项目交付智能体导出服务
"""

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from app.services.import_export_engine import ExcelExportEngine


class OTDExportService:
    """OTD 导出服务"""

    @classmethod
    def export_scan_to_excel(
        cls,
        scan_data: Dict[str, Any],
        detail_level: str = "summary",
    ) -> io.BytesIO:
        """
        导出 OTD 扫描结果到 Excel

        Args:
            scan_data: 扫描结果数据
            detail_level: summary 或 full

        Returns:
            Excel 文件流
        """
        projects = scan_data.get("projects", [])

        if not projects:
            # 空数据
            columns = ExcelExportEngine.build_columns(
                ["项目编码", "项目名称", "阶段", "风险等级", "主因"],
                widths=[15, 30, 10, 12, 40],
            )
            return ExcelExportEngine.export_table(
                data=[],
                columns=columns,
                sheet_name="OTD风险扫描",
                title="OTD 项目交付风险扫描结果",
            )

        if detail_level == "summary":
            # 精简模式
            labels = [
                "项目编码",
                "项目名称",
                "阶段",
                "风险等级",
                "主因",
                "预警ID",
            ]
            widths = [15, 30, 10, 12, 50, 15]
            columns = ExcelExportEngine.build_columns(labels, widths=widths)

            data = []
            for p in projects:
                data.append({
                    "项目编码": p.get("project_code", ""),
                    "项目名称": p.get("project_name", ""),
                    "阶段": p.get("stage", ""),
                    "风险等级": p.get("severity", ""),
                    "主因": p.get("top_cause", ""),
                    "预警ID": p.get("alert_id", "") or "",
                })
        else:
            # 完整模式
            labels = [
                "项目编码",
                "项目名称",
                "阶段",
                "进度(%)",
                "计划结束",
                "风险等级",
                "主因",
                "AI建议",
                "命中维度",
                "预警ID",
            ]
            widths = [15, 30, 10, 10, 15, 12, 50, 50, 15, 15]
            columns = ExcelExportEngine.build_columns(labels, widths=widths)

            data = []
            for p in projects:
                risk_items = p.get("risk_items", [])
                dim_count = len([r for r in risk_items if r.get("severity") in ("HIGH", "CRITICAL")])

                data.append({
                    "项目编码": p.get("project_code", ""),
                    "项目名称": p.get("project_name", ""),
                    "阶段": p.get("stage", ""),
                    "进度(%)": p.get("progress", 0),
                    "计划结束": p.get("planned_end", ""),
                    "风险等级": p.get("severity", ""),
                    "主因": p.get("top_cause", ""),
                    "AI建议": p.get("suggestion", ""),
                    "命中维度": dim_count,
                    "预警ID": p.get("alert_id", "") or "",
                })

        # 按风险等级排序（CRITICAL > HIGH > MEDIUM > LOW）
        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        data.sort(key=lambda x: severity_order.get(x["风险等级"], 99))

        def post_process(worksheet, sheet_config):
            # 给风险等级列上色
            from openpyxl.styles import PatternFill

            fills = {
                "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                "HIGH": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
                "MEDIUM": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
                "LOW": PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid"),
            }

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                level = row[3].value  # 风险等级列
                if level in fills:
                    row[3].fill = fills[level]

        return ExcelExportEngine.export_multi_sheet(
            [
                {
                    "name": "OTD风险扫描",
                    "data": data,
                    "columns": columns,
                }
            ],
            sheet_post_process=post_process,
        )

    @classmethod
    def export_project_scan_to_excel(
        cls,
        profile: Dict[str, Any],
    ) -> io.BytesIO:
        """
        导出单项目 OTD 扫描结果到 Excel

        Args:
            profile: 单项目扫描结果

        Returns:
            Excel 文件流
        """
        # Sheet 1: 项目概览
        overview_labels = ["字段", "值"]
        overview_widths = [20, 60]
        overview_columns = ExcelExportEngine.build_columns(overview_labels, widths=overview_widths)

        overview_data = [
            {"字段": "项目编码", "值": profile.get("project_code", "")},
            {"字段": "项目名称", "值": profile.get("project_name", "")},
            {"字段": "阶段", "值": profile.get("stage", "")},
            {"字段": "进度(%)", "值": profile.get("progress", 0)},
            {"字段": "计划结束", "值": profile.get("planned_end", "")},
            {"字段": "风险等级", "值": profile.get("severity", "")},
            {"字段": "主因", "值": profile.get("top_cause", "")},
            {"字段": "AI建议", "值": profile.get("suggestion", "")},
        ]

        # Sheet 2: 风险维度详情
        risk_items = profile.get("risk_items", [])
        if risk_items:
            risk_labels = ["维度", "标签", "风险等级", "说明"]
            risk_widths = [20, 20, 12, 60]
            risk_columns = ExcelExportEngine.build_columns(risk_labels, widths=risk_widths)

            risk_data = []
            for item in risk_items:
                risk_data.append({
                    "维度": item.get("dim", ""),
                    "标签": item.get("label", ""),
                    "风险等级": item.get("severity", ""),
                    "说明": item.get("msg", ""),
                })

            sheets = [
                {"name": "项目概览", "data": overview_data, "columns": overview_columns},
                {"name": "风险维度", "data": risk_data, "columns": risk_columns},
            ]
        else:
            sheets = [
                {"name": "项目概览", "data": overview_data, "columns": overview_columns},
            ]

        return ExcelExportEngine.export_multi_sheet(sheets)

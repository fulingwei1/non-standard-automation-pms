# -*- coding: utf-8 -*-
"""
项目导出服务
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import desc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def get_excel_styles() -> Dict[str, Any]:
    """
    获取Excel样式
    
    Returns:
        Dict[str, Any]: 样式字典
    """
    if not OPENPYXL_AVAILABLE:
        return {}
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'title_font': title_font,
        'border': border
    }


def build_project_info_data(project) -> List[tuple]:
    """
    构建项目信息数据
    
    Returns:
        List[tuple]: (标签, 值) 元组列表
    """
    return [
        ('项目编码', project.project_code or ''),
        ('项目名称', project.project_name or ''),
        ('客户名称', project.customer_name or ''),
        ('合同编号', project.contract_no or ''),
        ('合同金额', f"{float(project.contract_amount or 0):,.2f}"),
        ('项目经理', project.pm_name or ''),
        ('项目类型', project.project_type or ''),
        ('阶段', project.stage or ''),
        ('状态', project.status or ''),
        ('健康度', project.health or ''),
        ('进度(%)', f"{float(project.progress_pct or 0):.2f}"),
        ('计划开始日期', project.planned_start_date.strftime('%Y-%m-%d') if project.planned_start_date else ''),
        ('计划结束日期', project.planned_end_date.strftime('%Y-%m-%d') if project.planned_end_date else ''),
        ('实际开始日期', project.actual_start_date.strftime('%Y-%m-%d') if project.actual_start_date else ''),
        ('实际结束日期', project.actual_end_date.strftime('%Y-%m-%d') if project.actual_end_date else ''),
    ]


def add_project_info_sheet(ws, project, styles: Dict[str, Any]) -> None:
    """
    添加项目基本信息工作表
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    title_font = styles['title_font']
    border = styles['border']
    
    # 标题
    ws.merge_cells('A1:B1')
    ws['A1'] = '项目基本信息'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # 项目信息
    row = 3
    project_info = build_project_info_data(project)
    
    for label, value in project_info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30


def add_tasks_sheet(wb, db: Session, project_id: int, styles: Dict[str, Any]) -> None:
    """
    添加任务列表工作表
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    from app.models.progress import Task
    
    tasks = db.query(Task).filter(
        Task.project_id == project_id
    ).order_by(desc(Task.created_at)).all()
    
    ws = wb.create_sheet("任务列表")
    
    header_fill = styles['header_fill']
    header_font = styles['header_font']
    border = styles['border']
    
    # 表头
    headers = ['任务编号', '任务名称', '任务类型', '优先级', '状态', '进度(%)', 
               '计划开始日期', '计划结束日期', '实际开始日期', '实际结束日期', 
               '负责人', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # 数据
    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=task.task_code or '')
        ws.cell(row=row_idx, column=2, value=task.task_name or '')
        ws.cell(row=row_idx, column=3, value=task.task_type or '')
        ws.cell(row=row_idx, column=4, value=task.priority or '')
        ws.cell(row=row_idx, column=5, value=task.status or '')
        ws.cell(row=row_idx, column=6, value=f"{float(task.progress_pct or 0):.2f}")
        ws.cell(row=row_idx, column=7, value=task.planned_start_date.strftime('%Y-%m-%d') if task.planned_start_date else '')
        ws.cell(row=row_idx, column=8, value=task.planned_end_date.strftime('%Y-%m-%d') if task.planned_end_date else '')
        ws.cell(row=row_idx, column=9, value=task.actual_start_date.strftime('%Y-%m-%d') if task.actual_start_date else '')
        ws.cell(row=row_idx, column=10, value=task.actual_end_date.strftime('%Y-%m-%d') if task.actual_end_date else '')
        ws.cell(row=row_idx, column=11, value=task.assignee_name or '')
        ws.cell(row=row_idx, column=12, value=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else '')
        
        # 设置边框
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = border
    
    # 设置列宽
    column_widths = [15, 30, 12, 10, 10, 10, 12, 12, 12, 12, 12, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width


def add_costs_sheet(wb, db: Session, project_id: int, styles: Dict[str, Any]) -> None:
    """
    添加成本列表工作表
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    from app.models.project import ProjectCost
    
    costs = db.query(ProjectCost).filter(
        ProjectCost.project_id == project_id
    ).order_by(desc(ProjectCost.cost_date)).all()
    
    ws = wb.create_sheet("成本列表")
    
    header_fill = styles['header_fill']
    header_font = styles['header_font']
    border = styles['border']
    
    # 表头
    headers = ['成本日期', '成本类型', '成本分类', '金额', '币种', '说明', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # 数据
    total_amount = 0
    for row_idx, cost in enumerate(costs, 2):
        amount = float(cost.amount or 0)
        total_amount += amount
        
        ws.cell(row=row_idx, column=1, value=cost.cost_date.strftime('%Y-%m-%d') if cost.cost_date else '')
        ws.cell(row=row_idx, column=2, value=cost.cost_type or '')
        ws.cell(row=row_idx, column=3, value=cost.cost_category or '')
        ws.cell(row=row_idx, column=4, value=amount)
        ws.cell(row=row_idx, column=5, value=cost.currency or 'CNY')
        ws.cell(row=row_idx, column=6, value=cost.description or '')
        ws.cell(row=row_idx, column=7, value=cost.created_at.strftime('%Y-%m-%d %H:%M:%S') if cost.created_at else '')
        
        # 设置边框
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border
    
    # 添加合计行
    if costs:
        total_row = len(costs) + 2
        ws.cell(row=total_row, column=3, value='合计').font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_amount).font = Font(bold=True)
        for col in range(1, 8):
            ws.cell(row=total_row, column=col).border = border
    
    # 设置列宽
    column_widths = [12, 12, 15, 15, 8, 40, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width


def create_project_detail_excel(
    db: Session,
    project,
    include_tasks: bool,
    include_costs: bool
) -> BytesIO:
    """
    创建项目详情Excel文件
    
    Returns:
        BytesIO: Excel文件的内存流
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Excel处理库未安装，请安装openpyxl")
    
    import io
    
    # 创建Excel工作簿
    output = io.BytesIO()
    wb = Workbook()
    
    # 获取样式
    styles = get_excel_styles()
    
    # ========== 项目基本信息 ==========
    ws1 = wb.active
    ws1.title = "项目基本信息"
    add_project_info_sheet(ws1, project, styles)
    
    # ========== 任务列表 ==========
    if include_tasks:
        add_tasks_sheet(wb, db, project.id, styles)
    
    # ========== 成本列表 ==========
    if include_costs:
        add_costs_sheet(wb, db, project.id, styles)
    
    wb.save(output)
    output.seek(0)
    
    return output

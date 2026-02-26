# -*- coding: utf-8 -*-
"""
报表 Excel 导出服务
功能：生成专业的 Excel 报表（总览 + 明细 + 图表）

NOTE: 通用 Excel 渲染请使用 ExcelRenderer (#39)。
"""

import logging
from pathlib import Path
from typing import Dict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl 未安装，Excel 导出功能不可用")

logger = logging.getLogger(__name__)


class ReportExcelService:
    """Excel 报表导出服务"""
    
    # 样式定义
    HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    
    TITLE_FONT = Font(name='微软雅黑', size=14, bold=True)
    TITLE_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    
    CELL_FONT = Font(name='微软雅黑', size=10)
    CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center')
    NUMBER_ALIGNMENT = Alignment(horizontal='right', vertical='center')
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def export_to_excel(data: Dict, template_name: str, output_dir: str = "reports") -> str:
        """
        导出为 Excel 文件
        
        Args:
            data: 报表数据
            template_name: 模板名称
            output_dir: 输出目录
            
        Returns:
            文件路径
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，无法导出 Excel")
        
        # 创建输出目录
        year = data['year']
        month = data['month']
        period = data['period']
        
        output_path = Path(output_dir) / str(year) / f"{month:02d}"
        output_path.mkdir(parents=True, exist_ok=True)
        
        # 文件名
        filename = f"{template_name}_{period}.xlsx"
        file_path = output_path / filename
        
        # 创建 Workbook
        wb = Workbook()
        
        # 删除默认 sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 创建总览 Sheet
        ws_summary = wb.create_sheet("总览")
        ReportExcelService._write_summary_sheet(ws_summary, data)
        
        # 创建明细 Sheet（如果有明细数据）
        if data.get('detail'):
            ws_detail = wb.create_sheet("明细")
            ReportExcelService._write_detail_sheet(ws_detail, data)
        
        # 创建图表 Sheet
        ws_chart = wb.create_sheet("图表")
        ReportExcelService._write_chart_sheet(ws_chart, data)
        
        # 保存文件
        wb.save(str(file_path))
        
        logger.info(f"✅ Excel 文件已生成: {file_path}")
        
        return str(file_path)
    
    @staticmethod
    def _write_summary_sheet(ws, data: Dict):
        """写入总览 Sheet"""
        
        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{data['period']} 工时报表总览"
        ws['A1'].font = ReportExcelService.TITLE_FONT
        ws['A1'].alignment = ReportExcelService.TITLE_ALIGNMENT
        
        # 生成时间
        ws.merge_cells('A2:H2')
        ws['A2'] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
        ws['A2'].font = Font(name='微软雅黑', size=9, italic=True)
        
        # 空行
        ws.append([])
        
        # 表头
        summary = data['summary']
        if not summary:
            ws.append(['无数据'])
            return
        
        # 根据数据确定列
        first_row = summary[0]
        headers = list(first_row.keys())
        
        # 写入表头
        header_row = []
        for header in headers:
            header_row.append(ReportExcelService._translate_header(header))
        
        ws.append(header_row)
        
        # 设置表头样式
        for col_num, cell in enumerate(ws[ws.max_row], start=1):
            cell.font = ReportExcelService.HEADER_FONT
            cell.fill = ReportExcelService.HEADER_FILL
            cell.alignment = ReportExcelService.HEADER_ALIGNMENT
            cell.border = ReportExcelService.THIN_BORDER
        
        # 写入数据
        for row_data in summary:
            row = []
            for header in headers:
                value = row_data.get(header, '')
                row.append(value)
            ws.append(row)
            
            # 设置数据行样式
            for col_num, cell in enumerate(ws[ws.max_row], start=1):
                cell.font = ReportExcelService.CELL_FONT
                cell.border = ReportExcelService.THIN_BORDER
                
                # 数字列右对齐
                if isinstance(cell.value, (int, float)):
                    cell.alignment = ReportExcelService.NUMBER_ALIGNMENT
                else:
                    cell.alignment = ReportExcelService.CELL_ALIGNMENT
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    @staticmethod
    def _write_detail_sheet(ws, data: Dict):
        """写入明细 Sheet"""
        
        # 标题
        ws.merge_cells('A1:I1')
        ws['A1'] = f"{data['period']} 工时明细"
        ws['A1'].font = ReportExcelService.TITLE_FONT
        ws['A1'].alignment = ReportExcelService.TITLE_ALIGNMENT
        
        # 空行
        ws.append([])
        
        # 表头
        detail = data['detail']
        if not detail:
            ws.append(['无数据'])
            return
        
        first_row = detail[0]
        headers = list(first_row.keys())
        
        # 写入表头
        header_row = []
        for header in headers:
            header_row.append(ReportExcelService._translate_header(header))
        
        ws.append(header_row)
        
        # 设置表头样式
        for cell in ws[ws.max_row]:
            cell.font = ReportExcelService.HEADER_FONT
            cell.fill = ReportExcelService.HEADER_FILL
            cell.alignment = ReportExcelService.HEADER_ALIGNMENT
            cell.border = ReportExcelService.THIN_BORDER
        
        # 写入数据
        for row_data in detail:
            row = []
            for header in headers:
                value = row_data.get(header, '')
                row.append(value)
            ws.append(row)
            
            # 设置数据行样式
            for cell in ws[ws.max_row]:
                cell.font = ReportExcelService.CELL_FONT
                cell.border = ReportExcelService.THIN_BORDER
                if isinstance(cell.value, (int, float)):
                    cell.alignment = ReportExcelService.NUMBER_ALIGNMENT
                else:
                    cell.alignment = ReportExcelService.CELL_ALIGNMENT
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    @staticmethod
    def _write_chart_sheet(ws, data: Dict):
        """写入图表 Sheet"""
        
        # 标题
        ws.merge_cells('A1:J1')
        ws['A1'] = f"{data['period']} 工时统计图表"
        ws['A1'].font = ReportExcelService.TITLE_FONT
        ws['A1'].alignment = ReportExcelService.TITLE_ALIGNMENT
        
        ws.append([])
        
        summary = data['summary']
        if not summary or len(summary) == 0:
            ws.append(['无数据'])
            return
        
        # 检查是否有足够的数据用于图表
        if len(summary) < 1:
            return
        
        # 准备图表数据
        # 柱状图：工时对比
        if 'total_hours' in summary[0]:
            ReportExcelService._add_bar_chart(ws, summary, data)
        
        # 饼图：部门分布（如果是部门报表）
        if 'department_name' in summary[0] and len(summary) > 1:
            ReportExcelService._add_pie_chart(ws, summary, data)
    
    @staticmethod
    def _add_bar_chart(ws, summary: list, data: Dict):
        """添加柱状图"""
        
        # 写入图表数据
        start_row = ws.max_row + 2
        
        # 写入数据表头
        if 'user_name' in summary[0]:
            ws.cell(start_row, 1, "姓名")
            name_key = 'user_name'
        elif 'department_name' in summary[0]:
            ws.cell(start_row, 1, "部门")
            name_key = 'department_name'
        elif 'project_name' in summary[0]:
            ws.cell(start_row, 1, "项目")
            name_key = 'project_name'
        else:
            name_key = list(summary[0].keys())[0]
            ws.cell(start_row, 1, "名称")
        
        ws.cell(start_row, 2, "总工时")
        
        # 写入数据
        for i, row in enumerate(summary, start=1):
            ws.cell(start_row + i, 1, row.get(name_key, ''))
            ws.cell(start_row + i, 2, row.get('total_hours', 0))
        
        # 创建柱状图
        chart = BarChart()
        chart.title = f"{data['period']} 工时统计"
        chart.x_axis.title = "人员/部门/项目"
        chart.y_axis.title = "工时（小时）"
        
        # 数据引用
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(summary))
        cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(summary))
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # 添加图表到工作表
        ws.add_chart(chart, f"D{start_row}")
    
    @staticmethod
    def _add_pie_chart(ws, summary: list, data: Dict):
        """添加饼图"""
        
        # 写入图表数据
        start_row = ws.max_row + 15  # 在柱状图下方
        
        ws.cell(start_row, 1, "部门")
        ws.cell(start_row, 2, "工时占比")
        
        # 写入数据
        for i, row in enumerate(summary, start=1):
            ws.cell(start_row + i, 1, row.get('department_name', ''))
            ws.cell(start_row + i, 2, row.get('total_hours', 0))
        
        # 创建饼图
        chart = PieChart()
        chart.title = f"{data['period']} 部门工时分布"
        
        # 数据引用
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(summary))
        cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(summary))
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # 添加图表到工作表
        ws.add_chart(chart, f"D{start_row}")
    
    @staticmethod
    def _translate_header(header: str) -> str:
        """翻译字段名为中文表头"""
        translations = {
            'user_id': '用户ID',
            'user_name': '姓名',
            'department': '部门',
            'department_id': '部门ID',
            'department_name': '部门名称',
            'total_hours': '总工时',
            'normal_hours': '正常工时',
            'overtime_hours': '加班工时',
            'weekend_hours': '周末加班',
            'holiday_hours': '节假日加班',
            'work_days': '工作天数',
            'avg_hours_per_day': '日均工时',
            'avg_hours_per_user': '人均工时',
            'user_count': '人数',
            'project_id': '项目ID',
            'project_name': '项目名称',
            'task_name': '任务名称',
            'work_date': '日期',
            'hours': '工时',
            'overtime_type': '类型',
            'work_content': '工作内容',
            'total_users': '总人数',
            'total_overtime': '总加班时长',
        }
        
        return translations.get(header, header)

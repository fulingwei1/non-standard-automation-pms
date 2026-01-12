# -*- coding: utf-8 -*-
"""
报表导出服务
支持 Excel、PDF、CSV 格式导出
"""

import os
import json
from datetime import datetime, date
from typing import Any, Dict, List, Optional
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import csv


class ReportExportService:
    """报表导出服务"""

    def __init__(self, export_dir: str = "exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)

        # 尝试注册中文字体
        self._register_chinese_font()

    def _register_chinese_font(self):
        """注册中文字体"""
        font_paths = [
            "/System/Library/Fonts/PingFang.ttc",  # macOS
            "/System/Library/Fonts/STHeiti Medium.ttc",  # macOS
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",  # Linux
            "C:/Windows/Fonts/simhei.ttf",  # Windows
        ]

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    self.chinese_font = 'ChineseFont'
                    return
                except Exception:
                    # Font registration may fail due to IOError, TTFError, etc.
                    continue

        self.chinese_font = 'Helvetica'

    def export_to_excel(
        self,
        data: Dict[str, Any],
        filename: str,
        title: str = "报表",
        sheets: Optional[List[Dict]] = None
    ) -> str:
        """
        导出为 Excel 文件

        Args:
            data: 报表数据
            filename: 文件名（不含扩展名）
            title: 报表标题
            sheets: 多工作表配置 [{ name: 'Sheet1', headers: [...], rows: [...] }, ...]

        Returns:
            文件路径
        """
        wb = Workbook()

        # 样式定义
        title_font = Font(name='微软雅黑', size=14, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        alt_fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        if sheets:
            # 多工作表模式
            for i, sheet_config in enumerate(sheets):
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_config.get('name', 'Sheet1')
                else:
                    ws = wb.create_sheet(title=sheet_config.get('name', f'Sheet{i+1}'))

                self._write_sheet(ws, sheet_config, title_font, header_font, cell_font,
                                 header_fill, alt_fill, border, center_align)
        else:
            # 单工作表模式
            ws = wb.active
            ws.title = "报表数据"

            # 写入标题
            ws.merge_cells('A1:H1')
            ws['A1'] = title
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # 写入生成时间
            ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=9, italic=True)

            # 写入数据
            if 'summary' in data:
                self._write_summary(ws, data['summary'], 4, header_font, cell_font, border)

            if 'details' in data and isinstance(data['details'], list):
                start_row = 10 if 'summary' in data else 4
                self._write_table(ws, data['details'], start_row, header_font, cell_font,
                                 header_fill, alt_fill, border, center_align)

        # 保存文件
        filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
        wb.save(filepath)

        return filepath

    def _write_sheet(self, ws, config: Dict, title_font, header_font, cell_font,
                     header_fill, alt_fill, border, center_align):
        """写入单个工作表"""
        headers = config.get('headers', [])
        rows = config.get('rows', [])
        sheet_title = config.get('title', '')

        row_num = 1

        # 写入标题
        if sheet_title:
            ws.merge_cells(f'A1:{get_column_letter(max(len(headers), 1))}1')
            ws['A1'] = sheet_title
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            row_num = 3

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        row_num += 1

        # 写入数据行
        for i, row_data in enumerate(rows):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=self._format_value(value))
                cell.font = cell_font
                cell.border = border
                if i % 2 == 1:
                    cell.fill = alt_fill
            row_num += 1

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_summary(self, ws, summary: Dict, start_row: int, header_font, cell_font, border):
        """写入摘要数据"""
        ws.cell(row=start_row, column=1, value="摘要").font = header_font
        row = start_row + 1
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key).font = cell_font
            ws.cell(row=row, column=2, value=self._format_value(value)).font = cell_font
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).border = border
            row += 1

    def _write_table(self, ws, data: List[Dict], start_row: int, header_font, cell_font,
                     header_fill, alt_fill, border, center_align):
        """写入表格数据"""
        if not data:
            return

        headers = list(data[0].keys())

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # 写入数据
        for i, row_data in enumerate(data):
            row = start_row + i + 1
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row, column=col, value=self._format_value(value))
                cell.font = cell_font
                cell.border = border
                if i % 2 == 1:
                    cell.fill = alt_fill

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def export_to_pdf(
        self,
        data: Dict[str, Any],
        filename: str,
        title: str = "报表",
        landscape_mode: bool = False
    ) -> str:
        """
        导出为 PDF 文件

        Args:
            data: 报表数据
            filename: 文件名
            title: 报表标题
            landscape_mode: 是否横向

        Returns:
            文件路径
        """
        filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")

        page_size = landscape(A4) if landscape_mode else A4
        doc = SimpleDocTemplate(filepath, pagesize=page_size,
                               leftMargin=1*cm, rightMargin=1*cm,
                               topMargin=1*cm, bottomMargin=1*cm)

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=self.chinese_font,
            fontSize=16,
            alignment=1,  # 居中
            spaceAfter=20
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=self.chinese_font,
            fontSize=10
        )

        elements = []

        # 标题
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Spacer(1, 20))

        # 摘要
        if 'summary' in data:
            summary_data = [[k, str(self._format_value(v))] for k, v in data['summary'].items()]
            if summary_data:
                summary_table = Table(summary_data, colWidths=[100, 200])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
                    ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))

        # 详细数据表格
        if 'details' in data and isinstance(data['details'], list) and data['details']:
            headers = list(data['details'][0].keys())
            table_data = [headers]

            for row in data['details']:
                table_data.append([str(self._format_value(row.get(h, ''))) for h in headers])

            # 计算列宽
            available_width = page_size[0] - 2*cm
            col_width = available_width / len(headers)

            detail_table = Table(table_data, colWidths=[col_width] * len(headers))
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ]))
            elements.append(detail_table)

        doc.build(elements)
        return filepath

    def export_to_csv(
        self,
        data: Dict[str, Any],
        filename: str
    ) -> str:
        """
        导出为 CSV 文件

        Args:
            data: 报表数据
            filename: 文件名

        Returns:
            文件路径
        """
        filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv")

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            if 'details' in data and isinstance(data['details'], list) and data['details']:
                headers = list(data['details'][0].keys())
                writer.writerow(headers)

                for row in data['details']:
                    writer.writerow([self._format_value(row.get(h, '')) for h in headers])
            elif isinstance(data, list) and data:
                headers = list(data[0].keys())
                writer.writerow(headers)

                for row in data:
                    writer.writerow([self._format_value(row.get(h, '')) for h in headers])

        return filepath

    def _format_value(self, value: Any) -> Any:
        """格式化值"""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if isinstance(value, dict) or isinstance(value, list):
            return json.dumps(value, ensure_ascii=False)
        return value


# 报表生成器
class ReportGenerator:
    """报表数据生成器"""

    @staticmethod
    def generate_project_report(projects: List[Dict]) -> Dict:
        """生成项目报表数据"""
        return {
            "summary": {
                "总项目数": len(projects),
                "进行中": sum(1 for p in projects if p.get('status') == 'EXECUTING'),
                "已完成": sum(1 for p in projects if p.get('status') == 'COMPLETED'),
            },
            "details": projects
        }

    @staticmethod
    def generate_financial_report(financial_data: List[Dict]) -> Dict:
        """生成财务报表数据"""
        total_revenue = sum(d.get('revenue', 0) for d in financial_data)
        total_cost = sum(d.get('cost', 0) for d in financial_data)

        return {
            "summary": {
                "总营收": f"¥{total_revenue:,.2f}",
                "总成本": f"¥{total_cost:,.2f}",
                "总利润": f"¥{total_revenue - total_cost:,.2f}",
                "利润率": f"{((total_revenue - total_cost) / total_revenue * 100) if total_revenue else 0:.1f}%",
            },
            "details": financial_data
        }

    @staticmethod
    def generate_utilization_report(utilization_data: List[Dict]) -> Dict:
        """生成人员利用率报表"""
        avg_rate = sum(d.get('rate', 0) for d in utilization_data) / len(utilization_data) if utilization_data else 0

        return {
            "summary": {
                "人员总数": len(utilization_data),
                "平均利用率": f"{avg_rate:.1f}%",
                "饱和人数(>=80%)": sum(1 for d in utilization_data if d.get('rate', 0) >= 80),
                "空闲人数(<60%)": sum(1 for d in utilization_data if d.get('rate', 0) < 60),
            },
            "details": utilization_data
        }


# 创建单例
report_export_service = ReportExportService()

# -*- coding: utf-8 -*-
"""
Excel 渲染器

使用 openpyxl 生成 Excel 报告
"""

import os
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.services.report_framework.renderers.base import Renderer, ReportResult, RenderError
from app.services.report_framework.renderers.excel_styles import (
    HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT,
    DATA_FONT, DATA_ALIGNMENT,
    METRIC_LABEL_FONT, METRIC_VALUE_FONT,
    THIN_BORDER, ALT_ROW_FILL,
)


class ExcelRenderer(Renderer):
    """
    Excel 渲染器 — canonical Excel renderer (#39).
    共享样式定义见 ``excel_styles.py``。
    """

    def __init__(self, output_dir: str = "reports/excel"):
        self.output_dir = output_dir
        self.header_font = HEADER_FONT
        self.header_fill = HEADER_FILL
        self.header_alignment = HEADER_ALIGNMENT
        self.data_font = DATA_FONT
        self.data_alignment = DATA_ALIGNMENT
        self.metric_label_font = METRIC_LABEL_FONT
        self.metric_value_font = METRIC_VALUE_FONT
        self.border = THIN_BORDER
        self.alt_row_fill = ALT_ROW_FILL

    @property
    def format_name(self) -> str:
        return "excel"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def render(
        self,
        sections: List[Dict[str, Any]],
        metadata: Dict[str, Any],
    ) -> ReportResult:
        """
        渲染报告为 Excel

        Args:
            sections: 渲染后的 section 数据
            metadata: 报告元数据

        Returns:
            ReportResult: Excel 格式的报告结果
        """
        try:
            # 确保输出目录存在
            os.makedirs(self.output_dir, exist_ok=True)

            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_code = metadata.get("code", "report")
            file_name = f"{report_code}_{timestamp}.xlsx"
            file_path = os.path.join(self.output_dir, file_name)

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = metadata.get("name", "报告")[:31]  # Excel sheet name max 31 chars

            # 写入标题
            current_row = self._write_header(ws, metadata)

            # 渲染各 section
            for section in sections:
                current_row = self._render_section(ws, section, current_row)

            # 调整列宽
            self._auto_adjust_column_width(ws)

            # 保存文件
            wb.save(file_path)

            return ReportResult(
                data={"file_path": file_path, "sections": sections},
                format=self.format_name,
                file_path=file_path,
                file_name=file_name,
                content_type=self.content_type,
                metadata=metadata,
            )

        except Exception as e:
            raise RenderError(f"Excel rendering failed: {e}")

    def _write_header(self, ws, metadata: Dict[str, Any]) -> int:
        """写入报告头部"""
        # 标题
        title = metadata.get("name", "报告")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        # 生成时间
        gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        time_cell = ws.cell(row=2, column=1, value=f"生成时间: {gen_time}")
        time_cell.font = Font(size=10, color="666666")
        time_cell.alignment = Alignment(horizontal="center")

        return 4  # 从第4行开始写内容

    def _render_section(self, ws, section: Dict[str, Any], start_row: int) -> int:
        """渲染单个 section"""
        section_type = section.get("type")
        title = section.get("title")
        current_row = start_row

        # section 标题
        if title:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            title_cell = ws.cell(row=current_row, column=1, value=title)
            title_cell.font = Font(bold=True, size=12)
            current_row += 1

        if section_type == "metrics":
            current_row = self._render_metrics(ws, section, current_row)
        elif section_type == "table":
            current_row = self._render_table(ws, section, current_row)
        elif section_type == "chart":
            ws.cell(row=current_row, column=1, value="[图表: Excel 原生图表需要额外实现]")
            current_row += 1

        # 空行分隔
        return current_row + 1

    def _render_metrics(self, ws, section: Dict[str, Any], start_row: int) -> int:
        """渲染指标卡片"""
        items = section.get("items", [])
        if not items:
            return start_row

        # 每行 4 个指标
        cols = 4
        current_row = start_row

        for i, item in enumerate(items):
            col = (i % cols) * 2 + 1
            if i > 0 and i % cols == 0:
                current_row += 2

            label = str(item.get("label", ""))
            value = str(item.get("value", ""))

            # 标签
            label_cell = ws.cell(row=current_row, column=col, value=label)
            label_cell.font = self.metric_label_font

            # 值
            value_cell = ws.cell(row=current_row + 1, column=col, value=value)
            value_cell.font = self.metric_value_font

        return current_row + 3

    def _render_table(self, ws, section: Dict[str, Any], start_row: int) -> int:
        """渲染数据表格"""
        data = section.get("data", [])
        columns = section.get("columns", [])

        if not data or not columns:
            ws.cell(row=start_row, column=1, value="无数据")
            return start_row + 1

        current_row = start_row

        # 写入表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(
                row=current_row,
                column=col_idx,
                value=col.get("label", col.get("field", ""))
            )
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        current_row += 1

        # 写入数据行（限制最多 1000 行）
        for row_idx, row in enumerate(data[:1000]):
            for col_idx, col in enumerate(columns, 1):
                field = col.get("field", "")
                value = row.get(field, "")
                if value is None:
                    value = ""

                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.border

                # 斑马纹
                if row_idx % 2 == 1:
                    cell.fill = self.alt_row_fill

            current_row += 1

        # 显示数据量提示
        if len(data) > 1000:
            ws.cell(
                row=current_row,
                column=1,
                value=f"(显示前 1000 条，共 {len(data)} 条)"
            )
            current_row += 1

        return current_row

    def _auto_adjust_column_width(self, ws) -> None:
        """自动调整列宽"""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_idx)

            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                # 跳过合并单元格
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        # 中文字符按 2 个单位计算
                        cell_length = sum(
                            2 if ord(c) > 127 else 1
                            for c in str(cell.value)
                        )
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass

            # 设置列宽（最小 8，最大 50）
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[column].width = adjusted_width

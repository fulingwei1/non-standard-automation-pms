# -*- coding: utf-8 -*-
"""
Excel 共享样式常量

所有 Excel 导出服务共用的样式定义，消除跨模块的样式重复。
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 表头样式 ──────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="336699", end_color="336699", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── 数据行样式 ────────────────────────────────────────
DATA_FONT = Font(size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

# ── 标题样式 ──────────────────────────────────────────
TITLE_FONT = Font(bold=True, size=16)
TITLE_ALIGNMENT = Alignment(horizontal="center")
SUBTITLE_FONT = Font(size=10, color="666666")

# ── 指标样式 ──────────────────────────────────────────
METRIC_LABEL_FONT = Font(size=9, color="666666")
METRIC_VALUE_FONT = Font(size=14, bold=True)

# ── 边框 ─────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

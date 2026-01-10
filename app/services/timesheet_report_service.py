# -*- coding: utf-8 -*-
"""
工时报表生成服务
负责生成HR、财务、研发、项目等多格式报表
"""

import io
from typing import Optional, Dict, List, Any
from datetime import date, datetime
from decimal import Decimal
from sqlalchemy.orm import Session

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.services.timesheet_aggregation_service import TimesheetAggregationService
from app.services.overtime_calculation_service import OvertimeCalculationService
from app.services.hourly_rate_service import HourlyRateService
from app.models.organization import Department
from app.models.user import User


class TimesheetReportService:
    """工时报表生成服务"""
    
    def __init__(self, db: Session):
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel处理库未安装，请安装pandas和openpyxl: pip install pandas openpyxl")
        self.db = db
        self.aggregation_service = TimesheetAggregationService(db)
        self.overtime_service = OvertimeCalculationService(db)
    
    def generate_hr_report_excel(
        self,
        year: int,
        month: int,
        department_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        生成HR加班工资报表（Excel格式）
        
        Args:
            year: 年份
            month: 月份
            department_id: 部门ID（可选）
            
        Returns:
            Excel文件的内存流
        """
        # 获取HR报表数据
        hr_data = self.aggregation_service.generate_hr_report(year, month, department_id)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月加班工资表"
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=16, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:J1')
        ws['A1'] = f"{year}年{month}月加班工资汇总表"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # 写入生成时间
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, italic=True)
        
        # 表头
        headers = [
            '员工姓名', '部门', '正常工时', '工作日加班', '周末加班', 
            '节假日加班', '工作日加班工资', '周末加班工资', '节假日加班工资', '合计加班工资'
        ]
        
        row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # 写入数据
        row = 5
        for user_record in hr_data:
            # 计算加班工资
            user_id = user_record['user_id']
            user = self.db.query(User).filter(User.id == user_id).first()
            hourly_rate = HourlyRateService.get_user_hourly_rate(self.db, user_id, date(year, month, 1))
            
            # 获取部门名称
            department_name = ''
            if user and hasattr(user, 'department_id') and user.department_id:
                department = self.db.query(Department).filter(Department.id == user.department_id).first()
                if department:
                    department_name = department.name
            
            overtime_pay = Decimal(str(user_record['overtime_hours'])) * hourly_rate * Decimal("0.5")  # 1.5倍，额外0.5倍
            weekend_pay = Decimal(str(user_record['weekend_hours'])) * hourly_rate * Decimal("1.0")  # 2倍，额外1倍
            holiday_pay = Decimal(str(user_record['holiday_hours'])) * hourly_rate * Decimal("2.0")  # 3倍，额外2倍
            total_overtime_pay = overtime_pay + weekend_pay + holiday_pay
            
            ws.cell(row=row, column=1, value=user_record['user_name'])
            ws.cell(row=row, column=2, value=department_name)
            ws.cell(row=row, column=3, value=float(user_record['normal_hours']))
            ws.cell(row=row, column=4, value=float(user_record['overtime_hours']))
            ws.cell(row=row, column=5, value=float(user_record['weekend_hours']))
            ws.cell(row=row, column=6, value=float(user_record['holiday_hours']))
            ws.cell(row=row, column=7, value=float(overtime_pay))
            ws.cell(row=row, column=8, value=float(weekend_pay))
            ws.cell(row=row, column=9, value=float(holiday_pay))
            ws.cell(row=row, column=10, value=float(total_overtime_pay))
            
            # 应用样式
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                if col >= 3:  # 数字列右对齐
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = center_align
            
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        for col in range(3, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_finance_report_excel(
        self,
        year: int,
        month: int,
        project_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        生成财务报表（Excel格式）
        
        Args:
            year: 年份
            month: 月份
            project_id: 项目ID（可选）
            
        Returns:
            Excel文件的内存流
        """
        # 获取财务报表数据
        finance_data = self.aggregation_service.generate_finance_report(year, month, project_id)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月项目成本表"
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=16, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{year}年{month}月项目成本核算表"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # 写入生成时间
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, italic=True)
        
        # 表头
        headers = ['项目编号', '项目名称', '人员姓名', '工作日期', '工时', '时薪', '成本金额', '工作内容']
        
        row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # 写入数据
        row = 5
        for project_record in finance_data:
            project_code = project_record.get('project_code', '')
            project_name = project_record.get('project_name', '')
            
            for personnel_record in project_record.get('personnel_records', []):
                ws.cell(row=row, column=1, value=project_code)
                ws.cell(row=row, column=2, value=project_name)
                ws.cell(row=row, column=3, value=personnel_record['user_name'])
                ws.cell(row=row, column=4, value=personnel_record['date'])
                ws.cell(row=row, column=5, value=personnel_record['hours'])
                ws.cell(row=row, column=6, value=personnel_record['hourly_rate'])
                ws.cell(row=row, column=7, value=personnel_record['cost'])
                ws.cell(row=row, column=8, value=personnel_record.get('work_content', ''))
                
                # 应用样式
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = border
                    if col in [5, 6, 7]:  # 数字列右对齐
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = center_align
                
                row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_rd_report_excel(
        self,
        year: int,
        month: int,
        rd_project_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        生成研发报表（Excel格式）
        
        Args:
            year: 年份
            month: 月份
            rd_project_id: 研发项目ID（可选）
            
        Returns:
            Excel文件的内存流
        """
        # 获取研发报表数据
        rd_data = self.aggregation_service.generate_rd_report(year, month, rd_project_id)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月研发费用表"
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=16, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{year}年{month}月研发费用核算表"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # 写入生成时间
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, italic=True)
        
        # 表头
        headers = ['研发项目编号', '研发项目名称', '人员姓名', '工作日期', '工时', '时薪', '费用金额', '工作内容']
        
        row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # 写入数据
        row = 5
        for rd_project_record in rd_data:
            rd_project_code = rd_project_record.get('rd_project_code', '')
            rd_project_name = rd_project_record.get('rd_project_name', '')
            
            for personnel_record in rd_project_record.get('personnel_records', []):
                ws.cell(row=row, column=1, value=rd_project_code)
                ws.cell(row=row, column=2, value=rd_project_name)
                ws.cell(row=row, column=3, value=personnel_record['user_name'])
                ws.cell(row=row, column=4, value=personnel_record['date'])
                ws.cell(row=row, column=5, value=personnel_record['hours'])
                ws.cell(row=row, column=6, value=personnel_record['hourly_rate'])
                ws.cell(row=row, column=7, value=personnel_record['cost'])
                ws.cell(row=row, column=8, value=personnel_record.get('work_content', ''))
                
                # 应用样式
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = border
                    if col in [5, 6, 7]:  # 数字列右对齐
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = center_align
                
                row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_project_report_excel(
        self,
        project_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> io.BytesIO:
        """
        生成项目报表（Excel格式）
        
        Args:
            project_id: 项目ID
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            
        Returns:
            Excel文件的内存流
        """
        # 获取项目报表数据
        project_data = self.aggregation_service.generate_project_report(
            project_id, start_date, end_date
        )
        
        if 'error' in project_data:
            # 创建错误提示的Excel
            wb = Workbook()
            ws = wb.active
            ws['A1'] = project_data['error']
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # 创建工作簿
        wb = Workbook()
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=16, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Sheet 1: 人员贡献度统计
        ws1 = wb.active
        ws1.title = "人员贡献度"
        
        ws1.merge_cells('A1:C1')
        ws1['A1'] = f"{project_data['project_name']} - 人员贡献度统计"
        ws1['A1'].font = title_font
        ws1['A1'].alignment = center_align
        
        headers1 = ['人员姓名', '总工时', '贡献度(%)']
        row = 3
        for col_idx, header in enumerate(headers1, start=1):
            cell = ws1.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        row = 4
        for personnel in project_data.get('personnel_stats', []):
            ws1.cell(row=row, column=1, value=personnel['user_name'])
            ws1.cell(row=row, column=2, value=personnel['total_hours'])
            ws1.cell(row=row, column=3, value=f"{personnel['contribution_rate']:.2f}%")
            
            for col in range(1, 4):
                cell = ws1.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                cell.alignment = center_align if col == 1 else Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 15
        
        # Sheet 2: 每日工时分布
        ws2 = wb.create_sheet(title="每日工时分布")
        
        ws2.merge_cells('A1:D1')
        ws2['A1'] = f"{project_data['project_name']} - 每日工时分布"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = center_align
        
        headers2 = ['日期', '总工时', '参与人数', '人员列表']
        row = 3
        for col_idx, header in enumerate(headers2, start=1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        row = 4
        for daily in project_data.get('daily_stats', []):
            personnel_list = ', '.join([p['user_name'] for p in daily.get('personnel', [])])
            
            ws2.cell(row=row, column=1, value=daily['date'])
            ws2.cell(row=row, column=2, value=daily['hours'])
            ws2.cell(row=row, column=3, value=daily.get('personnel_count', 0))
            ws2.cell(row=row, column=4, value=personnel_list)
            
            for col in range(1, 5):
                cell = ws2.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                cell.alignment = center_align if col in [1, 3, 4] else Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 40
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

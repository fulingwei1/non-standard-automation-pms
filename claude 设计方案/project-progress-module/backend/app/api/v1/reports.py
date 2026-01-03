"""
报表模块 API
支持多角色视角的报表生成和导出
"""
from fastapi import APIRouter, Query, HTTPException, Depends
from fastapi.responses import FileResponse
from typing import Optional, List
from datetime import date, datetime
from enum import Enum
import tempfile
import os

from app.services.report_engine import (
    ReportGenerator, 
    MockDataProvider,
    RoleType, 
    ReportType, 
    ReportPeriod,
    ROLE_CONFIGS,
    generate_report_for_role
)

router = APIRouter(prefix="/reports", tags=["报表模块"])


# ==================== 报表配置接口 ====================

@router.get("/roles", summary="获取支持的角色列表")
async def get_supported_roles():
    """
    获取系统支持的所有角色及其配置
    """
    roles = []
    for role_type, config in ROLE_CONFIGS.items():
        roles.append({
            "code": role_type.value,
            "name": config.display_name,
            "focus_areas": config.focus_areas,
            "data_scope": config.data_scope,
            "language_style": config.language_style,
            "chart_preferences": config.chart_preferences,
            "show_recommendations": config.show_recommendations,
            "show_trends": config.show_trends
        })
    return {"code": 200, "data": roles}


@router.get("/types", summary="获取报表类型列表")
async def get_report_types():
    """
    获取系统支持的所有报表类型
    """
    types = [
        {"code": "project_weekly", "name": "项目周报", "period": "weekly", "description": "项目周度进展汇报"},
        {"code": "project_monthly", "name": "项目月报", "period": "monthly", "description": "项目月度总结分析"},
        {"code": "dept_weekly", "name": "部门周报", "period": "weekly", "description": "部门周度工作汇总"},
        {"code": "dept_monthly", "name": "部门月报", "period": "monthly", "description": "部门月度经营分析"},
        {"code": "company_monthly", "name": "公司月报", "period": "monthly", "description": "公司经营分析报告"},
        {"code": "cost_analysis", "name": "成本分析", "period": "monthly", "description": "项目成本分析报表"},
        {"code": "workload_analysis", "name": "负荷分析", "period": "weekly", "description": "人员负荷分析报表"},
        {"code": "risk_report", "name": "风险报告", "period": "weekly", "description": "项目风险预警报告"},
    ]
    return {"code": 200, "data": types}


@router.get("/role-report-matrix", summary="获取角色-报表权限矩阵")
async def get_role_report_matrix():
    """
    获取不同角色可以查看的报表类型矩阵
    """
    matrix = {
        "gm": ["project_weekly", "project_monthly", "dept_monthly", "company_monthly", "cost_analysis", "workload_analysis", "risk_report"],
        "dept_manager": ["project_weekly", "project_monthly", "dept_weekly", "dept_monthly", "workload_analysis", "risk_report"],
        "pm": ["project_weekly", "project_monthly", "cost_analysis", "workload_analysis", "risk_report"],
        "te": ["project_weekly", "project_monthly", "workload_analysis", "risk_report"],
        "engineer": ["project_weekly"],
        "pmc": ["project_weekly", "workload_analysis"],
        "finance": ["project_monthly", "cost_analysis"],
        "qa": ["project_weekly", "project_monthly"],
    }
    return {"code": 200, "data": matrix}


# ==================== 报表生成接口 ====================

@router.post("/generate", summary="生成报表")
async def generate_report(
    report_type: str = Query(..., description="报表类型"),
    role: str = Query(..., description="查看角色"),
    period: str = Query("weekly", description="周期: daily/weekly/monthly/quarterly"),
    project_id: Optional[int] = Query(None, description="项目ID（项目报表需要）"),
    dept_id: Optional[int] = Query(None, description="部门ID（部门报表需要）"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD")
):
    """
    根据角色和类型生成报表
    
    不同角色查看同一报表会得到不同的：
    - 数据范围（总经理看全局，工程师看个人）
    - 关注指标（管理层看经营指标，技术看质量指标）
    - 语言风格（正式/简洁/详细）
    - 建议内容（是否显示、行动级别）
    """
    # 验证角色
    try:
        role_type = RoleType(role)
    except ValueError:
        valid_roles = [r.value for r in RoleType]
        raise HTTPException(status_code=400, detail=f"无效的角色，可选值: {valid_roles}")
    
    # 验证报表类型
    try:
        rpt_type = ReportType(report_type)
    except ValueError:
        valid_types = [t.value for t in ReportType]
        raise HTTPException(status_code=400, detail=f"无效的报表类型，可选值: {valid_types}")
    
    # 验证周期
    try:
        prd = ReportPeriod(period)
    except ValueError:
        valid_periods = [p.value for p in ReportPeriod]
        raise HTTPException(status_code=400, detail=f"无效的周期，可选值: {valid_periods}")
    
    # 解析日期
    parsed_start = None
    parsed_end = None
    if start_date:
        try:
            parsed_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为 YYYY-MM-DD")
    if end_date:
        try:
            parsed_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为 YYYY-MM-DD")
    
    # 确定 scope_id
    scope_id = project_id or dept_id
    
    # 生成报表
    generator = ReportGenerator(MockDataProvider())
    report = generator.generate_report(
        report_type=rpt_type,
        role=role_type,
        period=prd,
        scope_id=scope_id,
        start_date=parsed_start,
        end_date=parsed_end
    )
    
    return {
        "code": 200,
        "message": "报表生成成功",
        "data": report.to_dict()
    }


@router.get("/preview/{report_type}", summary="预览报表（简化版）")
async def preview_report(
    report_type: str,
    role: str = Query("pm", description="查看角色")
):
    """
    快速预览报表，返回简化数据用于前端展示
    """
    result = generate_report_for_role(role, report_type)
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    
    # 返回简化版本用于预览
    preview = {
        "title": result["title"],
        "subtitle": result["subtitle"],
        "generated_at": result["generated_at"],
        "metrics_count": len(result["metrics"]),
        "alerts_count": len(result["alerts"]),
        "sections_count": len(result["sections"]),
        "top_metrics": result["metrics"][:4],
        "critical_alerts": [a for a in result["alerts"] if a["level"] in ["high", "critical"]][:3]
    }
    
    return {"code": 200, "data": preview}


# ==================== 报表比较接口 ====================

@router.post("/compare-roles", summary="比较不同角色视角的报表")
async def compare_role_reports(
    report_type: str = Query(..., description="报表类型"),
    roles: List[str] = Query(..., description="要比较的角色列表")
):
    """
    生成同一报表在不同角色视角下的对比
    
    用于展示同一数据如何为不同决策者呈现不同重点
    """
    if len(roles) < 2:
        raise HTTPException(status_code=400, detail="至少需要2个角色进行比较")
    if len(roles) > 5:
        raise HTTPException(status_code=400, detail="最多支持5个角色同时比较")
    
    comparison = {
        "report_type": report_type,
        "compared_roles": [],
        "metrics_comparison": {},
        "focus_areas_comparison": {}
    }
    
    generator = ReportGenerator(MockDataProvider())
    
    for role in roles:
        try:
            role_type = RoleType(role)
            rpt_type = ReportType(report_type)
        except ValueError:
            continue
        
        report = generator.generate_report(
            report_type=rpt_type,
            role=role_type
        )
        
        config = ROLE_CONFIGS[role_type]
        
        comparison["compared_roles"].append({
            "role": role,
            "role_name": config.display_name,
            "focus_areas": config.focus_areas,
            "data_scope": config.data_scope,
            "title": report.title,
            "metrics": [{"name": m.name, "value": m.value, "unit": m.unit} for m in report.metrics],
            "alerts_count": len(report.alerts),
            "recommendations_count": len(report.recommendations),
            "language_style": config.language_style
        })
    
    return {"code": 200, "data": comparison}


# ==================== 报表导出接口 ====================

@router.post("/export", summary="导出报表")
async def export_report(
    report_type: str = Query(..., description="报表类型"),
    role: str = Query(..., description="查看角色"),
    format: str = Query("xlsx", description="导出格式: xlsx/pdf/html"),
    period: str = Query("weekly", description="周期")
):
    """
    导出报表为指定格式
    
    支持格式：
    - xlsx: Excel格式，适合数据分析
    - pdf: PDF格式，适合正式报告
    - html: HTML格式，适合邮件发送
    """
    if format not in ["xlsx", "pdf", "html"]:
        raise HTTPException(status_code=400, detail="不支持的导出格式，可选: xlsx/pdf/html")
    
    # 生成报表数据
    result = generate_report_for_role(role, report_type, period)
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    
    # 根据格式生成文件
    temp_dir = tempfile.mkdtemp()
    
    if format == "xlsx":
        filepath = os.path.join(temp_dir, f"{result['title']}.xlsx")
        _export_to_excel(result, filepath)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif format == "pdf":
        filepath = os.path.join(temp_dir, f"{result['title']}.pdf")
        _export_to_pdf(result, filepath)
        media_type = "application/pdf"
    else:  # html
        filepath = os.path.join(temp_dir, f"{result['title']}.html")
        _export_to_html(result, filepath)
        media_type = "text/html"
    
    return FileResponse(
        path=filepath,
        filename=os.path.basename(filepath),
        media_type=media_type
    )


def _export_to_excel(report_data: dict, filepath: str):
    """导出为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "报表概览"
    
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = report_data['title']
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A2'] = report_data['subtitle']
    ws['A3'] = f"生成时间: {report_data['generated_at']}"
    
    # 核心指标
    ws['A5'] = "核心指标"
    ws['A5'].font = Font(bold=True, size=12)
    
    row = 6
    for metric in report_data.get('metrics', []):
        ws.cell(row=row, column=1, value=metric['name'])
        ws.cell(row=row, column=2, value=f"{metric['value']}{metric.get('unit', '')}")
        ws.cell(row=row, column=3, value=metric.get('status', ''))
        row += 1
    
    # 预警信息
    row += 1
    ws.cell(row=row, column=1, value="预警信息").font = Font(bold=True, size=12)
    row += 1
    
    for alert in report_data.get('alerts', []):
        ws.cell(row=row, column=1, value=alert['level'])
        ws.cell(row=row, column=2, value=alert['title'])
        ws.cell(row=row, column=3, value=alert['description'])
        row += 1
    
    wb.save(filepath)


def _export_to_pdf(report_data: dict, filepath: str):
    """导出为PDF（简化版，实际需要reportlab或weasyprint）"""
    # 简化实现：生成HTML后转PDF
    html_path = filepath.replace('.pdf', '.html')
    _export_to_html(report_data, html_path)
    
    # 实际项目中应使用 weasyprint 或 reportlab
    # 这里简化处理，直接复制HTML
    import shutil
    shutil.copy(html_path, filepath)


def _export_to_html(report_data: dict, filepath: str):
    """导出为HTML"""
    html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_data['title']}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 40px; background: #f5f5f5; }}
        .report {{ max-width: 900px; margin: 0 auto; background: white; border-radius: 12px; padding: 40px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }}
        .header {{ border-bottom: 1px solid #eee; padding-bottom: 20px; margin-bottom: 30px; }}
        .header h1 {{ font-size: 28px; color: #1a1a2e; margin-bottom: 8px; }}
        .header p {{ color: #666; font-size: 14px; }}
        .section {{ margin-bottom: 30px; }}
        .section-title {{ font-size: 18px; font-weight: 600; color: #333; margin-bottom: 16px; padding-left: 12px; border-left: 4px solid #6366F1; }}
        .metrics-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; }}
        .metric-card {{ background: #f8fafc; border-radius: 10px; padding: 20px; }}
        .metric-card .value {{ font-size: 28px; font-weight: 700; color: #1a1a2e; }}
        .metric-card .name {{ font-size: 13px; color: #666; margin-top: 4px; }}
        .metric-card .status {{ font-size: 11px; padding: 2px 8px; border-radius: 4px; display: inline-block; margin-top: 8px; }}
        .metric-card .status.good {{ background: #dcfce7; color: #166534; }}
        .metric-card .status.warning {{ background: #fef3c7; color: #92400e; }}
        .metric-card .status.critical {{ background: #fee2e2; color: #991b1b; }}
        .alert {{ padding: 16px; border-radius: 8px; margin-bottom: 12px; border-left: 4px solid; }}
        .alert.high {{ background: #fef2f2; border-color: #ef4444; }}
        .alert.medium {{ background: #fffbeb; border-color: #f59e0b; }}
        .alert.low {{ background: #f0fdf4; border-color: #22c55e; }}
        .alert-title {{ font-weight: 600; color: #333; margin-bottom: 4px; }}
        .alert-desc {{ font-size: 13px; color: #666; }}
        .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; text-align: center; color: #999; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="report">
        <div class="header">
            <h1>{report_data['title']}</h1>
            <p>{report_data['subtitle']} · 生成时间: {report_data['generated_at']}</p>
        </div>
        
        <div class="section">
            <h2 class="section-title">核心指标</h2>
            <div class="metrics-grid">
                {''.join([f'''
                <div class="metric-card">
                    <div class="value">{m['value']}{m.get('unit', '')}</div>
                    <div class="name">{m['name']}</div>
                    <span class="status {m.get('status', 'normal')}">{m.get('status', '').upper() or '正常'}</span>
                </div>
                ''' for m in report_data.get('metrics', [])])}
            </div>
        </div>
        
        <div class="section">
            <h2 class="section-title">预警信息</h2>
            {''.join([f'''
            <div class="alert {a['level']}">
                <div class="alert-title">{a['title']}</div>
                <div class="alert-desc">{a['description']}</div>
            </div>
            ''' for a in report_data.get('alerts', [])])}
            {'' if report_data.get('alerts') else '<p style="color: #666; font-size: 14px;">暂无预警信息</p>'}
        </div>
        
        <div class="footer">
            <p>本报表由项目进度管理系统自动生成</p>
        </div>
    </div>
</body>
</html>
    """
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html_content)


# ==================== 报表模板管理接口 ====================

@router.get("/templates", summary="获取报表模板列表")
async def get_report_templates():
    """
    获取可用的报表模板
    """
    templates = [
        {
            "id": "tpl_project_weekly_standard",
            "name": "标准项目周报",
            "type": "project_weekly",
            "description": "包含进度、任务、风险、下周计划",
            "sections": ["progress_overview", "task_summary", "risk_alerts", "next_week_plan"],
            "default_for_roles": ["pm", "te"]
        },
        {
            "id": "tpl_project_weekly_executive",
            "name": "项目周报-管理层版",
            "type": "project_weekly",
            "description": "精简版，聚焦关键指标和风险",
            "sections": ["kpi_dashboard", "critical_alerts", "key_decisions"],
            "default_for_roles": ["gm", "dept_manager"]
        },
        {
            "id": "tpl_cost_analysis_detail",
            "name": "详细成本分析",
            "type": "cost_analysis",
            "description": "包含成本构成、趋势、预测",
            "sections": ["cost_breakdown", "cost_trend", "cost_forecast", "variance_analysis"],
            "default_for_roles": ["finance"]
        },
    ]
    
    return {"code": 200, "data": templates}


@router.post("/templates/apply", summary="应用报表模板")
async def apply_report_template(
    template_id: str = Query(..., description="模板ID"),
    role: str = Query(..., description="查看角色"),
    scope_id: Optional[int] = Query(None, description="作用域ID")
):
    """
    根据模板生成报表
    """
    # 模拟模板应用
    template_type_map = {
        "tpl_project_weekly_standard": "project_weekly",
        "tpl_project_weekly_executive": "project_weekly",
        "tpl_cost_analysis_detail": "cost_analysis",
    }
    
    report_type = template_type_map.get(template_id)
    if not report_type:
        raise HTTPException(status_code=404, detail="模板不存在")
    
    result = generate_report_for_role(role, report_type)
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    
    result["template_id"] = template_id
    
    return {"code": 200, "data": result}

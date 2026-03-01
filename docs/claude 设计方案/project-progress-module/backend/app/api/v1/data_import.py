"""
数据导入导出API
"""
from fastapi import APIRouter, UploadFile, File, Query, HTTPException
from fastapi.responses import FileResponse
from typing import Optional
import tempfile
import os
import shutil

from app.utils.data_importer import (
    ExcelTemplateGenerator, ExcelImporter, ImportType
)

router = APIRouter(prefix="/import", tags=["数据导入"])


@router.get("/templates/{import_type}", summary="下载导入模板")
async def download_template(import_type: str):
    """
    下载指定类型的导入模板
    
    Args:
        import_type: 导入类型 (project/task/user/timesheet/customer/department)
    """
    try:
        it = ImportType(import_type)
    except ValueError:
        valid_types = [t.value for t in ImportType]
        raise HTTPException(
            status_code=400, 
            detail=f"不支持的导入类型: {import_type}，可选值: {valid_types}"
        )
    
    # 生成模板到临时目录
    temp_dir = tempfile.mkdtemp()
    filename = f"导入模板_{import_type}.xlsx"
    filepath = os.path.join(temp_dir, filename)
    
    generator = ExcelTemplateGenerator()
    generator.generate_template(it, filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/templates", summary="获取所有模板类型")
async def list_templates():
    """获取所有可用的导入模板类型"""
    templates = []
    for it in ImportType:
        templates.append({
            "type": it.value,
            "name": {
                "project": "项目导入",
                "task": "任务导入",
                "user": "人员导入",
                "timesheet": "工时导入",
                "customer": "客户导入",
                "department": "部门导入"
            }.get(it.value, it.value),
            "download_url": f"/api/v1/import/templates/{it.value}"
        })
    return {"code": 200, "data": templates}


@router.post("/preview", summary="预览导入数据")
async def preview_import(
    file: UploadFile = File(...),
    import_type: str = Query(..., description="导入类型"),
    sheet_name: Optional[str] = Query(None, description="工作表名称"),
    max_rows: int = Query(20, description="预览行数")
):
    """
    预览Excel文件数据
    
    上传文件后返回前N行数据预览，用于确认数据格式
    """
    try:
        it = ImportType(import_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"不支持的导入类型: {import_type}")
    
    # 保存上传文件到临时目录
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, file.filename)
    
    try:
        with open(temp_file, "wb") as f:
            shutil.copyfileobj(file.file, f)
        
        importer = ExcelImporter()
        result = importer.preview_import(temp_file, it, sheet_name, max_rows)
        
        if result["success"]:
            return {"code": 200, "data": result}
        else:
            return {"code": 400, "message": result.get("message", "预览失败")}
    finally:
        # 清理临时文件
        shutil.rmtree(temp_dir, ignore_errors=True)


@router.post("/upload", summary="上传并导入数据")
async def upload_and_import(
    file: UploadFile = File(...),
    import_type: str = Query(..., description="导入类型"),
    sheet_name: Optional[str] = Query(None, description="工作表名称"),
    update_existing: bool = Query(False, description="是否更新已存在的数据")
):
    """
    上传Excel文件并导入数据
    
    Args:
        file: Excel文件
        import_type: 导入类型 (project/task/user/timesheet/customer/department)
        sheet_name: 工作表名称，默认第一个
        update_existing: 是否更新已存在的数据，默认False
        
    Returns:
        导入结果，包含成功/失败数量和错误详情
    """
    # 验证文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件(.xlsx, .xls)")
    
    try:
        it = ImportType(import_type)
    except ValueError:
        valid_types = [t.value for t in ImportType]
        raise HTTPException(
            status_code=400, 
            detail=f"不支持的导入类型: {import_type}，可选值: {valid_types}"
        )
    
    # 保存上传文件
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, file.filename)
    
    try:
        with open(temp_file, "wb") as f:
            shutil.copyfileobj(file.file, f)
        
        # 执行导入
        importer = ExcelImporter()
        result = importer.import_from_excel(temp_file, it, sheet_name, 0, update_existing)
        
        return {
            "code": 200 if result.success else 400,
            "message": "导入成功" if result.success else "导入失败，请检查错误信息",
            "data": result.to_dict()
        }
    finally:
        # 清理临时文件
        shutil.rmtree(temp_dir, ignore_errors=True)


@router.post("/validate", summary="验证导入数据")
async def validate_import(
    file: UploadFile = File(...),
    import_type: str = Query(..., description="导入类型"),
    sheet_name: Optional[str] = Query(None, description="工作表名称")
):
    """
    验证Excel文件数据格式，不执行实际导入
    
    用于在正式导入前检查数据是否符合要求
    """
    try:
        it = ImportType(import_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"不支持的导入类型: {import_type}")
    
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, file.filename)
    
    try:
        with open(temp_file, "wb") as f:
            shutil.copyfileobj(file.file, f)
        
        # 只验证不导入
        importer = ExcelImporter()
        result = importer.import_from_excel(temp_file, it, sheet_name, 0, False)
        
        # 不管验证是否通过，都返回结果
        return {
            "code": 200,
            "data": {
                "valid": result.error_count == 0,
                "total_rows": result.total_rows,
                "error_count": result.error_count,
                "errors": result.errors[:50],  # 最多返回50条错误
                "warnings": result.warnings[:20]
            }
        }
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# ==================== 数据导出 ====================

@router.get("/export/{export_type}", summary="导出数据")
async def export_data(
    export_type: str,
    project_id: Optional[int] = Query(None, description="项目ID"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期")
):
    """
    导出数据为Excel文件
    
    Args:
        export_type: 导出类型 (project_list/project_detail/task_list/timesheet/workload)
        project_id: 项目ID（导出项目详情时需要）
        start_date: 开始日期（导出工时等需要）
        end_date: 结束日期
    """
    valid_types = ['project_list', 'project_detail', 'task_list', 'timesheet', 'workload', 'user_list']
    if export_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"不支持的导出类型，可选值: {valid_types}")
    
    # TODO: 实现实际的导出逻辑
    # 这里先返回一个示例模板
    temp_dir = tempfile.mkdtemp()
    filename = f"导出_{export_type}_{start_date or 'all'}.xlsx"
    filepath = os.path.join(temp_dir, filename)
    
    # 生成示例导出文件
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "导出数据"
    ws.cell(row=1, column=1, value="导出功能开发中...")
    ws.cell(row=2, column=1, value=f"导出类型: {export_type}")
    ws.cell(row=3, column=1, value=f"项目ID: {project_id}")
    ws.cell(row=4, column=1, value=f"日期范围: {start_date} ~ {end_date}")
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

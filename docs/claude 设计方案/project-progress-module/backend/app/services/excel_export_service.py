# ===========================================
# Excel报表导出服务
# ===========================================

from fastapi import APIRouter, Query, Response
from typing import Optional, List, Dict, Any
from datetime import date, datetime
from enum import Enum
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart

router = APIRouter(prefix="/api/v1/export", tags=["报表导出"])


# ===========================================
# 样式定义
# ===========================================

class ExcelStyles:
    """Excel样式定义"""
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')
    
    # 汇总行样式
    summary_font = Font(name='微软雅黑', size=11, bold=True)
    summary_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    # 状态颜色
    status_colors = {
        'success': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'warning': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'danger': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'info': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    }
    
    # 数字格式
    number_format = '#,##0'
    decimal_format = '#,##0.00'
    percent_format = '0.00%'
    date_format = 'YYYY-MM-DD'
    datetime_format = 'YYYY-MM-DD HH:MM:SS'


class ExcelExporter:
    """Excel导出器基类"""
    
    def __init__(self):
        self.wb = Workbook()
        self.styles = ExcelStyles()
    
    def create_sheet(self, title: str):
        """创建工作表"""
        if self.wb.active.title == 'Sheet':
            ws = self.wb.active
            ws.title = title
        else:
            ws = self.wb.create_sheet(title)
        return ws
    
    def set_column_widths(self, ws, widths: Dict[str, int]):
        """设置列宽"""
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def write_title(self, ws, title: str, row: int = 1, col_span: int = 1):
        """写入标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.title_font
        cell.alignment = self.styles.title_alignment
        if col_span > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    
    def write_header(self, ws, headers: List[str], row: int = 3):
        """写入表头"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.thin_border
    
    def write_data_row(self, ws, data: List[Any], row: int, formats: List[str] = None):
        """写入数据行"""
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.data_font
            cell.border = self.styles.thin_border
            
            # 根据数据类型设置对齐方式
            if isinstance(value, (int, float)):
                cell.alignment = self.styles.data_alignment_right
            elif isinstance(value, (date, datetime)):
                cell.alignment = self.styles.data_alignment_center
            else:
                cell.alignment = self.styles.data_alignment
            
            # 应用格式
            if formats and col <= len(formats):
                if formats[col-1] == 'number':
                    cell.number_format = self.styles.number_format
                elif formats[col-1] == 'decimal':
                    cell.number_format = self.styles.decimal_format
                elif formats[col-1] == 'percent':
                    cell.number_format = self.styles.percent_format
                    if isinstance(value, (int, float)):
                        cell.value = value / 100  # 转换为小数
                elif formats[col-1] == 'date':
                    cell.number_format = self.styles.date_format
    
    def write_summary_row(self, ws, data: List[Any], row: int):
        """写入汇总行"""
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.summary_font
            cell.fill = self.styles.summary_fill
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.data_alignment_right if isinstance(value, (int, float)) else self.styles.data_alignment
    
    def apply_status_color(self, cell, status: str):
        """应用状态颜色"""
        color_map = {
            'complete': 'success',
            'completed': 'success',
            'resolved': 'success',
            'fulfilled': 'success',
            'on_time': 'success',
            'partial': 'warning',
            'handling': 'warning',
            'pending': 'warning',
            'delayed': 'danger',
            'shortage': 'danger',
            'overdue': 'danger',
            'level3': 'warning',
            'level4': 'danger'
        }
        style_key = color_map.get(status.lower() if status else '', 'info')
        if style_key in self.styles.status_colors:
            cell.fill = self.styles.status_colors[style_key]
    
    def add_chart(self, ws, chart_type: str, data_range: str, title: str, 
                  position: str = 'A1', width: int = 15, height: int = 10):
        """添加图表"""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            return
        
        chart.title = title
        chart.width = width
        chart.height = height
        ws.add_chart(chart, position)
        return chart
    
    def save_to_bytes(self) -> bytes:
        """保存到字节流"""
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()


# ===========================================
# 缺料报表导出
# ===========================================

class ShortageReportExporter(ExcelExporter):
    """缺料报表导出器"""
    
    def export_kit_rate_report(self, data: Dict) -> bytes:
        """导出齐套率报表"""
        ws = self.create_sheet('齐套率分析')
        
        # 标题
        self.write_title(ws, f"齐套率分析报表 ({data['period']})", col_span=8)
        
        # 汇总信息
        ws.cell(row=2, column=1, value=f"统计周期: {data['start_date']} 至 {data['end_date']}")
        
        # 表头
        headers = ['日期', '检查工单数', '齐套工单', '部分齐套', '缺料工单', '齐套率(%)', '环比', '同比']
        self.write_header(ws, headers, row=4)
        
        # 数据
        formats = ['date', 'number', 'number', 'number', 'number', 'decimal', 'percent', 'percent']
        for idx, row_data in enumerate(data['trend_data'], 5):
            self.write_data_row(ws, [
                row_data['date'],
                row_data['total'],
                row_data['complete'],
                row_data['partial'],
                row_data['shortage'],
                row_data['kit_rate'],
                row_data.get('mom', 0),
                row_data.get('yoy', 0)
            ], idx, formats)
        
        # 汇总行
        summary_row = 5 + len(data['trend_data'])
        self.write_summary_row(ws, [
            '合计/平均',
            data['summary']['total_orders'],
            data['summary']['complete_orders'],
            data['summary']['partial_orders'],
            data['summary']['shortage_orders'],
            data['summary']['avg_kit_rate'],
            '-',
            '-'
        ], summary_row)
        
        # 车间对比表
        ws2 = self.create_sheet('车间齐套率对比')
        self.write_title(ws2, '车间齐套率对比', col_span=6)
        
        headers2 = ['车间', '检查工单数', '齐套工单', '缺料工单', '齐套率(%)', '排名']
        self.write_header(ws2, headers2, row=3)
        
        for idx, ws_data in enumerate(data['by_workshop'], 4):
            row_values = [
                ws_data['name'],
                ws_data['total'],
                ws_data['complete'],
                ws_data['shortage'],
                ws_data['kit_rate'],
                idx - 3
            ]
            self.write_data_row(ws2, row_values, idx, ['', 'number', 'number', 'number', 'decimal', 'number'])
            
            # 状态颜色
            rate_cell = ws2.cell(row=idx, column=5)
            if ws_data['kit_rate'] >= 90:
                self.apply_status_color(rate_cell, 'complete')
            elif ws_data['kit_rate'] >= 80:
                self.apply_status_color(rate_cell, 'partial')
            else:
                self.apply_status_color(rate_cell, 'shortage')
        
        # 设置列宽
        self.set_column_widths(ws, {'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 10, 'H': 10})
        self.set_column_widths(ws2, {'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 8})
        
        return self.save_to_bytes()
    
    def export_shortage_alert_report(self, data: Dict) -> bytes:
        """导出缺料预警报表"""
        ws = self.create_sheet('缺料预警明细')
        
        # 标题
        self.write_title(ws, f"缺料预警报表 ({data['start_date']} 至 {data['end_date']})", col_span=12)
        
        # 汇总统计
        ws.cell(row=2, column=1, value=f"预警总数: {data['summary']['total']}  "
                f"待处理: {data['summary']['pending']}  "
                f"处理中: {data['summary']['handling']}  "
                f"已解决: {data['summary']['resolved']}")
        
        # 表头
        headers = ['预警编号', '预警级别', '工单号', '项目名称', '物料编码', '物料名称', 
                   '缺料数量', '需求日期', '状态', '处理人', '响应时间', '解决时间']
        self.write_header(ws, headers, row=4)
        
        # 数据
        level_map = {'level1': '一级', 'level2': '二级', 'level3': '三级', 'level4': '四级'}
        status_map = {'pending': '待处理', 'handling': '处理中', 'resolved': '已解决', 
                      'escalated': '已升级', 'closed': '已关闭'}
        
        for idx, alert in enumerate(data['alerts'], 5):
            row_values = [
                alert['alert_no'],
                level_map.get(alert['alert_level'], alert['alert_level']),
                alert['work_order_no'],
                alert['project_name'],
                alert['material_code'],
                alert['material_name'],
                alert['shortage_qty'],
                alert['required_date'],
                status_map.get(alert['status'], alert['status']),
                alert.get('handler_name', '-'),
                alert.get('response_time', '-'),
                alert.get('resolve_time', '-')
            ]
            self.write_data_row(ws, row_values, idx)
            
            # 级别颜色
            level_cell = ws.cell(row=idx, column=2)
            self.apply_status_color(level_cell, alert['alert_level'])
            
            # 状态颜色
            status_cell = ws.cell(row=idx, column=9)
            self.apply_status_color(status_cell, alert['status'])
        
        # 设置列宽
        self.set_column_widths(ws, {
            'A': 15, 'B': 10, 'C': 15, 'D': 25, 'E': 15, 'F': 20,
            'G': 10, 'H': 12, 'I': 10, 'J': 10, 'K': 18, 'L': 18
        })
        
        # 添加统计表
        ws2 = self.create_sheet('预警统计')
        self.write_title(ws2, '预警统计分析', col_span=5)
        
        # 按级别统计
        ws2.cell(row=3, column=1, value='按预警级别统计')
        ws2.cell(row=3, column=1).font = Font(bold=True)
        
        headers_level = ['级别', '数量', '占比', '平均响应时间', '平均解决时间']
        self.write_header(ws2, headers_level, row=4)
        
        for idx, level_stat in enumerate(data['by_level'], 5):
            self.write_data_row(ws2, [
                level_map.get(level_stat['level'], level_stat['level']),
                level_stat['count'],
                level_stat['percent'],
                level_stat['avg_response_time'],
                level_stat['avg_resolve_time']
            ], idx, ['', 'number', 'percent', '', ''])
        
        # 按原因统计
        reason_start = 5 + len(data['by_level']) + 2
        ws2.cell(row=reason_start, column=1, value='按缺料原因统计')
        ws2.cell(row=reason_start, column=1).font = Font(bold=True)
        
        headers_reason = ['原因', '数量', '占比']
        self.write_header(ws2, headers_reason, row=reason_start + 1)
        
        for idx, reason_stat in enumerate(data['by_reason'], reason_start + 2):
            self.write_data_row(ws2, [
                reason_stat['reason'],
                reason_stat['count'],
                reason_stat['percent']
            ], idx, ['', 'number', 'percent'])
        
        self.set_column_widths(ws2, {'A': 20, 'B': 10, 'C': 10, 'D': 15, 'E': 15})
        
        return self.save_to_bytes()
    
    def export_supplier_delivery_report(self, data: Dict) -> bytes:
        """导出供应商交期报表"""
        ws = self.create_sheet('供应商交期统计')
        
        # 标题
        self.write_title(ws, f"供应商交期统计报表 ({data['period']})", col_span=9)
        
        # 表头
        headers = ['排名', '供应商名称', '订单总数', '准时订单', '延迟订单', 
                   '准时率(%)', '平均延迟(天)', '趋势', '评级']
        self.write_header(ws, headers, row=3)
        
        # 数据
        for idx, supplier in enumerate(data['suppliers'], 4):
            trend_text = '↑' if supplier['trend'] == 'up' else ('↓' if supplier['trend'] == 'down' else '→')
            row_values = [
                idx - 3,
                supplier['name'],
                supplier['total_orders'],
                supplier['on_time_orders'],
                supplier['delayed_orders'],
                supplier['on_time_rate'],
                supplier['avg_delay_days'],
                trend_text,
                supplier['rating']
            ]
            self.write_data_row(ws, row_values, idx, 
                ['number', '', 'number', 'number', 'number', 'decimal', 'decimal', '', ''])
            
            # 准时率颜色
            rate_cell = ws.cell(row=idx, column=6)
            if supplier['on_time_rate'] >= 90:
                self.apply_status_color(rate_cell, 'complete')
            elif supplier['on_time_rate'] >= 80:
                self.apply_status_color(rate_cell, 'partial')
            else:
                self.apply_status_color(rate_cell, 'shortage')
        
        # 延迟订单明细
        ws2 = self.create_sheet('延迟订单明细')
        self.write_title(ws2, '延迟订单明细', col_span=10)
        
        headers2 = ['采购单号', '供应商', '物料编码', '物料名称', '订购数量', 
                    '承诺交期', '实际到货', '延迟天数', '延迟原因', '关联项目']
        self.write_header(ws2, headers2, row=3)
        
        for idx, order in enumerate(data['delayed_orders'], 4):
            self.write_data_row(ws2, [
                order['po_no'],
                order['supplier_name'],
                order['material_code'],
                order['material_name'],
                order['order_qty'],
                order['promised_date'],
                order['actual_date'],
                order['delay_days'],
                order['delay_reason'],
                order['project_name']
            ], idx, ['', '', '', '', 'number', 'date', 'date', 'number', '', ''])
        
        self.set_column_widths(ws, {'A': 6, 'B': 25, 'C': 10, 'D': 10, 'E': 10, 'F': 12, 'G': 12, 'H': 8, 'I': 8})
        self.set_column_widths(ws2, {'A': 15, 'B': 20, 'C': 15, 'D': 20, 'E': 10, 'F': 12, 'G': 12, 'H': 10, 'I': 20, 'J': 25})
        
        return self.save_to_bytes()


# ===========================================
# 项目报表导出
# ===========================================

class ProjectReportExporter(ExcelExporter):
    """项目报表导出器"""
    
    def export_project_overview(self, data: Dict) -> bytes:
        """导出项目总览报表"""
        ws = self.create_sheet('项目总览')
        
        # 标题
        self.write_title(ws, f"项目总览报表 ({data['year']}年)", col_span=10)
        
        # 汇总统计
        ws.cell(row=2, column=1, value=f"项目总数: {data['summary']['total']}  "
                f"进行中: {data['summary']['active']}  "
                f"已完成: {data['summary']['completed']}  "
                f"延期: {data['summary']['delayed']}")
        
        # 表头
        headers = ['项目编号', '项目名称', '客户', '项目经理', '计划开始', '计划结束',
                   '当前进度(%)', '状态', '健康度', '备注']
        self.write_header(ws, headers, row=4)
        
        status_map = {'planning': '规划中', 'active': '进行中', 'completed': '已完成',
                      'suspended': '暂停', 'delayed': '延期'}
        health_map = {'good': '良好', 'warning': '警告', 'danger': '危险'}
        
        for idx, project in enumerate(data['projects'], 5):
            row_values = [
                project['project_no'],
                project['project_name'],
                project['customer_name'],
                project['pm_name'],
                project['plan_start_date'],
                project['plan_end_date'],
                project['progress'],
                status_map.get(project['status'], project['status']),
                health_map.get(project['health'], project['health']),
                project.get('remark', '')
            ]
            self.write_data_row(ws, row_values, idx, 
                ['', '', '', '', 'date', 'date', 'decimal', '', '', ''])
            
            # 状态颜色
            status_cell = ws.cell(row=idx, column=8)
            self.apply_status_color(status_cell, project['status'])
            
            # 健康度颜色
            health_cell = ws.cell(row=idx, column=9)
            if project['health'] == 'good':
                self.apply_status_color(health_cell, 'complete')
            elif project['health'] == 'warning':
                self.apply_status_color(health_cell, 'partial')
            else:
                self.apply_status_color(health_cell, 'shortage')
        
        self.set_column_widths(ws, {
            'A': 15, 'B': 30, 'C': 20, 'D': 12, 'E': 12, 
            'F': 12, 'G': 12, 'H': 10, 'I': 10, 'J': 20
        })
        
        return self.save_to_bytes()
    
    def export_progress_report(self, data: Dict) -> bytes:
        """导出进度分析报表"""
        ws = self.create_sheet('进度分析')
        
        self.write_title(ws, '项目进度对比分析', col_span=8)
        
        headers = ['项目名称', '计划进度(%)', '实际进度(%)', '偏差(%)', 
                   '计划完成', '预计完成', '延期天数', '风险等级']
        self.write_header(ws, headers, row=3)
        
        for idx, project in enumerate(data['projects'], 4):
            deviation = project['actual_progress'] - project['plan_progress']
            row_values = [
                project['project_name'],
                project['plan_progress'],
                project['actual_progress'],
                deviation,
                project['plan_end_date'],
                project['forecast_end_date'],
                project['delay_days'],
                project['risk_level']
            ]
            self.write_data_row(ws, row_values, idx,
                ['', 'decimal', 'decimal', 'decimal', 'date', 'date', 'number', ''])
            
            # 偏差颜色
            deviation_cell = ws.cell(row=idx, column=4)
            if deviation >= 0:
                self.apply_status_color(deviation_cell, 'complete')
            elif deviation >= -10:
                self.apply_status_color(deviation_cell, 'partial')
            else:
                self.apply_status_color(deviation_cell, 'shortage')
        
        self.set_column_widths(ws, {
            'A': 30, 'B': 12, 'C': 12, 'D': 10, 'E': 12, 'F': 12, 'G': 10, 'H': 10
        })
        
        return self.save_to_bytes()
    
    def export_workload_report(self, data: Dict) -> bytes:
        """导出工时报表"""
        ws = self.create_sheet('工时统计')
        
        self.write_title(ws, f"工时统计报表 ({data['period']})", col_span=8)
        
        # 汇总
        ws.cell(row=2, column=1, value=f"总工时: {data['summary']['total_hours']}h  "
                f"计划工时: {data['summary']['plan_hours']}h  "
                f"加班工时: {data['summary']['overtime_hours']}h  "
                f"利用率: {data['summary']['utilization']}%")
        
        # 部门工时
        ws.cell(row=4, column=1, value='部门工时统计').font = Font(bold=True)
        headers = ['部门', '人数', '计划工时', '实际工时', '加班工时', '利用率(%)', '人均工时']
        self.write_header(ws, headers, row=5)
        
        for idx, dept in enumerate(data['by_department'], 6):
            self.write_data_row(ws, [
                dept['name'],
                dept['headcount'],
                dept['plan_hours'],
                dept['actual_hours'],
                dept['overtime_hours'],
                dept['utilization'],
                dept['avg_hours']
            ], idx, ['', 'number', 'number', 'number', 'number', 'decimal', 'decimal'])
        
        # 人员明细
        ws2 = self.create_sheet('人员工时明细')
        self.write_title(ws2, '人员工时明细', col_span=9)
        
        headers2 = ['姓名', '部门', '项目', '计划工时', '实际工时', '加班工时', 
                    '利用率(%)', '本周工时', '本月工时']
        self.write_header(ws2, headers2, row=3)
        
        for idx, person in enumerate(data['by_person'], 4):
            self.write_data_row(ws2, [
                person['name'],
                person['department'],
                person['project_name'],
                person['plan_hours'],
                person['actual_hours'],
                person['overtime_hours'],
                person['utilization'],
                person['week_hours'],
                person['month_hours']
            ], idx, ['', '', '', 'number', 'number', 'number', 'decimal', 'number', 'number'])
        
        self.set_column_widths(ws, {'A': 15, 'B': 8, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})
        self.set_column_widths(ws2, {'A': 10, 'B': 12, 'C': 25, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10})
        
        return self.save_to_bytes()


# ===========================================
# 生产报表导出
# ===========================================

class ProductionReportExporter(ExcelExporter):
    """生产报表导出器"""
    
    def export_production_report(self, data: Dict) -> bytes:
        """导出生产统计报表"""
        ws = self.create_sheet('生产统计')
        
        self.write_title(ws, f"生产统计报表 ({data['period']})", col_span=10)
        
        # 汇总
        ws.cell(row=2, column=1, value=f"工单总数: {data['summary']['total_orders']}  "
                f"完成: {data['summary']['completed']}  "
                f"完成率: {data['summary']['completion_rate']}%  "
                f"准时率: {data['summary']['on_time_rate']}%")
        
        # 车间统计
        headers = ['车间', '工单数', '完成数', '进行中', '延期', '完成率(%)', '准时率(%)', '产能利用率(%)']
        self.write_header(ws, headers, row=4)
        
        for idx, ws_data in enumerate(data['by_workshop'], 5):
            self.write_data_row(ws, [
                ws_data['name'],
                ws_data['total'],
                ws_data['completed'],
                ws_data['in_progress'],
                ws_data['delayed'],
                ws_data['completion_rate'],
                ws_data['on_time_rate'],
                ws_data['capacity_rate']
            ], idx, ['', 'number', 'number', 'number', 'number', 'decimal', 'decimal', 'decimal'])
        
        # 工单明细
        ws2 = self.create_sheet('工单明细')
        self.write_title(ws2, '工单完成明细', col_span=12)
        
        headers2 = ['工单号', '项目', '任务名称', '车间', '负责人', '计划开始', '计划结束',
                    '实际开始', '实际结束', '状态', '进度(%)', '备注']
        self.write_header(ws2, headers2, row=3)
        
        status_map = {'pending': '待开始', 'in_progress': '进行中', 'completed': '已完成',
                      'suspended': '暂停', 'delayed': '延期'}
        
        for idx, order in enumerate(data['work_orders'], 4):
            self.write_data_row(ws2, [
                order['work_order_no'],
                order['project_name'],
                order['task_name'],
                order['workshop_name'],
                order['assignee_name'],
                order['plan_start_date'],
                order['plan_end_date'],
                order.get('actual_start_date', '-'),
                order.get('actual_end_date', '-'),
                status_map.get(order['status'], order['status']),
                order['progress'],
                order.get('remark', '')
            ], idx, ['', '', '', '', '', 'date', 'date', 'date', 'date', '', 'decimal', ''])
            
            # 状态颜色
            status_cell = ws2.cell(row=idx, column=10)
            self.apply_status_color(status_cell, order['status'])
        
        self.set_column_widths(ws, {'A': 12, 'B': 10, 'C': 10, 'D': 10, 'E': 8, 'F': 12, 'G': 12, 'H': 14})
        self.set_column_widths(ws2, {
            'A': 15, 'B': 25, 'C': 20, 'D': 12, 'E': 10, 'F': 12, 
            'G': 12, 'H': 12, 'I': 12, 'J': 10, 'K': 10, 'L': 15
        })
        
        return self.save_to_bytes()
    
    def export_quality_report(self, data: Dict) -> bytes:
        """导出质量报表"""
        ws = self.create_sheet('质量统计')
        
        self.write_title(ws, f"质量统计报表 ({data['period']})", col_span=8)
        
        ws.cell(row=2, column=1, value=f"一次合格率: {data['summary']['pass_rate']}%  "
                f"返工次数: {data['summary']['rework_count']}  "
                f"返工率: {data['summary']['rework_rate']}%")
        
        # 车间质量
        headers = ['车间', '检验数', '合格数', '不合格数', '合格率(%)', '返工次数', '返工率(%)']
        self.write_header(ws, headers, row=4)
        
        for idx, ws_data in enumerate(data['by_workshop'], 5):
            self.write_data_row(ws, [
                ws_data['name'],
                ws_data['inspected'],
                ws_data['passed'],
                ws_data['failed'],
                ws_data['pass_rate'],
                ws_data['rework_count'],
                ws_data['rework_rate']
            ], idx, ['', 'number', 'number', 'number', 'decimal', 'number', 'decimal'])
        
        # 问题类型分析
        ws2 = self.create_sheet('问题类型分析')
        headers2 = ['问题类型', '发生次数', '占比(%)', '主要原因', '改进措施']
        self.write_header(ws2, headers2, row=3)
        
        for idx, problem in enumerate(data['problem_types'], 4):
            self.write_data_row(ws2, [
                problem['type'],
                problem['count'],
                problem['percent'],
                problem['main_cause'],
                problem['improvement']
            ], idx, ['', 'number', 'decimal', '', ''])
        
        self.set_column_widths(ws, {'A': 12, 'B': 10, 'C': 10, 'D': 10, 'E': 12, 'F': 10, 'G': 12})
        self.set_column_widths(ws2, {'A': 15, 'B': 12, 'C': 10, 'D': 30, 'E': 30})
        
        return self.save_to_bytes()


# ===========================================
# API接口
# ===========================================

class ReportType(str, Enum):
    KIT_RATE = "kit_rate"
    SHORTAGE_ALERT = "shortage_alert"
    SUPPLIER_DELIVERY = "supplier_delivery"
    PROJECT_OVERVIEW = "project_overview"
    PROJECT_PROGRESS = "project_progress"
    WORKLOAD = "workload"
    PRODUCTION = "production"
    QUALITY = "quality"


@router.get("/excel/{report_type}")
async def export_excel_report(
    report_type: ReportType,
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    workshop_id: Optional[int] = Query(None, description="车间ID"),
    project_id: Optional[int] = Query(None, description="项目ID"),
    supplier_id: Optional[int] = Query(None, description="供应商ID")
):
    """
    导出Excel报表
    
    report_type:
    - kit_rate: 齐套率报表
    - shortage_alert: 缺料预警报表
    - supplier_delivery: 供应商交期报表
    - project_overview: 项目总览报表
    - project_progress: 项目进度报表
    - workload: 工时报表
    - production: 生产报表
    - quality: 质量报表
    """
    
    # 根据报表类型获取数据并导出
    if report_type == ReportType.KIT_RATE:
        data = await get_kit_rate_data(start_date, end_date, workshop_id)
        exporter = ShortageReportExporter()
        content = exporter.export_kit_rate_report(data)
        filename = f"齐套率报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.SHORTAGE_ALERT:
        data = await get_shortage_alert_data(start_date, end_date, workshop_id)
        exporter = ShortageReportExporter()
        content = exporter.export_shortage_alert_report(data)
        filename = f"缺料预警报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.SUPPLIER_DELIVERY:
        data = await get_supplier_delivery_data(start_date, end_date, supplier_id)
        exporter = ShortageReportExporter()
        content = exporter.export_supplier_delivery_report(data)
        filename = f"供应商交期报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.PROJECT_OVERVIEW:
        data = await get_project_overview_data(start_date.year)
        exporter = ProjectReportExporter()
        content = exporter.export_project_overview(data)
        filename = f"项目总览报表_{start_date.year}.xlsx"
        
    elif report_type == ReportType.PROJECT_PROGRESS:
        data = await get_project_progress_data(project_id)
        exporter = ProjectReportExporter()
        content = exporter.export_progress_report(data)
        filename = f"项目进度报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.WORKLOAD:
        data = await get_workload_data(start_date, end_date)
        exporter = ProjectReportExporter()
        content = exporter.export_workload_report(data)
        filename = f"工时报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.PRODUCTION:
        data = await get_production_data(start_date, end_date, workshop_id)
        exporter = ProductionReportExporter()
        content = exporter.export_production_report(data)
        filename = f"生产报表_{start_date}_{end_date}.xlsx"
        
    elif report_type == ReportType.QUALITY:
        data = await get_quality_data(start_date, end_date, workshop_id)
        exporter = ProductionReportExporter()
        content = exporter.export_quality_report(data)
        filename = f"质量报表_{start_date}_{end_date}.xlsx"
    
    else:
        raise HTTPException(status_code=400, detail="不支持的报表类型")
    
    # 返回文件
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


# ===========================================
# 数据获取函数（示例）
# ===========================================

async def get_kit_rate_data(start_date: date, end_date: date, workshop_id: Optional[int]) -> Dict:
    """获取齐套率数据"""
    # 实际项目中从数据库查询
    return {
        'period': '2025年1月',
        'start_date': str(start_date),
        'end_date': str(end_date),
        'summary': {
            'total_orders': 156,
            'complete_orders': 128,
            'partial_orders': 15,
            'shortage_orders': 13,
            'avg_kit_rate': 87.5
        },
        'trend_data': [
            {'date': '2025-01-01', 'total': 20, 'complete': 17, 'partial': 2, 'shortage': 1, 'kit_rate': 85.0, 'mom': 2.5, 'yoy': 5.0},
            {'date': '2025-01-02', 'total': 22, 'complete': 19, 'partial': 2, 'shortage': 1, 'kit_rate': 86.4, 'mom': 1.4, 'yoy': 4.2},
            {'date': '2025-01-03', 'total': 25, 'complete': 22, 'partial': 2, 'shortage': 1, 'kit_rate': 88.0, 'mom': 1.6, 'yoy': 6.0},
        ],
        'by_workshop': [
            {'name': '装配车间', 'total': 50, 'complete': 46, 'shortage': 4, 'kit_rate': 92.0},
            {'name': '机加车间', 'total': 45, 'complete': 40, 'shortage': 5, 'kit_rate': 88.9},
            {'name': '调试车间', 'total': 35, 'complete': 28, 'shortage': 7, 'kit_rate': 80.0},
            {'name': '电气车间', 'total': 26, 'complete': 20, 'shortage': 6, 'kit_rate': 76.9},
        ]
    }


async def get_shortage_alert_data(start_date: date, end_date: date, workshop_id: Optional[int]) -> Dict:
    """获取缺料预警数据"""
    return {
        'start_date': str(start_date),
        'end_date': str(end_date),
        'summary': {
            'total': 45,
            'pending': 8,
            'handling': 12,
            'resolved': 25
        },
        'alerts': [
            {
                'alert_no': 'ALT-20250103-001',
                'alert_level': 'level3',
                'work_order_no': 'WO-0103-001',
                'project_name': 'XX汽车传感器测试设备',
                'material_code': 'M-0123',
                'material_name': '传动轴',
                'shortage_qty': 2,
                'required_date': '2025-01-03',
                'status': 'handling',
                'handler_name': '张采购',
                'response_time': '2025-01-03 09:30:00',
                'resolve_time': None
            }
        ],
        'by_level': [
            {'level': 'level1', 'count': 20, 'percent': 44.4, 'avg_response_time': '4.5h', 'avg_resolve_time': '12h'},
            {'level': 'level2', 'count': 15, 'percent': 33.3, 'avg_response_time': '2.1h', 'avg_resolve_time': '6h'},
            {'level': 'level3', 'count': 8, 'percent': 17.8, 'avg_response_time': '0.8h', 'avg_resolve_time': '3h'},
            {'level': 'level4', 'count': 2, 'percent': 4.4, 'avg_response_time': '0.3h', 'avg_resolve_time': '1.5h'}
        ],
        'by_reason': [
            {'reason': '供应商延迟', 'count': 18, 'percent': 40},
            {'reason': '需求变更', 'count': 12, 'percent': 26.7},
            {'reason': '库存不准', 'count': 8, 'percent': 17.8},
            {'reason': '质量问题', 'count': 5, 'percent': 11.1},
            {'reason': '其他', 'count': 2, 'percent': 4.4}
        ]
    }


async def get_supplier_delivery_data(start_date: date, end_date: date, supplier_id: Optional[int]) -> Dict:
    """获取供应商交期数据"""
    return {
        'period': f'{start_date} 至 {end_date}',
        'suppliers': [
            {'name': '西门子代理', 'total_orders': 15, 'on_time_orders': 14, 'delayed_orders': 1, 
             'on_time_rate': 93.3, 'avg_delay_days': 0.5, 'trend': 'up', 'rating': 'A'},
            {'name': 'ZZ自动化', 'total_orders': 22, 'on_time_orders': 20, 'delayed_orders': 2, 
             'on_time_rate': 90.9, 'avg_delay_days': 1.2, 'trend': 'stable', 'rating': 'A'},
            {'name': 'BB五金', 'total_orders': 22, 'on_time_orders': 16, 'delayed_orders': 6, 
             'on_time_rate': 72.7, 'avg_delay_days': 3.5, 'trend': 'down', 'rating': 'C'}
        ],
        'delayed_orders': [
            {
                'po_no': 'PO-20250101-001',
                'supplier_name': 'BB五金',
                'material_code': 'M-0456',
                'material_name': '联轴器',
                'order_qty': 10,
                'promised_date': '2025-01-02',
                'actual_date': '2025-01-05',
                'delay_days': 3,
                'delay_reason': '原材料短缺',
                'project_name': 'YY新能源电池检测线'
            }
        ]
    }


async def get_project_overview_data(year: int) -> Dict:
    """获取项目总览数据"""
    return {
        'year': year,
        'summary': {'total': 45, 'active': 18, 'completed': 22, 'delayed': 5},
        'projects': [
            {
                'project_no': 'PRJ-2025-001',
                'project_name': 'XX汽车传感器测试设备',
                'customer_name': 'XX汽车',
                'pm_name': '张工',
                'plan_start_date': '2024-11-01',
                'plan_end_date': '2025-02-28',
                'progress': 75,
                'status': 'active',
                'health': 'good',
                'remark': ''
            }
        ]
    }


async def get_project_progress_data(project_id: Optional[int]) -> Dict:
    """获取项目进度数据"""
    return {
        'projects': [
            {
                'project_name': 'XX汽车传感器测试设备',
                'plan_progress': 80,
                'actual_progress': 75,
                'plan_end_date': '2025-02-28',
                'forecast_end_date': '2025-03-05',
                'delay_days': 5,
                'risk_level': '中'
            }
        ]
    }


async def get_workload_data(start_date: date, end_date: date) -> Dict:
    """获取工时数据"""
    return {
        'period': f'{start_date} 至 {end_date}',
        'summary': {
            'total_hours': 12580,
            'plan_hours': 11500,
            'overtime_hours': 1080,
            'utilization': 85
        },
        'by_department': [
            {'name': '机械部', 'headcount': 25, 'plan_hours': 4000, 'actual_hours': 4200, 
             'overtime_hours': 200, 'utilization': 105, 'avg_hours': 168},
        ],
        'by_person': [
            {'name': '张三', 'department': '机械部', 'project_name': 'XX汽车项目',
             'plan_hours': 160, 'actual_hours': 175, 'overtime_hours': 15,
             'utilization': 109, 'week_hours': 45, 'month_hours': 175}
        ]
    }


async def get_production_data(start_date: date, end_date: date, workshop_id: Optional[int]) -> Dict:
    """获取生产数据"""
    return {
        'period': f'{start_date} 至 {end_date}',
        'summary': {
            'total_orders': 156,
            'completed': 128,
            'completion_rate': 82,
            'on_time_rate': 88
        },
        'by_workshop': [
            {'name': '装配车间', 'total': 52, 'completed': 45, 'in_progress': 5, 
             'delayed': 2, 'completion_rate': 86.5, 'on_time_rate': 91.1, 'capacity_rate': 90}
        ],
        'work_orders': [
            {
                'work_order_no': 'WO-0103-001',
                'project_name': 'XX汽车传感器测试设备',
                'task_name': '支架装配',
                'workshop_name': '装配车间',
                'assignee_name': '李师傅',
                'plan_start_date': '2025-01-02',
                'plan_end_date': '2025-01-05',
                'actual_start_date': '2025-01-02',
                'actual_end_date': None,
                'status': 'in_progress',
                'progress': 60,
                'remark': ''
            }
        ]
    }


async def get_quality_data(start_date: date, end_date: date, workshop_id: Optional[int]) -> Dict:
    """获取质量数据"""
    return {
        'period': f'{start_date} 至 {end_date}',
        'summary': {
            'pass_rate': 94,
            'rework_count': 18,
            'rework_rate': 6
        },
        'by_workshop': [
            {'name': '装配车间', 'inspected': 500, 'passed': 485, 'failed': 15,
             'pass_rate': 97, 'rework_count': 5, 'rework_rate': 1}
        ],
        'problem_types': [
            {'type': '尺寸超差', 'count': 8, 'percent': 35, 
             'main_cause': '刀具磨损', 'improvement': '增加刀具检测频率'},
            {'type': '装配不良', 'count': 5, 'percent': 22, 
             'main_cause': '操作不规范', 'improvement': '加强培训'}
        ]
    }

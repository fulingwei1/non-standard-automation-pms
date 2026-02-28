"""
数据导入工具
支持从Excel批量导入项目、任务、人员、工时等数据
"""
import os
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
from enum import Enum
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment


class ImportType(Enum):
    """导入类型"""
    PROJECT = "project"
    TASK = "task"
    USER = "user"
    TIMESHEET = "timesheet"
    CUSTOMER = "customer"
    DEPARTMENT = "department"


@dataclass
class ImportResult:
    """导入结果"""
    success: bool = True
    total_rows: int = 0
    success_count: int = 0
    error_count: int = 0
    skip_count: int = 0
    errors: List[Dict] = field(default_factory=list)
    warnings: List[Dict] = field(default_factory=list)
    imported_ids: List[int] = field(default_factory=list)
    
    def add_error(self, row: int, field: str, message: str):
        self.error_count += 1
        self.errors.append({
            "row": row,
            "field": field,
            "message": message
        })
    
    def add_warning(self, row: int, field: str, message: str):
        self.warnings.append({
            "row": row,
            "field": field,
            "message": message
        })
    
    def to_dict(self) -> Dict:
        return {
            "success": self.success and self.error_count == 0,
            "total_rows": self.total_rows,
            "success_count": self.success_count,
            "error_count": self.error_count,
            "skip_count": self.skip_count,
            "errors": self.errors[:100],  # 最多返回100条错误
            "warnings": self.warnings[:50],
            "imported_ids": self.imported_ids[:100]
        }


class DataValidator:
    """数据验证器"""
    
    @staticmethod
    def validate_required(value: Any, field_name: str) -> Optional[str]:
        """验证必填字段"""
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return f"{field_name}不能为空"
        if pd.isna(value):
            return f"{field_name}不能为空"
        return None
    
    @staticmethod
    def validate_date(value: Any, field_name: str) -> Optional[str]:
        """验证日期格式"""
        if value is None or pd.isna(value):
            return None
        if isinstance(value, (date, datetime)):
            return None
        if isinstance(value, str):
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                try:
                    datetime.strptime(value.strip(), fmt)
                    return None
                except ValueError:
                    continue
        return f"{field_name}日期格式不正确，应为YYYY-MM-DD"
    
    @staticmethod
    def validate_number(value: Any, field_name: str, 
                       min_val: float = None, max_val: float = None,
                       allow_decimal: bool = True) -> Optional[str]:
        """验证数字"""
        if value is None or pd.isna(value):
            return None
        try:
            num = float(value)
            if not allow_decimal and num != int(num):
                return f"{field_name}必须是整数"
            if min_val is not None and num < min_val:
                return f"{field_name}不能小于{min_val}"
            if max_val is not None and num > max_val:
                return f"{field_name}不能大于{max_val}"
            return None
        except (ValueError, TypeError):
            return f"{field_name}必须是数字"
    
    @staticmethod
    def validate_enum(value: Any, field_name: str, valid_values: List[str]) -> Optional[str]:
        """验证枚举值"""
        if value is None or pd.isna(value):
            return None
        str_value = str(value).strip()
        if str_value not in valid_values:
            return f"{field_name}值无效，可选值: {', '.join(valid_values)}"
        return None
    
    @staticmethod
    def validate_length(value: Any, field_name: str, 
                       min_length: int = None, max_length: int = None) -> Optional[str]:
        """验证长度"""
        if value is None or pd.isna(value):
            return None
        str_value = str(value)
        if min_length is not None and len(str_value) < min_length:
            return f"{field_name}长度不能少于{min_length}个字符"
        if max_length is not None and len(str_value) > max_length:
            return f"{field_name}长度不能超过{max_length}个字符"
        return None
    
    @staticmethod
    def validate_email(value: Any, field_name: str) -> Optional[str]:
        """验证邮箱格式"""
        if value is None or pd.isna(value):
            return None
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(pattern, str(value).strip()):
            return f"{field_name}格式不正确"
        return None
    
    @staticmethod
    def validate_phone(value: Any, field_name: str) -> Optional[str]:
        """验证手机号"""
        if value is None or pd.isna(value):
            return None
        import re
        phone = str(value).strip().replace('-', '').replace(' ', '')
        if not re.match(r'^1[3-9]\d{9}$', phone):
            return f"{field_name}格式不正确"
        return None


class ExcelTemplateGenerator:
    """Excel模板生成器"""
    
    # 样式定义
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    REQUIRED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 模板配置
    TEMPLATES = {
        ImportType.PROJECT: {
            "sheet_name": "项目导入",
            "columns": [
                {"field": "project_code", "title": "项目编号*", "width": 15, "required": True, "comment": "唯一标识，如：PRJ2025001"},
                {"field": "project_name", "title": "项目名称*", "width": 30, "required": True, "comment": "项目全称"},
                {"field": "customer_name", "title": "客户名称", "width": 20, "required": False, "comment": "客户公司名称"},
                {"field": "project_level", "title": "项目级别", "width": 10, "required": False, "comment": "A/B/C/D", "validation": ["A", "B", "C", "D"]},
                {"field": "pm_name", "title": "项目经理", "width": 12, "required": False, "comment": "项目经理姓名"},
                {"field": "te_name", "title": "技术负责人", "width": 12, "required": False, "comment": "技术负责人姓名"},
                {"field": "plan_start_date", "title": "计划开始日期", "width": 15, "required": False, "comment": "格式：2025-01-01"},
                {"field": "plan_end_date", "title": "计划结束日期", "width": 15, "required": False, "comment": "格式：2025-06-30"},
                {"field": "plan_manhours", "title": "计划工时", "width": 12, "required": False, "comment": "单位：人时"},
                {"field": "budget_amount", "title": "预算金额", "width": 15, "required": False, "comment": "单位：元"},
                {"field": "description", "title": "项目描述", "width": 40, "required": False, "comment": "项目简要说明"},
                {"field": "remark", "title": "备注", "width": 30, "required": False, "comment": "其他说明"},
            ],
            "sample_data": [
                ["PRJ2025001", "XX公司自动化测试设备", "XX科技有限公司", "B", "张经理", "李工", "2025-01-15", "2025-04-30", 800, 500000, "自动化功能测试设备", "首期合作"],
                ["PRJ2025002", "YY产线改造项目", "YY电子", "C", "王经理", "赵工", "2025-02-01", "2025-03-31", 400, 200000, "产线自动化改造", ""],
            ]
        },
        ImportType.TASK: {
            "sheet_name": "任务导入",
            "columns": [
                {"field": "project_code", "title": "项目编号*", "width": 15, "required": True, "comment": "所属项目编号"},
                {"field": "wbs_code", "title": "WBS编码*", "width": 15, "required": True, "comment": "任务编码，如：1.1.1"},
                {"field": "task_name", "title": "任务名称*", "width": 30, "required": True, "comment": "任务名称"},
                {"field": "phase", "title": "所属阶段", "width": 12, "required": False, "comment": "项目阶段", 
                 "validation": ["立项启动", "方案设计", "结构设计", "电气设计", "采购制造", "装配调试", "验收交付"]},
                {"field": "parent_wbs_code", "title": "父任务编码", "width": 15, "required": False, "comment": "上级任务WBS编码"},
                {"field": "task_type", "title": "任务类型", "width": 10, "required": False, "comment": "task/milestone", "validation": ["task", "milestone"]},
                {"field": "owner_name", "title": "负责人", "width": 12, "required": False, "comment": "任务负责人姓名"},
                {"field": "plan_start_date", "title": "计划开始", "width": 12, "required": False, "comment": "格式：2025-01-01"},
                {"field": "plan_end_date", "title": "计划结束", "width": 12, "required": False, "comment": "格式：2025-01-15"},
                {"field": "plan_manhours", "title": "计划工时", "width": 10, "required": False, "comment": "单位：人时"},
                {"field": "weight", "title": "权重%", "width": 8, "required": False, "comment": "0-100"},
                {"field": "deliverable", "title": "交付物", "width": 25, "required": False, "comment": "任务交付物"},
                {"field": "predecessors", "title": "前置任务", "width": 15, "required": False, "comment": "前置任务WBS，多个用逗号分隔"},
                {"field": "remark", "title": "备注", "width": 25, "required": False, "comment": ""},
            ],
            "sample_data": [
                ["PRJ2025001", "1", "立项启动", "立项启动", "", "task", "张经理", "2025-01-15", "2025-01-17", 24, 5, "立项报告", "", ""],
                ["PRJ2025001", "1.1", "需求确认", "立项启动", "1", "task", "张经理", "2025-01-15", "2025-01-16", 16, 3, "需求确认单", "", ""],
                ["PRJ2025001", "1.2", "立项评审", "立项启动", "1", "milestone", "张经理", "2025-01-17", "2025-01-17", 8, 2, "评审记录", "1.1", "里程碑"],
            ]
        },
        ImportType.USER: {
            "sheet_name": "人员导入",
            "columns": [
                {"field": "employee_code", "title": "工号*", "width": 12, "required": True, "comment": "员工工号"},
                {"field": "user_name", "title": "姓名*", "width": 12, "required": True, "comment": "员工姓名"},
                {"field": "dept_name", "title": "部门", "width": 15, "required": False, "comment": "所属部门"},
                {"field": "position", "title": "岗位", "width": 15, "required": False, "comment": "岗位名称"},
                {"field": "role", "title": "角色", "width": 12, "required": False, "comment": "系统角色",
                 "validation": ["admin", "pm", "te", "me_leader", "ee_leader", "te_leader", "engineer", "buyer", "viewer"]},
                {"field": "email", "title": "邮箱", "width": 25, "required": False, "comment": "公司邮箱"},
                {"field": "mobile", "title": "手机", "width": 15, "required": False, "comment": "手机号码"},
                {"field": "wechat_userid", "title": "企微ID", "width": 20, "required": False, "comment": "企业微信用户ID"},
                {"field": "status", "title": "状态", "width": 10, "required": False, "comment": "在职/离职", "validation": ["在职", "离职", "试用"]},
            ],
            "sample_data": [
                ["E001", "张三", "机械组", "机械工程师", "engineer", "zhangsan@company.com", "13800138001", "zhangsan", "在职"],
                ["E002", "李四", "电气组", "电气工程师", "engineer", "lisi@company.com", "13800138002", "lisi", "在职"],
            ]
        },
        ImportType.TIMESHEET: {
            "sheet_name": "工时导入",
            "columns": [
                {"field": "employee_code", "title": "工号*", "width": 12, "required": True, "comment": "员工工号"},
                {"field": "project_code", "title": "项目编号*", "width": 15, "required": True, "comment": "项目编号"},
                {"field": "task_wbs_code", "title": "任务编码", "width": 15, "required": False, "comment": "WBS编码"},
                {"field": "work_date", "title": "日期*", "width": 12, "required": True, "comment": "格式：2025-01-15"},
                {"field": "hours", "title": "工时*", "width": 10, "required": True, "comment": "0-24小时"},
                {"field": "work_content", "title": "工作内容", "width": 40, "required": False, "comment": "工作内容描述"},
                {"field": "overtime_type", "title": "类型", "width": 10, "required": False, "comment": "正常/加班", "validation": ["正常", "加班", "调休"]},
            ],
            "sample_data": [
                ["E001", "PRJ2025001", "1.1", "2025-01-15", 8, "完成需求分析文档", "正常"],
                ["E001", "PRJ2025001", "1.1", "2025-01-16", 6, "参加评审会议", "正常"],
                ["E001", "PRJ2025001", "1.1", "2025-01-16", 2, "修改评审意见", "加班"],
            ]
        },
        ImportType.CUSTOMER: {
            "sheet_name": "客户导入",
            "columns": [
                {"field": "customer_code", "title": "客户编号", "width": 15, "required": False, "comment": "可自动生成"},
                {"field": "customer_name", "title": "客户名称*", "width": 30, "required": True, "comment": "客户公司全称"},
                {"field": "short_name", "title": "简称", "width": 15, "required": False, "comment": "客户简称"},
                {"field": "contact_name", "title": "联系人", "width": 12, "required": False, "comment": "主要联系人"},
                {"field": "contact_phone", "title": "联系电话", "width": 15, "required": False, "comment": "联系电话"},
                {"field": "address", "title": "地址", "width": 40, "required": False, "comment": "公司地址"},
                {"field": "remark", "title": "备注", "width": 25, "required": False, "comment": ""},
            ],
            "sample_data": [
                ["CUS001", "深圳XX科技有限公司", "XX科技", "王总", "13800138000", "深圳市南山区XX路XX号", "重点客户"],
                ["CUS002", "东莞YY电子有限公司", "YY电子", "李经理", "13900139000", "东莞市长安镇XX工业区", ""],
            ]
        },
        ImportType.DEPARTMENT: {
            "sheet_name": "部门导入",
            "columns": [
                {"field": "dept_code", "title": "部门编码", "width": 15, "required": False, "comment": "可自动生成"},
                {"field": "dept_name", "title": "部门名称*", "width": 20, "required": True, "comment": "部门名称"},
                {"field": "parent_dept_name", "title": "上级部门", "width": 20, "required": False, "comment": "上级部门名称"},
                {"field": "leader_name", "title": "部门负责人", "width": 12, "required": False, "comment": "负责人姓名"},
                {"field": "sort_order", "title": "排序", "width": 8, "required": False, "comment": "显示顺序"},
            ],
            "sample_data": [
                ["D001", "技术中心", "", "技术总监", 1],
                ["D002", "机械组", "技术中心", "机械主管", 1],
                ["D003", "电气组", "技术中心", "电气主管", 2],
                ["D004", "测试组", "技术中心", "测试主管", 3],
            ]
        }
    }
    
    def generate_template(self, import_type: ImportType, output_path: str) -> str:
        """生成导入模板"""
        config = self.TEMPLATES.get(import_type)
        if not config:
            raise ValueError(f"不支持的导入类型: {import_type}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = config["sheet_name"]
        
        # 写入表头
        for col_idx, col_config in enumerate(config["columns"], 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config["title"])
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            ws.column_dimensions[cell.column_letter].width = col_config["width"]
            
            # 添加批注
            if col_config.get("comment"):
                cell.comment = Comment(col_config["comment"], "系统")
            
            # 添加数据验证
            if col_config.get("validation"):
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(col_config["validation"])}"',
                    allow_blank=True
                )
                dv.add(f'{cell.column_letter}2:{cell.column_letter}1000')
                ws.add_data_validation(dv)
        
        # 写入示例数据
        for row_idx, sample_row in enumerate(config.get("sample_data", []), 2):
            for col_idx, value in enumerate(sample_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                # 必填列加背景色
                if config["columns"][col_idx-1].get("required"):
                    cell.fill = self.REQUIRED_FILL
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 添加说明sheet
        ws_help = wb.create_sheet("填写说明")
        ws_help.cell(row=1, column=1, value="导入模板填写说明").font = Font(bold=True, size=14)
        ws_help.cell(row=3, column=1, value="1. 带*的列为必填项")
        ws_help.cell(row=4, column=1, value="2. 日期格式为：YYYY-MM-DD（如2025-01-15）")
        ws_help.cell(row=5, column=1, value="3. 橙色背景的单元格为必填")
        ws_help.cell(row=6, column=1, value="4. 请勿修改表头名称")
        ws_help.cell(row=7, column=1, value="5. 示例数据可以删除或覆盖")
        ws_help.cell(row=9, column=1, value="字段说明：").font = Font(bold=True)
        
        for idx, col_config in enumerate(config["columns"], 10):
            required_mark = "（必填）" if col_config.get("required") else ""
            ws_help.cell(row=idx, column=1, value=f"• {col_config['title']}: {col_config.get('comment', '')}{required_mark}")
        
        wb.save(output_path)
        return output_path
    
    def generate_all_templates(self, output_dir: str) -> List[str]:
        """生成所有模板"""
        os.makedirs(output_dir, exist_ok=True)
        files = []
        for import_type in ImportType:
            filename = f"导入模板_{import_type.value}.xlsx"
            filepath = os.path.join(output_dir, filename)
            self.generate_template(import_type, filepath)
            files.append(filepath)
        return files


class ExcelImporter:
    """Excel数据导入器"""
    
    # 字段映射配置
    FIELD_MAPPINGS = {
        ImportType.PROJECT: {
            '项目编号': 'project_code', '项目编号*': 'project_code',
            '项目名称': 'project_name', '项目名称*': 'project_name',
            '客户名称': 'customer_name',
            '项目级别': 'project_level',
            '项目经理': 'pm_name',
            '技术负责人': 'te_name',
            '计划开始日期': 'plan_start_date', '计划开始': 'plan_start_date',
            '计划结束日期': 'plan_end_date', '计划结束': 'plan_end_date',
            '计划工时': 'plan_manhours',
            '预算金额': 'budget_amount',
            '项目描述': 'description',
            '备注': 'remark'
        },
        ImportType.TASK: {
            '项目编号': 'project_code', '项目编号*': 'project_code',
            'WBS编码': 'wbs_code', 'WBS编码*': 'wbs_code',
            '任务名称': 'task_name', '任务名称*': 'task_name',
            '所属阶段': 'phase',
            '父任务编码': 'parent_wbs_code',
            '任务类型': 'task_type',
            '负责人': 'owner_name',
            '计划开始': 'plan_start_date', '计划开始日期': 'plan_start_date',
            '计划结束': 'plan_end_date', '计划结束日期': 'plan_end_date',
            '计划工时': 'plan_manhours',
            '权重': 'weight', '权重%': 'weight',
            '交付物': 'deliverable',
            '前置任务': 'predecessors',
            '备注': 'remark'
        },
        ImportType.USER: {
            '工号': 'employee_code', '工号*': 'employee_code',
            '姓名': 'user_name', '姓名*': 'user_name',
            '部门': 'dept_name',
            '岗位': 'position',
            '角色': 'role',
            '邮箱': 'email',
            '手机': 'mobile',
            '企微ID': 'wechat_userid',
            '状态': 'status'
        },
        ImportType.TIMESHEET: {
            '工号': 'employee_code', '工号*': 'employee_code',
            '项目编号': 'project_code', '项目编号*': 'project_code',
            '任务编码': 'task_wbs_code',
            '日期': 'work_date', '日期*': 'work_date',
            '工时': 'hours', '工时*': 'hours',
            '工作内容': 'work_content',
            '类型': 'overtime_type', '加班类型': 'overtime_type'
        },
        ImportType.CUSTOMER: {
            '客户编号': 'customer_code',
            '客户名称': 'customer_name', '客户名称*': 'customer_name',
            '简称': 'short_name', '客户简称': 'short_name',
            '联系人': 'contact_name',
            '联系电话': 'contact_phone',
            '地址': 'address',
            '备注': 'remark'
        },
        ImportType.DEPARTMENT: {
            '部门编码': 'dept_code',
            '部门名称': 'dept_name', '部门名称*': 'dept_name',
            '上级部门': 'parent_dept_name',
            '部门负责人': 'leader_name',
            '排序': 'sort_order'
        }
    }
    
    # 必填字段
    REQUIRED_FIELDS = {
        ImportType.PROJECT: ['project_code', 'project_name'],
        ImportType.TASK: ['project_code', 'wbs_code', 'task_name'],
        ImportType.USER: ['employee_code', 'user_name'],
        ImportType.TIMESHEET: ['employee_code', 'project_code', 'work_date', 'hours'],
        ImportType.CUSTOMER: ['customer_name'],
        ImportType.DEPARTMENT: ['dept_name']
    }
    
    # 枚举值定义
    ENUM_VALUES = {
        'project_level': ['A', 'B', 'C', 'D'],
        'task_type': ['task', 'milestone'],
        'phase': ['立项启动', '方案设计', '结构设计', '电气设计', '采购制造', '装配调试', '验收交付'],
        'overtime_type': ['正常', '加班', '调休'],
        'status': ['在职', '离职', '试用'],
        'role': ['admin', 'gm', 'dept_manager', 'pm', 'te', 'me_leader', 'ee_leader', 'te_leader', 
                'me_engineer', 'ee_engineer', 'sw_engineer', 'te_engineer', 'buyer', 'warehouse', 
                'assembler', 'qa', 'pmc', 'sales', 'finance', 'viewer', 'engineer']
    }
    
    def __init__(self, db_session=None):
        self.db = db_session
        self.validator = DataValidator()
        self._cache = {}
    
    def import_from_excel(self, file_path: str, import_type: ImportType,
                          sheet_name: str = None, skip_rows: int = 0,
                          update_existing: bool = False) -> ImportResult:
        """
        从Excel文件导入数据
        
        Args:
            file_path: Excel文件路径
            import_type: 导入类型
            sheet_name: 工作表名称（默认第一个）
            skip_rows: 跳过的行数
            update_existing: 是否更新已存在的数据
            
        Returns:
            ImportResult: 导入结果
        """
        result = ImportResult()
        
        try:
            # 读取Excel
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, skiprows=skip_rows)
            
            # 去除空行
            df = df.dropna(how='all')
            result.total_rows = len(df)
            
            if result.total_rows == 0:
                result.add_error(0, '', '文件中没有数据')
                result.success = False
                return result
            
            # 字段映射
            field_mapping = self.FIELD_MAPPINGS.get(import_type, {})
            df = self._map_columns(df, field_mapping)
            
            # 检查必填列是否存在
            required_fields = self.REQUIRED_FIELDS.get(import_type, [])
            missing_fields = [f for f in required_fields if f not in df.columns]
            if missing_fields:
                result.add_error(0, '', f'缺少必填列: {", ".join(missing_fields)}')
                result.success = False
                return result
            
            # 数据验证和导入
            validated_data = []
            for idx, row in df.iterrows():
                row_num = idx + 2 + skip_rows  # Excel行号
                row_data = row.to_dict()
                row_errors = self._validate_row(row_data, import_type, row_num)
                
                if row_errors:
                    for err in row_errors:
                        result.add_error(err['row'], err['field'], err['message'])
                else:
                    # 清理数据
                    cleaned_data = self._clean_row_data(row_data, import_type)
                    validated_data.append(cleaned_data)
            
            # 如果有验证错误，不执行导入
            if result.error_count > 0:
                result.success = False
                return result
            
            # 执行导入
            imported_ids = self._do_import(validated_data, import_type, update_existing)
            result.success_count = len(imported_ids)
            result.imported_ids = imported_ids
            
            return result
            
        except FileNotFoundError:
            result.add_error(0, '', f'文件不存在: {file_path}')
            result.success = False
            return result
        except Exception as e:
            result.add_error(0, '', f'导入失败: {str(e)}')
            result.success = False
            return result
    
    def _map_columns(self, df: pd.DataFrame, field_mapping: Dict[str, str]) -> pd.DataFrame:
        """映射列名"""
        column_mapping = {}
        for col in df.columns:
            col_str = str(col).strip()
            if col_str in field_mapping:
                column_mapping[col] = field_mapping[col_str]
        return df.rename(columns=column_mapping)
    
    def _validate_row(self, row_data: Dict, import_type: ImportType, row_num: int) -> List[Dict]:
        """验证单行数据"""
        errors = []
        required_fields = self.REQUIRED_FIELDS.get(import_type, [])
        
        # 验证必填字段
        for field in required_fields:
            error = self.validator.validate_required(row_data.get(field), field)
            if error:
                errors.append({"row": row_num, "field": field, "message": error})
        
        # 验证日期字段
        date_fields = ['plan_start_date', 'plan_end_date', 'work_date']
        for field in date_fields:
            if field in row_data:
                error = self.validator.validate_date(row_data.get(field), field)
                if error:
                    errors.append({"row": row_num, "field": field, "message": error})
        
        # 验证数字字段
        number_validations = {
            'plan_manhours': {'min_val': 0, 'max_val': 100000},
            'hours': {'min_val': 0, 'max_val': 24},
            'weight': {'min_val': 0, 'max_val': 100},
            'budget_amount': {'min_val': 0},
            'sort_order': {'min_val': 0, 'allow_decimal': False}
        }
        for field, params in number_validations.items():
            if field in row_data and row_data.get(field) is not None:
                error = self.validator.validate_number(
                    row_data.get(field), field, 
                    params.get('min_val'), params.get('max_val'),
                    params.get('allow_decimal', True)
                )
                if error:
                    errors.append({"row": row_num, "field": field, "message": error})
        
        # 验证枚举字段
        for field, valid_values in self.ENUM_VALUES.items():
            if field in row_data and row_data.get(field) is not None and not pd.isna(row_data.get(field)):
                error = self.validator.validate_enum(row_data.get(field), field, valid_values)
                if error:
                    errors.append({"row": row_num, "field": field, "message": error})
        
        # 验证邮箱
        if 'email' in row_data:
            error = self.validator.validate_email(row_data.get('email'), 'email')
            if error:
                errors.append({"row": row_num, "field": 'email', "message": error})
        
        # 验证手机号
        if 'mobile' in row_data:
            error = self.validator.validate_phone(row_data.get('mobile'), 'mobile')
            if error:
                errors.append({"row": row_num, "field": 'mobile', "message": error})
        
        return errors
    
    def _clean_row_data(self, row_data: Dict, import_type: ImportType) -> Dict:
        """清理行数据"""
        cleaned = {}
        for key, value in row_data.items():
            if pd.isna(value):
                cleaned[key] = None
            elif isinstance(value, str):
                cleaned[key] = value.strip()
            elif isinstance(value, datetime):
                cleaned[key] = value.date()
            else:
                cleaned[key] = value
        return cleaned
    
    def _do_import(self, data_list: List[Dict], import_type: ImportType, 
                   update_existing: bool) -> List[int]:
        """执行实际导入（模拟）"""
        # 这里应该调用实际的数据库操作
        # 目前返回模拟的ID列表
        return list(range(1, len(data_list) + 1))
    
    def preview_import(self, file_path: str, import_type: ImportType,
                      sheet_name: str = None, max_rows: int = 10) -> Dict:
        """预览导入数据"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=max_rows)
            field_mapping = self.FIELD_MAPPINGS.get(import_type, {})
            df = self._map_columns(df, field_mapping)
            
            return {
                "success": True,
                "total_rows": len(df),
                "columns": list(df.columns),
                "preview_data": df.head(max_rows).to_dict('records')
            }
        except Exception as e:
            return {
                "success": False,
                "message": str(e)
            }


# 导出接口
def generate_import_template(import_type: str, output_path: str) -> str:
    """生成导入模板"""
    generator = ExcelTemplateGenerator()
    try:
        it = ImportType(import_type)
        return generator.generate_template(it, output_path)
    except ValueError:
        raise ValueError(f"不支持的导入类型: {import_type}，可选值: {[t.value for t in ImportType]}")


def import_data_from_excel(file_path: str, import_type: str, 
                           sheet_name: str = None, 
                           update_existing: bool = False) -> Dict:
    """从Excel导入数据"""
    importer = ExcelImporter()
    try:
        it = ImportType(import_type)
        result = importer.import_from_excel(file_path, it, sheet_name, 0, update_existing)
        return result.to_dict()
    except ValueError:
        return {
            "success": False,
            "message": f"不支持的导入类型: {import_type}"
        }


if __name__ == "__main__":
    # 测试生成模板
    generator = ExcelTemplateGenerator()
    
    import tempfile
    output_dir = tempfile.mkdtemp()
    
    for import_type in ImportType:
        filepath = f"{output_dir}/导入模板_{import_type.value}.xlsx"
        generator.generate_template(import_type, filepath)
        print(f"已生成: {filepath}")
    
    print(f"\n模板已生成到: {output_dir}")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成系统重复问题审查报告的Excel表格
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_duplication_report_excel():
    """创建重复问题审查报告的Excel文件"""

    # 创建工作簿
    wb = Workbook()

    # ========== Sheet 1: API端点重复问题 ==========
    ws1 = wb.active
    ws1.title = "API端点重复"

    api_duplication_data = [
        {
            "序号": 1,
            "问题类型": "供应商API重复",
            "重复位置1": "app/api/v1/endpoints/suppliers.py",
            "重复位置2": "app/api/v1/endpoints/materials/suppliers.py",
            "重复内容": "read_suppliers() - 获取供应商列表",
            "重复代码行数": "~200行",
            "严重程度": "🔴 高",
            "影响": "维护成本高，需要同步修改两处",
            "建议方案": "删除materials/suppliers.py，统一使用/suppliers/路由",
            "优先级": "P0 - 立即处理",
            "预计工作量": "2小时",
        },
        {
            "序号": 2,
            "问题类型": "项目角色API重复",
            "重复位置1": "app/api/v1/endpoints/project_roles/",
            "重复位置2": "app/api/v1/endpoints/projects/roles/",
            "重复内容": "项目角色管理功能",
            "重复代码行数": "待确认",
            "严重程度": "🟡 中",
            "影响": "路由混乱，职责边界不清",
            "建议方案": "统一为项目中心设计：/projects/{id}/roles/",
            "优先级": "P1 - 短期处理",
            "预计工作量": "4小时",
        },
        {
            "序号": 3,
            "问题类型": "里程碑API重复",
            "重复位置1": "app/api/v1/endpoints/milestones/",
            "重复位置2": "app/api/v1/endpoints/projects/milestones/",
            "重复内容": "里程碑CRUD和流程管理",
            "重复代码行数": "~300行",
            "严重程度": "🟡 中",
            "影响": "路由混乱，功能重复",
            "建议方案": "统一为项目中心设计：/projects/{id}/milestones/",
            "优先级": "P1 - 短期处理",
            "预计工作量": "4小时",
        },
    ]

    df1 = pd.DataFrame(api_duplication_data)
    write_dataframe_to_sheet(ws1, df1, "API端点重复问题汇总")

    # ========== Sheet 2: CRUD操作重复 ==========
    ws2 = wb.create_sheet("CRUD操作重复")

    crud_duplication_data = [
        {
            "问题类型": "分页查询模式",
            "重复位置": "几乎所有列表接口",
            "涉及文件数": "456个文件",
            "重复函数数": "~400个",
            "重复代码行数": "~4000行",
            "严重程度": "🟡 中",
            "重复模式": "分页计算、关键词搜索、筛选逻辑",
            "建议方案": "创建通用CRUD基类或工具函数",
            "优先级": "P1 - 短期优化",
            "预计工作量": "16小时",
            "预期收益": "减少30-40%重复代码",
        },
        {
            "问题类型": "详情查询模式",
            "重复位置": "几乎所有详情接口",
            "涉及文件数": "456个文件",
            "重复函数数": "~300个",
            "重复代码行数": "~2000行",
            "严重程度": "🟡 中",
            "重复模式": "ID查询、404检查、响应构建",
            "建议方案": "创建通用详情查询基类",
            "优先级": "P1 - 短期优化",
            "预计工作量": "8小时",
            "预期收益": "统一错误处理逻辑",
        },
        {
            "问题类型": "创建操作模式",
            "重复位置": "几乎所有创建接口",
            "涉及文件数": "456个文件",
            "重复函数数": "~300个",
            "重复代码行数": "~3000行",
            "严重程度": "🟡 中",
            "重复模式": "唯一性检查、数据创建、事务提交",
            "建议方案": "创建通用创建操作基类",
            "优先级": "P1 - 短期优化",
            "预计工作量": "12小时",
            "预期收益": "统一验证和错误处理",
        },
        {
            "问题类型": "更新操作模式",
            "重复位置": "几乎所有更新接口",
            "涉及文件数": "456个文件",
            "重复函数数": "~200个",
            "重复代码行数": "~2500行",
            "严重程度": "🟡 中",
            "重复模式": "存在性检查、字段更新、事务提交",
            "建议方案": "创建通用更新操作基类",
            "优先级": "P1 - 短期优化",
            "预计工作量": "10小时",
            "预期收益": "统一更新逻辑",
        },
        {
            "问题类型": "删除操作模式",
            "重复位置": "几乎所有删除接口",
            "涉及文件数": "456个文件",
            "重复函数数": "~97个",
            "重复代码行数": "~1500行",
            "严重程度": "🟡 中",
            "重复模式": "存在性检查、关联检查、软删除/硬删除",
            "建议方案": "创建通用删除操作基类",
            "优先级": "P1 - 短期优化",
            "预计工作量": "8小时",
            "预期收益": "统一删除策略",
        },
    ]

    df2 = pd.DataFrame(crud_duplication_data)
    write_dataframe_to_sheet(ws2, df2, "CRUD操作重复模式汇总")

    # ========== Sheet 3: 统计功能重复 ==========
    ws3 = wb.create_sheet("统计功能重复")

    statistics_duplication_data = [
        {
            "模块": "统一工作台统计",
            "文件路径": "app/api/v1/endpoints/dashboard_stats.py",
            "功能": "多角色统计数据",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "客服统计",
            "文件路径": "app/api/v1/endpoints/service/statistics.py",
            "功能": "服务工单统计",
            "重复模式": "计数统计、平均响应时间",
            "严重程度": "🟡 中",
        },
        {
            "模块": "预警统计",
            "文件路径": "app/api/v1/endpoints/alerts/statistics/",
            "功能": "预警数据统计（3个文件）",
            "重复模式": "分布统计、趋势分析、性能指标",
            "严重程度": "🟡 中",
        },
        {
            "模块": "缺料统计",
            "文件路径": "app/api/v1/endpoints/shortage/analytics/dashboard.py",
            "功能": "缺料数据分析",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "生产统计",
            "文件路径": "app/api/v1/endpoints/production/dashboard.py",
            "功能": "生产数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "售前统计",
            "文件路径": "app/api/v1/endpoints/presales_integration/dashboard.py",
            "功能": "售前数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "人员匹配统计",
            "文件路径": "app/api/v1/endpoints/staff_matching/dashboard.py",
            "功能": "人员匹配数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "商务支持统计",
            "文件路径": "app/api/v1/endpoints/business_support/dashboard.py",
            "功能": "商务支持数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "装配套件统计",
            "文件路径": "app/api/v1/endpoints/assembly_kit/dashboard.py",
            "功能": "装配套件数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "配套率统计",
            "文件路径": "app/api/v1/endpoints/kit_rate/dashboard.py",
            "功能": "配套率数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "管理节奏统计",
            "文件路径": "app/api/v1/endpoints/management_rhythm/dashboard.py",
            "功能": "管理节奏数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "人事统计",
            "文件路径": "app/api/v1/endpoints/hr_management/dashboard.py",
            "功能": "人事数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
        {
            "模块": "PMO驾驶舱",
            "文件路径": "app/api/v1/endpoints/pmo/cockpit.py",
            "功能": "PMO数据统计",
            "重复模式": "计数统计、趋势分析",
            "严重程度": "🟡 中",
        },
    ]

    df3 = pd.DataFrame(statistics_duplication_data)
    write_dataframe_to_sheet(ws3, df3, "统计功能重复汇总")

    # ========== Sheet 4: 前端组件重复 ==========
    ws4 = wb.create_sheet("前端组件重复")

    frontend_duplication_data = [
        {
            "问题类型": "数据加载模式",
            "重复位置": "几乎所有页面",
            "涉及文件数": "405个文件",
            "重复代码行数": "~8000行",
            "严重程度": "🟢 低",
            "重复模式": "useState(loading/data/error), useEffect, try-catch",
            "建议方案": "创建useDataLoader通用hook",
            "优先级": "P2 - 长期优化",
            "预计工作量": "16小时",
            "预期收益": "减少40-50%重复代码",
        },
        {
            "问题类型": "筛选和分页模式",
            "重复位置": "所有列表页面",
            "涉及文件数": "~200个文件",
            "重复代码行数": "~4000行",
            "严重程度": "🟢 低",
            "重复模式": "filters状态、pagination状态、handleFilterChange",
            "建议方案": "创建usePaginatedData通用hook",
            "优先级": "P2 - 长期优化",
            "预计工作量": "12小时",
            "预期收益": "统一分页逻辑",
        },
        {
            "问题类型": "表单处理模式",
            "重复位置": "所有表单页面",
            "涉及文件数": "~150个文件",
            "重复代码行数": "~3000行",
            "严重程度": "🟢 低",
            "重复模式": "formData状态、submitting状态、handleSubmit",
            "建议方案": "创建useForm通用hook",
            "优先级": "P2 - 长期优化",
            "预计工作量": "10小时",
            "预期收益": "统一表单处理逻辑",
        },
        {
            "问题类型": "表格展示模式",
            "重复位置": "所有列表页面",
            "涉及文件数": "~200个文件",
            "重复代码行数": "~5000行",
            "严重程度": "🟢 低",
            "重复模式": "Table组件、columns定义、rowSelection",
            "建议方案": "创建通用DataTable组件",
            "优先级": "P2 - 长期优化",
            "预计工作量": "20小时",
            "预期收益": "统一表格样式和功能",
        },
    ]

    df4 = pd.DataFrame(frontend_duplication_data)
    write_dataframe_to_sheet(ws4, df4, "前端组件重复模式汇总")

    # ========== Sheet 5: 总体统计 ==========
    ws5 = wb.create_sheet("总体统计")

    summary_data = [
        {
            "类别": "API端点重复",
            "问题数量": 3,
            "涉及文件数": "6个文件",
            "重复代码行数": "~500行",
            "严重程度": "高/中",
            "优先级": "P0-P1",
            "预计工作量": "10小时",
            "预期收益": "统一路由，降低维护成本",
        },
        {
            "类别": "CRUD操作重复",
            "问题数量": 5,
            "涉及文件数": "456个文件",
            "重复代码行数": "~13000行",
            "严重程度": "中",
            "优先级": "P1",
            "预计工作量": "54小时",
            "预期收益": "减少30-40%重复代码",
        },
        {
            "类别": "统计功能重复",
            "问题数量": 13,
            "涉及文件数": "262个文件",
            "重复代码行数": "~8000行",
            "严重程度": "中",
            "优先级": "P1",
            "预计工作量": "40小时",
            "预期收益": "减少50%重复代码",
        },
        {
            "类别": "前端组件重复",
            "问题数量": 4,
            "涉及文件数": "405个文件",
            "重复代码行数": "~20000行",
            "严重程度": "低",
            "优先级": "P2",
            "预计工作量": "58小时",
            "预期收益": "减少40-50%重复代码",
        },
        {
            "类别": "总计",
            "问题数量": 25,
            "涉及文件数": "1129个文件",
            "重复代码行数": "~41500行",
            "严重程度": "-",
            "优先级": "-",
            "预计工作量": "162小时",
            "预期收益": "减少约39%重复代码（~16200行）",
        },
    ]

    df5 = pd.DataFrame(summary_data)
    write_dataframe_to_sheet(ws5, df5, "重复问题总体统计")

    # ========== Sheet 6: 行动计划 ==========
    ws6 = wb.create_sheet("行动计划")

    action_plan_data = [
        {
            "阶段": "第一阶段（立即处理）",
            "任务": "删除materials/suppliers.py，统一使用/suppliers/路由",
            "负责人": "待分配",
            "预计完成时间": "本周",
            "优先级": "P0",
            "状态": "待开始",
        },
        {
            "阶段": "第一阶段（立即处理）",
            "任务": "审查并统一project_roles/和projects/roles/的职责",
            "负责人": "待分配",
            "预计完成时间": "本周",
            "优先级": "P0",
            "状态": "待开始",
        },
        {
            "阶段": "第一阶段（立即处理）",
            "任务": "审查并统一milestones/和projects/milestones/的职责",
            "负责人": "待分配",
            "预计完成时间": "本周",
            "优先级": "P0",
            "状态": "待开始",
        },
        {
            "阶段": "第二阶段（短期优化）",
            "任务": "创建通用CRUD基类，减少重复代码",
            "负责人": "待分配",
            "预计完成时间": "本月",
            "优先级": "P1",
            "状态": "待开始",
        },
        {
            "阶段": "第二阶段（短期优化）",
            "任务": "创建统一统计服务层",
            "负责人": "待分配",
            "预计完成时间": "本月",
            "优先级": "P1",
            "状态": "待开始",
        },
        {
            "阶段": "第二阶段（短期优化）",
            "任务": "创建通用前端hooks（useDataLoader, usePaginatedData）",
            "负责人": "待分配",
            "预计完成时间": "本月",
            "优先级": "P1",
            "状态": "待开始",
        },
        {
            "阶段": "第三阶段（长期优化）",
            "任务": "重构所有列表页面使用通用DataTable组件",
            "负责人": "待分配",
            "预计完成时间": "本季度",
            "优先级": "P2",
            "状态": "待开始",
        },
        {
            "阶段": "第三阶段（长期优化）",
            "任务": "重构所有表单页面使用通用Form组件",
            "负责人": "待分配",
            "预计完成时间": "本季度",
            "优先级": "P2",
            "状态": "待开始",
        },
        {
            "阶段": "第三阶段（长期优化）",
            "任务": "建立代码审查机制，防止新的重复代码",
            "负责人": "待分配",
            "预计完成时间": "本季度",
            "优先级": "P2",
            "状态": "待开始",
        },
    ]

    df6 = pd.DataFrame(action_plan_data)
    write_dataframe_to_sheet(ws6, df6, "重复问题优化行动计划")

    # 保存文件
    filename = f"reports/系统重复问题审查报告_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Excel报告已生成: {filename}")
    return filename


def write_dataframe_to_sheet(ws, df, title):
    """将DataFrame写入工作表并应用样式"""

    # 写入标题
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头（从第3行开始）
    header_row = 3
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # 写入数据
    for row_num, row_data in enumerate(df.values, header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 根据严重程度设置背景色
            if isinstance(value, str):
                if "🔴" in value or "高" in value:
                    cell.fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
                elif "🟡" in value or "中" in value:
                    cell.fill = PatternFill(
                        start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"
                    )
                elif "🟢" in value or "低" in value:
                    cell.fill = PatternFill(
                        start_color="E6F7E6", end_color="E6F7E6", fill_type="solid"
                    )

    # 自动调整列宽
    for col_num, column_title in enumerate(df.columns, 1):
        max_length = len(str(column_title))
        for row_num in range(header_row + 1, header_row + len(df) + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 25


if __name__ == "__main__":
    create_duplication_report_excel()

# -*- coding: utf-8 -*-
"""
生成项目四算体系设计文档 Excel 版本
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_style():
    """创建样式"""
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="left", vertical="center")

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'title_font': title_font,
        'title_alignment': title_alignment,
        'cell_alignment': cell_alignment,
        'border': thin_border
    }

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_header_row(ws, row, headers, styles):
    """添加表头行"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']

def add_data_row(ws, row, data, styles):
    """添加数据行"""
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = styles['cell_alignment']
        cell.border = styles['border']

def create_overview_sheet(wb, styles):
    """创建概述页"""
    ws = wb.active
    ws.title = "1.概述"
    set_column_widths(ws, [20, 60])

    # 标题
    ws.merge_cells('A1:B1')
    ws['A1'] = "项目四算体系设计文档"
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:B2')
    ws['A2'] = f"文档日期: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal="center")

    # 四算定义
    row = 4
    ws.cell(row=row, column=1, value="一、四算定义").font = styles['title_font']
    row += 1

    add_header_row(ws, row, ["阶段", "含义与价值"], styles)
    row += 1

    definitions = [
        ("概算", "设计项目利润的过程 - 从目标利润反推成本上限"),
        ("预算", "管理增收节支的过程 - 承接概算，细化到可执行的成本控制"),
        ("核算", "管理增收节支的过程 - 归集实际成本，实时对比预算"),
        ("决算", "传承经验的过程 - 总结经验，反馈到概算模板"),
    ]
    for item in definitions:
        add_data_row(ws, row, item, styles)
        row += 1

    # 核心理念
    row += 1
    ws.cell(row=row, column=1, value="二、核心理念").font = styles['title_font']
    row += 1

    principles = [
        ("项目是细胞", "项目是最小核算单元，项目清楚了，部门和公司就清楚了"),
        ("一把手是概算经理", "部门负责人对项目概算利润负责"),
        ("四算拉通", "概算→预算→核算→决算形成闭环，经验传承到下个项目"),
        ("预算分到前方", "项目预算与平台预算分开，机关为项目服务"),
    ]
    add_header_row(ws, row, ["理念", "说明"], styles)
    row += 1
    for item in principles:
        add_data_row(ws, row, item, styles)
        row += 1

    # 设计目标
    row += 1
    ws.cell(row=row, column=1, value="三、设计目标").font = styles['title_font']
    row += 1
    ws.cell(row=row, column=1, value="建立分层四算体系：")
    row += 1

    layers = [
        ("第一层", "公司级经营看板 - 年度目标/实际利润/概算准确率"),
        ("第二层", "部门级汇总 - 部门概算/预算/核算汇总"),
        ("第三层", "项目级四算 - 概算→预算→核算→决算完整闭环"),
        ("第四层", "机台级明细 - 单台设备成本归集"),
    ]
    add_header_row(ws, row, ["层级", "内容"], styles)
    row += 1
    for item in layers:
        add_data_row(ws, row, item, styles)
        row += 1

def create_estimate_sheet(wb, styles):
    """创建概算模块页"""
    ws = wb.create_sheet("2.概算模块")
    set_column_widths(ws, [25, 15, 50])

    row = 1
    ws.cell(row=row, column=1, value="概算模块设计").font = Font(bold=True, size=16)

    # 概算流程
    row += 2
    ws.cell(row=row, column=1, value="一、概算流程").font = styles['title_font']
    row += 1

    flow = [
        ("步骤1", "确定合同金额（或报价金额）"),
        ("步骤2", "设定目标利润率（如25%）"),
        ("步骤3", "计算允许成本上限 = 金额 × (1 - 利润率)"),
        ("步骤4", "拆解成本（物料/人工/外协/其他）"),
        ("步骤5", "概算审批 → 成为预算基准"),
    ]
    add_header_row(ws, row, ["步骤", "说明"], styles)
    row += 1
    for item in flow:
        add_data_row(ws, row, item, styles)
        row += 1

    # 数据模型
    row += 1
    ws.cell(row=row, column=1, value="二、数据模型 - project_estimates").font = styles['title_font']
    row += 1

    fields = [
        ("id", "Integer", "主键"),
        ("estimate_no", "String(50)", "概算编号（GS+年月日+序号）"),
        ("project_id", "Integer", "项目ID（可空，中标后关联）"),
        ("opportunity_id", "Integer", "商机ID"),
        ("quote_id", "Integer", "报价ID"),
        ("contract_amount", "Numeric(14,2)", "合同/报价金额"),
        ("target_margin_rate", "Numeric(5,2)", "目标利润率"),
        ("target_profit", "Numeric(14,2)", "目标利润"),
        ("allowed_cost", "Numeric(14,2)", "允许成本上限"),
        ("estimated_cost", "Numeric(14,2)", "概算成本"),
        ("estimated_profit", "Numeric(14,2)", "概算利润"),
        ("estimated_margin", "Numeric(5,2)", "概算利润率"),
        ("estimate_manager_id", "Integer", "概算经理ID"),
        ("estimate_manager_name", "String(50)", "概算经理姓名"),
        ("status", "String(20)", "状态：DRAFT/SUBMITTED/APPROVED/REJECTED"),
        ("version", "String(20)", "版本号"),
        ("approved_by", "Integer", "审批人"),
        ("approved_at", "DateTime", "审批时间"),
    ]
    add_header_row(ws, row, ["字段名", "类型", "说明"], styles)
    row += 1
    for item in fields:
        add_data_row(ws, row, item, styles)
        row += 1

    # 审批层级
    row += 1
    ws.cell(row=row, column=1, value="三、概算审批层级").font = styles['title_font']
    row += 1

    approval = [
        ("≥25%", "一级", "销售经理"),
        ("20%-25%", "二级", "部门负责人（概算经理）"),
        ("15%-20%", "三级", "销售总监"),
        ("<15%", "四级", "财务总监 + 总经理"),
    ]
    add_header_row(ws, row, ["概算利润率", "审批层级", "审批人"], styles)
    row += 1
    for item in approval:
        add_data_row(ws, row, item, styles)
        row += 1

def create_budget_sheet(wb, styles):
    """创建预算模块页"""
    ws = wb.create_sheet("3.预算模块")
    set_column_widths(ws, [25, 15, 50])

    row = 1
    ws.cell(row=row, column=1, value="预算模块设计").font = Font(bold=True, size=16)

    # 预算流程
    row += 2
    ws.cell(row=row, column=1, value="一、预算流程").font = styles['title_font']
    row += 1

    flow = [
        ("步骤1", "概算审批通过"),
        ("步骤2", "自动/手动生成预算草稿"),
        ("步骤3", "细化预算明细（按成本类别、机台分解）"),
        ("步骤4", "预算审批"),
        ("步骤5", "预算生效（控制采购/外协）"),
    ]
    add_header_row(ws, row, ["步骤", "说明"], styles)
    row += 1
    for item in flow:
        add_data_row(ws, row, item, styles)
        row += 1

    # 扩展字段
    row += 1
    ws.cell(row=row, column=1, value="二、现有表扩展字段 - project_budgets").font = styles['title_font']
    row += 1

    fields = [
        ("estimate_id", "Integer", "关联概算ID"),
        ("estimate_no", "String(50)", "概算编号（冗余）"),
        ("budget_source", "String(20)", "来源：ESTIMATE/MANUAL/TEMPLATE"),
        ("budget_category", "String(20)", "类型：PROJECT/PLATFORM/SHARED"),
        ("control_mode", "String(20)", "控制模式：HARD/SOFT/NONE"),
        ("warning_threshold", "Numeric(5,2)", "预警阈值（如0.80）"),
        ("original_amount", "Numeric(14,2)", "原始预算"),
        ("adjusted_amount", "Numeric(14,2)", "调整后预算"),
        ("adjustment_count", "Integer", "调整次数"),
    ]
    add_header_row(ws, row, ["字段名", "类型", "说明"], styles)
    row += 1
    for item in fields:
        add_data_row(ws, row, item, styles)
        row += 1

    # 预算控制
    row += 1
    ws.cell(row=row, column=1, value="三、预算控制机制").font = styles['title_font']
    row += 1

    control = [
        ("<80%", "正常提交", "正常提交"),
        ("80%-100%", "预警，可提交", "预警，可提交"),
        (">100%", "阻止提交", "预警，可提交"),
    ]
    add_header_row(ws, row, ["预算使用率", "HARD模式", "SOFT模式"], styles)
    row += 1
    for item in control:
        add_data_row(ws, row, item, styles)
        row += 1

def create_accounting_sheet(wb, styles):
    """创建核算模块页"""
    ws = wb.create_sheet("4.核算模块")
    set_column_widths(ws, [25, 15, 50])

    row = 1
    ws.cell(row=row, column=1, value="核算模块设计").font = Font(bold=True, size=16)

    # 成本归集
    row += 2
    ws.cell(row=row, column=1, value="一、成本自动归集").font = styles['title_font']
    row += 1

    sources = [
        ("采购入库", "自动归集", "物料成本"),
        ("外协验收", "自动归集", "外协成本"),
        ("工时审批", "自动计算", "人工成本（工时×时薪）"),
        ("费用报销", "自动归集", "差旅/其他费用"),
    ]
    add_header_row(ws, row, ["数据来源", "归集方式", "成本类型"], styles)
    row += 1
    for item in sources:
        add_data_row(ws, row, item, styles)
        row += 1

    # 数据模型
    row += 1
    ws.cell(row=row, column=1, value="二、数据模型 - project_cost_ledgers").font = styles['title_font']
    row += 1

    fields = [
        ("id", "Integer", "主键"),
        ("project_id", "Integer", "项目ID"),
        ("machine_id", "Integer", "机台ID（可空）"),
        ("period_type", "String(20)", "期间类型：MONTHLY/QUARTERLY/YEARLY/CUMULATIVE"),
        ("period_value", "String(20)", "期间值（如2025-01）"),
        ("material_cost", "Numeric(14,2)", "物料成本"),
        ("labor_cost", "Numeric(14,2)", "人工成本"),
        ("outsource_cost", "Numeric(14,2)", "外协成本"),
        ("travel_cost", "Numeric(14,2)", "差旅费用"),
        ("other_cost", "Numeric(14,2)", "其他费用"),
        ("total_cost", "Numeric(14,2)", "合计"),
        ("budget_amount", "Numeric(14,2)", "对应预算"),
        ("variance_amount", "Numeric(14,2)", "预算差异"),
        ("variance_rate", "Numeric(5,2)", "差异率"),
        ("status", "String(20)", "状态：DRAFT/CONFIRMED/LOCKED"),
    ]
    add_header_row(ws, row, ["字段名", "类型", "说明"], styles)
    row += 1
    for item in fields:
        add_data_row(ws, row, item, styles)
        row += 1

def create_settlement_sheet(wb, styles):
    """创建决算模块页"""
    ws = wb.create_sheet("5.决算模块")
    set_column_widths(ws, [25, 15, 50])

    row = 1
    ws.cell(row=row, column=1, value="决算模块设计").font = Font(bold=True, size=16)

    # 决算流程
    row += 2
    ws.cell(row=row, column=1, value="一、决算流程").font = styles['title_font']
    row += 1

    flow = [
        ("步骤1", "项目完工"),
        ("步骤2", "财务决算（结清账目）"),
        ("步骤3", "偏差分析（找原因）"),
        ("步骤4", "经验沉淀（入知识库）"),
        ("步骤5", "反馈概算模板"),
    ]
    add_header_row(ws, row, ["步骤", "说明"], styles)
    row += 1
    for item in flow:
        add_data_row(ws, row, item, styles)
        row += 1

    # 偏差原因分类
    row += 1
    ws.cell(row=row, column=1, value="二、偏差根本原因分类").font = styles['title_font']
    row += 1

    causes = [
        ("ESTIMATE_ERROR", "概算估计不准", "概算时对工作量、价格估计偏差"),
        ("PRICE_CHANGE", "价格变动", "物料/外协价格上涨或下降"),
        ("SCOPE_CHANGE", "范围变更", "项目范围变更（ECN）导致"),
        ("EFFICIENCY", "效率问题", "执行效率高于或低于预期"),
        ("QUALITY_ISSUE", "质量问题返工", "因质量问题导致返工"),
        ("EXTERNAL", "外部因素", "客户原因、不可抗力等"),
    ]
    add_header_row(ws, row, ["原因代码", "原因名称", "说明"], styles)
    row += 1
    for item in causes:
        add_data_row(ws, row, item, styles)
        row += 1

    # 经验传承闭环
    row += 1
    ws.cell(row=row, column=1, value="三、经验传承闭环").font = styles['title_font']
    row += 1

    feedback = [
        ("1", "决算完成", "项目完工后编制决算报告"),
        ("2", "偏差分析", "分析概算vs实际的差异及原因"),
        ("3", "经验沉淀", "将经验总结存入经验库"),
        ("4", "模板更新", "根据经验更新概算模板"),
        ("5", "新项目引用", "新项目概算时参考历史经验"),
    ]
    add_header_row(ws, row, ["序号", "环节", "说明"], styles)
    row += 1
    for item in feedback:
        add_data_row(ws, row, item, styles)
        row += 1

def create_roles_sheet(wb, styles):
    """创建角色权限页"""
    ws = wb.create_sheet("6.角色权限")
    set_column_widths(ws, [18, 18, 50])

    row = 1
    ws.cell(row=row, column=1, value="角色与权限设计").font = Font(bold=True, size=16)

    # 角色定义
    row += 2
    ws.cell(row=row, column=1, value="一、四算相关角色").font = styles['title_font']
    row += 1

    roles = [
        ("estimate_manager", "概算经理", "对项目概算负责，通常是部门负责人"),
        ("project_fc", "项目财经经理(PFC)", "项目级财务管理，负责四算执行"),
        ("dept_fc", "部门财经经理(BFC)", "部门级财务汇总与分析"),
        ("cfo", "财务总监", "公司级财务管理"),
    ]
    add_header_row(ws, row, ["角色代码", "角色名称", "职责说明"], styles)
    row += 1
    for item in roles:
        add_data_row(ws, row, item, styles)
        row += 1

    # 权限矩阵
    row += 1
    ws.cell(row=row, column=1, value="二、权限矩阵").font = styles['title_font']
    row += 1

    ws_perm = wb.create_sheet("6.1权限矩阵")
    set_column_widths(ws_perm, [20, 12, 12, 12, 12, 12])

    perm_row = 1
    add_header_row(ws_perm, perm_row, ["功能", "项目经理", "PFC", "概算经理", "BFC", "CFO"], styles)
    perm_row += 1

    permissions = [
        ("创建概算", "✓", "✓", "✓", "-", "-"),
        ("编辑概算", "✓", "✓", "✓", "-", "-"),
        ("审批概算(L1)", "-", "-", "-", "✓", "-"),
        ("审批概算(L2)", "-", "-", "✓", "-", "-"),
        ("审批概算(L3)", "-", "-", "-", "-", "✓"),
        ("创建预算", "-", "✓", "-", "-", "-"),
        ("审批预算", "-", "-", "-", "✓", "✓"),
        ("查看成本", "✓", "✓", "✓", "✓", "✓"),
        ("确认核算", "-", "✓", "-", "✓", "-"),
        ("创建决算", "-", "✓", "-", "-", "-"),
        ("审核决算", "-", "-", "✓", "✓", "-"),
        ("审批决算", "-", "-", "-", "-", "✓"),
        ("管理模板", "-", "-", "-", "-", "✓"),
    ]
    for item in permissions:
        add_data_row(ws_perm, perm_row, item, styles)
        perm_row += 1

    # 数据权限
    row += 1
    ws.cell(row=row, column=1, value="三、数据权限范围").font = styles['title_font']
    row += 1

    data_scope = [
        ("项目经理", "own_projects", "只能看自己负责的项目"),
        ("PFC", "assigned_projects", "分配给自己的项目"),
        ("BFC", "dept_projects", "部门所有项目"),
        ("CFO", "all_projects", "所有项目"),
    ]
    add_header_row(ws, row, ["角色", "数据范围", "说明"], styles)
    row += 1
    for item in data_scope:
        add_data_row(ws, row, item, styles)
        row += 1

def create_integration_sheet(wb, styles):
    """创建集成与检查点页"""
    ws = wb.create_sheet("7.LTC集成")
    set_column_widths(ws, [15, 20, 25, 20])

    row = 1
    ws.cell(row=row, column=1, value="四算与LTC业务流程集成").font = Font(bold=True, size=16)

    # 检查点
    row += 2
    ws.cell(row=row, column=1, value="一、预置检查点").font = styles['title_font']
    row += 1

    checkpoints = [
        ("CP01", "报价提交审批", "概算是否完成", "阻止提交"),
        ("CP02", "报价审批", "概算利润率是否达标", "升级审批"),
        ("CP03", "合同签订", "概算是否审批通过", "阻止签订"),
        ("CP04", "项目启动", "预算是否生成", "提醒生成"),
        ("CP05", "采购申请", "是否超预算", "预警/阻止"),
        ("CP06", "项目完工", "决算是否完成", "阻止结项"),
        ("CP07", "质保结束", "决算是否审批", "提醒完成"),
    ]
    add_header_row(ws, row, ["检查点", "触发时机", "检查内容", "失败处理"], styles)
    row += 1
    for item in checkpoints:
        add_data_row(ws, row, item, styles)
        row += 1

    # 状态流转
    row += 1
    ws.cell(row=row, column=1, value="二、四算状态流转").font = styles['title_font']
    row += 1

    states = [
        ("投标中", "概算编制中", "创建/编辑概算"),
        ("投标中", "概算待审批", "提交概算审批"),
        ("投标中", "概算已审批", "可以签合同"),
        ("已签约", "预算生成中", "概算转预算"),
        ("已签约", "预算待审批", "提交预算审批"),
        ("已签约", "预算已生效", "可以执行采购"),
        ("执行中", "核算进行中", "成本自动归集"),
        ("执行中", "核算月度确认", "月度核算确认"),
        ("已完工", "决算编制中", "创建决算报告"),
        ("已完工", "决算待审核", "提交决算审批"),
        ("已完工", "决算已完成", "可以结项"),
        ("已结项", "四算归档", "经验沉淀"),
    ]
    add_header_row(ws, row, ["项目状态", "四算状态", "允许操作"], styles)
    row += 1
    for item in states:
        add_data_row(ws, row, item, styles)
        row += 1

def create_implementation_sheet(wb, styles):
    """创建实施计划页"""
    ws = wb.create_sheet("8.实施计划")
    set_column_widths(ws, [15, 35, 12, 15])

    row = 1
    ws.cell(row=row, column=1, value="实施路径建议").font = Font(bold=True, size=16)

    # Phase 1
    row += 2
    ws.cell(row=row, column=1, value="Phase 1: 基础闭环（优先）").font = styles['title_font']
    row += 1

    phase1 = [
        ("概算模块", "简化版概算，支持目标利润设定和成本拆解", "P0", "中"),
        ("预算扩展", "概算→预算自动转化，预算审批", "P0", "小"),
        ("核算归集", "采购/外协成本自动归集到项目", "P0", "中"),
        ("决算模块", "简化版决算，支持决算编制和审批", "P1", "中"),
    ]
    add_header_row(ws, row, ["模块", "内容", "优先级", "工作量"], styles)
    row += 1
    for item in phase1:
        add_data_row(ws, row, item, styles)
        row += 1

    # Phase 2
    row += 1
    ws.cell(row=row, column=1, value="Phase 2: 深化控制").font = styles['title_font']
    row += 1

    phase2 = [
        ("预算控制", "采购提交时预算校验，超预算预警/阻止", "P1", "中"),
        ("核算月度确认", "月度核算确认流程", "P1", "小"),
        ("决算偏差分析", "偏差原因分析，根因分类", "P1", "中"),
        ("经验库", "经验沉淀和检索", "P2", "中"),
    ]
    add_header_row(ws, row, ["模块", "内容", "优先级", "工作量"], styles)
    row += 1
    for item in phase2:
        add_data_row(ws, row, item, styles)
        row += 1

    # Phase 3
    row += 1
    ws.cell(row=row, column=1, value="Phase 3: 管理提升").font = styles['title_font']
    row += 1

    phase3 = [
        ("部门级汇总", "部门财务汇总报表", "P2", "小"),
        ("概算模板管理", "概算模板维护，历史数据参考", "P2", "中"),
        ("四算检查点", "LTC流程检查点配置", "P2", "小"),
        ("经营分析报表", "公司级/部门级经营分析", "P3", "中"),
    ]
    add_header_row(ws, row, ["模块", "内容", "优先级", "工作量"], styles)
    row += 1
    for item in phase3:
        add_data_row(ws, row, item, styles)
        row += 1

    # 系统改造
    row += 1
    ws.cell(row=row, column=1, value="现有系统改造点").font = styles['title_font']
    row += 1

    changes = [
        ("报价模块", "增加概算关联，报价提交前必须完成概算", "P0", "小"),
        ("合同模块", "签约时检查概算审批状态", "P0", "小"),
        ("项目模块", "增加四算状态字段，关联概算/预算/决算", "P0", "小"),
        ("采购模块", "增加预算校验，成本自动归集", "P0", "中"),
        ("外协模块", "增加预算校验，成本自动归集", "P0", "中"),
        ("工时模块", "人工成本自动归集到项目", "P1", "中"),
        ("角色权限", "新增 PFC/BFC/概算经理角色", "P1", "小"),
    ]
    add_header_row(ws, row, ["模块", "改造内容", "优先级", "工作量"], styles)
    row += 1
    for item in changes:
        add_data_row(ws, row, item, styles)
        row += 1

def create_gap_sheet(wb, styles):
    """创建差距分析页"""
    ws = wb.create_sheet("9.差距分析")
    set_column_widths(ws, [15, 25, 25, 15])

    row = 1
    ws.cell(row=row, column=1, value="现有系统与四算目标差距分析").font = Font(bold=True, size=16)

    row += 2
    ws.cell(row=row, column=1, value="一、需要新增的模块").font = styles['title_font']
    row += 1

    new_modules = [
        ("概算模块", "project_estimates, project_estimate_items", "关联 Quote（报价）", "P0"),
        ("预算扩展", "扩展 project_budgets，新增 budget_adjustments", "增加概算关联", "P0"),
        ("核算台账", "project_cost_ledgers, cost_allocation_rules", "汇总现有成本数据", "P0"),
        ("决算模块", "project_settlement_finals, settlement_variance_analyses", "新增", "P1"),
        ("经验库", "project_lessons_learned", "新增", "P2"),
        ("部门汇总", "dept_financial_summaries", "新增", "P1"),
        ("检查点", "four_estimate_checkpoints", "新增", "P2"),
    ]
    add_header_row(ws, row, ["模块", "核心表", "与现有系统关系", "优先级"], styles)
    row += 1
    for item in new_modules:
        add_data_row(ws, row, item, styles)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="二、现有功能覆盖度").font = styles['title_font']
    row += 1

    coverage = [
        ("概算", "报价成本模板、报价成本审批、毛利率预警", "★★☆☆☆", "缺少目标利润驱动、概算经理"),
        ("预算", "项目预算表、预算明细、成本分摊规则", "★★★☆☆", "缺少概算关联、预算控制"),
        ("核算", "项目成本记录、收款计划、财务工作台", "★★★☆☆", "缺少自动归集、台账汇总"),
        ("决算", "项目结算页面（基础功能）", "★★☆☆☆", "缺少偏差分析、经验传承"),
    ]
    add_header_row(ws, row, ["四算阶段", "现有功能", "覆盖度", "主要差距"], styles)
    row += 1
    for item in coverage:
        add_data_row(ws, row, item, styles)
        row += 1

def main():
    """主函数"""
    wb = Workbook()
    styles = create_style()

    # 创建各个 sheet
    create_overview_sheet(wb, styles)
    create_estimate_sheet(wb, styles)
    create_budget_sheet(wb, styles)
    create_accounting_sheet(wb, styles)
    create_settlement_sheet(wb, styles)
    create_roles_sheet(wb, styles)
    create_integration_sheet(wb, styles)
    create_implementation_sheet(wb, styles)
    create_gap_sheet(wb, styles)

    # 保存文件
    output_path = "docs/plans/项目四算体系设计文档.xlsx"
    wb.save(output_path)
    print(f"Excel 文件已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    main()

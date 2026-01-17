# -*- coding: utf-8 -*-
"""
生成四算差距分析 Excel 报告
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_header_style():
    """创建表头样式"""
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_cell_style():
    """创建单元格样式"""
    return {
        'alignment': Alignment(vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_header_style(cell):
    """应用表头样式"""
    style = create_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def apply_cell_style(cell):
    """应用单元格样式"""
    style = create_cell_style()
    cell.alignment = style['alignment']
    cell.border = style['border']


def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_executive_summary(wb):
    """创建执行摘要"""
    ws = wb.create_sheet("1-执行摘要")

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = "四算体系差距分析 - 执行摘要"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 覆盖率分析
    ws['A3'] = "模块覆盖率分析"
    ws['A3'].font = Font(bold=True, size=12)

    headers = ["模块", "现有覆盖率", "差距评估", "核心缺失"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)

    data = [
        ["概算（估算利润）", "30%", "有基础（报价成本模板）", "概算主表、审批流程、概算经理"],
        ["预算（成本控制）", "60%", "有预算表和明细", "控制机制、概算关联、预警阈值"],
        ["核算（成本归集）", "50%", "有成本记录", "台账汇总、自动归集、期间确认"],
        ["决算（经验传承）", "0%", "完全缺失", "决算表、偏差分析、经验库"],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)
            # 覆盖率列标红/黄/绿
            if col_idx == 2:
                if "0%" in value:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif "30%" in value or "50%" in value:
                    cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
                elif "60%" in value:
                    cell.fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")

    # 工作量估算
    ws['A10'] = "工作量估算"
    ws['A10'].font = Font(bold=True, size=12)

    headers2 = ["任务类型", "数量", "优先级"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=11, column=col, value=header)
        apply_header_style(cell)

    data2 = [
        ["新增数据表", "8张", "高"],
        ["修改数据表", "4张", "高"],
        ["新增API端点", "35+个", "高"],
        ["新增前端页面", "8个", "中"],
        ["新增服务类", "6个", "高"],
    ]

    for row_idx, row_data in enumerate(data2, 12):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [20, 15, 25, 35])


def create_existing_assets(wb):
    """创建现有资产分析"""
    ws = wb.create_sheet("2-现有资产")

    ws.merge_cells('A1:E1')
    ws['A1'] = "现有系统资产分析"
    ws['A1'].font = Font(bold=True, size=14)

    # 数据库模型
    ws['A3'] = "已有的四算相关数据库模型"
    ws['A3'].font = Font(bold=True, size=12)

    headers = ["模型名称", "文件位置", "四算关联", "覆盖功能", "评估"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)

    data = [
        ["Project", "models/project.py", "核心", "项目基础+金额字段", "✓ 可用"],
        ["ProjectCost", "models/project.py", "核算", "成本记录(类型/分类/金额/来源)", "✓ 可用"],
        ["FinancialProjectCost", "models/project.py", "核算", "财务上传成本", "✓ 可用"],
        ["ProjectPaymentPlan", "models/project.py", "核算", "收款计划", "✓ 可用"],
        ["ProjectBudget", "models/budget.py", "预算", "预算主表(编号/版本/审批)", "△ 需扩展"],
        ["ProjectBudgetItem", "models/budget.py", "预算", "预算明细", "✓ 可用"],
        ["ProjectCostAllocationRule", "models/budget.py", "核算", "成本分摊规则", "✓ 可用"],
        ["QuoteCostTemplate", "models/sales.py", "概算", "报价成本模板", "△ 需扩展"],
        ["Quote/QuoteVersion", "models/sales.py", "概算", "报价管理", "△ 需扩展"],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    # API端点
    ws['A16'] = "已有的四算相关API端点"
    ws['A16'].font = Font(bold=True, size=12)

    headers2 = ["端点模块", "文件位置", "功能覆盖"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=17, column=col, value=header)
        apply_header_style(cell)

    data2 = [
        ["/budgets", "endpoints/budget.py", "预算CRUD、审批、明细管理"],
        ["/budgets/allocation-rules", "endpoints/budget.py", "成本分摊规则管理"],
        ["/costs", "endpoints/costs/basic.py", "成本记录CRUD"],
        ["/costs/analysis", "endpoints/costs/analysis.py", "成本分析统计"],
        ["/costs/allocation", "endpoints/costs/allocation.py", "成本分摊计算"],
        ["/sales/cost-templates", "endpoints/sales/cost_management.py", "成本模板管理"],
        ["/projects/{id}/payment-plans", "endpoints/projects/payment_plans.py", "收款计划管理"],
        ["/timesheet/reports/finance", "endpoints/timesheet/reports.py", "工时→成本报表"],
    ]

    for row_idx, row_data in enumerate(data2, 18):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [25, 35, 15, 35, 15])


def create_new_tables(wb):
    """创建新增表清单"""
    ws = wb.create_sheet("3-新增数据表")

    ws.merge_cells('A1:E1')
    ws['A1'] = "需要新增的数据表（8张）"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["序号", "表名", "模块", "用途", "主要字段"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    data = [
        ["1", "project_estimates", "概算", "概算主表", "estimate_no, project_id, opportunity_id, target_margin_rate, allowed_cost, estimated_cost, status, version"],
        ["2", "project_estimate_items", "概算", "概算明细", "estimate_id, cost_category, cost_item, estimated_amount, reference_source, reference_project_id"],
        ["3", "project_budget_adjustments", "预算", "预算调整记录", "budget_id, adjustment_no, adjustment_type, before_amount, adjust_amount, reason_category, related_ecn_id, status"],
        ["4", "project_cost_ledgers", "核算", "项目成本台账（按期间汇总）", "project_id, machine_id, period_type, period_value, material_cost, labor_cost, outsource_cost, total_cost, budget_amount, variance_amount, status"],
        ["5", "dept_financial_summaries", "核算", "部门财务汇总", "dept_id, period_type, period_value, project_count, total_estimate_revenue, total_actual_cost, budget_variance, estimate_accuracy"],
        ["6", "project_settlement_finals", "决算", "决算主表", "settlement_no, project_id, estimate_id, budget_id, final_revenue, total_cost, gross_profit, cost_variance, estimate_accuracy, status"],
        ["7", "settlement_variance_analyses", "决算", "决算偏差分析", "settlement_id, variance_category, variance_type, estimated_amount, actual_amount, root_cause, lesson_learned, feedback_to_template"],
        ["8", "project_lessons_learned", "决算", "项目经验库", "settlement_id, project_type, lesson_type, title, description, estimated_value, actual_value, recommendation, reference_count"],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [8, 30, 10, 30, 80])


def create_modify_tables(wb):
    """创建修改表清单"""
    ws = wb.create_sheet("4-修改数据表")

    ws.merge_cells('A1:D1')
    ws['A1'] = "需要修改的数据表（4张）"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["序号", "表名", "修改内容", "SQL语句"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    data = [
        ["1", "projects", "新增: estimate_id, four_estimate_status, settlement_id",
         "ALTER TABLE projects ADD COLUMN estimate_id INT;\nALTER TABLE projects ADD COLUMN four_estimate_status VARCHAR(50) DEFAULT 'ESTIMATE_PENDING';\nALTER TABLE projects ADD COLUMN settlement_id INT;"],
        ["2", "project_budgets", "新增: estimate_id, estimate_no, budget_source, budget_category, control_mode, warning_threshold, original_amount, adjusted_amount, adjustment_count",
         "ALTER TABLE project_budgets ADD COLUMN estimate_id INT;\nALTER TABLE project_budgets ADD COLUMN estimate_no VARCHAR(50);\nALTER TABLE project_budgets ADD COLUMN budget_source VARCHAR(20) DEFAULT 'MANUAL';\nALTER TABLE project_budgets ADD COLUMN control_mode VARCHAR(20) DEFAULT 'SOFT';\nALTER TABLE project_budgets ADD COLUMN warning_threshold DECIMAL(5,2) DEFAULT 0.80;"],
        ["3", "quotes / quote_versions", "新增: estimate_id",
         "ALTER TABLE quotes ADD COLUMN estimate_id INT;\nALTER TABLE quote_versions ADD COLUMN estimate_id INT;"],
        ["4", "quote_cost_templates", "新增: reference_project_ids (JSON), source_settlement_ids (JSON)",
         "ALTER TABLE quote_cost_templates ADD COLUMN reference_project_ids JSON;\nALTER TABLE quote_cost_templates ADD COLUMN source_settlement_ids JSON;"],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)
        ws.row_dimensions[row_idx].height = 80

    set_column_widths(ws, [8, 25, 60, 80])


def create_new_apis(wb):
    """创建新增API清单"""
    ws = wb.create_sheet("5-新增API端点")

    ws.merge_cells('A1:D1')
    ws['A1'] = "需要新增的API端点（35+个）"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["模块", "端点", "方法", "功能描述"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    data = [
        # 概算模块
        ["概算", "/estimates", "GET", "概算列表（支持按商机/项目筛选）"],
        ["概算", "/estimates", "POST", "创建概算"],
        ["概算", "/estimates/{id}", "GET/PUT/DELETE", "概算详情/更新/删除"],
        ["概算", "/estimates/{id}/items", "GET/POST", "概算明细列表/创建"],
        ["概算", "/estimates/items/{id}", "PUT/DELETE", "概算明细更新/删除"],
        ["概算", "/estimates/{id}/submit", "POST", "提交概算审批"],
        ["概算", "/estimates/{id}/approve", "POST", "审批概算"],
        ["概算", "/estimates/{id}/transfer-to-budget", "POST", "概算转预算"],
        ["概算", "/estimates/templates", "GET", "获取概算模板（基于历史项目）"],
        ["概算", "/estimates/{id}/copy", "POST", "复制概算"],
        # 预算扩展
        ["预算", "/budgets/{id}/adjustments", "GET/POST", "预算调整记录"],
        ["预算", "/budgets/adjustments/{id}/approve", "POST", "审批预算调整"],
        ["预算", "/budgets/{id}/check", "GET", "检查预算执行情况"],
        ["预算", "/budgets/{id}/warnings", "GET", "获取预算预警列表"],
        ["预算", "/budgets/from-estimate/{estimate_id}", "POST", "从概算生成预算"],
        # 核算模块
        ["核算", "/cost-ledgers", "GET", "成本台账列表"],
        ["核算", "/cost-ledgers/project/{project_id}", "GET", "项目成本台账"],
        ["核算", "/cost-ledgers/project/{project_id}/period/{period}", "GET", "项目期间台账详情"],
        ["核算", "/cost-ledgers/aggregate", "POST", "手动触发成本归集"],
        ["核算", "/cost-ledgers/confirm", "POST", "确认月度核算"],
        ["核算", "/cost-ledgers/lock", "POST", "锁定核算期间"],
        ["核算", "/dept-summaries", "GET", "部门财务汇总列表"],
        ["核算", "/dept-summaries/{dept_id}/period/{period}", "GET", "部门期间汇总"],
        ["核算", "/cost-analysis/budget-variance", "GET", "预算偏差分析"],
        ["核算", "/cost-analysis/estimate-variance", "GET", "概算偏差分析"],
        # 决算模块
        ["决算", "/settlements", "GET/POST", "决算列表/创建"],
        ["决算", "/settlements/{id}", "GET/PUT", "决算详情/更新"],
        ["决算", "/settlements/{id}/submit", "POST", "提交决算审核"],
        ["决算", "/settlements/{id}/review", "POST", "审核决算"],
        ["决算", "/settlements/{id}/approve", "POST", "审批决算"],
        ["决算", "/settlements/{id}/variance-analyses", "GET/POST", "偏差分析列表/创建"],
        ["决算", "/settlements/variance-analyses/{id}", "PUT/DELETE", "偏差分析更新/删除"],
        ["决算", "/lessons-learned", "GET/POST", "经验库列表/创建"],
        ["决算", "/lessons-learned/{id}", "GET/PUT", "经验详情/更新"],
        ["决算", "/lessons-learned/search", "POST", "搜索经验（用于概算参考）"],
        ["决算", "/lessons-learned/feedback-to-template", "POST", "反馈到概算模板"],
        # 四算集成
        ["四算集成", "/four-estimate/checkpoints", "GET/POST", "检查点配置管理"],
        ["四算集成", "/four-estimate/checkpoints/{id}", "PUT/DELETE", "检查点更新/删除"],
        ["四算集成", "/four-estimate/project/{project_id}/status", "GET", "项目四算状态"],
        ["四算集成", "/four-estimate/project/{project_id}/timeline", "GET", "项目四算时间线"],
        ["四算集成", "/four-estimate/dashboard", "GET", "四算经营看板"],
        ["四算集成", "/four-estimate/dept/{dept_id}/summary", "GET", "部门四算汇总"],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [12, 50, 20, 40])


def create_implementation_plan(wb):
    """创建实施计划"""
    ws = wb.create_sheet("6-实施优先级")

    ws.merge_cells('A1:D1')
    ws['A1'] = "实施优先级建议"
    ws['A1'].font = Font(bold=True, size=14)

    # Phase 1
    ws['A3'] = "Phase 1: 基础闭环（优先实施）"
    ws['A3'].font = Font(bold=True, size=12, color="FF0000")

    headers = ["任务", "工作量", "优先级", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)

    data1 = [
        ["创建 project_estimates 表和 API", "3天", "P0", "概算核心功能"],
        ["扩展 project_budgets 表（概算关联）", "1天", "P0", "字段扩展"],
        ["概算→预算转换 API", "1天", "P0", "四算衔接"],
        ["创建 project_settlement_finals 表和 API", "3天", "P0", "决算核心功能"],
        ["项目四算状态字段", "0.5天", "P0", "状态管理"],
    ]

    for row_idx, row_data in enumerate(data1, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    # Phase 2
    ws['A12'] = "Phase 2: 控制机制（次优先）"
    ws['A12'].font = Font(bold=True, size=12, color="FF8C00")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_header_style(cell)

    data2 = [
        ["预算控制机制（采购/外协校验）", "2天", "P1", "采购时检查预算"],
        ["预算调整记录表和 API", "2天", "P1", "预算变更管理"],
        ["成本台账表和 API", "3天", "P1", "核算汇总视图"],
        ["决算偏差分析", "2天", "P1", "决算深度分析"],
    ]

    for row_idx, row_data in enumerate(data2, 14):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    # Phase 3
    ws['A20'] = "Phase 3: 管理提升（后续）"
    ws['A20'].font = Font(bold=True, size=12, color="008000")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=21, column=col, value=header)
        apply_header_style(cell)

    data3 = [
        ["经验库和模板反馈", "3天", "P2", "知识沉淀"],
        ["部门财务汇总", "2天", "P2", "管理报表"],
        ["四算检查点配置", "2天", "P2", "流程控制"],
        ["四算经营看板", "3天", "P2", "可视化展示"],
    ]

    for row_idx, row_data in enumerate(data3, 22):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [40, 12, 12, 25])


def create_services(wb):
    """创建服务类清单"""
    ws = wb.create_sheet("7-新增服务类")

    ws.merge_cells('A1:C1')
    ws['A1'] = "需要新增的服务类"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["服务类", "职责", "位置"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    data = [
        ["EstimateService", "概算创建、审批、转预算", "app/services/estimate/"],
        ["BudgetControlService", "预算控制、预警检查", "app/services/budget/"],
        ["CostLedgerService", "成本台账归集、确认", "app/services/cost/"],
        ["SettlementService", "决算创建、审批、偏差分析", "app/services/settlement/"],
        ["LessonsLearnedService", "经验库管理、模板反馈", "app/services/settlement/"],
        ["FourEstimateIntegrationService", "四算检查点、状态管理", "app/services/four_estimate/"],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    set_column_widths(ws, [35, 35, 35])


def main():
    """主函数"""
    wb = Workbook()

    # 删除默认sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建各个sheet
    create_executive_summary(wb)
    create_existing_assets(wb)
    create_new_tables(wb)
    create_modify_tables(wb)
    create_new_apis(wb)
    create_implementation_plan(wb)
    create_services(wb)

    # 保存文件
    output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              "docs/plans/项目四算差距分析报告.xlsx")
    wb.save(output_path)
    print(f"Excel报告已生成: {output_path}")

    # 复制到桌面
    desktop_path = "/Users/flw/Desktop/AI 做的方案/项目四算差距分析报告.xlsx"
    import shutil
    shutil.copy(output_path, desktop_path)
    print(f"已复制到: {desktop_path}")


if __name__ == "__main__":
    main()

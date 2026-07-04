#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
售前弹药库：导入历史采购数据 → 部件级标准成本库

价值：采购价是"实际成交价"，比报价更权威。导入后 AI 能说：
  "电控柜对外报9.8万，实际采购PLC 8500+电源350+触摸屏2800≈1.2万，毛利有保证"

数据流：
  Excel（采购历史）→ 解析 → purchase_material_costs 表
  → 聚合成标准件价格（按物料名归一化，给中位价/最低价/供应商分布/交期）
  → 反哺 standard_costs + ToolRegistry，让 AI 能查"这个件我们历史买多少钱"

列名容错：支持中英文列名模糊匹配（material_name/物料名称/品名 都能识别）。
去重：按 (material_name + brand + supplier_name + unit_cost) 去重。

运行：
    python scripts/import_purchase_costs.py /path/to/purchase.xlsx
    python scripts/import_purchase_costs.py /path/to/purchase.xlsx --dry-run   # 先看解析结果
    python scripts/import_purchase_costs.py /path/to/dir/                       # 导入目录下所有 xlsx
"""
from __future__ import annotations

import argparse
import logging
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_DB_PATH = ROOT_DIR / "data" / "app.db"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("import_purchase")

# ============= 列名模糊匹配表 =============
# 键是 purchase_material_costs 表的标准字段，值是 Excel 里可能的列名（按优先级）
COLUMN_ALIASES = {
    "material_name": [
        "material_name", "物料名称", "品名", "名称", "材料名称", "零件名称",
        "设备名称", "部件名称", "product_name", "item_name", "description",
    ],
    "specification": [
        "specification", "规格", "规格型号", "型号", "规格参数", "spec", "model",
    ],
    "brand": ["brand", "品牌", "厂家", "制造商", "牌号", "manufacturer"],
    "unit": ["unit", "单位", "计量单位", "uom"],
    "material_type": [
        "material_type", "物料类型", "类别", "分类", "类型", "category", "material_category",
    ],
    "unit_cost": [
        "unit_cost", "单价", "采购单价", "价格", "price", "cost", "含税单价", "unit_price",
    ],
    "currency": ["currency", "币种", "currency_code"],
    "supplier_name": [
        "supplier_name", "供应商", "供应商名称", "供货商", "厂家名称", "vendor",
        "supplier", "supplier_name_en",
    ],
    "purchase_date": [
        "purchase_date", "采购日期", "下单日期", "日期", "order_date", "date",
    ],
    "purchase_order_no": [
        "purchase_order_no", "采购订单号", "订单号", "PO号", "单号", "order_no", "po_number",
    ],
    "purchase_quantity": [
        "purchase_quantity", "采购数量", "数量", "qty", "quantity", "amount",
    ],
    "lead_time_days": [
        "lead_time_days", "交期", "交货周期", "交货期", "lead_time", "delivery_days",
    ],
    "material_code": [
        "material_code", "物料编码", "编码", "料号", "item_code", "code", "part_number",
    ],
    "remark": ["remark", "备注", "说明", "note", "comment"],
}

# 必填字段（缺了这些的行会跳过）
REQUIRED_FIELDS = ["material_name", "unit_cost"]


def find_column(headers: list, standard_field: str) -> str | None:
    """从 Excel 表头里找匹配的列名。"""
    aliases = COLUMN_ALIASES.get(standard_field, [])
    headers_lower = {str(h).strip().lower(): str(h) for h in headers}
    for alias in aliases:
        if alias.lower() in headers_lower:
            return headers_lower[alias.lower()]
    # 模糊包含匹配
    for alias in aliases:
        for h_lower, h_orig in headers_lower.items():
            if alias.lower() in h_lower or h_lower in alias.lower():
                return h_orig
    return None


def parse_excel(file_path: Path) -> list[dict]:
    """解析 Excel，返回标准化的采购记录列表。"""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 找表头（第一个非空行）
        header_row = None
        for r in rows[:5]:
            if any(c and str(c).strip() for c in r):
                header_row = [str(c).strip() if c else "" for c in r]
                break
        if not header_row:
            continue

        # 建立列映射
        col_map: dict[str, int] = {}  # standard_field -> col_index
        for std_field in COLUMN_ALIASES:
            col_name = find_column(header_row, std_field)
            if col_name:
                col_map[std_field] = header_row.index(col_name)

        # 检查必填字段
        missing = [f for f in REQUIRED_FIELDS if f not in col_map]
        if missing:
            logger.warning(f"Sheet '{sheet_name}' 缺少必填字段 {missing}，跳过")
            continue

        logger.info(
            f"Sheet '{sheet_name}': 识别列映射 {col_map}（{len(header_row)} 列，"
            f"匹配 {len(col_map)} 个标准字段）"
        )

        # 解析数据行（从表头下一行开始）
        data_start = header_row.index(next(h for h in header_row if h))
        for row in rows[1:]:
            if not row or not any(c and str(c).strip() for c in row):
                continue
            record = {}
            for std_field, col_idx in col_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    record[std_field] = _clean_value(val, std_field)
            # 校验必填
            if not record.get("material_name") or record.get("unit_cost") is None:
                continue
            record["_sheet"] = sheet_name
            all_rows.append(record)

    wb.close()
    return all_rows


def _clean_value(val, field: str):
    """清洗字段值。"""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    s = str(val).strip()
    if field in ("unit_cost", "purchase_quantity"):
        # 去掉货币符号和千分位
        s = s.replace("¥", "").replace("￥", "").replace(",", "").replace(" ", "")
        try:
            return float(s)
        except ValueError:
            return None
    if field == "lead_time_days":
        # "30天" → 30
        digits = "".join(c for c in s if c.isdigit())
        return int(digits) if digits else None
    if field == "purchase_date":
        if isinstance(val, (datetime, date)):
            return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
        # 尝试解析字符串日期
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        return s
    return s


def dedup_key(record: dict) -> str:
    """去重键：物料名+品牌+供应商+单价。"""
    return "|".join([
        str(record.get("material_name", "")),
        str(record.get("brand", "")),
        str(record.get("supplier_name", "")),
        str(record.get("unit_cost", "")),
    ])


def import_to_db(records: list[dict], db_path: Path, dry_run: bool = False) -> dict:
    """导入数据库。返回统计。"""
    if dry_run:
        # dry-run 只打印预览
        logger.info("（dry-run）解析结果预览（前 10 条）：")
        for r in records[:10]:
            print(
                f"  {r.get('material_name','')[:25]:<25} "
                f"{r.get('brand','')[:10]:<10} "
                f"¥{r.get('unit_cost',0):>10} /{r.get('unit','')} "
                f"供应商:{r.get('supplier_name','')[:12]} "
                f"交期:{r.get('lead_time_days','')}天"
            )
        return {"total": len(records), "inserted": 0, "skipped": 0}

    conn = sqlite3.connect(db_path)
    # 已存在的去重键
    existing_keys = set()
    for r in conn.execute(
        "SELECT material_name, brand, supplier_name, unit_cost FROM purchase_material_costs "
        "WHERE material_name IS NOT NULL"
    ).fetchall():
        existing_keys.add("|".join(str(x or "") for x in r))

    inserted = 0
    skipped = 0
    today = date.today().isoformat()

    for record in records:
        key = dedup_key(record)
        if key in existing_keys:
            skipped += 1
            continue
        # 判断是否标准件（有规格型号 + 单价合理的）
        is_standard = 1 if record.get("specification") and record.get("unit_cost", 0) > 0 else 0
        conn.execute(
            """INSERT INTO purchase_material_costs
            (material_name, material_code, specification, brand, unit, material_type,
             is_standard_part, unit_cost, currency, supplier_name, purchase_date,
             purchase_order_no, purchase_quantity, lead_time_days, is_active,
             match_priority, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)""",
            (
                record.get("material_name"), record.get("material_code"),
                record.get("specification"), record.get("brand"), record.get("unit") or "件",
                record.get("material_type"), is_standard, record.get("unit_cost"),
                record.get("currency") or "CNY", record.get("supplier_name"),
                record.get("purchase_date"), record.get("purchase_order_no"),
                record.get("purchase_quantity"), record.get("lead_time_days"),
            ),
        )
        existing_keys.add(key)
        inserted += 1

    conn.commit()
    conn.close()
    return {"total": len(records), "inserted": inserted, "skipped": skipped}


def aggregate_to_standard_costs(db_path: Path):
    """导入后聚合到 standard_costs（部件级标准成本，供 AI 对标用）。"""
    conn = sqlite3.connect(db_path)
    from statistics import median

    rows = conn.execute(
        "SELECT material_name, brand, unit_cost, lead_time_days, supplier_name, "
        "specification, unit FROM purchase_material_costs "
        "WHERE unit_cost IS NOT NULL AND unit_cost > 0 AND material_name NOT LIKE '%_name_%'"
    ).fetchall()

    if not rows:
        logger.info("无采购数据可聚合")
        return

    # 按 material_name 归一化聚合
    groups: dict[str, list] = {}
    for r in rows:
        groups.setdefault(r[0], []).append(r)

    today = date.today().isoformat()
    upserted = 0
    for name, items in groups.items():
        if len(items) < 1:
            continue
        prices = [float(r[2]) for r in items]
        med = median(prices)
        suppliers = list(set(r[4] for r in items if r[4]))[:3]
        lead_times = [r[3] for r in items if r[3]]
        brand = items[0][1] or ""
        spec = items[0][5] or ""

        cost_code = f"PC-{name[:20].replace(' ', '').replace('/', '')}"
        source_desc = (
            f"采购历史聚合：样本{len(items)}，中位{med}，"
            f"区间{min(prices)}~{max(prices)}，"
            f"供应商：{'、'.join(suppliers) or '未知'}，"
            f"交期：{median(lead_times) if lead_times else '未知'}天"
        )
        existing = conn.execute(
            "SELECT id FROM standard_costs WHERE cost_code = ?", (cost_code,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE standard_costs SET cost_name=?, cost_category='PURCHASED', specification=?, "
                "unit=?, standard_cost=?, cost_source='PURCHASE_HISTORY', source_description=?, "
                "effective_date=?, is_active=1 WHERE cost_code=?",
                (name, spec, items[0][6] or "件", med, source_desc, today, cost_code),
            )
        else:
            conn.execute(
                "INSERT INTO standard_costs (cost_code, cost_name, cost_category, specification, "
                "unit, standard_cost, currency, cost_source, source_description, effective_date, "
                "version, is_active, created_at, updated_at) "
                "VALUES (?,?, 'PURCHASED', ?, ?, ?, 'CNY', 'PURCHASE_HISTORY', ?, ?, 1,1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
                (cost_code, name, spec, items[0][6] or "件", med, source_desc, today),
            )
        upserted += 1

    conn.commit()
    conn.close()
    logger.info(f"聚合 {upserted} 个部件级标准成本到 standard_costs（来源：采购历史）")


def main():
    parser = argparse.ArgumentParser(description="导入历史采购数据 → 部件级标准成本库")
    parser.add_argument("source", type=Path, help="Excel 文件或目录路径")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--dry-run", action="store_true", help="只解析不写库")
    parser.add_argument("--no-aggregate", action="store_true", help="跳过聚合到 standard_costs")
    args = parser.parse_args()

    # 收集 Excel 文件
    if args.source.is_dir():
        files = sorted(list(args.source.glob("*.xlsx")) + list(args.source.glob("*.xls")))
    elif args.source.is_file():
        files = [args.source]
    else:
        logger.error(f"路径不存在: {args.source}")
        sys.exit(1)

    if not files:
        logger.error("未找到 Excel 文件")
        sys.exit(1)

    logger.info(f"待导入文件: {[f.name for f in files]}")

    all_records = []
    for f in files:
        logger.info(f"解析 {f.name}...")
        records = parse_excel(f)
        logger.info(f"  {f.name}: 解析出 {len(records)} 条")
        all_records.extend(records)

    logger.info(f"总计解析: {len(all_records)} 条采购记录")

    # 去重（文件间 + 文件内）
    seen = set()
    unique_records = []
    for r in all_records:
        key = dedup_key(r)
        if key not in seen:
            seen.add(key)
            unique_records.append(r)
    if len(unique_records) < len(all_records):
        logger.info(f"去重: {len(all_records)} → {len(unique_records)} 条")

    # 导入
    stats = import_to_db(unique_records, args.db, dry_run=args.dry_run)
    logger.info(f"导入完成: {stats}")

    # 聚合到标准成本库
    if not args.dry_run and not args.no_aggregate:
        logger.info("聚合到 standard_costs...")
        aggregate_to_standard_costs(args.db)

    if not args.dry_run:
        logger.info(
            f"✓ 完成。purchase_material_costs 现在有 "
            f"{sqlite3.connect(args.db).execute('SELECT COUNT(*) FROM purchase_material_costs').fetchone()[0]} 条数据"
        )


if __name__ == "__main__":
    main()

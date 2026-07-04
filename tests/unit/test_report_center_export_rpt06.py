# -*- coding: utf-8 -*-
"""RPT-06: report-center XLSX export must render detail rows."""

from types import SimpleNamespace
from unittest.mock import MagicMock

from openpyxl import load_workbook

from app.api.v1.endpoints.report_center.generate import export as export_module
from app.schemas.report_center import ReportExportRequest


def test_legacy_xlsx_export_writes_details_rows(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    generation = SimpleNamespace(
        id=123,
        report_code=None,
        report_type="SALES",
        report_title="销售报表",
        report_data={
            "summary": {"总数": 2},
            "details": [
                {"客户": "客户A", "金额": 100},
                {"客户": "客户B", "金额": 200},
            ],
        },
        export_format=None,
        export_path=None,
        exported_at=None,
    )

    monkeypatch.setattr(export_module, "get_or_404", lambda *args, **kwargs: generation)

    response = export_module.export_report(
        db=MagicMock(),
        export_in=ReportExportRequest(report_id=123, export_format="xlsx"),
        current_user=MagicMock(),
    )

    workbook_path = tmp_path / response.data["file_path"]
    workbook = load_workbook(workbook_path)
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]

    assert "无数据" not in values
    assert "客户" in values
    assert "金额" in values
    assert "客户A" in values
    assert 100 in values

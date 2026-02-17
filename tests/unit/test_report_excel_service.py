# -*- coding: utf-8 -*-
"""
I2组 - 报表Excel导出服务 单元测试
覆盖: app/services/report_excel_service.py
"""
from unittest.mock import MagicMock, patch, call
import pytest


# ─── _translate_header ────────────────────────────────────────────────────────

class TestTranslateHeader:
    def setup_method(self):
        from app.services.report_excel_service import ReportExcelService
        self.svc = ReportExcelService

    def test_known_key(self):
        assert self.svc._translate_header("user_name") == "姓名"

    def test_department_name(self):
        assert self.svc._translate_header("department_name") == "部门名称"

    def test_total_hours(self):
        assert self.svc._translate_header("total_hours") == "总工时"

    def test_unknown_key(self):
        assert self.svc._translate_header("unknown_field") == "unknown_field"

    def test_all_translations(self):
        translations = {
            'user_id': '用户ID',
            'user_name': '姓名',
            'department': '部门',
            'total_hours': '总工时',
            'normal_hours': '正常工时',
            'overtime_hours': '加班工时',
            'work_days': '工作天数',
            'user_count': '人数',
            'project_name': '项目名称',
            'task_name': '任务名称',
            'work_date': '日期',
        }
        for key, expected in translations.items():
            assert self.svc._translate_header(key) == expected


# ─── export_to_excel (mock openpyxl) ─────────────────────────────────────────

class TestExportToExcel:
    """测试 Excel 导出主流程（mock openpyxl 写入）"""

    def _make_data(self, with_detail=False):
        data = {
            "year": 2024,
            "month": 1,
            "period": "2024-01",
            "summary": [
                {"user_name": "张三", "total_hours": 160, "overtime_hours": 8},
            ],
        }
        if with_detail:
            data["detail"] = [
                {"user_name": "张三", "work_date": "2024-01-02", "hours": 8},
            ]
        return data

    def test_export_without_detail(self, tmp_path):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")

        mock_wb = MagicMock()
        mock_ws_summary = MagicMock()
        mock_ws_chart = MagicMock()

        mock_wb.sheetnames = []
        mock_wb.create_sheet.side_effect = [mock_ws_summary, mock_ws_chart]

        with patch("app.services.report_excel_service.Workbook", return_value=mock_wb):
            with patch.object(ReportExcelService, "_write_summary_sheet"):
                with patch.object(ReportExcelService, "_write_chart_sheet"):
                    result = ReportExcelService.export_to_excel(
                        self._make_data(),
                        "test_report",
                        str(tmp_path)
                    )
        assert "test_report" in result

    def test_export_with_detail(self, tmp_path):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")

        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_wb.create_sheet.return_value = MagicMock()

        with patch("app.services.report_excel_service.Workbook", return_value=mock_wb):
            with patch.object(ReportExcelService, "_write_summary_sheet"):
                with patch.object(ReportExcelService, "_write_detail_sheet") as mock_detail:
                    with patch.object(ReportExcelService, "_write_chart_sheet"):
                        result = ReportExcelService.export_to_excel(
                            self._make_data(with_detail=True),
                            "test_detail",
                            str(tmp_path)
                        )
        assert "test_detail" in result

    def test_raises_if_no_openpyxl(self):
        from app.services.report_excel_service import ReportExcelService
        with patch("app.services.report_excel_service.OPENPYXL_AVAILABLE", False):
            with pytest.raises(ImportError):
                ReportExcelService.export_to_excel({}, "test", "/tmp")

    def test_default_sheet_removed(self, tmp_path):
        """有 'Sheet' 默认页时应删除"""
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")

        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_wb.sheetnames = ["Sheet"]
        mock_wb.__getitem__ = MagicMock(return_value=mock_sheet)
        mock_wb.create_sheet.return_value = MagicMock()

        with patch("app.services.report_excel_service.Workbook", return_value=mock_wb):
            with patch.object(ReportExcelService, "_write_summary_sheet"):
                with patch.object(ReportExcelService, "_write_chart_sheet"):
                    ReportExcelService.export_to_excel(
                        self._make_data(), "rpt", str(tmp_path)
                    )
        mock_wb.remove.assert_called_once()


# ─── _write_summary_sheet ────────────────────────────────────────────────────

class TestWriteSummarySheet:
    def test_empty_summary(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {"period": "2024-01", "summary": []}
        # 不应抛出异常
        ReportExcelService._write_summary_sheet(ws, data)
        # 检查 "无数据" 被写入
        values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, 3)]
        assert "无数据" in values

    def test_with_summary_data(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {
            "period": "2024-01",
            "summary": [{"user_name": "张三", "total_hours": 160}]
        }
        ReportExcelService._write_summary_sheet(ws, data)
        # 有数据时表头和数据应被写入
        assert ws.max_row >= 4  # 标题 + 时间 + 空行 + 表头 + 数据

    def test_column_width_adjusted(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {
            "period": "2024-01",
            "summary": [{"user_name": "张三", "total_hours": 160, "overtime_hours": 8}]
        }
        ReportExcelService._write_summary_sheet(ws, data)
        # 列宽应被设置（不为默认 None）
        assert ws.column_dimensions['A'].width is not None


# ─── _write_detail_sheet ─────────────────────────────────────────────────────

class TestWriteDetailSheet:
    def test_empty_detail(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {"period": "2024-01", "detail": []}
        ReportExcelService._write_detail_sheet(ws, data)
        values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, 3)]
        assert "无数据" in values

    def test_with_detail_data(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {
            "period": "2024-01",
            "detail": [{"user_name": "李四", "hours": 8}]
        }
        ReportExcelService._write_detail_sheet(ws, data)
        assert ws.max_row >= 3


# ─── _write_chart_sheet ───────────────────────────────────────────────────────

class TestWriteChartSheet:
    def test_empty_summary_no_crash(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {"period": "2024-01", "summary": []}
        # 不应抛出异常
        ReportExcelService._write_chart_sheet(ws, data)

    def test_with_total_hours(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {
            "period": "2024-01",
            "summary": [{"user_name": "张三", "total_hours": 160}]
        }
        with patch.object(ReportExcelService, "_add_bar_chart") as mock_bar:
            ReportExcelService._write_chart_sheet(ws, data)
            mock_bar.assert_called_once()

    def test_with_department_data(self):
        from app.services.report_excel_service import ReportExcelService, OPENPYXL_AVAILABLE
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl 未安装")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        data = {
            "period": "2024-01",
            "summary": [
                {"department_name": "研发部", "total_hours": 320},
                {"department_name": "测试部", "total_hours": 160},
            ]
        }
        with patch.object(ReportExcelService, "_add_bar_chart"):
            with patch.object(ReportExcelService, "_add_pie_chart") as mock_pie:
                ReportExcelService._write_chart_sheet(ws, data)
                mock_pie.assert_called_once()

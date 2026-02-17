# -*- coding: utf-8 -*-
"""
I4组：app/utils/spec_extractor/formats.py 深度单元测试

覆盖目标：4% → 60%+
策略：mock openpyxl / python-docx / PyPDF2，覆盖 extract_from_excel/word/pdf 代码路径
"""

import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch, call, PropertyMock


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_service():
    """SpecExtractor mock"""
    svc = MagicMock()
    return svc


@pytest.fixture
def mock_db():
    return MagicMock()


# ===========================================================================
# extract_from_excel
# ===========================================================================

class TestExtractFromExcel:
    """extract_from_excel 测试"""

    def _make_cell(self, value):
        c = MagicMock()
        c.value = value
        return c

    def _make_row(self, cells):
        row = MagicMock()
        # ws[row_idx] 返回一个可迭代的 cell 列表
        row.__iter__ = MagicMock(return_value=iter(cells))
        return cells  # 直接返回 cell list

    def test_import_error_raises(self, mock_service, mock_db):
        """缺少 openpyxl 时抛出 ImportError"""
        from app.utils.spec_extractor.formats import extract_from_excel

        with patch.dict('sys.modules', {'openpyxl': None}):
            with pytest.raises((ImportError, Exception)):
                extract_from_excel(
                    mock_service, mock_db,
                    Path('/tmp/test.xlsx'), 1, 1, 1
                )

    def test_exception_raises(self, mock_service, mock_db):
        """文件读取异常时抛出Exception"""
        from app.utils.spec_extractor.formats import extract_from_excel

        # openpyxl 在函数内 lazy import，patch openpyxl.load_workbook
        import openpyxl
        with patch.object(openpyxl, 'load_workbook', side_effect=FileNotFoundError("not found")):
            with pytest.raises(Exception):
                extract_from_excel(
                    mock_service, mock_db,
                    Path('/tmp/nonexistent.xlsx'), 1, 1, 1
                )

    def test_empty_sheet_returns_empty_list(self, mock_service, mock_db):
        """空表格返回空列表"""
        from app.utils.spec_extractor.formats import extract_from_excel
        import openpyxl

        mock_ws = MagicMock()
        mock_ws.max_row = 1  # 只有表头行，无数据
        mock_ws.__getitem__ = MagicMock(return_value=[
            self._make_cell('物料名称'),
            self._make_cell('规格')
        ])

        mock_wb = MagicMock()
        mock_wb.active = mock_ws

        with patch.object(openpyxl, 'load_workbook', return_value=mock_wb):
            result = extract_from_excel(
                mock_service, mock_db,
                Path('/tmp/test.xlsx'), 1, 1, 1
            )

        assert result == []

    def test_with_valid_data_rows(self, mock_service, mock_db):
        """有效数据行正确提取"""
        from app.utils.spec_extractor.formats import extract_from_excel
        import openpyxl

        # 构建 header row（row 1）
        header_cells = [
            self._make_cell('物料名称'),
            self._make_cell('规格型号'),
            self._make_cell('品牌'),
        ]

        # 构建 data row（row 2）
        data_cell_name = MagicMock()
        data_cell_name.value = '伺服电机'
        data_cell_spec = MagicMock()
        data_cell_spec.value = 'AC220V 1.5kW'
        data_cell_brand = MagicMock()
        data_cell_brand.value = 'Siemens'
        data_row = [data_cell_name, data_cell_spec, data_cell_brand]

        row_map = {1: header_cells, 2: data_row}
        mock_ws = MagicMock()
        mock_ws.max_row = 2
        mock_ws.__getitem__ = MagicMock(side_effect=lambda idx: row_map[idx])

        mock_wb = MagicMock()
        mock_wb.active = mock_ws

        mock_req = MagicMock()

        with patch.object(openpyxl, 'load_workbook', return_value=mock_wb), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req):
            result = extract_from_excel(
                mock_service, mock_db,
                Path('/tmp/test.xlsx'), 1, 1, 1
            )

        assert len(result) == 1

    def test_row_without_material_name_skipped(self, mock_service, mock_db):
        """没有物料名称的行被跳过"""
        from app.utils.spec_extractor.formats import extract_from_excel
        import openpyxl

        header_cells = [self._make_cell('物料名称')]

        data_cell = MagicMock()
        data_cell.value = None  # 无物料名称
        data_row = [data_cell]

        row_map = {1: header_cells, 2: data_row}
        mock_ws = MagicMock()
        mock_ws.max_row = 2
        mock_ws.__getitem__ = MagicMock(side_effect=lambda idx: row_map[idx])

        mock_wb = MagicMock()
        mock_wb.active = mock_ws

        with patch.object(openpyxl, 'load_workbook', return_value=mock_wb), \
             patch('app.utils.spec_extractor.formats.create_requirement'):
            result = extract_from_excel(
                mock_service, mock_db,
                Path('/tmp/test.xlsx'), 1, 1, 1
            )

        assert result == []

    def test_all_column_types_detected(self, mock_service, mock_db):
        """所有列类型（material_code/name/spec/brand/model）均能正确识别"""
        from app.utils.spec_extractor.formats import extract_from_excel
        import openpyxl

        header_cells = [
            self._make_cell('物料编码'),
            self._make_cell('物料名称'),
            self._make_cell('规格'),
            self._make_cell('品牌'),
            self._make_cell('型号'),
        ]

        cells = [
            MagicMock(value='MC-001'),
            MagicMock(value='传感器'),
            MagicMock(value='5V/100mA'),
            MagicMock(value='Honeywell'),
            MagicMock(value='HX500'),
        ]

        row_map = {1: header_cells, 2: cells}
        mock_ws = MagicMock()
        mock_ws.max_row = 2
        mock_ws.__getitem__ = MagicMock(side_effect=lambda idx: row_map[idx])

        mock_wb = MagicMock()
        mock_wb.active = mock_ws

        mock_req = MagicMock()

        with patch.object(openpyxl, 'load_workbook', return_value=mock_wb), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req) as mock_cr:
            result = extract_from_excel(
                mock_service, mock_db,
                Path('/tmp/test.xlsx'), 1, 1, 1
            )

        assert len(result) == 1
        # 验证 create_requirement 被调用且传入了正确参数
        call_kwargs = mock_cr.call_args[1]
        assert call_kwargs['material_code'] == 'MC-001'
        assert call_kwargs['material_name'] == '传感器'
        assert call_kwargs['brand'] == 'Honeywell'
        assert call_kwargs['model'] == 'HX500'


# ===========================================================================
# extract_from_word
# ===========================================================================

class TestExtractFromWord:
    """extract_from_word 测试"""

    def test_import_error_raises(self, mock_service, mock_db):
        """缺少 python-docx 时抛出 ImportError"""
        from app.utils.spec_extractor.formats import extract_from_word

        with patch.dict('sys.modules', {'docx': None}):
            with pytest.raises((ImportError, Exception)):
                extract_from_word(
                    mock_service, mock_db,
                    Path('/tmp/test.docx'), 1, 1, 1
                )

    def test_exception_raises(self, mock_service, mock_db):
        """文件读取异常时抛出 Exception"""
        from app.utils.spec_extractor.formats import extract_from_word

        mock_docx = MagicMock()
        mock_docx.Document.side_effect = Exception("file error")

        with patch.dict('sys.modules', {'docx': mock_docx}):
            with pytest.raises(Exception):
                extract_from_word(
                    mock_service, mock_db,
                    Path('/tmp/test.docx'), 1, 1, 1
                )

    def test_empty_doc_returns_empty_list(self, mock_service, mock_db):
        """空文档（无表格）返回空列表"""
        from app.utils.spec_extractor.formats import extract_from_word

        mock_doc = MagicMock()
        mock_doc.tables = []
        mock_doc.paragraphs = []

        mock_docx_module = MagicMock()
        mock_docx_module.Document.return_value = mock_doc

        with patch.dict('sys.modules', {'docx': mock_docx_module}):
            result = extract_from_word(
                mock_service, mock_db,
                Path('/tmp/test.docx'), 1, 1, 1
            )

        assert result == []

    def test_word_table_with_valid_data(self, mock_service, mock_db):
        """Word 表格有效数据行正确提取"""
        from app.utils.spec_extractor.formats import extract_from_word

        # 构建表格：header row + data row
        header_cell_name = MagicMock()
        header_cell_name.text = '物料名称'
        header_cell_spec = MagicMock()
        header_cell_spec.text = '规格型号'

        data_cell_name = MagicMock()
        data_cell_name.text = 'PLC控制器'
        data_cell_spec = MagicMock()
        data_cell_spec.text = 'AC220V CPU1214C'

        header_row = MagicMock()
        header_row.cells = [header_cell_name, header_cell_spec]

        data_row = MagicMock()
        data_row.cells = [data_cell_name, data_cell_spec]

        mock_table = MagicMock()
        mock_table.rows = [header_row, data_row]

        mock_doc = MagicMock()
        mock_doc.tables = [mock_table]
        mock_doc.paragraphs = []

        mock_docx_module = MagicMock()
        mock_docx_module.Document.return_value = mock_doc

        mock_req = MagicMock()

        with patch.dict('sys.modules', {'docx': mock_docx_module}), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req):
            result = extract_from_word(
                mock_service, mock_db,
                Path('/tmp/test.docx'), 1, 1, 1
            )

        assert len(result) == 1

    def test_word_table_header_no_material_name_col(self, mock_service, mock_db):
        """表格无物料名称列时跳过数据行"""
        from app.utils.spec_extractor.formats import extract_from_word

        # 表头没有 material_name 列
        header_cell = MagicMock()
        header_cell.text = '其他列'
        data_cell = MagicMock()
        data_cell.text = '数据'

        header_row = MagicMock()
        header_row.cells = [header_cell]
        data_row = MagicMock()
        data_row.cells = [data_cell]

        mock_table = MagicMock()
        mock_table.rows = [header_row, data_row]

        mock_doc = MagicMock()
        mock_doc.tables = [mock_table]
        mock_doc.paragraphs = []

        mock_docx_module = MagicMock()
        mock_docx_module.Document.return_value = mock_doc

        with patch.dict('sys.modules', {'docx': mock_docx_module}):
            result = extract_from_word(
                mock_service, mock_db,
                Path('/tmp/test.docx'), 1, 1, 1
            )

        assert result == []

    def test_word_empty_material_name_skipped(self, mock_service, mock_db):
        """物料名称为空的行被跳过"""
        from app.utils.spec_extractor.formats import extract_from_word

        header_cell_name = MagicMock()
        header_cell_name.text = '物料名称'

        data_cell_name = MagicMock()
        data_cell_name.text = ''  # 空物料名称

        header_row = MagicMock()
        header_row.cells = [header_cell_name]
        data_row = MagicMock()
        data_row.cells = [data_cell_name]

        mock_table = MagicMock()
        mock_table.rows = [header_row, data_row]

        mock_doc = MagicMock()
        mock_doc.tables = [mock_table]
        mock_doc.paragraphs = []

        mock_docx_module = MagicMock()
        mock_docx_module.Document.return_value = mock_doc

        with patch.dict('sys.modules', {'docx': mock_docx_module}):
            result = extract_from_word(
                mock_service, mock_db,
                Path('/tmp/test.docx'), 1, 1, 1
            )

        assert result == []

    def test_word_all_columns_parsed(self, mock_service, mock_db):
        """Word 表格所有列均能识别"""
        from app.utils.spec_extractor.formats import extract_from_word

        col_texts = ['物料编码', '物料名称', '规格', '品牌', '型号']
        header_cells = [MagicMock(text=t) for t in col_texts]

        data_vals = ['C001', '传感器', '5V', 'ABB', 'HX100']
        data_cells = [MagicMock(text=v) for v in data_vals]

        header_row = MagicMock()
        header_row.cells = header_cells
        data_row = MagicMock()
        data_row.cells = data_cells

        mock_table = MagicMock()
        mock_table.rows = [header_row, data_row]

        mock_doc = MagicMock()
        mock_doc.tables = [mock_table]
        mock_doc.paragraphs = []

        mock_docx_module = MagicMock()
        mock_docx_module.Document.return_value = mock_doc

        mock_req = MagicMock()

        with patch.dict('sys.modules', {'docx': mock_docx_module}), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req) as mock_cr:
            result = extract_from_word(
                mock_service, mock_db,
                Path('/tmp/test.docx'), 1, 1, 1
            )

        assert len(result) == 1


# ===========================================================================
# extract_from_pdf
# ===========================================================================

class TestExtractFromPdf:
    """extract_from_pdf 测试"""

    def test_import_error_raises(self, mock_service, mock_db):
        """缺少 PyPDF2 时抛出 ImportError"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        with patch.dict('sys.modules', {'PyPDF2': None}):
            with pytest.raises((ImportError, Exception)):
                extract_from_pdf(
                    mock_service, mock_db,
                    Path('/tmp/test.pdf'), 1, 1, 1
                )

    def test_file_error_raises(self, mock_service, mock_db):
        """文件打开异常时抛出 Exception"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        mock_pypdf2 = MagicMock()

        with patch.dict('sys.modules', {'PyPDF2': mock_pypdf2}), \
             patch('builtins.open', side_effect=FileNotFoundError("not found")):
            with pytest.raises(Exception):
                extract_from_pdf(
                    mock_service, mock_db,
                    Path('/tmp/nonexistent.pdf'), 1, 1, 1
                )

    def test_empty_pdf_returns_empty_list(self, mock_service, mock_db):
        """空 PDF 返回空列表"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        mock_page = MagicMock()
        mock_page.extract_text.return_value = ''

        mock_reader = MagicMock()
        mock_reader.pages = [mock_page]

        mock_pypdf2 = MagicMock()
        mock_pypdf2.PdfReader.return_value = mock_reader

        mock_file = MagicMock()
        mock_file.__enter__ = MagicMock(return_value=mock_file)
        mock_file.__exit__ = MagicMock(return_value=False)

        with patch.dict('sys.modules', {'PyPDF2': mock_pypdf2}), \
             patch('builtins.open', return_value=mock_file):
            result = extract_from_pdf(
                mock_service, mock_db,
                Path('/tmp/test.pdf'), 1, 1, 1
            )

        assert result == []

    def test_pdf_with_material_line(self, mock_service, mock_db):
        """PDF 中包含物料行时正确提取"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        mock_page = MagicMock()
        mock_page.extract_text.return_value = '物料 伺服电机 AC220V 1.5kW Siemens'

        mock_reader = MagicMock()
        mock_reader.pages = [mock_page]

        mock_pypdf2 = MagicMock()
        mock_pypdf2.PdfReader.return_value = mock_reader

        mock_file = MagicMock()
        mock_file.__enter__ = MagicMock(return_value=mock_file)
        mock_file.__exit__ = MagicMock(return_value=False)

        mock_req = MagicMock()

        with patch.dict('sys.modules', {'PyPDF2': mock_pypdf2}), \
             patch('builtins.open', return_value=mock_file), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req):
            result = extract_from_pdf(
                mock_service, mock_db,
                Path('/tmp/test.pdf'), 1, 1, 1
            )

        assert len(result) >= 1

    def test_pdf_non_material_lines_ignored(self, mock_service, mock_db):
        """PDF 中非物料行被忽略"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        mock_page = MagicMock()
        mock_page.extract_text.return_value = '这是一行普通文本\n另一行不相关内容'

        mock_reader = MagicMock()
        mock_reader.pages = [mock_page]

        mock_pypdf2 = MagicMock()
        mock_pypdf2.PdfReader.return_value = mock_reader

        mock_file = MagicMock()
        mock_file.__enter__ = MagicMock(return_value=mock_file)
        mock_file.__exit__ = MagicMock(return_value=False)

        with patch.dict('sys.modules', {'PyPDF2': mock_pypdf2}), \
             patch('builtins.open', return_value=mock_file):
            result = extract_from_pdf(
                mock_service, mock_db,
                Path('/tmp/test.pdf'), 1, 1, 1
            )

        assert result == []

    def test_pdf_multiple_pages(self, mock_service, mock_db):
        """多页 PDF 文本累积处理"""
        from app.utils.spec_extractor.formats import extract_from_pdf

        page1 = MagicMock()
        page1.extract_text.return_value = '普通文本第一页'
        page2 = MagicMock()
        page2.extract_text.return_value = '材料 传感器 5V/100mA'

        mock_reader = MagicMock()
        mock_reader.pages = [page1, page2]

        mock_pypdf2 = MagicMock()
        mock_pypdf2.PdfReader.return_value = mock_reader

        mock_file = MagicMock()
        mock_file.__enter__ = MagicMock(return_value=mock_file)
        mock_file.__exit__ = MagicMock(return_value=False)

        mock_req = MagicMock()

        with patch.dict('sys.modules', {'PyPDF2': mock_pypdf2}), \
             patch('builtins.open', return_value=mock_file), \
             patch('app.utils.spec_extractor.formats.create_requirement', return_value=mock_req):
            result = extract_from_pdf(
                mock_service, mock_db,
                Path('/tmp/test.pdf'), 1, 1, 1
            )

        # page2 有 "材料" 关键词，应当产生至少1条记录
        assert len(result) >= 1

# -*- coding: utf-8 -*-
"""
工时报表生成服务测试
"""

import io
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from app.services.timesheet_report_service import TimesheetReportService


@pytest.fixture
def mock_aggregation_service():
    """Mock TimesheetAggregationService"""
    service = MagicMock()
    return service


@pytest.fixture
def mock_overtime_service():
    """Mock OvertimeCalculationService"""
    service = MagicMock()
    return service


@pytest.fixture
def mock_hourly_rate_service():
    """Mock HourlyRateService"""
    return MagicMock()


@pytest.fixture
def timesheet_report_service(
    db_session, mock_aggregation_service, mock_overtime_service
):
    """创建 TimesheetReportService 实例"""
    with patch(
        "app.services.timesheet_report_service.TimesheetAggregationService",
        return_value=mock_aggregation_service,
    ):
        with patch(
        "app.services.timesheet_report_service.OvertimeCalculationService",
        return_value=mock_overtime_service,
        ):
        service = TimesheetReportService(db_session)
        service.aggregation_service = mock_aggregation_service
        service.overtime_service = mock_overtime_service
        return service


class TestTimesheetReportServiceInit:
    """测试初始化"""

    def test_init_success(self, db_session):
        """测试成功初始化"""
        with patch("app.services.timesheet_report_service.EXCEL_AVAILABLE", True):
            service = TimesheetReportService(db_session)
            assert service.db == db_session
            assert service.aggregation_service is not None
            assert service.overtime_service is not None

    def test_init_excel_not_available(self, db_session):
        """测试 Excel 库未安装时抛出异常"""
        with patch("app.services.timesheet_report_service.EXCEL_AVAILABLE", False):
            with pytest.raises(ImportError, match="Excel处理库未安装"):
                TimesheetReportService(db_session)


class TestGenerateHrReportExcel:
    """测试生成HR报表"""

    def test_generate_hr_report_success(
        self, timesheet_report_service, mock_aggregation_service, db_session
    ):
        """测试成功生成HR报表"""
        # Mock 数据
        hr_data = [
        {
        "user_id": 1,
        "user_name": "张三",
        "normal_hours": 160.0,
        "overtime_hours": 10.0,
        "weekend_hours": 5.0,
        "holiday_hours": 2.0,
        }
        ]
        mock_aggregation_service.generate_hr_report.return_value = hr_data

        # Mock HourlyRateService
        with patch(
        "app.services.timesheet_report_service.HourlyRateService"
        ) as mock_rate_service_class:
        mock_rate_service = MagicMock()
        mock_rate_service.get_user_hourly_rate.return_value = Decimal("50.0")
        mock_rate_service_class.get_user_hourly_rate = (
        mock_rate_service.get_user_hourly_rate
        )

            # Mock User 和 Department
        mock_user = MagicMock()
        mock_user.department_id = 1
        mock_department = MagicMock()
        mock_department.name = "工程部"

        with patch.object(
        timesheet_report_service.db.query(), "first"
        ) as mock_query:
        mock_query.return_value = mock_user

        with patch.object(
        timesheet_report_service.db.query(), "filter"
        ) as mock_filter:
        mock_filter.return_value.first.return_value = mock_department

            # 生成报表
        result = timesheet_report_service.generate_hr_report_excel(
        2024, 1, department_id=None
        )

            # 验证
        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_hr_report.assert_called_once_with(
        2024, 1, None
        )

            # 验证 Excel 文件可以加载
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "2024年1月加班工资表"

            # 验证表头
        headers = [cell.value for cell in ws[4]]
        expected_headers = [
        "员工姓名",
        "部门",
        "正常工时",
        "工作日加班",
        "周末加班",
        "节假日加班",
        "工作日加班工资",
        "周末加班工资",
        "节假日加班工资",
        "合计加班工资",
        ]
        assert headers == expected_headers

            # 验证数据行
        data_row = [cell.value for cell in ws[5]]
        assert data_row[0] == "张三"
        assert data_row[1] == "工程部"
        assert data_row[2] == 160.0
        assert data_row[3] == 10.0
        assert data_row[4] == 5.0
        assert data_row[5] == 2.0

            # 验证加班工资计算
            # 工作日加班: 10 * 50 * 0.5 = 250 (1.5倍，额外0.5倍)
            # 周末加班: 5 * 50 * 1.0 = 250 (2倍，额外1倍)
            # 节假日加班: 2 * 50 * 2.0 = 200 (3倍，额外2倍)
        assert data_row[6] == 250.0
        assert data_row[7] == 250.0
        assert data_row[8] == 200.0
        assert data_row[9] == 700.0

    def test_generate_hr_report_empty_data(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试生成空数据的HR报表"""
        mock_aggregation_service.generate_hr_report.return_value = []

        with patch("app.services.timesheet_report_service.HourlyRateService"):
            result = timesheet_report_service.generate_hr_report_excel(2024, 1)

            assert isinstance(result, io.BytesIO)
            result.seek(0)
            wb = load_workbook(result)
            ws = wb.active

            # 只有表头，没有数据行
        assert ws.max_row == 4  # 标题行(1-2) + 空行(3) + 表头行(4)

    def test_generate_hr_report_with_department_filter(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试带部门过滤的HR报表"""
        mock_aggregation_service.generate_hr_report.return_value = []

        with patch("app.services.timesheet_report_service.HourlyRateService"):
            timesheet_report_service.generate_hr_report_excel(2024, 1, department_id=5)

            mock_aggregation_service.generate_hr_report.assert_called_once_with(
            2024, 1, 5
            )


class TestGenerateFinanceReportExcel:
    """测试生成财务报表"""

    def test_generate_finance_report_success(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试成功生成财务报表"""
        # Mock 数据
        finance_data = [
        {
        "project_code": "PJ001",
        "project_name": "测试项目",
        "personnel_records": [
        {
        "user_name": "张三",
        "date": "2024-01-15",
        "hours": 8.0,
        "hourly_rate": 50.0,
        "cost": 400.0,
        "work_content": "系统测试",
        }
        ],
        }
        ]
        mock_aggregation_service.generate_finance_report.return_value = finance_data

        # 生成报表
        result = timesheet_report_service.generate_finance_report_excel(
        2024, 1, project_id=None
        )

        # 验证
        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_finance_report.assert_called_once_with(
        2024, 1, None
        )

        # 验证 Excel 文件
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "2024年1月项目成本表"

        # 验证表头
        headers = [cell.value for cell in ws[4]]
        expected_headers = [
        "项目编号",
        "项目名称",
        "人员姓名",
        "工作日期",
        "工时",
        "时薪",
        "成本金额",
        "工作内容",
        ]
        assert headers == expected_headers

        # 验证数据行
        data_row = [cell.value for cell in ws[5]]
        assert data_row[0] == "PJ001"
        assert data_row[1] == "测试项目"
        assert data_row[2] == "张三"
        assert data_row[3] == "2024-01-15"
        assert data_row[4] == 8.0
        assert data_row[5] == 50.0
        assert data_row[6] == 400.0
        assert data_row[7] == "系统测试"

    def test_generate_finance_report_with_project_filter(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试带项目过滤的财务报表"""
        mock_aggregation_service.generate_finance_report.return_value = []

        result = timesheet_report_service.generate_finance_report_excel(
        2024, 1, project_id=10
        )

        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_finance_report.assert_called_once_with(
        2024, 1, 10
        )


class TestGenerateRdReportExcel:
    """测试生成研发报表"""

    def test_generate_rd_report_success(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试成功生成研发报表"""
        # Mock 数据
        rd_data = [
        {
        "project_code": "RD001",
        "project_name": "研发项目A",
        "personnel_records": [
        {
        "user_name": "李四",
        "date": "2024-01-20",
        "hours": 6.0,
        "task_name": "功能开发",
        "cost": 300.0,
        }
        ],
        }
        ]
        mock_aggregation_service.generate_rd_report.return_value = rd_data

        # 生成报表
        result = timesheet_report_service.generate_rd_report_excel(
        2024, 1, rd_project_id=None
        )

        # 验证
        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_rd_report.assert_called_once_with(
        2024, 1, None
        )

        # 验证 Excel 文件
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "2024年1月研发费用表"


class TestEdgeCases:
    """测试边缘情况"""

    def test_zero_overtime_hours(
        self, timesheet_report_service, mock_aggregation_service, db_session
    ):
        """测试加班工时为零的情况"""
        hr_data = [
        {
        "user_id": 1,
        "user_name": "王五",
        "normal_hours": 160.0,
        "overtime_hours": 0.0,
        "weekend_hours": 0.0,
        "holiday_hours": 0.0,
        }
        ]
        mock_aggregation_service.generate_hr_report.return_value = hr_data

        with patch(
        "app.services.timesheet_report_service.HourlyRateService"
        ) as mock_rate_service_class:
        mock_rate_service = MagicMock()
        mock_rate_service.get_user_hourly_rate.return_value = Decimal("50.0")
        mock_rate_service_class.get_user_hourly_rate = (
        mock_rate_service.get_user_hourly_rate
        )

        mock_user = MagicMock()
        mock_user.department_id = None

        with patch.object(
        timesheet_report_service.db.query(), "first"
        ) as mock_query:
        mock_query.return_value = mock_user

        result = timesheet_report_service.generate_hr_report_excel(2024, 1)

        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active

            # 验证加班工资都为零
        data_row = [cell.value for cell in ws[5]]
        assert data_row[6] == 0.0  # 工作日加班工资
        assert data_row[7] == 0.0  # 周末加班工资
        assert data_row[8] == 0.0  # 节假日加班工资
        assert data_row[9] == 0.0  # 合计加班工资

    def test_large_overtime_hours(
        self, timesheet_report_service, mock_aggregation_service, db_session
    ):
        """测试大额加班工时的情况"""
        hr_data = [
        {
        "user_id": 1,
        "user_name": "赵六",
        "normal_hours": 160.0,
        "overtime_hours": 100.0,
        "weekend_hours": 50.0,
        "holiday_hours": 10.0,
        }
        ]
        mock_aggregation_service.generate_hr_report.return_value = hr_data

        with patch(
        "app.services.timesheet_report_service.HourlyRateService"
        ) as mock_rate_service_class:
        mock_rate_service = MagicMock()
        mock_rate_service.get_user_hourly_rate.return_value = Decimal("100.0")
        mock_rate_service_class.get_user_hourly_rate = (
        mock_rate_service.get_user_hourly_rate
        )

        mock_user = MagicMock()
        mock_user.department_id = None

        with patch.object(
        timesheet_report_service.db.query(), "first"
        ) as mock_query:
        mock_query.return_value = mock_user

        result = timesheet_report_service.generate_hr_report_excel(2024, 1)

        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active

        data_row = [cell.value for cell in ws[5]]
            # 工作日: 100 * 100 * 0.5 = 5000
            # 周末: 50 * 100 * 1.0 = 5000
            # 节假日: 10 * 100 * 2.0 = 2000
            # 合计: 12000
        assert data_row[6] == 5000.0
        assert data_row[7] == 5000.0
        assert data_row[8] == 2000.0
        assert data_row[9] == 12000.0


class TestGenerateProjectReportExcel:
    """测试生成项目报表"""

    def test_generate_project_report_success(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试成功生成项目报表"""
        from datetime import date
        
        # Mock 数据
        project_data = {
        "project_name": "测试项目A",
        "personnel_stats": [
        {
        "user_name": "张三",
        "total_hours": 160.0,
        "contribution_rate": 50.0
        }
        ],
        "daily_stats": [
        {
        "date": "2024-01-15",
        "hours": 8.0,
        "personnel_count": 1,
        "personnel": [{"user_name": "张三"}]
        }
        ]
        }
        mock_aggregation_service.generate_project_report.return_value = project_data

        # 生成报表
        result = timesheet_report_service.generate_project_report_excel(
        project_id=1,
        start_date=date(2024, 1, 1),
        end_date=date(2024, 1, 31)
        )

        # 验证
        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_project_report.assert_called_once_with(
        1, date(2024, 1, 1), date(2024, 1, 31)
        )

        # 验证 Excel 文件
        result.seek(0)
        wb = load_workbook(result)
        
        # 验证有两个工作表
        assert len(wb.sheetnames) == 2
        assert "人员贡献度" in wb.sheetnames
        assert "每日工时分布" in wb.sheetnames

        # 验证人员贡献度工作表
        ws1 = wb["人员贡献度"]
        headers1 = [cell.value for cell in ws1[3]]
        expected_headers1 = ["人员姓名", "总工时", "贡献度(%)"]
        assert headers1 == expected_headers1

        # 验证每日工时分布工作表
        ws2 = wb["每日工时分布"]
        headers2 = [cell.value for cell in ws2[3]]
        expected_headers2 = ["日期", "总工时", "参与人数", "人员列表"]
        assert headers2 == expected_headers2

    def test_generate_project_report_error(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试项目报表生成错误场景"""
        from datetime import date
        
        # Mock 错误数据
        project_data = {"error": "项目不存在"}
        mock_aggregation_service.generate_project_report.return_value = project_data

        # 生成报表
        result = timesheet_report_service.generate_project_report_excel(
        project_id=999,
        start_date=date(2024, 1, 1),
        end_date=date(2024, 1, 31)
        )

        # 验证返回错误提示的Excel
        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws['A1'].value == "项目不存在"


class TestGenerateProjectReportExcel:
    """测试生成项目报表"""

    def test_generate_project_report_success(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试成功生成项目报表"""
        # Mock 数据
        project_data = [
        {
        "project_code": "PJ001",
        "project_name": "测试项目A",
        "personnel_records": [
        {
        "user_name": "张三",
        "date": "2024-01-15",
        "hours": 8.0,
        "task_name": "功能开发",
        "work_content": "实现用户登录功能",
        }
        ],
        }
        ]
        mock_aggregation_service.generate_project_report.return_value = project_data

        # 生成报表
        result = timesheet_report_service.generate_project_report_excel(
        2024, 1, project_id=1
        )

        # 验证
        assert isinstance(result, io.BytesIO)
        mock_aggregation_service.generate_project_report.assert_called_once_with(
        2024, 1, 1
        )

        # 验证 Excel 文件
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "2024年1月项目工时表"

        # 验证表头
        headers = [cell.value for cell in ws[4]]
        expected_headers = [
        "项目编号",
        "项目名称",
        "人员姓名",
        "工作日期",
        "工时",
        "任务名称",
        "工作内容",
        ]
        assert headers == expected_headers

        # 验证数据行
        data_row = [cell.value for cell in ws[5]]
        assert data_row[0] == "PJ001"
        assert data_row[1] == "测试项目A"
        assert data_row[2] == "张三"
        assert data_row[3] == "2024-01-15"
        assert data_row[4] == 8.0
        assert data_row[5] == "功能开发"
        assert data_row[6] == "实现用户登录功能"

    def test_generate_project_report_empty_data(
        self, timesheet_report_service, mock_aggregation_service
    ):
        """测试生成空数据的项目报表"""
        mock_aggregation_service.generate_project_report.return_value = []

        result = timesheet_report_service.generate_project_report_excel(2024, 1, project_id=1)

        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active

        # 只有表头，没有数据行
        assert ws.max_row == 4  # 标题行(1-2) + 空行(3) + 表头行(4)
